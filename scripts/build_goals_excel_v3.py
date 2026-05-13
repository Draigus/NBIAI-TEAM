"""Build Goals Studio Pricing Engagement Excel v3 — all 6 sheets from scratch."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = 'Clients/Goals/goals_pricing_engagement_plan_v3.xlsx'

# ── Shared styles ──────────────────────────────────────────────────────────────
HDR_FONT   = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
HDR_FILL   = PatternFill(start_color='1F3864', end_color='1F3864', fill_type='solid')
TITLE_FONT = Font(name='Calibri', size=14, bold=True, color='1F3864')
BODY_FONT  = Font(name='Calibri', size=10)
BOLD_FONT  = Font(name='Calibri', size=10, bold=True)
WRAP       = Alignment(wrap_text=True, vertical='top')
CENTER     = Alignment(horizontal='center', vertical='center', wrap_text=True)
TOP_LEFT   = Alignment(wrap_text=True, vertical='top')

ALT_FILL   = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')
WHITE_FILL = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
DIVIDER_FILL = PatternFill(start_color='1F3864', end_color='1F3864', fill_type='solid')

PRIORITY_FILLS = {
    'Critical': PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid'),
    'High':     PatternFill(start_color='FF9900', end_color='FF9900', fill_type='solid'),
    'Medium':   PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid'),
    'Low':      PatternFill(start_color='92D050', end_color='92D050', fill_type='solid'),
}
PRIORITY_FONTS = {
    'Critical': Font(name='Calibri', size=10, bold=True, color='FFFFFF'),
    'High':     Font(name='Calibri', size=10, bold=True, color='FFFFFF'),
    'Medium':   Font(name='Calibri', size=10, bold=True, color='000000'),
    'Low':      Font(name='Calibri', size=10, bold=True, color='000000'),
}

THIN = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)

CLIENT_ID = '6975460f-c302-42c5-a586-1d04c5fcb929'


def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width


def hdr_cell(ws, coord, value, width=None):
    c = ws[coord]
    c.value = value
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = CENTER
    c.border = THIN
    return c


def data_cell(ws, row, col, value, fill=None, font=None, align=None, border=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font or BODY_FONT
    c.alignment = align or WRAP
    if fill:
        c.fill = fill
    if border:
        c.border = THIN
    return c


def title_row(ws, row_num, text, cols_span, height=32):
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=cols_span)
    c = ws.cell(row=row_num, column=1, value=text)
    c.font = TITLE_FONT
    c.alignment = Alignment(vertical='center')
    ws.row_dimensions[row_num].height = height


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — OVERVIEW
# ══════════════════════════════════════════════════════════════════════════════
def build_overview(wb):
    ws = wb.create_sheet('Overview')

    # Column widths
    widths = {'A': 30, 'B': 55, 'C': 22, 'D': 22, 'E': 22}
    for col, w in widths.items():
        set_col_width(ws, col, w)

    title_row(ws, 1, 'Goals Studio Pricing Engagement — Project Overview', 5)

    # Project info block
    info = [
        ('Project',   'Goals Studio Pricing Engagement'),
        ('Client',    'Goals AB (Goals Studio), Stockholm, Sweden'),
        ('Contacts',  'Jonas Rundberg (Chief Executive Officer), Julius (Live Ops Lead)'),
        ('Budget',    '100,000 SEK (~$10,000 USD)'),
        ('Total Estimated Hours', '~116 hours'),
        ('Hard Deadline', '27 April 2026 (platform submission to PlayStation/Xbox)'),
        ('Game Launch', '14 May 2026'),
        ('Execution Model', 'Option A: Artificial Intelligence-Assisted (5 calendar days)'),
        ('Key Question from Client', '"Are we pricing too low?"'),
        ('Core Pricing Principle', 'Launch high, adjust downward if needed. Console platforms make price increases nearly impossible after launch — Sony and Xbox re-certification is slow and generates negative press coverage.'),
    ]

    r = 3
    # Sub-heading
    ws.merge_cells(f'A{r}:E{r}')
    c = ws.cell(row=r, column=1, value='Project Details')
    c.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    c.fill = HDR_FILL
    c.alignment = CENTER
    ws.row_dimensions[r].height = 22
    r += 1

    for label, value in info:
        ws.cell(row=r, column=1, value=label).font = BOLD_FONT
        ws.cell(row=r, column=1).fill = ALT_FILL if r % 2 == 0 else WHITE_FILL
        ws.cell(row=r, column=1).border = THIN
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        c = ws.cell(row=r, column=2, value=value)
        c.font = BODY_FONT
        c.fill = ALT_FILL if r % 2 == 0 else WHITE_FILL
        c.alignment = WRAP
        c.border = THIN
        ws.row_dimensions[r].height = 30
        r += 1

    r += 1  # spacer

    # Feature summary table
    ws.merge_cells(f'A{r}:E{r}')
    c = ws.cell(row=r, column=1, value='Feature Summary — Estimated Hours and Budget Allocation')
    c.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    c.fill = HDR_FILL
    c.alignment = CENTER
    ws.row_dimensions[r].height = 22
    r += 1

    headers = ['Feature', 'Description', 'Estimated Hours', '% of Budget', 'Days']
    for ci, h in enumerate(headers, 1):
        hdr_cell(ws, ws.cell(row=r, column=ci).coordinate, h)
    ws.row_dimensions[r].height = 25
    r += 1

    features = [
        ('F1: Competitive Hard Currency (HC) Pricing Benchmark',
         'Show Goals where they sit relative to the sports/football Free-to-Play (F2P) market. '
         'Independent cross-genre validation and pricing psychology assessment.',
         38, '33%', 'Days 1-2'),
        ('F2: Regional Pricing Strategy',
         'Validate Goals\' 40+ country pricing matrix. Adds competitive regional context, '
         'Purchasing Power Parity (PPP)-adjusted analysis, and platform compliance verification.',
         34, '29%', 'Days 2-4'),
        ('F3: Risk Assessment, Scenarios and Recommendations',
         'Synthesise all analysis into clear, specific, actionable guidance with modelled revenue impact.',
         16, '14%', 'Day 4'),
        ('F4: Value-Add (Conditional — only if F1-3 complete on time)',
         'Supplementary tools: reusable pricing methodology, World Cup tactical observations, '
         'Phase 2 positioning. Conditional on F1-3 finishing within budget.',
         8, '7%', 'Day 4'),
        ('F5: Deliverable Assembly and Client Review',
         'Package all analysis into the contracted deliverable: written summary (12-18 pages) '
         'plus a 30-45 minute remote review call with Jonas and Julius.',
         18, '16%', 'Days 5-6'),
        ('TOTAL', '', 116, '100%', '6 days (22-27 Apr)'),
    ]

    for i, (feat, desc, hrs, pct, days) in enumerate(features):
        fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
        is_total = feat == 'TOTAL'
        if is_total:
            fill = PatternFill(start_color='1F3864', end_color='1F3864', fill_type='solid')
            fnt = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        else:
            fnt = BOLD_FONT if i == 0 else BODY_FONT

        for ci, val in enumerate([feat, desc, hrs, pct, days], 1):
            c = ws.cell(row=r, column=ci, value=val)
            c.font = fnt if is_total else (BOLD_FONT if ci == 1 else BODY_FONT)
            c.fill = fill
            c.alignment = WRAP
            c.border = THIN
        ws.row_dimensions[r].height = 45 if not is_total else 25
        r += 1

    ws.freeze_panes = 'A3'
    return ws


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — TASK TRACKER
# ══════════════════════════════════════════════════════════════════════════════
TASKS = [
    # (task_id, feature_label, story, title, est_hours, priority, depends, notes)
    ('T1.0.1', 'F1', 'S1.0: Foundation — Map the Hard Currency Economy',
     'Map Hard Currency (HC) Value Chain', 3, 'High',
     'None — Day 1 start task',
     'Document the complete path from real money to in-game value. '
     '1 HC coin = 50 points = $0.012 USD at base rate. Map what each of the 6 HC tiers actually buys. '
     'Flag any unconfirmed item prices (e.g. Mech/Pulse kits) for Julius to confirm. '
     'Identify the cheapest satisfying purchase a new player can make.'),

    ('T1.0.2', 'F1', 'S1.0: Foundation — Map the Hard Currency Economy',
     'Establish Payer Persona Benchmarks', 5, 'High',
     'None — Day 1 start task',
     'Goals\' targets: 10.4% overall payer rate, Average Revenue Per Paying User (ARPPU) $32.48, '
     'Average Revenue Per User (ARPU) $3.38. '
     'Benchmark against published data from eFootball (Konami quarterly reports), '
     'EA FC (EA quarterly earnings), Sensor Tower/data.ai. '
     'Assess whether 10.4% is realistic for month 1 or a steady-state figure. '
     'Note: Goals has no seasonal resets or battle pass, so the Hard Currency ladder must do all conversion work.'),

    ('T1.1.1', 'F1', 'S1.1: Competitive Pricing Position',
     'Build Normalised Price Position Map', 8, 'High',
     'T1.0.1 (value chain must be established first)',
     'Using Goals\' existing competitive data, create a position map showing Goals\' effective USD per HC '
     'against the competitive range at each price tier. '
     'Calculate market minimum, median, maximum, and Goals\' percentile at each tier. '
     'Spot-check 2-3 competitor prices to confirm data is still current. '
     'Output: one clear sentence summarising overall position.'),

    ('T1.1.2', 'F1', 'S1.1: Competitive Pricing Position',
     'Discount Curve Comparison', 4, 'High',
     'T1.1.1',
     'Goals\' bonus curve: Tier 1 (0%), Tier 2 (~2.5%), Tier 3 (~7.7%), Tier 4 (~9.4%), '
     'Tier 5 (~15.4%), Tier 6 (~18.1%). Competitor max discounts: EA FC ~39%, Fortnite ~24%, Apex ~13%. '
     'Assess whether 18% max discount is enough to incentivise high-spend commitment at launch. '
     'New launches should be conservative — can always increase discounts later, never reduce without backlash.'),

    ('T1.1.3', 'F1', 'S1.1: Competitive Pricing Position',
     'Price Tier (SKU) Structure and Entry Point Audit', 5, 'High',
     'T1.1.1, T1.1.2',
     'Assess structural health of 6-tier ladder: $5.99/$9.99/$19.99/$29.99/$49.99/$99.99. '
     'Compare with EA FC (8 tiers, starts $0.99), Fortnite (4 tiers, starts $7.99), Apex (6 tiers, starts $4.99). '
     'Entry point: Does $5.99 create first-purchase friction? What can 380 HC buy? '
     'Note: a $0.99 tier nets only $0.69 after 30% platform fee — may not be worth the transaction cost. '
     'Whale ceiling: $99.99 cap with no purchase limits is correct for anti-Pay-to-Win (P2W) positioning.'),

    ('T1.1.4', 'F1', 'S1.1: Competitive Pricing Position',
     'Platform Fee Impact and Net Revenue Table', 2, 'Medium',
     'T1.1.3',
     'Build net revenue table: 6 tiers x 4 platforms (Steam 30%, Epic 12%, PlayStation 30%, Xbox 30%). '
     'Epic Games advantage: Goals nets ~26% more per transaction on Epic vs Steam/PlayStation/Xbox. '
     'Verify alignment with PlayStation wholesale rates in Goals\' pricing matrix. '
     'Note for regional work: these are USD nets; regional pricing may differ.'),

    ('T1.2.1', 'F1', 'S1.2: Conversion Rate and Elasticity Context',
     'Price Elasticity Scenario Modelling', 8, 'High',
     'T1.0.2, T1.1.1',
     'Model 4 scenarios using Goals\' own revenue model (805,250 Monthly Active Users (MAU), '
     '10.4% payer rate, $32.48 ARPPU = $2.72M total). '
     'Scenario A: Raise all prices 10% (assume conversion drops 5-15%). '
     'Scenario B: Raise all prices 20% (conversion drops 10-25%). '
     'Scenario C: Raise entry tier only from $5.99 to $6.99. '
     'Scenario D: Keep prices, steepen top-tier discount to 25%. '
     'Present as conservative-to-optimistic ranges, not point estimates.'),

    ('T1.3.1', 'F1', 'S1.3: Promotional and First-Purchase Pricing',
     'First-Purchase and Starter Pack Recommendations', 5, 'Medium',
     'T1.0.1, T1.1.3',
     'Research what EA FC, Fortnite, Apex, eFootball currently offer as first-purchase incentives. '
     'Do NOT cite conversion improvement percentages without a verifiable source. '
     'Recommend: (1) First Purchase Bonus — double Hard Currency on first-ever purchase at any tier. '
     '(2) Limited Starter Pack — one-time $3.99 offer with 500 HC + Rare+ player + exclusive kit. '
     '(3) World Cup Launch Bundle tied to June event. '
     'Frame as complementary to core pricing, not a change to it.'),

    ('T2.1.1', 'F2', 'S2.1: Regional Deviation and Risk Identification',
     'Categorise All Countries by Risk Level', 8, 'High',
     'T1.1.1 (need competitive baseline before comparing regionally)',
     'Using Goals\' existing deviation data, categorise every country: '
     'Green (deviation <15% from Foreign Exchange (FX) rate AND Goals matrix says proceed), '
     'Amber (deviation 15-25% OR flagged in Goals\' matrix), '
     'Red (deviation >25% OR large player base with >15% over-pricing). '
     'Red candidates: Poland (+25%), Switzerland (+26%), Ukraine (-26%), Turkey, Brazil. '
     'Cross-reference with beta community size: UK 1,274 players, France 1,026, Turkey 667, Brazil 495, Poland 392.'),

    ('T2.1.2', 'F2', 'S2.1: Regional Deviation and Risk Identification',
     'Competitor Regional Pricing for Red and Amber Markets', 15, 'High',
     'T2.1.1 (need country risk list first)',
     'For 5-10 red/amber countries, research EA FC and Fortnite actual prices. '
     'Sources: PlayStation Store regional web store, SteamDB (historical and current), Epic Games Store. '
     'For each market: Goals proposed price vs EA FC actual vs Fortnite actual vs market median. '
     'Priority markets: Brazil (large player base, low Purchasing Power Parity), '
     'Turkey (5th largest beta community, extreme PPP differential), '
     'Poland (large community, +25% above FX rate). '
     'This is the single largest research task in the engagement.'),

    ('T2.2.1', 'F2', 'S2.2: Platform Compliance and Submission Guidance',
     'Verify Platform Tier Compliance', 3, 'Critical',
     'T2.1.1',
     'Cross-reference Goals\' proposed prices against valid PlayStation/Xbox tier selections. '
     'Any price that does not align to a standard platform tier will be rejected by Sony or Microsoft. '
     'Submission checklist: all prices in valid tiers, all territories covered, USD fallback for unsupported currencies. '
     'Review period is ~9 days (per Julius). '
     'Timeline: deliver 27 Apr > Goals submits ~5 May > 9-day review > 14 May launch. '
     'CRITICAL PATH — any rejection delays launch.'),

    ('T2.3.1', 'F2', 'S2.3: Country-Specific Recommendations',
     'Write Country Recommendation Cards', 8, 'High',
     'T2.1.1, T2.1.2, T2.2.1',
     'For each red/amber country write a self-contained recommendation card: '
     'current proposed price | issue | recommended price | rationale | revenue impact | risk if unchanged. '
     'Prioritise by: player base size multiplied by deviation severity. '
     'Group by action type: reduce price, can increase (revenue being left on the table), or platform tier adjustment. '
     'Expected output: 8-12 cards. Julius must be able to implement each card without reading the full report.'),

    ('T3.1.1', 'F3', 'S3.1: "Are We Too Low?" — The Central Answer',
     'Build the Definitive Position Statement', 4, 'Critical',
     'T1.1.1, T1.0.2, T1.2.1 (all Feature 1 tasks must be complete)',
     'Synthesise competitive position, payer benchmarks, and scenario modelling into one clear paragraph. '
     'State Goals\' percentile at each tier, whether conversion targets are realistic, '
     'and give a specific recommendation (a number, not a range). '
     'Take a position. No hedging. '
     'Core principle: you can always reduce prices post-launch; you cannot raise them on console.'),

    ('T3.2.1', 'F3', 'S3.2: Risk Register',
     'Compile and Rank Risks', 5, 'High',
     'T3.1.1, T2.3.1',
     'Document minimum 10 risks with: description, severity, likelihood, markets affected, mitigation. '
     'Rank by severity x likelihood. Top 3 risks must have specific, implementable mitigation steps. '
     'Includes: Pay-to-Win perception, price increase impossibility, over-pricing in price-sensitive markets, '
     'platform rejection, shallow discount curve, entry-tier friction, '
     'Netherlands gambling authority (Kansspelautoriteit — KSA) regulatory risk, '
     'South Korea drop rate disclosure (legal requirement), non-seasonal demand sustainability.'),

    ('T3.3.1', 'F3', 'S3.3: Top 5 Recommendations',
     'Write Top 5 Recommendations', 7, 'High',
     'T3.1.1, T3.2.1',
     'Five specific, numbered, actionable recommendations. Each must include: '
     'exact change to make, evidence from analysis, revenue impact as a conservative-to-optimistic range, '
     'implementation steps (which cells in Goals\' matrix, which platform to resubmit), '
     'and risk if skipped. '
     'Must pass the test: would a studio economy designer read this and act same day?'),

    ('T4.1.1', 'F4', 'S4.1: Reusable Regional Pricing Methodology',
     'Write "How to Price a New SKU (Price Tier)" Guide', 3, 'Medium',
     'F3 complete (conditional on F1-3 finishing on time)',
     'One-page 6-step guide Julius can follow for any future price tier without NBI help: '
     '(1) Set USD base price anchored against EA FC equivalent. '
     '(2) Convert to local currencies using current Foreign Exchange rate. '
     '(3) Check Purchasing Power Parity factor using lookup table. '
     '(4) Compare against EA FC in that market (+/-15% is acceptable). '
     '(5) Round to nearest platform tier. '
     '(6) Verify discount curve is maintained. '
     'Include worked example: pricing a new Celebration at $7.00 for Brazil. CONDITIONAL.'),

    ('T4.2.1', 'F4', 'S4.2: World Cup Pricing Opportunity Notes',
     'Write World Cup Quick-Hit Observations', 2, 'Low',
     'F3 complete (conditional)',
     '3-5 tactical pricing observations for the June World Cup window. '
     'Keep to half a page — this is a teaser for Phase 2 live service consulting, not a full plan. '
     'Ideas: national kit premium during group stage, discount in knockout rounds, '
     'time-limited "World Cup Coin Pack" at 10% bonus value. '
     'Do NOT cite revenue uplift multipliers without a verifiable published source. CONDITIONAL.'),

    ('T4.3.1', 'F4', 'S4.3: Phase 2 Positioning',
     'Write "What\'s Next" Section', 3, 'Medium',
     'F3 complete (conditional)',
     'Half-page section positioning Phase 2 naturally — helpful, not a sales pitch. '
     'Natural next steps: post-launch telemetry review 2-3 weeks after 14 May, '
     'A/B test design, store optimisation, World Cup monetisation strategy, ongoing live service consulting. '
     'Close with recommendation for a check-in call 2-3 weeks post-launch. '
     'Do NOT include NBI internal Phase 2 pricing in the client document. CONDITIONAL.'),

    ('T5.1.1', 'F5', 'S5.1: Written Summary Document',
     'Assemble the Written Summary', 10, 'High',
     'All F1, F2, F3 tasks complete; F4 tasks complete if conditional scope was triggered',
     'Final client-facing document (12-18 pages, British English, no em dashes, numbers-backed). '
     '10 sections: Executive Summary, Competitive Position, Regional Pricing Assessment, Pricing Scenarios, '
     'Risk Register, Recommendations, Pricing Framework (conditional), World Cup Opportunities (conditional), '
     'Next Steps, Appendix. Written for both Jonas (strategic) and Julius (practitioner).'),

    ('T5.2.1', 'F5', 'S5.2: Internal Review and Polish',
     'Quality Gate — Internal Review', 4, 'High',
     'T5.1.1',
     'Internal review covering: domain accuracy (do recommendations make sense for this game?), '
     'tone (expertise without condescension), completeness (does it answer Jonas\'s core questions?), '
     'commercial positioning (does it set up Phase 2 naturally?), NBI brand consistency, contractual issues. '
     'Address all feedback and produce final version. Send to Jonas/Julius 24 hours before review call.'),

    ('T5.3.1', 'F5', 'S5.3: Remote Review Session',
     'Conduct Client Review Call', 4, 'High',
     'T5.2.1 (final document must be sent to client first)',
     '30-45 minute call: 5 min scope recap, 10 min top 5 recommendations, '
     '10 min regional pricing highlights, 5 min pricing framework, 5 min next steps, 10 min Q&A. '
     'Follow-up within 24 hours: summary email, adjusted recommendations if needed, confirmed next steps. '
     'Prepared answers: "Should we delay?" — No, these are configuration changes. '
     '"Can we raise prices later?" — Down only. Launch high, adjust downward. '
     'Deadline day — 27 April 2026.'),
]

FEATURE_MAP = {
    'F1': 'F1: Competitive Hard Currency (HC) Pricing Benchmark',
    'F2': 'F2: Regional Pricing Strategy',
    'F3': 'F3: Risk Assessment, Scenarios and Recommendations',
    'F4': 'F4: Value-Add (Conditional)',
    'F5': 'F5: Deliverable Assembly and Client Review',
}

TASK_DAY_MAP = {
    'T1.0.1': 'Day 1 (22 Apr)',
    'T1.0.2': 'Day 1 (22 Apr)',
    'T1.1.1': 'Day 1 (22 Apr)',
    'T1.1.2': 'Day 1 (22 Apr)',
    'T1.1.3': 'Day 2 (23 Apr)',
    'T1.1.4': 'Day 2 (23 Apr)',
    'T1.2.1': 'Day 2 (23 Apr)',
    'T1.3.1': 'Day 2 (23 Apr)',
    'T2.1.1': 'Days 2-3 (23-24 Apr)',
    'T2.1.2': 'Day 3 (24 Apr)',
    'T2.2.1': 'Day 3 (24 Apr)',
    'T2.3.1': 'Days 3-4 (24-25 Apr)',
    'T3.1.1': 'Day 4 (25 Apr)',
    'T3.2.1': 'Day 4 (25 Apr)',
    'T3.3.1': 'Day 4 (25 Apr)',
    'T4.1.1': 'Day 4 (25 Apr) — conditional',
    'T4.2.1': 'Day 4 (25 Apr) — conditional',
    'T4.3.1': 'Day 4 (25 Apr) — conditional',
    'T5.1.1': 'Day 5 (26 Apr)',
    'T5.2.1': 'Day 5 (26 Apr)',
    'T5.3.1': 'Day 6 (27 Apr) — deadline',
}


def build_task_tracker(wb):
    ws = wb.create_sheet('Task Tracker')

    cols = [
        ('A', 'Task ID', 10),
        ('B', 'Feature', 22),
        ('C', 'Story', 38),
        ('D', 'Task Title', 48),
        ('E', 'Estimated Hours', 11),
        ('F', 'Status', 14),
        ('G', 'Priority', 12),
        ('H', 'Scheduled Day', 22),
        ('I', 'Dependencies', 40),
        ('J', 'Task Notes', 100),
    ]

    for col_letter, _, width in cols:
        set_col_width(ws, col_letter, width)

    title_row(ws, 1, 'Goals Studio Pricing Engagement — Task Tracker (21 Tasks)', len(cols))
    ws.row_dimensions[1].height = 28

    # Headers row 2
    for ci, (_, label, _) in enumerate(cols, 1):
        c = ws.cell(row=2, column=ci, value=label)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = CENTER
        c.border = THIN
    ws.row_dimensions[2].height = 25

    ws.freeze_panes = 'A3'

    r = 3
    current_feature = None

    for task in TASKS:
        task_id, feat_key, story, title, hours, priority, deps, notes = task

        # Feature divider row
        if feat_key != current_feature:
            current_feature = feat_key
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(cols))
            c = ws.cell(row=r, column=1, value=FEATURE_MAP[feat_key])
            c.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            c.fill = DIVIDER_FILL
            c.alignment = Alignment(horizontal='left', vertical='center')
            c.border = THIN
            ws.row_dimensions[r].height = 22
            r += 1

        fill = ALT_FILL if r % 2 == 0 else WHITE_FILL

        vals = [task_id, FEATURE_MAP[feat_key], story, title, hours, 'Not Started',
                priority, TASK_DAY_MAP.get(task_id, ''), deps, notes]

        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=r, column=ci, value=val)
            c.border = THIN
            c.alignment = WRAP
            if ci == 7 and priority in PRIORITY_FILLS:  # Priority column
                c.fill = PRIORITY_FILLS[priority]
                c.font = PRIORITY_FONTS[priority]
            else:
                c.fill = fill
                c.font = BODY_FONT
        ws.row_dimensions[r].height = 30
        r += 1

    # Total row
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    c = ws.cell(row=r, column=1, value='TOTAL ESTIMATED HOURS')
    c.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    c.fill = DIVIDER_FILL
    c.alignment = CENTER
    ws.cell(row=r, column=5, value=116).font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    ws.cell(row=r, column=5).fill = DIVIDER_FILL
    ws.cell(row=r, column=5).alignment = CENTER
    for ci in range(6, len(cols)+1):
        ws.cell(row=r, column=ci).fill = DIVIDER_FILL
    ws.row_dimensions[r].height = 22

    return ws


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — TIMELINE
# ══════════════════════════════════════════════════════════════════════════════
def build_timeline(wb):
    ws = wb.create_sheet('Timeline')

    cols = [
        ('A', 'Day', 14),
        ('B', 'Date', 14),
        ('C', 'Tasks in Scope', 38),
        ('D', 'Focus Area', 55),
        ('E', 'Estimated Hours', 12),
        ('F', 'Artificial Intelligence (AI) Support', 65),
        ('G', 'Quality Gate / Output', 55),
    ]
    for col_letter, _, width in cols:
        set_col_width(ws, col_letter, width)

    title_row(ws, 1, 'Goals Studio Pricing Engagement — Day-by-Day Timeline (Option A: AI-Assisted)', len(cols))

    for ci, (_, label, _) in enumerate(cols, 1):
        c = ws.cell(row=2, column=ci, value=label)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = CENTER; c.border = THIN
    ws.row_dimensions[2].height = 25
    ws.freeze_panes = 'A3'

    days = [
        ('Day 1', '22 Apr 2026',
         'T1.0.1, T1.0.2, T1.1.1, T1.1.2',
         'Foundation: map Hard Currency value chain, establish payer persona benchmarks, '
         'begin building normalised price position map, start discount curve comparison.',
         8,
         'AI gathers raw competitive data for the price position map and begins '
         'categorising the 40+ country matrix. Human team validates outputs.',
         'End of Day 1: Can we answer "what does each HC tier buy?" '
         'If not, escalate to Julius immediately.'),

        ('Day 2', '23 Apr 2026',
         'T1.1.3, T1.1.4, T1.2.1, T1.3.1, T2.1.1 (start)',
         'Complete Feature 1 competitive analysis: price tier structure audit, '
         'platform fee net revenue table, price elasticity scenario modelling, '
         'first-purchase recommendations. Begin country risk categorisation.',
         20,
         'AI continues country categorisation from Goals\' matrix data. '
         'AI supports net revenue table calculations and scenario modelling inputs.',
         'End of Day 2: Competitive position map draft complete. '
         'Do we know where Goals sits vs market? Draft top-level answer to Jonas\'s question.'),

        ('Day 3', '24 Apr 2026',
         'T2.1.1 (finish), T2.1.2, T2.2.1, T2.3.1 (start)',
         'Complete country risk categorisation. '
         'Research competitor regional prices for all red/amber markets (largest research task). '
         'Verify platform tier compliance for PlayStation and Xbox. '
         'Begin writing country recommendation cards.',
         20,
         'AI scrapes SteamDB and PlayStation Store for competitor regional pricing data. '
         'AI drafts initial country recommendation cards for human review.',
         'End of Day 3 (CRITICAL): Internal review of draft top 5 recommendations. '
         'Go/no-go on direction before synthesis begins.'),

        ('Day 4', '25 Apr 2026',
         'T2.3.1 (finish), T3.1.1, T3.2.1, T3.3.1, T4.1.1, T4.2.1, T4.3.1',
         'Finish country recommendation cards. Write definitive position statement answering '
         '"Are we too low?". Compile and rank risk register. Write top 5 recommendations. '
         'If F1-3 on track: complete conditional Feature 4 (pricing methodology guide, '
         'World Cup observations, Phase 2 positioning).',
         19,
         'AI drafts Feature 4 conditional content (pricing guide, World Cup observations, '
         'Phase 2 section) for human review and editing.',
         'End of Day 4: Full deliverable content complete pending assembly. '
         'Feature 4 status confirmed: complete or cut.'),

        ('Day 5', '26 Apr 2026',
         'T5.1.1, T5.2.1',
         'Assemble the full written summary document (12-18 pages). '
         'Internal quality gate review: domain accuracy, tone, completeness, '
         'commercial positioning, brand consistency. '
         'Produce final version and send to Jonas and Julius.',
         14,
         'Minimal AI involvement — this is human synthesis and editorial work. '
         'AI may assist with formatting and consistency checking.',
         'End of Day 5: Final document sent to client 24 hours before review call. '
         'Call scheduled with Jonas and Julius.'),

        ('Day 6', '27 Apr 2026',
         'T5.3.1',
         '30-45 minute remote review call with Jonas and Julius. '
         'Walk through findings and top 5 recommendations. '
         'Capture any adjustments needed. Send follow-up email within 24 hours.',
         4,
         'None required for the call itself.',
         'DEADLINE DAY. Call completed, follow-up sent. '
         'Phase 2 interest level documented.'),
    ]

    for i, row_data in enumerate(days):
        fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=3+i, column=ci, value=val)
            c.font = BOLD_FONT if ci == 1 else BODY_FONT
            c.fill = fill
            c.alignment = WRAP
            c.border = THIN
        ws.row_dimensions[3+i].height = 80

    return ws


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — RISK REGISTER
# ══════════════════════════════════════════════════════════════════════════════
RISKS = [
    ('R01',
     'Pay-to-Win (P2W) perception at high price points',
     'Critical', 'Medium', 'All markets',
     'Beta community data shows "pay to win" is the #1 cited reason players quit. '
     'If premium packs feel like they buy wins, pricing becomes irrelevant because players leave. '
     'The 80/20 skill-to-purchase design mitigates this structurally, but player perception may not match reality.',
     'Frame all Hard Currency purchases as time-saving or cosmetic, never competitive advantage. '
     'Store messaging must emphasise "customise" not "dominate". '
     'Monitor community sentiment in the first 2 weeks post-launch.'),

    ('R02',
     'Post-launch price increase is practically impossible',
     'High', 'Medium', 'All console markets',
     'If current prices prove too low after launch, raising them on PlayStation and Xbox requires '
     're-certification, generates player backlash, and often gets press coverage. '
     'This is a structural platform constraint, not just a commercial risk.',
     'Launch at the upper end of the acceptable range. '
     'Use "introductory pricing" framing so any future adjustments are framed as expected, not a penalty. '
     'Core principle: launch high, adjust downward if needed.'),

    ('R03',
     'Over-pricing in price-sensitive markets',
     'High', 'High', 'Turkey, Brazil, Poland (and potentially other emerging markets)',
     'Large player bases in markets with low purchasing power. '
     'Poland is already +25% above the foreign exchange rate in Goals\' own matrix. '
     'Turkey has the 5th largest beta community with extreme purchasing power differences vs the USD. '
     'Over-pricing these markets = low conversion = missed World Cup revenue opportunity.',
     'Implement country-specific price adjustments per the recommendation cards in Feature 2. '
     'Prioritise by: player base size multiplied by degree of over-pricing.'),

    ('R04',
     'Platform pricing rejection causes launch delay',
     'Critical', 'Low', 'PlayStation, Xbox',
     'Sony and Microsoft reject submitted pricing if any tier does not align to a valid platform price point. '
     'A rejection at submission adds at minimum 1-2 weeks to the timeline, '
     'potentially missing the 14 May launch window.',
     'Task T2.2.1 verifies all tiers against valid platform price points before submission. '
     'Submit pricing 2+ weeks before the launch deadline (target: 5 May or earlier). '
     'Ask Julius whether a test pricing configuration has already been submitted.'),

    ('R05',
     'Discount curve too shallow to drive high-spend commitment',
     'Medium', 'Medium', 'All markets',
     'Goals\' maximum discount at the top tier is ~18%. '
     'EA FC offers ~39%, Fortnite ~24%. '
     'A shallow curve may not provide enough value incentive for players willing to spend $50-$100 at launch, '
     'reducing revenue from the highest-spending segment.',
     'Consider steepening the top-tier discount to 22-25% for the first 90 days after launch, '
     'then normalising. This creates a time-limited reason to buy high early '
     'without permanently reducing margins.'),

    ('R06',
     'Entry-tier friction reduces first-time buyer conversion',
     'Medium', 'Medium', 'All markets',
     'The minimum purchase is $5.99 (380 Hard Currency). '
     'If 380 HC cannot buy a complete, satisfying item, new players may abandon their first purchase. '
     'EA FC starts at $0.99, providing a much lower commitment barrier.',
     'First-purchase incentive (double HC on first purchase) addresses this '
     'without adding a micro-tier that has poor margins after the 30% platform fee. '
     'Ensure 380 HC buys at least one complete item with some leftover.'),

    ('R07',
     'No pricing framework for Originals and Celebrity packs post-launch',
     'Low', 'N/A — post-launch consideration', 'Post-launch content drops',
     'Premium Hard Currency-only items (Originals, Celebrity packs) are planned for '
     'post-launch content drops but have no pricing methodology in the current documentation. '
     'Without a framework, pricing decisions will be ad hoc and may be inconsistent.',
     'Feature 4 (conditional) provides a reusable pricing methodology guide. '
     'Phase 2 engagement to build a full automated pricing model. '
     'Not launch-blocking but flag for Julius.'),

    ('R08',
     'Netherlands gambling regulation risk for player packs',
     'High', 'Medium', 'Netherlands',
     'Player packs that grant athletes with meaningful attribute differences '
     'may be classified as gambling-adjacent under Dutch Kansspelautoriteit (KSA — Dutch gambling authority) rulings. '
     'EA FC has been fined in the Netherlands for similar mechanics. '
     'The Netherlands uses Euro pricing, so this affects the European tier broadly.',
     'Legal review of pack mechanic before Netherlands launch. '
     'Consider offering cosmetic-only packs in the Netherlands, '
     'or ensure pack contents are predetermined and visible before purchase. '
     'Research current KSA enforcement stance on Free-to-Play sports games.'),

    ('R09',
     'South Korea drop rate disclosure is a legal requirement',
     'High', 'Certain (if launching in South Korea)', 'South Korea',
     'Korean law mandates disclosure of exact probabilities for all randomised item purchases. '
     'Non-compliance blocks the Korean launch — this is not a risk of fine but of outright rejection. '
     'Goals\' player tier distribution data provides the rates.',
     'Confirm drop rate disclosure is built into the pack purchase user interface '
     'before submitting to the Korean app stores. '
     'Verify with the development team that rates shown match the actual distribution in Goals\' player data.'),

    ('R10',
     'Non-seasonal economy creates flat rather than pulsing revenue',
     'Medium', 'High', 'All markets (long-term revenue planning)',
     'Without seasonal resets, a battle pass, or Fear of Missing Out (FOMO)-driven content wipes, '
     'there is no guaranteed per-player spend floor per quarter. '
     'Goals\' economy relies on player aging (gradual stat decline) as a spend driver, '
     'which is predictable but slow. '
     'Comparable non-seasonal game: eFootball (closer model than EA FC).',
     'Plan a regular cosmetic refresh cadence (monthly kit and celebration drops). '
     'Leverage real-world football events as demand spikes (World Cup, domestic seasons). '
     'Use game mode variety to drive squad depth demand. '
     'Not launch-blocking — Phase 2 work — but should be acknowledged in the deliverable.'),
]


def build_risk_register(wb):
    ws = wb.create_sheet('Risk Register')

    cols = [
        ('A', 'Risk ID', 8),
        ('B', 'Risk Description', 42),
        ('C', 'Severity', 12),
        ('D', 'Likelihood', 14),
        ('E', 'Markets Affected', 38),
        ('F', 'Detail', 90),
        ('G', 'Mitigation', 90),
    ]
    for col_letter, _, width in cols:
        set_col_width(ws, col_letter, width)

    title_row(ws, 1, 'Goals Studio Pricing Engagement — Risk Register (10 Risks)', len(cols))

    for ci, (_, label, _) in enumerate(cols, 1):
        c = ws.cell(row=2, column=ci, value=label)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = CENTER; c.border = THIN
    ws.row_dimensions[2].height = 25
    ws.freeze_panes = 'A3'

    for i, (rid, desc, severity, likelihood, markets, detail, mitigation) in enumerate(RISKS):
        fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
        r = 3 + i

        ws.cell(row=r, column=1, value=rid).fill = fill
        ws.cell(row=r, column=1).font = BOLD_FONT
        ws.cell(row=r, column=1).alignment = CENTER
        ws.cell(row=r, column=1).border = THIN

        ws.cell(row=r, column=2, value=desc).fill = fill
        ws.cell(row=r, column=2).font = BOLD_FONT
        ws.cell(row=r, column=2).alignment = WRAP
        ws.cell(row=r, column=2).border = THIN

        # Severity cell — colour coded
        sev_fill = PRIORITY_FILLS.get(severity, fill)
        sev_font = PRIORITY_FONTS.get(severity, BODY_FONT)
        ws.cell(row=r, column=3, value=severity).fill = sev_fill
        ws.cell(row=r, column=3).font = sev_font
        ws.cell(row=r, column=3).alignment = CENTER
        ws.cell(row=r, column=3).border = THIN

        # Likelihood
        ws.cell(row=r, column=4, value=likelihood).fill = fill
        ws.cell(row=r, column=4).font = BODY_FONT
        ws.cell(row=r, column=4).alignment = CENTER
        ws.cell(row=r, column=4).border = THIN

        ws.cell(row=r, column=5, value=markets).fill = fill
        ws.cell(row=r, column=5).font = BODY_FONT
        ws.cell(row=r, column=5).alignment = WRAP
        ws.cell(row=r, column=5).border = THIN

        ws.cell(row=r, column=6, value=detail).fill = fill
        ws.cell(row=r, column=6).font = BODY_FONT
        ws.cell(row=r, column=6).alignment = WRAP
        ws.cell(row=r, column=6).border = THIN

        ws.cell(row=r, column=7, value=mitigation).fill = fill
        ws.cell(row=r, column=7).font = BODY_FONT
        ws.cell(row=r, column=7).alignment = WRAP
        ws.cell(row=r, column=7).border = THIN

        ws.row_dimensions[r].height = 90

    return ws


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — DELIVERABLE STRUCTURE
# ══════════════════════════════════════════════════════════════════════════════
DELIVERABLE_SECTIONS = [
    ('1', 'Executive Summary',
     '~1 page',
     'F1, F2, F3',
     'Headline finding answering "Are we too low?", top 5 recommendations, '
     'immediate actions required before platform submission. '
     'Written for Jonas — strategic, numbers-backed, one-page maximum.',
     'Jonas can brief his board on pricing in 2 minutes after reading this section alone.'),

    ('2', 'Your Competitive Position',
     '2-3 pages',
     'F1 (all tasks)',
     'Where Goals sits relative to the sports/football Free-to-Play market. '
     'Normalised price position map by tier, discount curve comparison, '
     'price tier structure assessment, platform fee net revenue table. '
     'Payer persona benchmarks vs published industry data.',
     'Goals can explain their pricing position to any investor or partner with evidence, '
     'not just internal assumptions.'),

    ('3', 'Regional Pricing Assessment',
     '3-4 pages + appendix',
     'F2 (all tasks)',
     'Country risk map (green/amber/red) for all 40+ markets. '
     'Competitor regional pricing for all red and amber markets. '
     'Platform tier compliance verification. '
     '8-12 self-contained country recommendation cards.',
     'Julius can implement every regional price change needed directly from the recommendation cards, '
     'without reading the full report.'),

    ('4', 'Pricing Scenarios',
     '1-2 pages',
     'F1 — S1.2 (T1.2.1)',
     'Four modelled scenarios showing revenue impact of price adjustments: '
     '(A) raise all prices 10%, (B) raise all prices 20%, '
     '(C) raise entry tier only, (D) keep prices but steepen discount curve. '
     'Conservative-to-optimistic ranges. Clear recommendation on which scenario is best for launch.',
     'Goals understands the revenue impact of each pricing option before committing.'),

    ('5', 'Risk Register',
     '2 pages',
     'F3 — S3.2 (T3.2.1)',
     '10 identified pricing risks ranked by severity and likelihood. '
     'Includes: Pay-to-Win perception, price increase impossibility, over-pricing in emerging markets, '
     'platform rejection, Netherlands regulatory risk, South Korea drop rate disclosure. '
     'Top 3 risks have specific, immediately implementable mitigation steps.',
     'Goals has a complete audit trail of pricing risks reviewed before submission, '
     'protecting against post-launch "we should have known" situations.'),

    ('6', 'Recommendations',
     '2-3 pages',
     'F3 — S3.3 (T3.3.1)',
     'Top 5 specific, numbered, actionable recommendations. '
     'Each recommendation includes: exact change, evidence, revenue impact range, '
     'implementation steps (which cells in the pricing matrix, which platform to resubmit), '
     'and the risk of not acting.',
     'A studio economy designer reads these and acts same day.'),

    ('7', 'Pricing Framework (Conditional)',
     '1 page',
     'F4 — S4.1 (T4.1.1) — conditional on F1-3 completing on time',
     '6-step reusable methodology for pricing any future item (kit, celebration, player pack) '
     'across all launch markets. Includes worked example (pricing a new Celebration at $7.00 USD in Brazil) '
     'and a Purchasing Power Parity lookup table for all launch markets.',
     'Julius can price any new item without needing to consult NBI, '
     'reducing dependency and building internal capability.'),

    ('8', 'World Cup Opportunities (Conditional)',
     '0.5 pages',
     'F4 — S4.2 (T4.2.1) — conditional on F1-3 completing on time',
     '3-5 tactical pricing observations for the June World Cup window: '
     'national kit pricing strategy, time-limited World Cup coin pack, '
     'group stage vs knockout round pricing dynamics.',
     'Goals has a starting point for World Cup monetisation planning — '
     'teaser for the Phase 2 live service engagement.'),

    ('9', 'Next Steps',
     '0.5 pages',
     'F4 — S4.3 (T4.3.1)',
     'Post-launch telemetry review timeline, A/B test design opportunities, '
     'store optimisation options, World Cup full strategy (Phase 2), '
     'ongoing live service consulting. '
     'Closes with recommendation for a check-in call 2-3 weeks after 14 May launch.',
     'Jonas and Julius have a clear picture of what comes after launch and why it matters.'),

    ('10', 'Appendix: Supporting Data',
     'Variable',
     'All features',
     'Full normalised comparison tables for all 7 competitor titles. '
     'Complete country-by-country pricing matrix with risk ratings. '
     'Competitor regional pricing data with sources. '
     'Platform tier reference tables for PlayStation and Xbox. '
     'Payer persona benchmarks with cited sources.',
     'All data underpinning the report is available for Goals\' internal records '
     'and future reference without needing to re-research.'),
]


def build_deliverable_structure(wb):
    ws = wb.create_sheet('Deliverable Structure')

    cols = [
        ('A', 'Section', 8),
        ('B', 'Section Title', 38),
        ('C', 'Estimated Length', 16),
        ('D', 'Source Features / Tasks', 30),
        ('E', 'Content Description', 100),
        ('F', 'Done When (Success Measure)', 75),
    ]
    for col_letter, _, width in cols:
        set_col_width(ws, col_letter, width)

    title_row(ws, 1,
              'Goals Studio Pricing Engagement — Final Deliverable Structure (10 Sections)', len(cols))

    for ci, (_, label, _) in enumerate(cols, 1):
        c = ws.cell(row=2, column=ci, value=label)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = CENTER; c.border = THIN
    ws.row_dimensions[2].height = 25
    ws.freeze_panes = 'A3'

    for i, row_data in enumerate(DELIVERABLE_SECTIONS):
        fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
        r = 3 + i
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=r, column=ci, value=val)
            c.fill = fill
            c.border = THIN
            c.alignment = CENTER if ci in (1, 3) else WRAP
            c.font = BOLD_FONT if ci in (1, 2) else BODY_FONT
        ws.row_dimensions[r].height = 80

    # Notes row
    r2 = 3 + len(DELIVERABLE_SECTIONS) + 1
    ws.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=len(cols))
    c = ws.cell(row=r2, column=1,
                value='Note: Sections 7 and 8 are conditional on Feature 4 being completed within budget. '
                      'Core contracted deliverable (Sections 1-6, 9, 10) totals approximately 106 estimated hours. '
                      'Style throughout: British English, no em dashes, direct and numbers-backed.')
    c.font = Font(name='Calibri', size=9, italic=True, color='666666')
    c.alignment = WRAP
    ws.row_dimensions[r2].height = 35

    return ws


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 6 — WORKSAGE IMPORT
# ══════════════════════════════════════════════════════════════════════════════

WS_ROWS = [
    # PROJECT
    {
        '_temp_id': 'P1', '_temp_parent_id': '',
        'item_type': 'project',
        'title': 'Goals Studio Pricing Engagement',
        'description': (
            'Deliver a competitive pricing benchmark, regional pricing strategy, and risk assessment '
            'for the Goals Studio football game launching 14 May 2026.\n\n'
            'Budget: 100,000 SEK (~$10,000 USD) | ~116 estimated hours\n'
            'Client: Goals AB (Goals Studio), Stockholm, Sweden\n'
            'Contacts: Jonas Rundberg (Chief Executive Officer), Julius (Live Ops Lead)\n\n'
            'Statement of Work (SOW) scope:\n'
            '- Review competitive pricing positioning in the sports and football game space\n'
            '- Price point selection, currency pack structure, conversion rate guidance\n'
            '- Regional pricing for all launch markets\n'
            '- Assessment of current pricing plans against market norms\n'
            '- Recommendations on adjustments and areas of risk\n'
            '- Deliverable: Written summary (12-18 pages) plus a remote review call\n\n'
            'Architecture: Foundation (map Hard Currency economy) > '
            'Competitive benchmark > Regional analysis > '
            'Synthesis (recommendations and scenarios) > Deliverable\n\n'
            'Hard deadline: Must deliver by 27 April 2026 before Goals submits pricing to PlayStation and Xbox.\n'
            'Key question from Jonas: "Are we pricing too low?"\n'
            'Core principle: Launch high, adjust downward if needed. '
            'Console platforms make price increases nearly impossible after launch.'
        ),
        'status': 'In Progress',
        'priority': 'High',
        'hours_estimated': 116,
        'start_date': '2026-04-22',
        'due_date': '2026-04-27',
        'success_factor': '',
        'notes': 'Feature 4 is conditional — only delivered if Features 1-3 complete within budget.',
    },

    # FEATURE 1
    {
        '_temp_id': 'F1', '_temp_parent_id': 'P1',
        'item_type': 'feature',
        'title': 'F1: Competitive Hard Currency (HC) Pricing Benchmark',
        'description': (
            'Show Goals where their currency pack pricing sits relative to the sports and football '
            'Free-to-Play (F2P) market. Extend their existing competitive analysis with insights '
            'their internal position does not give them.\n\n'
            'What is added: independent cross-genre validation, pricing psychology assessment, '
            'conversion rate benchmarks from comparable live titles, and external expert validation.'
        ),
        'status': 'Not Started', 'priority': 'High', 'hours_estimated': 38,
        'start_date': '2026-04-22', 'due_date': '2026-04-23',
        'success_factor': '', 'notes': '33% of total budget',
    },

    # S1.0
    {
        '_temp_id': 'S1.0', '_temp_parent_id': 'F1',
        'item_type': 'story',
        'title': 'S1.0: Foundation — Map the Hard Currency (HC) Economy',
        'description': 'Before comparing prices, establish exactly what Goals\' Hard Currency buys '
                       'and what the player value chain looks like. Prerequisite for everything else.',
        'status': 'Not Started', 'priority': 'High', 'hours_estimated': 8,
        'start_date': '2026-04-22', 'due_date': '2026-04-22',
        'success_factor': '', 'notes': '',
    },
    {
        '_temp_id': 'T1.0.1', '_temp_parent_id': 'S1.0',
        'item_type': 'task',
        'title': 'T1.0.1: Map Hard Currency (HC) Value Chain',
        'description': (
            'Document the complete path from real money to in-game value in Goals.\n\n'
            'Key conversion: 1 Hard Currency coin = 50 points = $0.012 USD at base rate '
            '(25,000 points = 500 coins = $6.00 USD, from Pricing Model conversion rates sheet).\n\n'
            'Map what each Hard Currency tier actually buys:\n'
            '- 380 HC ($5.99): 1 Internal Kit (333 coins) with 47 coins leftover\n'
            '- 650 HC ($9.99): 1 Internal Kit plus 317 coins leftover (not enough for a second kit)\n\n'
            'Gap to confirm with Julius: Mech and Pulse kit Hard Currency pricing is not documented in client materials.\n\n'
            'Identify the minimum meaningful purchase — the cheapest satisfying thing a new player can buy.\n\n'
            'Sources: Goals Studio Monetisation Design Doc, Pricing Model spreadsheet, player_pricing data.'
        ),
        'status': 'Not Started', 'priority': 'High', 'hours_estimated': 3,
        'start_date': '2026-04-22', 'due_date': '2026-04-22',
        'success_factor': (
            'One-page document showing all 6 Hard Currency tiers and what each one buys. '
            'Any unconfirmed prices are clearly flagged. '
            'The minimum meaningful purchase is identified with its cost.'
        ),
        'notes': 'Day 1',
    },
    {
        '_temp_id': 'T1.0.2', '_temp_parent_id': 'S1.0',
        'item_type': 'task',
        'title': 'T1.0.2: Establish Payer Persona Benchmarks',
        'description': (
            'Benchmark Goals\' 3-persona model against published industry data.\n\n'
            'Goals\' targets (from Pricing Model Overview sheet):\n'
            '- Persona 1 (core player, every day): 20% of players, 30% pay, '
            'Average Revenue Per Paying User (ARPPU) $46\n'
            '- Persona 2 (dedicated, 2-3 times per week): 35% of players, 10% pay, ARPPU $15.81\n'
            '- Persona 3 (casual, few times per month): 45% of players, 2% pay, ARPPU $7.24\n'
            '- Overall: 10.4% payer rate, Average Revenue Per User (ARPU) $3.38, ARPPU $32.48\n\n'
            'Research method: pull published conversion and ARPPU data from eFootball '
            '(Konami quarterly reports), EA FC (EA quarterly earnings), Sensor Tower or data.ai, '
            'and Game Developers Conference (GDC) presentations with cited benchmarks.\n\n'
            'Critical question: is the 10.4% target a month-1 assumption or a steady-state figure? '
            'The plan does not specify.\n\n'
            'Seasonless economy note: Goals has no seasonal resets, no battle pass, and no '
            'Fear of Missing Out (FOMO)-driven content wipes, so the Hard Currency ladder must do '
            'all conversion work that seasonal games spread across time-limited events.'
        ),
        'status': 'Not Started', 'priority': 'High', 'hours_estimated': 5,
        'start_date': '2026-04-22', 'due_date': '2026-04-22',
        'success_factor': (
            'Written assessment with at least 2 cited sources comparing Goals\' assumptions to '
            'real published data. Clear verdict: realistic, aggressive, or overstated. '
            'If overstated, state what the implication is for pricing strategy.'
        ),
        'notes': 'Day 1',
    },

    # S1.1
    {
        '_temp_id': 'S1.1', '_temp_parent_id': 'F1',
        'item_type': 'story',
        'title': 'S1.1: Competitive Pricing Position',
        'description': 'Show Goals exactly where their Hard Currency pricing sits relative to the market — '
                       'by synthesising their existing competitive data into a clear positioning statement '
                       'they cannot write themselves.',
        'status': 'Not Started', 'priority': 'High', 'hours_estimated': 19,
        'start_date': '2026-04-22', 'due_date': '2026-04-23',
        'success_factor': '', 'notes': '',
    },
    {
        '_temp_id': 'T1.1.1', '_temp_parent_id': 'S1.1',
        'item_type': 'task',
        'title': 'T1.1.1: Build Normalised Price Position Map',
        'description': (
            'Create a position map showing Goals\' effective USD per Hard Currency coin '
            'against the competitive range at each price tier.\n\n'
            'Goals\' data (from Pricing Model softhard currency comparison sheet):\n'
            '- Entry tier ($5-6): Goals = $0.01576 per HC coin\n'
            '- Mid tier (~$20): Goals = $0.01454 per HC coin\n'
            '- Top tier ($100): Goals = $0.01290 per HC coin\n\n'
            'For each tier: calculate market minimum, median, maximum, '
            'and Goals\' position as a percentile.\n\n'
            'Spot-check 2-3 competitor prices to confirm Goals\' data is still current.\n\n'
            'Output: one clear sentence summarising overall competitive position, e.g. '
            '"Goals is at the Xth percentile of market pricing — below/at/above median."'
        ),
        'status': 'Not Started', 'priority': 'High', 'hours_estimated': 8,
        'start_date': '2026-04-22', 'due_date': '2026-04-22',
        'success_factor': (
            'Position map exists for all 6 tiers. Each tier shows Goals price, '
            'market range (minimum, median, maximum), and Goals\' percentile. '
            'One clear sentence summarising overall position.'
        ),
        'notes': 'Day 1. Source: Goals Pricing Model.',
    },
    {
        '_temp_id': 'T1.1.2', '_temp_parent_id': 'S1.1',
        'item_type': 'task',
        'title': 'T1.1.2: Discount Curve Comparison',
        'description': (
            'Compare Goals\' bonus and discount scaling against competitors.\n\n'
            'Goals\' discount curve: Tier 1 (0%), Tier 2 (~2.5%), Tier 3 (~7.7%), '
            'Tier 4 (~9.4%), Tier 5 (~15.4%), Tier 6 (~18.1%).\n\n'
            'Competitor maximum discounts: EA FC ~39%, Fortnite ~24%, Apex ~13%.\n\n'
            'Assess: is 18% maximum discount enough to drive high-spend commitment at launch? '
            'Is the jump from tier to tier consistent? Is there a clear "best value" tier?\n\n'
            'Key principle: new launches should be conservative on discounts — '
            'discounts can always be increased later but never reduced without player backlash.'
        ),
        'status': 'Not Started', 'priority': 'High', 'hours_estimated': 4,
        'start_date': '2026-04-22', 'due_date': '2026-04-22',
        'success_factor': (
            'Comparison table showing discount curves for Goals and 4 competitors. '
            'Written recommendation: keep current curve, steepen it, or flatten it — with reasoning.'
        ),
        'notes': 'Day 1',
    },
    {
        '_temp_id': 'T1.1.3', '_temp_parent_id': 'S1.1',
        'item_type': 'task',
        'title': 'T1.1.3: Price Tier (SKU) Structure and Entry Point Audit',
        'description': (
            'Assess structural health of the 6-tier Hard Currency ladder: '
            '$5.99 / $9.99 / $19.99 / $29.99 / $49.99 / $99.99.\n\n'
            'Compare with: EA FC (8 tiers, starts at $0.99), '
            'Fortnite (4 tiers, starts at $7.99), Apex (6 tiers, starts at $4.99).\n\n'
            'Entry point analysis: does $5.99 create first-purchase friction? '
            'What can 380 Hard Currency buy? '
            'Counter-argument: a $0.99 tier nets only $0.69 after the 30% platform fee — '
            'may not justify the transaction cost for a small studio.\n\n'
            'Structural gaps: $9.99 to $19.99 is a 100% price jump; '
            '$49.99 to $99.99 is also a 100% jump. These are standard in the industry.\n\n'
            'Whale ceiling: $99.99 cap with no purchase limits is whale-friendly '
            'without looking optically aggressive. Correct for the anti-Pay-to-Win positioning.'
        ),
        'status': 'Not Started', 'priority': 'High', 'hours_estimated': 5,
        'start_date': '2026-04-23', 'due_date': '2026-04-23',
        'success_factor': (
            'Written audit covering: entry point assessment, tier gap analysis, whale ceiling analysis. '
            'Clear recommendation for each area: keep as-is, add a tier, or adjust. '
            'Supported by competitor comparison.'
        ),
        'notes': 'Day 2',
    },
    {
        '_temp_id': 'T1.1.4', '_temp_parent_id': 'S1.1',
        'item_type': 'task',
        'title': 'T1.1.4: Platform Fee Impact and Net Revenue Table',
        'description': (
            'Calculate actual net revenue per price tier per platform.\n\n'
            'Table (6 tiers x 4 platforms):\n'
            '380 HC $5.99: Steam $4.19 | Epic $5.27 | PlayStation $4.19 | Xbox $4.19\n'
            '650 HC $9.99: Steam $6.99 | Epic $8.79 | PlayStation $6.99 | Xbox $6.99\n'
            '1,375 HC $19.99: Steam $14.00 | Epic $17.59 | PlayStation $14.00 | Xbox $14.00\n'
            '2,100 HC $29.99: Steam $21.00 | Epic $26.39 | PlayStation $21.00 | Xbox $21.00\n'
            '3,750 HC $49.99: Steam $35.00 | Epic $43.99 | PlayStation $35.00 | Xbox $35.00\n'
            '7,750 HC $99.99: Steam $70.00 | Epic $87.99 | PlayStation $70.00 | Xbox $70.00\n\n'
            'Epic Games advantage: Goals nets approximately 26% more per transaction on Epic '
            'vs Steam, PlayStation, and Xbox at the same price point.\n\n'
            'Verify alignment with PlayStation wholesale rates already in Goals\' pricing matrix.'
        ),
        'status': 'Not Started', 'priority': 'Medium', 'hours_estimated': 2,
        'start_date': '2026-04-23', 'due_date': '2026-04-23',
        'success_factor': (
            'Complete net revenue table (6 tiers x 4 platforms = 24 data cells). '
            'Any margin concerns flagged. Epic Games advantage quantified as a percentage.'
        ),
        'notes': 'Day 2',
    },

    # S1.2
    {
        '_temp_id': 'S1.2', '_temp_parent_id': 'F1',
        'item_type': 'story',
        'title': 'S1.2: Conversion Rate and Elasticity Context',
        'description': 'Provide conversion and revenue modelling context that turns a static price comparison '
                       'into actionable strategy.',
        'status': 'Not Started', 'priority': 'High', 'hours_estimated': 8,
        'start_date': '2026-04-23', 'due_date': '2026-04-23',
        'success_factor': '', 'notes': '',
    },
    {
        '_temp_id': 'T1.2.1', '_temp_parent_id': 'S1.2',
        'item_type': 'task',
        'title': 'T1.2.1: Price Elasticity Scenario Modelling',
        'description': (
            'Model 4 scenarios using Goals\' own revenue model: '
            '805,250 Monthly Active Users (MAU), 10.4% payer rate, '
            '$32.48 Average Revenue Per Paying User (ARPPU) = $2.72M total revenue.\n\n'
            'Scenario A: Raise all prices 10% — assume conversion drops 5-15%.\n'
            'Scenario B: Raise all prices 20% — assume conversion drops 10-25%.\n'
            'Scenario C: Raise entry tier only (from $5.99 to $6.99) — '
            'entry conversion drops 5-10%, mid and high tiers unaffected.\n'
            'Scenario D: Keep prices, steepen top-tier discount to 25% maximum.\n\n'
            'Present as conservative-to-optimistic ranges, not single point estimates.\n\n'
            'Key insight to surface: if the 10.4% conversion target is optimistic for month 1, '
            'lowering prices will not help — time and content build conversion. '
            'Therefore, launching at a slightly higher price and reducing later is the correct strategy.'
        ),
        'status': 'Not Started', 'priority': 'High', 'hours_estimated': 8,
        'start_date': '2026-04-23', 'due_date': '2026-04-23',
        'success_factor': (
            '4 scenarios modelled with revenue impact ranges. Each shows: price change, '
            'assumed conversion impact, net revenue change. '
            'Clear recommendation on which scenario maximises launch revenue while preserving flexibility.'
        ),
        'notes': 'Day 2',
    },

    # S1.3
    {
        '_temp_id': 'S1.3', '_temp_parent_id': 'F1',
        'item_type': 'story',
        'title': 'S1.3: Promotional and First-Purchase Pricing',
        'description': 'Address gap in static currency pack analysis. '
                       'First-purchase incentives and promotional pricing drive early conversion.',
        'status': 'Not Started', 'priority': 'Medium', 'hours_estimated': 5,
        'start_date': '2026-04-23', 'due_date': '2026-04-23',
        'success_factor': '', 'notes': '',
    },
    {
        '_temp_id': 'T1.3.1', '_temp_parent_id': 'S1.3',
        'item_type': 'task',
        'title': 'T1.3.1: First-Purchase and Starter Pack Recommendations',
        'description': (
            'Recommend complementary offers that improve conversion on the core Hard Currency ladder '
            'without changing the ladder itself.\n\n'
            '1. First Purchase Bonus: double Hard Currency on first-ever purchase at any tier. '
            'Research what EA FC, Fortnite, Apex, and eFootball currently offer as first-purchase incentives. '
            'Do NOT cite conversion improvement percentages without a verifiable source.\n\n'
            '2. Limited Starter Pack: one-time $3.99 offer (below the $5.99 entry tier) '
            'containing 500 Hard Currency + Rare+ player + exclusive kit. '
            'Breaks the first-purchase barrier.\n\n'
            '3. World Cup Launch Bundle: tied to June event, limited availability, '
            'higher value per dollar than standard packs. Creates Fear of Missing Out (FOMO) '
            'without permanently discounting.\n\n'
            'Frame all three as additions to the core pricing, not changes to it.'
        ),
        'status': 'Not Started', 'priority': 'Medium', 'hours_estimated': 5,
        'start_date': '2026-04-23', 'due_date': '2026-04-23',
        'success_factor': (
            '2-3 specific first-purchase recommendations. Each includes: what it is, '
            'what it costs Goals, expected benefit. '
            'Clearly framed as additional to core pricing, not a substitution.'
        ),
        'notes': 'Day 2. Value-add observation, not core scope.',
    },

    # FEATURE 2
    {
        '_temp_id': 'F2', '_temp_parent_id': 'P1',
        'item_type': 'feature',
        'title': 'F2: Regional Pricing Strategy',
        'description': (
            'Validate Goals\' 40+ country pricing matrix. What is added: competitive regional context, '
            'Purchasing Power Parity (PPP)-adjusted analysis, and platform compliance verification '
            'they cannot get from their own spreadsheet.\n\n'
            'Goals built their matrix using exchange rates and some PPP data. '
            'They cannot benchmark their regional pricing against what EA FC actually charges '
            'in Turkey or Brazil — that requires platform store access or competitor intelligence.'
        ),
        'status': 'Not Started', 'priority': 'High', 'hours_estimated': 34,
        'start_date': '2026-04-23', 'due_date': '2026-04-25',
        'success_factor': '', 'notes': '29% of total budget',
    },

    # S2.1
    {
        '_temp_id': 'S2.1', '_temp_parent_id': 'F2',
        'item_type': 'story',
        'title': 'S2.1: Regional Deviation and Risk Identification',
        'description': 'Flag countries where Goals\' pricing creates material risk: '
                       'too high (player backlash), too low (revenue leakage), or misaligned (platform rejection).',
        'status': 'Not Started', 'priority': 'High', 'hours_estimated': 23,
        'start_date': '2026-04-23', 'due_date': '2026-04-24',
        'success_factor': '', 'notes': '',
    },
    {
        '_temp_id': 'T2.1.1', '_temp_parent_id': 'S2.1',
        'item_type': 'task',
        'title': 'T2.1.1: Categorise All Countries by Risk Level',
        'description': (
            'Using Goals\' existing deviation data, categorise every country:\n\n'
            'Green: deviation less than 15% from the Foreign Exchange (FX) rate '
            'AND Goals\' own matrix says proceed.\n'
            'Amber: deviation 15-25% OR flagged in Goals\' matrix '
            'OR misaligned with competitor pricing in that market.\n'
            'Red: deviation greater than 25% OR Goals\' matrix recommends a price change '
            'OR large player market with more than 15% over-pricing.\n\n'
            'Red candidates from Goals\' data: Poland (+25%), Switzerland (+26%), '
            'Ukraine (-26%), Turkey, Brazil.\n'
            'Amber candidates: South Africa, Czech Republic, Norway, Denmark.\n\n'
            'Cross-reference with beta community size: UK 1,274 players, France 1,026, '
            'Turkey 667, Brazil 495, Poland 392.\n\n'
            'Note: Euro zone countries all use the same price tier but Purchasing Power Parity '
            'varies significantly within the Euro zone (Germany vs Portugal vs Greece).'
        ),
        'status': 'Not Started', 'priority': 'High', 'hours_estimated': 8,
        'start_date': '2026-04-23', 'due_date': '2026-04-24',
        'success_factor': (
            'Every country has a colour risk rating. Red countries have a one-sentence explanation '
            'of the specific problem. Output is a ranked list with highest-risk markets at the top.'
        ),
        'notes': 'Days 2-3',
    },
    {
        '_temp_id': 'T2.1.2', '_temp_parent_id': 'S2.1',
        'item_type': 'task',
        'title': 'T2.1.2: Competitor Regional Pricing for Red and Amber Markets',
        'description': (
            'For 5-10 red and amber countries, research what EA FC and Fortnite actually charge.\n\n'
            'Sources: PlayStation Store regional web store, SteamDB (historical and current pricing), '
            'Epic Games Store.\n\n'
            'For each market, build a comparison table: '
            'Country | Goals proposed price | EA FC actual price | Fortnite actual price | '
            'Market median | Goals position (above, below, or at market).\n\n'
            'Priority markets:\n'
            '- Brazil: large player base, low Purchasing Power Parity, '
            'critical for World Cup engagement\n'
            '- Turkey: 5th largest beta community, extreme PPP differential vs USD\n'
            '- Poland: large beta community, already +25% above the Foreign Exchange rate\n\n'
            'If direct store access is blocked by region-locking, use SteamDB or documented '
            'third-party pricing databases. Cite all sources.\n\n'
            'This is the single largest research task in the engagement.'
        ),
        'status': 'Not Started', 'priority': 'High', 'hours_estimated': 15,
        'start_date': '2026-04-24', 'due_date': '2026-04-24',
        'success_factor': (
            'Comparison table for all red and amber countries. '
            'Each row: country, Goals price, EA FC price, Fortnite price, percentage difference. '
            'Sources cited for every competitor price used.'
        ),
        'notes': 'Day 3. Largest single research task. Any overrun is absorbed from Feature 4 scope.',
    },

    # S2.2
    {
        '_temp_id': 'S2.2', '_temp_parent_id': 'F2',
        'item_type': 'story',
        'title': 'S2.2: Platform Compliance and Submission Guidance',
        'description': 'Ensure Goals\' pricing will pass PlayStation and Xbox review without rejection or delay.',
        'status': 'Not Started', 'priority': 'Critical', 'hours_estimated': 3,
        'start_date': '2026-04-24', 'due_date': '2026-04-24',
        'success_factor': '', 'notes': '',
    },
    {
        '_temp_id': 'T2.2.1', '_temp_parent_id': 'S2.2',
        'item_type': 'task',
        'title': 'T2.2.1: Verify Platform Tier Compliance',
        'description': (
            'Cross-reference Goals\' proposed prices against valid PlayStation and Xbox tier selections.\n\n'
            'Sources: Goals\' PS4/PS5 Digital sheet, Xbox wholesale pricing sheet in the Pricing Matrix.\n\n'
            'Any price that does not align to a standard platform tier will be rejected by Sony or Microsoft.\n\n'
            'Submission checklist:\n'
            '- All prices are valid platform tiers\n'
            '- All territories where the game is available are covered\n'
            '- USD fallback in place for currencies the platform does not support\n'
            '- Review period is approximately 9 days (per Julius from kickoff call)\n\n'
            'Timeline: deliver by 27 April > Goals submits by ~5 May > 9-day review > 14 May launch.\n\n'
            'Ask Julius: has a test pricing configuration been submitted previously? '
            'If yes, did it pass? CRITICAL PATH — any rejection delays the launch.'
        ),
        'status': 'Not Started', 'priority': 'Critical', 'hours_estimated': 3,
        'start_date': '2026-04-24', 'due_date': '2026-04-24',
        'success_factor': (
            'Written confirmation: all prices are valid platform tiers, OR a list of specific '
            'prices that need adjustment with the nearest valid tier for each one.'
        ),
        'notes': 'Day 3. Critical path — rejection delays launch.',
    },

    # S2.3
    {
        '_temp_id': 'S2.3', '_temp_parent_id': 'F2',
        'item_type': 'story',
        'title': 'S2.3: Country-Specific Recommendations',
        'description': 'Provide specific price adjustment recommendations for each red and amber country.',
        'status': 'Not Started', 'priority': 'High', 'hours_estimated': 8,
        'start_date': '2026-04-24', 'due_date': '2026-04-25',
        'success_factor': '', 'notes': '',
    },
    {
        '_temp_id': 'T2.3.1', '_temp_parent_id': 'S2.3',
        'item_type': 'task',
        'title': 'T2.3.1: Write Country Recommendation Cards',
        'description': (
            'For each red and amber country, write a self-contained recommendation card:\n\n'
            'Country | Current proposed price | Issue | Recommended price | '
            'Rationale | Revenue impact | Risk if unchanged\n\n'
            'Prioritise by: player base size multiplied by degree of deviation.\n\n'
            'Group by action type:\n'
            '- Reduce price (markets where Goals is over-priced)\n'
            '- Can increase (markets where revenue is being left on the table)\n'
            '- Platform tier adjustment (price needs rounding to the nearest valid tier)\n\n'
            'Expected output: 8-12 country cards.\n\n'
            'Each card must be self-contained — Julius can read one card and know exactly '
            'what to change in the pricing spreadsheet for that country, '
            'without reading the full report.'
        ),
        'status': 'Not Started', 'priority': 'High', 'hours_estimated': 8,
        'start_date': '2026-04-24', 'due_date': '2026-04-25',
        'success_factor': (
            '8-12 country cards completed. Each is self-contained. '
            'Julius can read one card and know exactly what to change in the pricing spreadsheet.'
        ),
        'notes': 'Days 3-4',
    },

    # FEATURE 3
    {
        '_temp_id': 'F3', '_temp_parent_id': 'P1',
        'item_type': 'feature',
        'title': 'F3: Risk Assessment, Scenarios and Recommendations',
        'description': 'Synthesise all analysis into clear, specific, actionable guidance with modelled revenue impact.',
        'status': 'Not Started', 'priority': 'High', 'hours_estimated': 16,
        'start_date': '2026-04-25', 'due_date': '2026-04-25',
        'success_factor': '', 'notes': '14% of total budget',
    },

    # S3.1
    {
        '_temp_id': 'S3.1', '_temp_parent_id': 'F3',
        'item_type': 'story',
        'title': 'S3.1: "Are We Too Low?" — The Central Answer',
        'description': 'Directly address Jonas\'s core concern with evidence.',
        'status': 'Not Started', 'priority': 'Critical', 'hours_estimated': 4,
        'start_date': '2026-04-25', 'due_date': '2026-04-25',
        'success_factor': '', 'notes': '',
    },
    {
        '_temp_id': 'T3.1.1', '_temp_parent_id': 'S3.1',
        'item_type': 'task',
        'title': 'T3.1.1: Build the Definitive Position Statement',
        'description': (
            'Synthesise the competitive position (T1.1.1), payer benchmarks (T1.0.2), '
            'and scenario modelling (T1.2.1) into one clear paragraph.\n\n'
            'Format: "Goals\' base Hard Currency pricing is at the Xth percentile of the competitive set. '
            'This means [interpretation]. Given the 80/20 skill-to-purchase positioning, '
            'this is [appropriate/low/high] because [reason]. '
            'The payer conversion target of 10.4% is [realistic/aggressive] for month 1. '
            'At realistic month-1 conversion of Y%, current pricing yields Z revenue. '
            'Our recommendation is [specific — a number, not a range]."\n\n'
            'Take a position. No hedging.\n\n'
            'Core principle: you can always reduce prices post-launch. '
            'You cannot raise them on console — re-certification is slow and generates press coverage.'
        ),
        'status': 'Not Started', 'priority': 'Critical', 'hours_estimated': 4,
        'start_date': '2026-04-25', 'due_date': '2026-04-25',
        'success_factor': (
            'One paragraph exists that Jonas can read and immediately know whether to raise prices or not. '
            'Contains a specific recommendation with a number. No hedging.'
        ),
        'notes': 'Day 4. Depends on T1.1.1, T1.0.2, T1.2.1 completion.',
    },

    # S3.2
    {
        '_temp_id': 'S3.2', '_temp_parent_id': 'F3',
        'item_type': 'story',
        'title': 'S3.2: Risk Register',
        'description': 'Document all identified pricing risks before launch, ranked by severity.',
        'status': 'Not Started', 'priority': 'High', 'hours_estimated': 5,
        'start_date': '2026-04-25', 'due_date': '2026-04-25',
        'success_factor': '', 'notes': '',
    },
    {
        '_temp_id': 'T3.2.1', '_temp_parent_id': 'S3.2',
        'item_type': 'task',
        'title': 'T3.2.1: Compile and Rank Risks',
        'description': (
            'Document each risk: Description | Severity | Likelihood | Markets Affected | Mitigation\n\n'
            '10 identified risks:\n'
            '1. Pay-to-Win (P2W) perception at high price points (Critical / Medium)\n'
            '2. Post-launch price increase is practically impossible (High / Medium)\n'
            '3. Over-pricing in price-sensitive markets — Turkey, Brazil, Poland (High / High)\n'
            '4. Platform pricing rejection causes launch delay (Critical / Low)\n'
            '5. Discount curve too shallow to drive high-spend commitment (Medium / Medium)\n'
            '6. Entry-tier friction at $5.99 reduces first-time buyer conversion (Medium / Medium)\n'
            '7. No pricing framework for Originals and Celebrity packs post-launch (Low / not applicable)\n'
            '8. Netherlands gambling authority (Kansspelautoriteit — KSA) regulatory risk '
            'for player packs (High / Medium)\n'
            '9. South Korea drop rate disclosure is a legal requirement (High / Certain if launching in KR)\n'
            '10. Non-seasonal economy creates flat rather than pulsing revenue (Medium / High)\n\n'
            'Ranked by severity multiplied by likelihood. '
            'Top 3 risks must have specific, immediately implementable mitigation steps.'
        ),
        'status': 'Not Started', 'priority': 'High', 'hours_estimated': 5,
        'start_date': '2026-04-25', 'due_date': '2026-04-25',
        'success_factor': (
            'Minimum 10 risks documented. Each has severity, likelihood, affected markets, '
            'and a mitigation action. '
            'Top 3 risks have specific, implementable steps Goals can take immediately.'
        ),
        'notes': 'Day 4',
    },

    # S3.3
    {
        '_temp_id': 'S3.3', '_temp_parent_id': 'F3',
        'item_type': 'story',
        'title': 'S3.3: Top 5 Recommendations',
        'description': 'The five highest-impact actions Goals should take before platform submission.',
        'status': 'Not Started', 'priority': 'High', 'hours_estimated': 7,
        'start_date': '2026-04-25', 'due_date': '2026-04-25',
        'success_factor': '', 'notes': '',
    },
    {
        '_temp_id': 'T3.3.1', '_temp_parent_id': 'S3.3',
        'item_type': 'task',
        'title': 'T3.3.1: Write Top 5 Recommendations',
        'description': (
            'Five specific, numbered, actionable recommendations. Each must include:\n'
            '- What to change (exact specification, a number)\n'
            '- Why (evidence from the analysis)\n'
            '- Revenue impact (conservative-to-optimistic range)\n'
            '- Implementation steps (which cells in Goals\' pricing matrix, which platform to resubmit)\n'
            '- Risk if skipped (what happens)\n\n'
            'Expected candidates:\n'
            '1. Pricing position adjustment — maintain or adjust base tier based on competitive percentile\n'
            '2. Regional correction — specific price changes for Poland, Turkey, Brazil\n'
            '3. Discount curve adjustment — consider steepening top-tier discount\n'
            '4. First-purchase incentive — implement double Hard Currency bonus on first purchase\n'
            '5. Launch framing — label current pricing as "Launch Pricing" to preserve future flexibility\n\n'
            'Must pass the test: would a studio economy designer read these and act same day?'
        ),
        'status': 'Not Started', 'priority': 'High', 'hours_estimated': 7,
        'start_date': '2026-04-25', 'due_date': '2026-04-25',
        'success_factor': (
            '5 recommendations written in full format. Each has: exact change to make, '
            'evidence for why, expected revenue impact as a range, and what happens if they skip it. '
            'A studio economy designer reads these and acts same day.'
        ),
        'notes': 'Day 4',
    },

    # FEATURE 4
    {
        '_temp_id': 'F4', '_temp_parent_id': 'P1',
        'item_type': 'feature',
        'title': 'F4: Value-Add (Conditional — Time Permitting)',
        'description': (
            'Supplementary tools demonstrating strategic depth and positioning Phase 2. '
            'ONLY executed if Features 1-3 complete within budget. '
            'Scope is reduced first if any earlier feature runs over time.'
        ),
        'status': 'Not Started', 'priority': 'Medium', 'hours_estimated': 8,
        'start_date': '2026-04-25', 'due_date': '2026-04-25',
        'success_factor': '', 'notes': '7% of total budget. CONDITIONAL.',
    },

    # S4.1
    {
        '_temp_id': 'S4.1', '_temp_parent_id': 'F4',
        'item_type': 'story',
        'title': 'S4.1: Reusable Regional Pricing Methodology',
        'description': 'One-page methodology Goals can use to price any future item without external help.',
        'status': 'Not Started', 'priority': 'Medium', 'hours_estimated': 3,
        'start_date': '2026-04-25', 'due_date': '2026-04-25',
        'success_factor': '', 'notes': 'Conditional on F1-3 completion',
    },
    {
        '_temp_id': 'T4.1.1', '_temp_parent_id': 'S4.1',
        'item_type': 'task',
        'title': 'T4.1.1: Write "How to Price a New Price Tier (SKU)" Guide',
        'description': (
            '6-step methodology Julius can follow for any future price tier:\n'
            '1. Set USD base price — anchor against the equivalent EA FC item price\n'
            '2. Convert to local currencies using current Foreign Exchange (FX) rates\n'
            '3. Check the Purchasing Power Parity (PPP) factor using the lookup table provided\n'
            '4. Compare against EA FC in that market — plus or minus 15% is acceptable\n'
            '5. Round to the nearest valid platform tier\n'
            '6. Verify the discount curve is maintained across all tiers\n\n'
            'Include a worked example: pricing a new Celebration item at $7.00 USD for the Brazil market.\n'
            'Include a PPP lookup table for all launch markets.\n\n'
            'Positioning: "This is the manual method. Phase 2 builds a model that does this automatically '
            'using live transaction data."'
        ),
        'status': 'Not Started', 'priority': 'Medium', 'hours_estimated': 3,
        'start_date': '2026-04-25', 'due_date': '2026-04-25',
        'success_factor': (
            'One-page guide exists. Julius can follow it without external help. '
            'Worked example included. PPP lookup table for all launch markets included.'
        ),
        'notes': 'Day 4. Conditional.',
    },

    # S4.2
    {
        '_temp_id': 'S4.2', '_temp_parent_id': 'F4',
        'item_type': 'story',
        'title': 'S4.2: World Cup Pricing Opportunity Notes',
        'description': '3-5 tactical pricing observations for the June World Cup window.',
        'status': 'Not Started', 'priority': 'Low', 'hours_estimated': 2,
        'start_date': '2026-04-25', 'due_date': '2026-04-25',
        'success_factor': '', 'notes': 'Conditional on F1-3 completion',
    },
    {
        '_temp_id': 'T4.2.1', '_temp_parent_id': 'S4.2',
        'item_type': 'task',
        'title': 'T4.2.1: Write World Cup Quick-Hit Observations',
        'description': (
            '3-5 tactical pricing observations for the June World Cup window. Half a page maximum.\n\n'
            'Ideas to include:\n'
            '- National kit pricing: charge a premium during the group stage '
            '(emotional spending peaks), consider discounting in the knockout rounds '
            '(fans of eliminated teams are less likely to spend on their nation\'s kit)\n'
            '- Time-limited "World Cup Coin Pack" at approximately 10% better value than standard packs — '
            'creates urgency without permanently reducing margins\n'
            '- Research what published data exists on revenue spikes during major sporting events '
            'in comparable Free-to-Play titles. '
            'Do NOT cite revenue multipliers without a verifiable published source.\n\n'
            'Positioned as a teaser for Phase 2 live service consulting — '
            'not a complete World Cup strategy.'
        ),
        'status': 'Not Started', 'priority': 'Low', 'hours_estimated': 2,
        'start_date': '2026-04-25', 'due_date': '2026-04-25',
        'success_factor': (
            '3-5 bullet points. Each is a specific, actionable pricing tactic for the World Cup window. '
            'Clearly positioned as a preview of what Phase 2 would cover in full detail.'
        ),
        'notes': 'Day 4. Conditional. Half page maximum.',
    },

    # S4.3
    {
        '_temp_id': 'S4.3', '_temp_parent_id': 'F4',
        'item_type': 'story',
        'title': 'S4.3: Phase 2 Positioning',
        'description': 'Brief section positioning the larger follow-on engagement.',
        'status': 'Not Started', 'priority': 'Medium', 'hours_estimated': 3,
        'start_date': '2026-04-25', 'due_date': '2026-04-25',
        'success_factor': '', 'notes': 'Conditional on F1-3 completion',
    },
    {
        '_temp_id': 'T4.3.1', '_temp_parent_id': 'S4.3',
        'item_type': 'task',
        'title': "T4.3.1: Write 'What's Next' Section",
        'description': (
            'Half-page section positioning Phase 2 as genuine advice, not a sales pitch.\n\n'
            'Natural next steps to include:\n'
            '- Post-launch telemetry review 2-3 weeks after 14 May: real transaction data replaces estimates\n'
            '- A/B test design: optimise conversion at each tier using live player cohorts\n'
            '- Store optimisation: offer sequencing, time-limited promotion design, personalised pricing\n'
            '- World Cup monetisation strategy: full tactical plan for the June revenue peak\n'
            '- Ongoing live service consulting: content cadence, engagement frameworks, team coaching\n\n'
            'Close with: "We recommend a check-in call 2-3 weeks post-launch to review initial '
            'transaction data and scope the next phase."\n\n'
            'Do NOT include internal Phase 2 pricing in the client document.'
        ),
        'status': 'Not Started', 'priority': 'Medium', 'hours_estimated': 3,
        'start_date': '2026-04-25', 'due_date': '2026-04-25',
        'success_factor': (
            'Half-page section written. Reads as genuine advice. '
            'Recommends a check-in call 2-3 weeks post-launch. '
            'Does not include internal pricing.'
        ),
        'notes': 'Day 4. Conditional.',
    },

    # FEATURE 5
    {
        '_temp_id': 'F5', '_temp_parent_id': 'P1',
        'item_type': 'feature',
        'title': 'F5: Deliverable Assembly and Client Review',
        'description': 'Package all analysis into the contracted deliverable: '
                       'a written summary (12-18 pages) plus a 30-45 minute remote review call '
                       'with Jonas and Julius.',
        'status': 'Not Started', 'priority': 'High', 'hours_estimated': 18,
        'start_date': '2026-04-26', 'due_date': '2026-04-27',
        'success_factor': '', 'notes': '16% of total budget',
    },

    # S5.1
    {
        '_temp_id': 'S5.1', '_temp_parent_id': 'F5',
        'item_type': 'story',
        'title': 'S5.1: Written Summary Document',
        'description': 'Create the final client-facing document: 12-18 pages, British English, numbers-backed.',
        'status': 'Not Started', 'priority': 'High', 'hours_estimated': 10,
        'start_date': '2026-04-26', 'due_date': '2026-04-26',
        'success_factor': '', 'notes': '',
    },
    {
        '_temp_id': 'T5.1.1', '_temp_parent_id': 'S5.1',
        'item_type': 'task',
        'title': 'T5.1.1: Assemble the Written Summary',
        'description': (
            'Final client-facing document structure (12-18 pages):\n'
            '1. Executive Summary (1 page) — headline finding, top 5 recommendations, immediate actions\n'
            '2. Your Competitive Position — where Goals sits vs market (Feature 1 findings)\n'
            '3. Regional Pricing Assessment — country risk map and recommendation cards (Feature 2)\n'
            '4. Pricing Scenarios — what-if modelling with revenue impact ranges (Story 1.2)\n'
            '5. Risk Register — all 10 identified risks with mitigation (Story 3.2)\n'
            '6. Recommendations — top 5 specific changes (Story 3.3)\n'
            '7. Pricing Framework — reusable methodology guide (Feature 4, if completed)\n'
            '8. World Cup Opportunities — tactical observations (Feature 4, if completed)\n'
            '9. Next Steps — Phase 2 positioning\n'
            '10. Appendix — full comparison tables, country matrix, competitor data\n\n'
            'Style: British English. No em dashes. Direct, specific, numbers-backed. '
            'Written for both Jonas (strategic audience) and Julius (practitioner) simultaneously.'
        ),
        'status': 'Not Started', 'priority': 'High', 'hours_estimated': 10,
        'start_date': '2026-04-26', 'due_date': '2026-04-26',
        'success_factor': (
            'Complete document draft ready for internal review. All sections filled with real findings — '
            'no placeholder text. Internally consistent. '
            'Written in British English, direct and numbers-backed.'
        ),
        'notes': 'Day 5',
    },

    # S5.2
    {
        '_temp_id': 'S5.2', '_temp_parent_id': 'F5',
        'item_type': 'story',
        'title': 'S5.2: Internal Review and Polish',
        'description': 'Internal quality gate review before client delivery.',
        'status': 'Not Started', 'priority': 'High', 'hours_estimated': 4,
        'start_date': '2026-04-26', 'due_date': '2026-04-26',
        'success_factor': '', 'notes': '',
    },
    {
        '_temp_id': 'T5.2.1', '_temp_parent_id': 'S5.2',
        'item_type': 'task',
        'title': 'T5.2.1: Quality Gate — Internal Review',
        'description': (
            'Internal review covering two perspectives:\n\n'
            'Perspective 1 — Domain accuracy:\n'
            '- Do the recommendations make sense for this specific game?\n'
            '- Is the tone right (demonstrates expertise without being condescending)?\n'
            '- Does it answer Jonas\'s core questions completely?\n\n'
            'Perspective 2 — Commercial and brand:\n'
            '- Does it position Phase 2 naturally, without a sales pitch feel?\n'
            '- Is it consistent with the NBI brand and quality standard?\n'
            '- Are there any contractual or legal concerns?\n\n'
            'Address all feedback. Produce the final version.\n'
            'Send to Jonas and Julius at least 24 hours before the review call.'
        ),
        'status': 'Not Started', 'priority': 'High', 'hours_estimated': 4,
        'start_date': '2026-04-26', 'due_date': '2026-04-26',
        'success_factor': (
            'Final document reviewed and approved. Sent to client at least 24 hours before the call. '
            'Review call scheduled with Jonas and Julius.'
        ),
        'notes': 'Day 5',
    },

    # S5.3
    {
        '_temp_id': 'S5.3', '_temp_parent_id': 'F5',
        'item_type': 'story',
        'title': 'S5.3: Remote Review Session',
        'description': '30-45 minute call walking Jonas and Julius through findings and recommendations.',
        'status': 'Not Started', 'priority': 'High', 'hours_estimated': 4,
        'start_date': '2026-04-27', 'due_date': '2026-04-27',
        'success_factor': '', 'notes': '',
    },
    {
        '_temp_id': 'T5.3.1', '_temp_parent_id': 'S5.3',
        'item_type': 'task',
        'title': 'T5.3.1: Conduct Client Review Call',
        'description': (
            'Call agenda (30-45 minutes):\n'
            '- 5 min: Recap scope and methodology used\n'
            '- 10 min: Top 5 recommendations (highest impact first)\n'
            '- 10 min: Regional pricing highlights (red countries and what needs to change)\n'
            '- 5 min: Pricing framework overview (if Feature 4 was completed)\n'
            '- 5 min: Next steps and Phase 2 opportunity\n'
            '- 10 min: Questions and answers\n\n'
            'Prepared answers for likely questions:\n'
            '- "Should we delay launch?" — No. These are configuration changes, not architecture changes.\n'
            '- "What if platforms reject our prices?" — Prices are built within the platform tier systems. '
            'Risk is low if compliance was verified.\n'
            '- "Can we change prices later?" — Downward only. Launch high, adjust downward. '
            'Raising prices post-launch on console is practically impossible.\n\n'
            'Follow-up within 24 hours: summary email, adjusted recommendations if any came up in the call, '
            'confirmed next steps.'
        ),
        'status': 'Not Started', 'priority': 'High', 'hours_estimated': 4,
        'start_date': '2026-04-27', 'due_date': '2026-04-27',
        'success_factor': (
            'Call completed. Follow-up email sent within 24 hours. '
            'Phase 2 interest level documented. Any adjusted recommendations captured in writing.'
        ),
        'notes': 'Day 6 — DEADLINE DAY (27 April 2026).',
    },
]


def build_worksage_import(wb):
    ws = wb.create_sheet('WorkSage Import')

    project_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    feature_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    story_fill   = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    task_fill    = WHITE_FILL
    type_fills   = {'project': project_fill, 'feature': feature_fill,
                    'story': story_fill, 'task': task_fill}

    headers = [
        ('A', '_temp_id', 12),
        ('B', '_temp_parent_id', 16),
        ('C', 'item_type', 12),
        ('D', 'title', 55),
        ('E', 'description', 120),
        ('F', 'status', 14),
        ('G', 'priority', 12),
        ('H', 'hours_estimated', 12),
        ('I', 'assignees', 14),
        ('J', 'client_id', 40),
        ('K', 'practice_area', 14),
        ('L', 'start_date', 14),
        ('M', 'due_date', 14),
        ('N', 'success_factor', 90),
        ('O', 'notes', 35),
    ]
    for col_letter, _, width in headers:
        set_col_width(ws, col_letter, width)

    # Title
    ws.merge_cells('A1:O1')
    ws['A1'] = 'WorkSage Import — Goals Studio Pricing Engagement'
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(vertical='center')
    ws.row_dimensions[1].height = 30

    # Instructions
    ws.merge_cells('A2:O2')
    ws['A2'] = (
        'Each row maps to one WorkSage item. Hierarchy: Project > Feature > Story > Task. '
        'The _temp_parent_id column links each child to its parent. '
        'All assignees fields are intentionally empty — assign within WorkSage after import. '
        'All dates are within the 27 April 2026 deadline.'
    )
    ws['A2'].font = Font(name='Calibri', size=9, italic=True, color='666666')
    ws['A2'].alignment = WRAP
    ws.row_dimensions[2].height = 30

    # Headers row 3
    for col_letter, label, _ in headers:
        c = ws[f'{col_letter}3']
        c.value = label
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = CENTER
        c.border = THIN
    ws.row_dimensions[3].height = 25

    # Data
    row_num = 4
    for item in WS_ROWS:
        fill = type_fills.get(item['item_type'], task_fill)
        values = [
            item['_temp_id'],
            item['_temp_parent_id'],
            item['item_type'],
            item['title'],
            item['description'],
            item['status'],
            item.get('priority', ''),
            item['hours_estimated'],
            '',               # assignees — always empty
            CLIENT_ID,
            'gaming',
            item.get('start_date', ''),
            item.get('due_date', ''),
            item.get('success_factor', ''),
            item.get('notes', ''),
        ]
        for col_idx, val in enumerate(values, 1):
            c = ws.cell(row=row_num, column=col_idx, value=val)
            c.font = BODY_FONT
            c.alignment = WRAP
            c.border = THIN
            c.fill = fill
        ws.row_dimensions[row_num].height = 30
        row_num += 1

    ws.freeze_panes = 'A4'
    ws.auto_filter.ref = f'A3:O{row_num - 1}'
    return ws, row_num - 4


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == '__main__':
    import os
    os.makedirs('Clients/Goals', exist_ok=True)

    wb = openpyxl.Workbook()
    # Remove default sheet
    del wb['Sheet']

    build_overview(wb)
    build_task_tracker(wb)
    build_timeline(wb)
    build_risk_register(wb)
    build_deliverable_structure(wb)
    ws_import, ws_rows = build_worksage_import(wb)

    wb.save(OUTPUT)
    print(f'Saved: {OUTPUT}')
    print(f'Sheets: {wb.sheetnames}')
    for name in wb.sheetnames:
        ws = wb[name]
        print(f'  {name}: {ws.max_row} rows, {ws.max_column} cols')
    print(f'WorkSage Import: {ws_rows} data rows')
