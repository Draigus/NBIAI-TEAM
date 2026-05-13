"""Add a WorkSage Import tab to the Goals pricing Excel — all fields, full hierarchy."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

XLSX = 'Clients/Goals/goals_pricing_engagement_plan_v3.xlsx'
CLIENT_ID = '6975460f-c302-42c5-a586-1d04c5fcb929'

wb = openpyxl.load_workbook(XLSX)
if 'WorkSage Import' in wb.sheetnames:
    del wb['WorkSage Import']
ws = wb.create_sheet('WorkSage Import')

# Styles
hdr_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
hdr_fill = PatternFill(start_color='1F3864', end_color='1F3864', fill_type='solid')
title_font = Font(name='Calibri', size=14, bold=True, color='1F3864')
wrap = Alignment(wrap_text=True, vertical='top')
thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)
project_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
feature_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
story_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
task_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
type_fills = {'project': project_fill, 'feature': feature_fill, 'story': story_fill, 'task': task_fill}

# Title row
ws.merge_cells('A1:O1')
ws['A1'] = 'WorkSage Import — Goals Studio Pricing Engagement'
ws['A1'].font = title_font
ws['A1'].alignment = Alignment(vertical='center')
ws.row_dimensions[1].height = 30

# Instructions row
ws.merge_cells('A2:O2')
ws['A2'] = 'Each row maps to one WorkSage item. Hierarchy: Project > Feature > Story > Task. Parent ID column links child to parent. Copy data rows (4+) into the WorkSage bulk import or use import_goals_project.py.'
ws['A2'].font = Font(name='Calibri', size=9, italic=True, color='666666')
ws['A2'].alignment = wrap

# Headers
headers = [
    ('A', '_temp_id', 12),
    ('B', '_temp_parent_id', 16),
    ('C', 'item_type', 12),
    ('D', 'title', 55),
    ('E', 'description', 120),
    ('F', 'status', 14),
    ('G', 'priority', 12),
    ('H', 'hours_estimated', 10),
    ('I', 'assignees', 14),
    ('J', 'client_id', 40),
    ('K', 'practice_area', 14),
    ('L', 'start_date', 14),
    ('M', 'due_date', 14),
    ('N', 'success_factor', 90),
    ('O', 'notes', 35),
]

for col_letter, header, width in headers:
    cell = ws[f'{col_letter}3']
    cell.value = header
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border
    ws.column_dimensions[col_letter].width = width

ws.row_dimensions[3].height = 25

# Data — complete hierarchy
rows_data = [
    # PROJECT
    {
        '_temp_id': 'P1',
        '_temp_parent_id': '',
        'item_type': 'project',
        'title': 'Goals Studio Pricing Engagement',
        'description': (
            'Deliver competitive pricing benchmark, regional pricing strategy, and risk assessment '
            'for GOALS F2P football game launching 14 May 2026.\n\n'
            'Budget: 100,000 SEK (~$10K USD) | ~116 billable hours\n'
            'Client: Goals AB (Goals Studio), Stockholm, Sweden\n'
            'Contacts: Jonas Rundberg (CEO), Julius (Live Ops)\n'
            'NBI team: To be assigned by team lead\n\n'
            'SOW scope:\n'
            '- Review competitive pricing positioning in sports/football game space\n'
            '- Price point selection, currency pack structure, conversion rate guidance\n'
            '- Regional pricing for base and regional markets\n'
            '- Assessment of current pricing plans against market norms\n'
            '- Recommendations on adjustments and areas of risk\n'
            '- Deliverable: Written summary (12-18 pages) + remote review call\n\n'
            'Architecture: Foundation (map HC economy) > Competitive benchmark > Regional analysis > '
            'Synthesis (recommendations + scenarios) > Deliverable\n\n'
            'Hard deadline: Must deliver by 27 April before Goals submits pricing to PS/Xbox\n'
            "Key question from Jonas: \"Are we pricing too low?\"\n"
            "Glen's principle: Launch high, patch down. Never raise prices post-launch."
        ),
        'status': 'In progress',
        'priority': 'High',
        'hours_estimated': 116,
        'assignees': '',
        'practice_area': 'gaming',
        'start_date': '2026-04-22',
        'due_date': '2026-04-27',
        'success_factor': '',
        'notes': 'Feature 4 is CONDITIONAL - only if F1-3 complete within budget',
    },

    # FEATURE 1
    {
        '_temp_id': 'F1', '_temp_parent_id': 'P1',
        'item_type': 'feature',
        'title': 'F1: Competitive HC Pricing Benchmark',
        'description': "Show Goals where they sit relative to the sports/football F2P market. Extend their existing competitive analysis with insights their internal position doesn't give them.\n\nWhat NBI adds: Independent cross-genre validation, pricing psychology assessment, conversion rate benchmarks, and the confidence that an external expert agrees (or disagrees) with their conclusions.",
        'status': 'Not started', 'priority': 'High', 'hours_estimated': 38,
        'assignees': '', 'practice_area': 'gaming',
        'start_date': '2026-04-22', 'due_date': '2026-04-23',
        'success_factor': '', 'notes': '33% of total budget',
    },

    # Story 1.0
    {
        '_temp_id': 'S1.0', '_temp_parent_id': 'F1',
        'item_type': 'story',
        'title': "S1.0: Foundation - Map the HC Economy",
        'description': "Before comparing prices, establish exactly what Goals' HC buys and what the player value chain looks like. Prerequisite for everything else.",
        'status': 'Not started', 'priority': 'High', 'hours_estimated': 8,
        'assignees': '', 'practice_area': 'gaming',
        'start_date': '2026-04-22', 'due_date': '2026-04-22',
        'success_factor': '', 'notes': '',
    },
    {
        '_temp_id': 'T1.0.1', '_temp_parent_id': 'S1.0',
        'item_type': 'task',
        'title': 'T1.0.1: Map HC Value Chain',
        'description': "Document the complete path from real money to in-game value in Goals.\n\nKey mapping: 1 HC coin = 50 points = $0.012 USD at base rate (25,000 points = 500 coins = $6.00 USD from Pricing Model Conversion rates sheet).\n\nMap what each HC tier ACTUALLY buys:\n- 380 HC ($5.99): 1 Internal Kit (333 coins) with 47 coins leftover\n- 650 HC ($9.99): 1 Internal Kit + 317 leftover (not enough for second kit)\n- Map all item categories with HC costs\n\nGAP: Mech and Pulse kit HC pricing not documented in client materials. Confirm with Julius.\n\nIdentify minimum meaningful purchase - cheapest satisfying thing a new player can buy.\n\nSources: Goals Studio Monetisation Design Doc, Pricing Model, player_pricing.md",
        'status': 'Not started', 'priority': 'High', 'hours_estimated': 3,
        'assignees': '', 'practice_area': 'gaming',
        'start_date': '2026-04-22', 'due_date': '2026-04-22',
        'success_factor': "One-page document exists showing all 6 HC tiers and what each one buys. Any unconfirmed prices are clearly flagged. Glen can explain the value chain to Jonas in plain English.",
        'notes': 'Day 1',
    },
    {
        '_temp_id': 'T1.0.2', '_temp_parent_id': 'S1.0',
        'item_type': 'task',
        'title': 'T1.0.2: Establish Payer Persona Benchmarks',
        'description': "Benchmark Goals' 3-persona model against industry norms.\n\nGoals targets (from Pricing Model Overview):\n- Persona 1 (Core FUT, every day): 20% of players, 30% pay, ARPPU $46\n- Persona 2 (Dedicated, 2-3x/week): 35%, 10% pay, ARPPU $15.81\n- Persona 3 (Casual, few times/month): 45%, 2% pay, ARPPU $7.24\n- Overall: 10.4% payer rate, ARPU $3.38, ARPPU $32.48\n\nResearch method: Pull published conversion and ARPPU from eFootball earnings (Konami quarterly), EA FC filings, Sensor Tower/data.ai, GDC presentations.\n\nGoals own Gemini deep research shows lower figures. Verify source.\n\nCritical question: Is 10.4% month-1 or steady-state? Plan doesn't specify.\n\nSeasonless economy: No resets/battle pass/FOMO means HC ladder must do ALL conversion work.",
        'status': 'Not started', 'priority': 'High', 'hours_estimated': 5,
        'assignees': '', 'practice_area': 'gaming',
        'start_date': '2026-04-22', 'due_date': '2026-04-22',
        'success_factor': "Written assessment with at least 2 cited sources comparing Goals' assumptions to real-world data. Clear verdict: realistic, aggressive, or fantasy. If aggressive, state what the implication is for pricing.",
        'notes': 'Day 1',
    },

    # Story 1.1
    {
        '_temp_id': 'S1.1', '_temp_parent_id': 'F1',
        'item_type': 'story',
        'title': 'S1.1: Competitive Pricing Position',
        'description': "Show Goals exactly where their HC pricing sits relative to market - by synthesising their existing competitive data into a clear positioning statement they can't write themselves.",
        'status': 'Not started', 'priority': 'High', 'hours_estimated': 19,
        'assignees': '', 'practice_area': 'gaming',
        'start_date': '2026-04-22', 'due_date': '2026-04-23',
        'success_factor': '', 'notes': '',
    },
    {
        '_temp_id': 'T1.1.1', '_temp_parent_id': 'S1.1',
        'item_type': 'task',
        'title': 'T1.1.1: Build Normalised Price Position Map',
        'description': "Create position map showing Goals' effective USD/HC against competitive range at each tier.\n\nGoals data (from Pricing Model Softhard currency comparison):\n- At entry ($5-6): Goals = $0.01576/HC\n- At mid ($20): Goals = $0.01454/HC\n- At top ($100): Goals = $0.01290/HC\n\nFor each tier: calculate market minimum, median, maximum, and Goals' position as percentile.\n\nSpot-check 2-3 competitor prices to confirm Goals' data is still current.",
        'status': 'Not started', 'priority': 'High', 'hours_estimated': 8,
        'assignees': '', 'practice_area': 'gaming',
        'start_date': '2026-04-22', 'due_date': '2026-04-22',
        'success_factor': "Position map exists for all 6 tiers. Each tier shows: Goals price, market range (min/median/max), Goals percentile. One clear sentence summarising overall position.",
        'notes': 'Day 1. Source: Goals Pricing Model',
    },
    {
        '_temp_id': 'T1.1.2', '_temp_parent_id': 'S1.1',
        'item_type': 'task',
        'title': 'T1.1.2: Discount Curve Comparison',
        'description': "Compare Goals' bonus/discount scaling against competitors.\n\nGoals curve: Tier 1 ($5.99) 0%, Tier 2 ($9.99) ~2.5%, Tier 3 ($19.99) ~7.7%, Tier 4 ($29.99) ~9.4%, Tier 5 ($49.99) ~15.4%, Tier 6 ($99.99) ~18.1%.\n\nCompetitor max discounts: EA FC ~39%, Fortnite ~24%, Apex ~13%. Goals at ~18%.\n\nAssess: Is 18% enough for whale incentive? Is tier-to-tier jump consistent? Is there a clear best-value tier?\n\nKey principle: New launches should be MORE conservative on discounts - can always increase later.",
        'status': 'Not started', 'priority': 'High', 'hours_estimated': 4,
        'assignees': '', 'practice_area': 'gaming',
        'start_date': '2026-04-22', 'due_date': '2026-04-22',
        'success_factor': "Comparison table showing discount curves for Goals + 4 competitors. Written recommendation: keep current curve, steepen it, or flatten it. With reasoning.",
        'notes': 'Day 1',
    },
    {
        '_temp_id': 'T1.1.3', '_temp_parent_id': 'S1.1',
        'item_type': 'task',
        'title': 'T1.1.3: SKU Structure and Entry Point Audit',
        'description': "Assess 6-tier HC ladder structural health: $5.99 / $9.99 / $19.99 / $29.99 / $49.99 / $99.99.\n\nCompare tier count: EA FC 8 tiers (starts $0.99), Fortnite 4 tiers (starts $7.99), Apex 6 tiers (starts $4.99).\n\nEntry point: Does $5.99 create first-purchase friction? What can 380 HC buy? Counter: $0.99 tier nets only $0.69 after 30% platform fee.\n\nStructural gaps: $9.99 to $19.99 (100% jump), $49.99 to $99.99 (100% jump) - standard in industry.\n\nWhale ceiling: $99.99 cap + no purchase limits = whale-friendly without optically aggressive pricing. Correct for anti-P2W positioning.",
        'status': 'Not started', 'priority': 'High', 'hours_estimated': 5,
        'assignees': '', 'practice_area': 'gaming',
        'start_date': '2026-04-23', 'due_date': '2026-04-23',
        'success_factor': "Written audit covering: entry point assessment, tier gap analysis, whale ceiling analysis. Clear recommendation for each: keep, add tier, or adjust. Supported by competitor comparison.",
        'notes': 'Day 2',
    },
    {
        '_temp_id': 'T1.1.4', '_temp_parent_id': 'S1.1',
        'item_type': 'task',
        'title': 'T1.1.4: Platform Fee Impact and Net Revenue Table',
        'description': "Calculate actual net revenue per SKU per platform.\n\nTable: SKU | Gross | Steam Net (30%) | Epic Net (12%) | PS/Xbox Net (30%)\n380 HC: $5.99 | $4.19 | $5.27 | $4.19\n650 HC: $9.99 | $6.99 | $8.79 | $6.99\n1,375 HC: $19.99 | $14.00 | $17.59 | $14.00\n2,100 HC: $29.99 | $21.00 | $26.39 | $21.00\n3,750 HC: $49.99 | $35.00 | $43.99 | $35.00\n7,750 HC: $99.99 | $70.00 | $87.99 | $70.00\n\nEpic advantage: ~26% more per transaction vs Steam/PS/Xbox.\nVerify alignment with PS wholesale rates in Goals matrix.",
        'status': 'Not started', 'priority': 'Medium', 'hours_estimated': 2,
        'assignees': '', 'practice_area': 'gaming',
        'start_date': '2026-04-23', 'due_date': '2026-04-23',
        'success_factor': "Complete net revenue table (6 tiers x 4 platforms = 24 cells). Any concerns about margins flagged. Epic advantage quantified as a percentage.",
        'notes': 'Day 2',
    },

    # Story 1.2
    {
        '_temp_id': 'S1.2', '_temp_parent_id': 'F1',
        'item_type': 'story',
        'title': 'S1.2: Conversion Rate and Elasticity Context',
        'description': 'Provide conversion/revenue modelling context that turns a static price comparison into actionable strategy.',
        'status': 'Not started', 'priority': 'High', 'hours_estimated': 8,
        'assignees': '', 'practice_area': 'gaming',
        'start_date': '2026-04-23', 'due_date': '2026-04-23',
        'success_factor': '', 'notes': '',
    },
    {
        '_temp_id': 'T1.2.1', '_temp_parent_id': 'S1.2',
        'item_type': 'task',
        'title': 'T1.2.1: Price Elasticity Scenario Modelling',
        'description': "Model 4 scenarios using Goals' revenue model (805,250 MAU, 10.4% payer rate, $32.48 ARPPU = $2.72M total).\n\nScenario A: Raise all prices 10% - assume conversion drops 5-15%\nScenario B: Raise all prices 20% - assume conversion drops 10-25%\nScenario C: Raise entry tier only ($5.99 to $6.99) - entry conversion drops 5-10%, mid/high unaffected\nScenario D: Keep prices, steepen discount curve to 25% max at top tier\n\nPresent as impact ranges (conservative/optimistic), not point estimates.\n\nKey insight: If 10.4% conversion is optimistic for month 1, lowering prices won't help - need time and content. Launch high, patch down.",
        'status': 'Not started', 'priority': 'High', 'hours_estimated': 8,
        'assignees': '', 'practice_area': 'gaming',
        'start_date': '2026-04-23', 'due_date': '2026-04-23',
        'success_factor': "4 scenarios modelled with revenue ranges. Each shows: price change, assumed conversion impact, net revenue change. Clear recommendation on which scenario is best for launch.",
        'notes': 'Day 2',
    },

    # Story 1.3
    {
        '_temp_id': 'S1.3', '_temp_parent_id': 'F1',
        'item_type': 'story',
        'title': 'S1.3: Promotional and First-Purchase Pricing',
        'description': 'Address gap in static HC pack analysis. First-purchase incentives and promotional pricing drive early conversion.',
        'status': 'Not started', 'priority': 'Medium', 'hours_estimated': 5,
        'assignees': '', 'practice_area': 'gaming',
        'start_date': '2026-04-23', 'due_date': '2026-04-23',
        'success_factor': '', 'notes': '',
    },
    {
        '_temp_id': 'T1.3.1', '_temp_parent_id': 'S1.3',
        'item_type': 'task',
        'title': 'T1.3.1: First-Purchase and Starter Pack Recommendations',
        'description': "Recommend complementary offers that improve conversion on core HC ladder.\n\n1. First Purchase Bonus: Double HC on first-ever purchase at any tier. Research what EA FC, Fortnite, Apex, eFootball currently offer. Do NOT cite conversion improvement percentages without a source.\n\n2. Limited Starter Pack: One-time $3.99 offer (below $5.99 entry) containing 500 HC + Rare+ player + exclusive kit. Breaks first-purchase barrier.\n\n3. World Cup Launch Bundle: Tied to June event, limited availability, higher value/dollar. Creates FOMO without permanently discounting.\n\nFrame as complementary to core pricing (not scope creep).",
        'status': 'Not started', 'priority': 'Medium', 'hours_estimated': 5,
        'assignees': '', 'practice_area': 'gaming',
        'start_date': '2026-04-23', 'due_date': '2026-04-23',
        'success_factor': "2-3 specific first-purchase recommendations. Each includes: what it is, what it costs Goals, expected benefit. Clearly framed as additional to the core pricing, not a change to it.",
        'notes': 'Day 2. Value-add observation, not core scope.',
    },

    # FEATURE 2
    {
        '_temp_id': 'F2', '_temp_parent_id': 'P1',
        'item_type': 'feature',
        'title': 'F2: Regional Pricing Strategy',
        'description': "Validate Goals' 40+ country pricing matrix. NBI adds: competitive regional context, PPP-adjusted analysis, and platform compliance verification they can't get from their own spreadsheet.\n\nGoals built their matrix using exchange rates and some PPP data. They CANNOT benchmark their regional pricing against what EA FC actually charges in Turkey or Brazil. NBI provides this.",
        'status': 'Not started', 'priority': 'High', 'hours_estimated': 34,
        'assignees': '', 'practice_area': 'gaming',
        'start_date': '2026-04-23', 'due_date': '2026-04-25',
        'success_factor': '', 'notes': '29% of total budget',
    },

    # Story 2.1
    {
        '_temp_id': 'S2.1', '_temp_parent_id': 'F2',
        'item_type': 'story',
        'title': 'S2.1: Regional Deviation and Risk Identification',
        'description': "Flag countries where Goals' pricing creates material risk (too high = player backlash, too low = revenue leakage, misaligned = platform rejection).",
        'status': 'Not started', 'priority': 'High', 'hours_estimated': 23,
        'assignees': '', 'practice_area': 'gaming',
        'start_date': '2026-04-23', 'due_date': '2026-04-24',
        'success_factor': '', 'notes': '',
    },
    {
        '_temp_id': 'T2.1.1', '_temp_parent_id': 'S2.1',
        'item_type': 'task',
        'title': 'T2.1.1: Categorise All Countries by Risk Level',
        'description': "Using Goals' existing deviation data, categorise every country:\n\nGreen: Deviation <15% from FX AND Goals matrix marks YES\nAmber: Deviation 15-25% OR matrix marks ! or ? OR misaligned with competitor pricing\nRed: Deviation >25% OR matrix recommends a price move OR large player market with >15% over-pricing\n\nRed candidates: Poland (+25%), Switzerland (+26%), Ukraine (-26%), Turkey, Brazil\nAmber candidates: South Africa, Czech Republic, Norway, Denmark\nEUR zone: Same price across all EUR countries but PPP varies (Germany vs Portugal vs Greece)\n\nCross-reference with beta community data: UK 1,274, France 1,026, Turkey 667, Brazil 495, Poland 392.",
        'status': 'Not started', 'priority': 'High', 'hours_estimated': 8,
        'assignees': '', 'practice_area': 'gaming',
        'start_date': '2026-04-23', 'due_date': '2026-04-24',
        'success_factor': "Every country has a colour rating. Red countries have a one-sentence explanation of the problem. Output is a ranked list with highest-risk markets at the top.",
        'notes': 'Day 2-3',
    },
    {
        '_temp_id': 'T2.1.2', '_temp_parent_id': 'S2.1',
        'item_type': 'task',
        'title': 'T2.1.2: Competitor Regional Pricing for Red/Amber Markets',
        'description': "For 5-10 red/amber countries, research EA FC and Fortnite actual pricing.\n\nSources: PS Store regional web store, SteamDB historical/current, Epic Games Store.\n\nFor each market build comparison: Country | Goals proposed | EA FC actual | Fortnite actual | Market median | Goals position\n\nPriority markets:\n- Brazil: Large player base, low PPP, critical for WC engagement\n- Turkey: 5th largest beta community, extreme PPP differential\n- Poland: Large beta community, +25% above FX\n- Russia/CIS: Extreme currency volatility\n\nFallback if stores region-locked: SteamDB for Steam, documented third-party sources for PS/Xbox.",
        'status': 'Not started', 'priority': 'High', 'hours_estimated': 15,
        'assignees': '', 'practice_area': 'gaming',
        'start_date': '2026-04-24', 'due_date': '2026-04-24',
        'success_factor': "Comparison table for all red/amber countries. Each row: country, Goals price, EA FC price, Fortnite price, percentage difference. Sources cited for every price.",
        'notes': 'Day 3. Largest single research task. Overrun absorbed from F4.',
    },

    # Story 2.2
    {
        '_temp_id': 'S2.2', '_temp_parent_id': 'F2',
        'item_type': 'story',
        'title': 'S2.2: Platform Compliance and Submission Guidance',
        'description': "Ensure Goals' pricing will pass PS/Xbox review without rejection or delay.",
        'status': 'Not started', 'priority': 'Critical', 'hours_estimated': 3,
        'assignees': '', 'practice_area': 'gaming',
        'start_date': '2026-04-24', 'due_date': '2026-04-24',
        'success_factor': '', 'notes': '',
    },
    {
        '_temp_id': 'T2.2.1', '_temp_parent_id': 'S2.2',
        'item_type': 'task',
        'title': 'T2.2.1: Verify Platform Tier Compliance',
        'description': "Cross-reference Goals' proposed prices against valid PS/Xbox tier selections.\n\nSources: Goals' PS4/PS5 Digital sheet, Xbox wholesale pricing sheet.\nFlag any non-standard tier prices (Sony/Xbox will reject).\n\nSubmission checklist:\n- All prices must be valid platform tiers\n- All territories covered where game is available\n- USD fallback for unsupported currencies\n- Review period: ~9 days (from Julius kickoff call)\n\nTimeline: deliver by 27 April > Goals submits by ~5 May > 9-day review > 14 May launch.\n\nAsk Julius: has a test pricing configuration been submitted previously?",
        'status': 'Not started', 'priority': 'Critical', 'hours_estimated': 3,
        'assignees': '', 'practice_area': 'gaming',
        'start_date': '2026-04-24', 'due_date': '2026-04-24',
        'success_factor': "Written confirmation: all prices are valid platform tiers, OR a list of specific prices that need adjustment with the nearest valid tier for each.",
        'notes': 'Day 3. Critical path - rejection delays launch.',
    },

    # Story 2.3
    {
        '_temp_id': 'S2.3', '_temp_parent_id': 'F2',
        'item_type': 'story',
        'title': 'S2.3: Country-Specific Recommendations',
        'description': 'Provide specific price adjustment recommendations for each red/amber country.',
        'status': 'Not started', 'priority': 'High', 'hours_estimated': 8,
        'assignees': '', 'practice_area': 'gaming',
        'start_date': '2026-04-24', 'due_date': '2026-04-25',
        'success_factor': '', 'notes': '',
    },
    {
        '_temp_id': 'T2.3.1', '_temp_parent_id': 'S2.3',
        'item_type': 'task',
        'title': 'T2.3.1: Write Country Recommendation Cards',
        'description': "For each red/amber country, write a recommendation card:\n\nCountry | Current proposed price | Issue | Recommended price | Rationale | Revenue impact | Risk if unchanged\n\nPrioritise by: player base size x deviation severity.\n\nGroup by action type:\n- Reduce price (over-priced markets)\n- Can increase (under-priced, revenue left on table)\n- Platform tier adjustment (needs rounding to valid tier)\n\nExpected output: 8-12 country cards.\n\nEach card must be self-contained - Julius can implement without reading the full report.",
        'status': 'Not started', 'priority': 'High', 'hours_estimated': 8,
        'assignees': '', 'practice_area': 'gaming',
        'start_date': '2026-04-24', 'due_date': '2026-04-25',
        'success_factor': "8-12 country cards completed. Each is self-contained. Julius can read one card and know exactly what to change in the pricing spreadsheet for that country.",
        'notes': 'Day 3-4',
    },

    # FEATURE 3
    {
        '_temp_id': 'F3', '_temp_parent_id': 'P1',
        'item_type': 'feature',
        'title': 'F3: Risk Assessment, Scenarios and Recommendations',
        'description': 'Synthesise all analysis into clear, specific, actionable guidance with modelled impact.',
        'status': 'Not started', 'priority': 'High', 'hours_estimated': 16,
        'assignees': '', 'practice_area': 'gaming',
        'start_date': '2026-04-25', 'due_date': '2026-04-25',
        'success_factor': '', 'notes': '14% of total budget',
    },

    # Story 3.1
    {
        '_temp_id': 'S3.1', '_temp_parent_id': 'F3',
        'item_type': 'story',
        'title': "S3.1: Are We Too Low? - The Central Answer",
        'description': "Directly address Jonas's core concern with evidence.",
        'status': 'Not started', 'priority': 'Critical', 'hours_estimated': 4,
        'assignees': '', 'practice_area': 'gaming',
        'start_date': '2026-04-25', 'due_date': '2026-04-25',
        'success_factor': '', 'notes': '',
    },
    {
        '_temp_id': 'T3.1.1', '_temp_parent_id': 'S3.1',
        'item_type': 'task',
        'title': 'T3.1.1: Build the Definitive Position Statement',
        'description': "Synthesise competitive position (T1.1.1), payer benchmarks (T1.0.2), and scenario modelling (T1.2.1) into one clear paragraph.\n\nFormat: \"Goals' base HC pricing is at the Xth percentile... This means... Given your 80/20 skill-to-purchase positioning, this is [appropriate/low/high] because... Your payer conversion target of 10.4% is [realistic/aggressive] for month 1... Our recommendation is [specific].\"\n\nTake a position. No hedging.\n\nCore principle: You can always reduce prices post-launch. You cannot raise them. Console platforms make price increases nearly impossible.",
        'status': 'Not started', 'priority': 'Critical', 'hours_estimated': 4,
        'assignees': '', 'practice_area': 'gaming',
        'start_date': '2026-04-25', 'due_date': '2026-04-25',
        'success_factor': "One paragraph exists that Jonas can read and immediately know whether to raise prices or not. Contains a specific recommendation with a number.",
        'notes': 'Day 4. Depends on T1.1.1, T1.0.2, T1.2.1 completion.',
    },

    # Story 3.2
    {
        '_temp_id': 'S3.2', '_temp_parent_id': 'F3',
        'item_type': 'story',
        'title': 'S3.2: Risk Register',
        'description': 'Document all identified pricing risks before launch, ranked by severity.',
        'status': 'Not started', 'priority': 'High', 'hours_estimated': 5,
        'assignees': '', 'practice_area': 'gaming',
        'start_date': '2026-04-25', 'due_date': '2026-04-25',
        'success_factor': '', 'notes': '',
    },
    {
        '_temp_id': 'T3.2.1', '_temp_parent_id': 'S3.2',
        'item_type': 'task',
        'title': 'T3.2.1: Compile and Rank Risks',
        'description': "Document each risk: Description | Severity | Likelihood | Markets Affected | Mitigation\n\n10 expected risks:\n1. P2W perception at high price points (Critical/Medium)\n2. Post-launch price increase impossibility (High/Medium)\n3. Over-pricing in price-sensitive markets - Turkey, Brazil, Poland (High/High)\n4. Platform pricing rejection (Critical/Low)\n5. Discount curve too shallow for launch (Medium/Medium)\n6. Missing entry-tier conversion at $5.99 (Medium/Medium)\n7. Originals/Celebrity pack pricing gap (Low/N/A)\n8. Netherlands regulatory risk for player packs (High/Medium)\n9. South Korea drop rate disclosure requirement (High/Certain if launching in KR)\n10. Non-seasonal demand sustainability (Medium/High)",
        'status': 'Not started', 'priority': 'High', 'hours_estimated': 5,
        'assignees': '', 'practice_area': 'gaming',
        'start_date': '2026-04-25', 'due_date': '2026-04-25',
        'success_factor': "Minimum 7 risks documented. Each has severity, likelihood, affected markets, and a mitigation action. Top 3 risks have specific, implementable steps Goals can take immediately.",
        'notes': 'Day 4',
    },

    # Story 3.3
    {
        '_temp_id': 'S3.3', '_temp_parent_id': 'F3',
        'item_type': 'story',
        'title': 'S3.3: Top 5 Recommendations',
        'description': 'The five highest-impact actions Goals should take before platform submission.',
        'status': 'Not started', 'priority': 'High', 'hours_estimated': 7,
        'assignees': '', 'practice_area': 'gaming',
        'start_date': '2026-04-25', 'due_date': '2026-04-25',
        'success_factor': '', 'notes': '',
    },
    {
        '_temp_id': 'T3.3.1', '_temp_parent_id': 'S3.3',
        'item_type': 'task',
        'title': 'T3.3.1: Write Top 5 Recommendations',
        'description': "Five specific numbered recommendations. Each must include:\n- What to change (exact specification)\n- Why (evidence from analysis)\n- Revenue impact (conservative-to-optimistic range)\n- Implementation (which cells in their matrix, which platform to resubmit)\n- Risk if skipped\n\nExpected candidates:\n1. Pricing position adjustment\n2. Regional correction (Poland/Turkey/Brazil)\n3. Discount curve adjustment\n4. First-purchase incentive\n5. Launch framing (label as 'Launch Pricing' to preserve future flexibility)\n\nMust pass test: would a studio economy designer respect this and act immediately?",
        'status': 'Not started', 'priority': 'High', 'hours_estimated': 7,
        'assignees': '', 'practice_area': 'gaming',
        'start_date': '2026-04-25', 'due_date': '2026-04-25',
        'success_factor': "5 recommendations written. Each has: exact change to make, evidence for why, expected revenue impact (range), and what happens if they skip it. A studio economy designer would read these and act same day.",
        'notes': 'Day 4',
    },

    # FEATURE 4
    {
        '_temp_id': 'F4', '_temp_parent_id': 'P1',
        'item_type': 'feature',
        'title': 'F4: Value-Add (Conditional - Time Permitting)',
        'description': "Supplementary tools demonstrating NBI's strategic depth and positioning Phase 2. ONLY executed if F1-3 complete within budget. Scope reduced first if any day overruns.",
        'status': 'Not started', 'priority': 'Medium', 'hours_estimated': 8,
        'assignees': '', 'practice_area': 'gaming',
        'start_date': '2026-04-25', 'due_date': '2026-04-25',
        'success_factor': '', 'notes': '7% of total budget. CONDITIONAL.',
    },

    # Story 4.1
    {
        '_temp_id': 'S4.1', '_temp_parent_id': 'F4',
        'item_type': 'story',
        'title': 'S4.1: Reusable Regional Pricing Methodology',
        'description': 'One-page methodology Goals can use for any future SKU without NBI.',
        'status': 'Not started', 'priority': 'Medium', 'hours_estimated': 3,
        'assignees': '', 'practice_area': 'gaming',
        'start_date': '2026-04-25', 'due_date': '2026-04-25',
        'success_factor': '', 'notes': 'Conditional on F1-3 completion',
    },
    {
        '_temp_id': 'T4.1.1', '_temp_parent_id': 'S4.1',
        'item_type': 'task',
        'title': "T4.1.1: Write How to Price a New SKU Guide",
        'description': "6-step methodology:\n1. Set USD base price (anchor against EA FC equivalent)\n2. Convert to local currencies using current FX\n3. Check PPP factor (provide lookup table)\n4. Compare against EA FC in that market (+/-15% acceptable)\n5. Round to nearest platform tier\n6. Verify discount curve is maintained\n\nInclude worked example: pricing a new Celebration at $7.00 USD for Brazil.\nInclude PPP factor table for all launch markets.\nPosition as: \"This is the manual method. Phase 2 builds a model that does this automatically.\"",
        'status': 'Not started', 'priority': 'Medium', 'hours_estimated': 3,
        'assignees': '', 'practice_area': 'gaming',
        'start_date': '2026-04-25', 'due_date': '2026-04-25',
        'success_factor': "One-page guide exists. Julius can follow it without NBI help. Worked example included. PPP lookup table for all launch markets included.",
        'notes': 'Day 4. Conditional.',
    },

    # Story 4.2
    {
        '_temp_id': 'S4.2', '_temp_parent_id': 'F4',
        'item_type': 'story',
        'title': 'S4.2: World Cup Pricing Opportunity Notes',
        'description': '3-5 tactical pricing observations for the June World Cup window.',
        'status': 'Not started', 'priority': 'Low', 'hours_estimated': 2,
        'assignees': '', 'practice_area': 'gaming',
        'start_date': '2026-04-25', 'due_date': '2026-04-25',
        'success_factor': '', 'notes': 'Conditional on F1-3 completion',
    },
    {
        '_temp_id': 'T4.2.1', '_temp_parent_id': 'S4.2',
        'item_type': 'task',
        'title': 'T4.2.1: Write World Cup Quick-Hit Observations',
        'description': "3-5 tactical pricing observations for June World Cup:\n\n- National kit pricing: premium during group stage (emotional spending), discount in knockout (losing teams = fire sale)\n- Time-limited World Cup Coin Pack at slightly better value (e.g., 10% bonus) - drives urgency\n- Research what published data exists on revenue spikes during major sporting events in comparable F2P titles. Do NOT cite multiplier ranges without a verifiable source.\n\nKeep SHORT. Half page max. Teaser for Phase 2 live service consulting, not a full plan.",
        'status': 'Not started', 'priority': 'Low', 'hours_estimated': 2,
        'assignees': '', 'practice_area': 'gaming',
        'start_date': '2026-04-25', 'due_date': '2026-04-25',
        'success_factor': "3-5 bullet points. Each is a specific, actionable pricing tactic for the World Cup window. Clearly positioned as a taste of what Phase 2 would cover in detail.",
        'notes': 'Day 4. Conditional. Half page max.',
    },

    # Story 4.3
    {
        '_temp_id': 'S4.3', '_temp_parent_id': 'F4',
        'item_type': 'story',
        'title': 'S4.3: Phase 2 Positioning',
        'description': 'Brief section positioning the larger engagement.',
        'status': 'Not started', 'priority': 'Medium', 'hours_estimated': 3,
        'assignees': '', 'practice_area': 'gaming',
        'start_date': '2026-04-25', 'due_date': '2026-04-25',
        'success_factor': '', 'notes': 'Conditional on F1-3 completion',
    },
    {
        '_temp_id': 'T4.3.1', '_temp_parent_id': 'S4.3',
        'item_type': 'task',
        'title': "T4.3.1: Write What's Next Section",
        'description': "Position Phase 2 naturally (helpful, not salesy):\n\n- Post-launch telemetry review (2-3 weeks after 14 May): real data replaces estimates\n- A/B test design: optimise conversion at each tier using live cohorts\n- Store optimisation: offer sequencing, time-limited promotion design\n- World Cup monetisation strategy: full tactical plan for June peak\n- Ongoing live service consulting: content cadence, engagement frameworks\n\nClose with: \"We recommend a check-in call 2-3 weeks post-launch to review initial transaction data and scope the next phase.\"\n\nDo NOT include NBI internal pricing for Phase 2.",
        'status': 'Not started', 'priority': 'Medium', 'hours_estimated': 3,
        'assignees': '', 'practice_area': 'gaming',
        'start_date': '2026-04-25', 'due_date': '2026-04-25',
        'success_factor': "Half-page section written. Reads as genuine advice, not upselling. Recommends a check-in call 2-3 weeks post-launch. Does NOT include NBI pricing.",
        'notes': 'Day 4. Conditional. Tom leads.',
    },

    # FEATURE 5
    {
        '_temp_id': 'F5', '_temp_parent_id': 'P1',
        'item_type': 'feature',
        'title': 'F5: Deliverable Assembly and Client Review',
        'description': 'Package all analysis into the contracted deliverable: written summary (12-18 pages) + remote review call.',
        'status': 'Not started', 'priority': 'High', 'hours_estimated': 18,
        'assignees': '', 'practice_area': 'gaming',
        'start_date': '2026-04-26', 'due_date': '2026-04-27',
        'success_factor': '', 'notes': '16% of total budget',
    },

    # Story 5.1
    {
        '_temp_id': 'S5.1', '_temp_parent_id': 'F5',
        'item_type': 'story',
        'title': 'S5.1: Written Summary Document',
        'description': 'Create the final client-facing document: 12-18 pages, British English, numbers-backed.',
        'status': 'Not started', 'priority': 'High', 'hours_estimated': 10,
        'assignees': '', 'practice_area': 'gaming',
        'start_date': '2026-04-26', 'due_date': '2026-04-26',
        'success_factor': '', 'notes': '',
    },
    {
        '_temp_id': 'T5.1.1', '_temp_parent_id': 'S5.1',
        'item_type': 'task',
        'title': 'T5.1.1: Assemble the Written Summary',
        'description': "Final client-facing document structure:\n1. Executive Summary (1 page) - headline finding, top 5 recs, immediate action\n2. Your Competitive Position - where Goals sits vs market (F1)\n3. Regional Pricing Assessment - country risk map + recommendation cards (F2)\n4. Pricing Scenarios - what-if modelling with revenue impact (S1.2)\n5. Risk Register - all risks with mitigation (S3.2)\n6. Recommendations - top 5 specific changes (S3.3)\n7. Pricing Framework - reusable methodology (F4, if completed)\n8. World Cup Opportunities - tactical observations (F4, if completed)\n9. Next Steps - Phase 2 positioning\n10. Appendix - full comparison tables, country matrix, competitor data\n\nStyle: British English. No em dashes. Direct, specific, numbers-backed. 12-18 pages.\nWritten for Jonas (strategic) and Julius (practitioner) simultaneously.",
        'status': 'Not started', 'priority': 'High', 'hours_estimated': 10,
        'assignees': '', 'practice_area': 'gaming',
        'start_date': '2026-04-26', 'due_date': '2026-04-26',
        'success_factor': "Complete document draft ready for internal review. All sections filled with real findings (no placeholders). Internally consistent. Written in British English, direct and numbers-backed.",
        'notes': 'Day 5',
    },

    # Story 5.2
    {
        '_temp_id': 'S5.2', '_temp_parent_id': 'F5',
        'item_type': 'story',
        'title': 'S5.2: Internal Review and Polish',
        'description': 'Glen and Tom review before client delivery.',
        'status': 'Not started', 'priority': 'High', 'hours_estimated': 4,
        'assignees': '', 'practice_area': 'gaming',
        'start_date': '2026-04-26', 'due_date': '2026-04-26',
        'success_factor': '', 'notes': '',
    },
    {
        '_temp_id': 'T5.2.1', '_temp_parent_id': 'S5.2',
        'item_type': 'task',
        'title': 'T5.2.1: Quality Gate - Internal Review',
        'description': "Glen reviews: domain accuracy, tone (expertise without condescension), completeness (answers Jonas's core questions).\n\nTom reviews: commercial positioning, NBI brand consistency, contractual/legal issues.\n\nAddress all feedback. Produce final version.\nSend to Jonas/Julius 24h before review call.",
        'status': 'Not started', 'priority': 'High', 'hours_estimated': 4,
        'assignees': '', 'practice_area': 'gaming',
        'start_date': '2026-04-26', 'due_date': '2026-04-26',
        'success_factor': "Final document approved by both Glen and Tom. Sent to client 24h before review call. Call scheduled.",
        'notes': 'Day 5',
    },

    # Story 5.3
    {
        '_temp_id': 'S5.3', '_temp_parent_id': 'F5',
        'item_type': 'story',
        'title': 'S5.3: Remote Review Session',
        'description': '30-45 minute call walking Jonas/Julius through findings and recommendations.',
        'status': 'Not started', 'priority': 'High', 'hours_estimated': 4,
        'assignees': '', 'practice_area': 'gaming',
        'start_date': '2026-04-27', 'due_date': '2026-04-27',
        'success_factor': '', 'notes': '',
    },
    {
        '_temp_id': 'T5.3.1', '_temp_parent_id': 'S5.3',
        'item_type': 'task',
        'title': 'T5.3.1: Conduct Client Review Call',
        'description': "Call agenda (30-45 min):\n- 5 min: Recap scope and methodology\n- 10 min: Top 5 recommendations (highest impact first)\n- 10 min: Regional pricing highlights (red countries, what to change)\n- 5 min: Pricing framework overview\n- 5 min: Next steps / Phase 2\n- 10 min: Q&A\n\nPrepared answers:\n- \"Should we delay?\" - No. Configuration changes, not architecture.\n- \"What if platforms reject?\" - Built within tier systems. Low risk.\n- \"Can we change later?\" - Down only. Launch high, patch down.\n\nFollow-up within 24h: summary email, adjusted recs, confirmed next steps.",
        'status': 'Not started', 'priority': 'High', 'hours_estimated': 4,
        'assignees': '', 'practice_area': 'gaming',
        'start_date': '2026-04-27', 'due_date': '2026-04-27',
        'success_factor': "Call completed. Follow-up email sent within 24 hours. Phase 2 interest level documented. Any adjusted recommendations captured.",
        'notes': 'Day 6. Deadline day.',
    },
]

# Write to sheet
row_num = 4
for item in rows_data:
    fill = type_fills.get(item['item_type'], task_fill)
    values = [
        item['_temp_id'],
        item['_temp_parent_id'],
        item['item_type'],
        item['title'],
        item['description'],
        item['status'],
        item.get('priority', ''),
        item['hours_estimated'],
        item.get('assignees', ''),
        CLIENT_ID,
        item.get('practice_area', ''),
        item.get('start_date', ''),
        item.get('due_date', ''),
        item.get('success_factor', ''),
        item.get('notes', ''),
    ]
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=val)
        cell.font = Font(name='Calibri', size=10)
        cell.alignment = wrap
        cell.border = thin_border
        cell.fill = fill
    ws.row_dimensions[row_num].height = 30
    row_num += 1

# Freeze panes
ws.freeze_panes = 'A4'

# Auto-filter
ws.auto_filter.ref = f'A3:O{row_num - 1}'

wb.save(XLSX)
print(f'Added WorkSage Import sheet with {row_num - 4} rows to {XLSX}')
