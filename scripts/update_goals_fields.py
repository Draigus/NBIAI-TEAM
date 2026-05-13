"""Update Goals Studio tasks with success_factor, practice_area, and timeline dates. Then export Excel."""
import json
import sys
try:
    import requests
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'requests'])
    import requests

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl'])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = 'http://localhost:8888'

resp = requests.post(f'{BASE}/api/auth/login', json={'username': 'glen', 'password': 'nbi2026'})
TOKEN = resp.json()['token']
HEADERS = {'Authorization': f'Bearer {TOKEN}', 'Content-Type': 'application/json'}

# Fetch all tasks
resp = requests.get(f'{BASE}/api/tasks?limit=50', headers=HEADERS)
data = resp.json()
rows = data['rows'] if isinstance(data, dict) else data

# Build lookup by title prefix (T1.0.1, S1.0, F1, etc.)
by_prefix = {}
for r in rows:
    title = r['title']
    if title.startswith('T') or title.startswith('S') or title.startswith('F'):
        prefix = title.split(':')[0].strip()
        by_prefix[prefix] = r

# Success factors (from plan "Done when" sections)
success_factors = {
    'T1.0.1': 'Complete HC value chain documented. Every store item category mapped to HC cost. Minimum meaningful purchase identified and costed.',
    'T1.0.2': 'Goals payer assumptions benchmarked against industry. Clear assessment of whether conversion/ARPPU targets are realistic at launch. Implication for pricing strategy documented.',
    'T1.1.1': 'Position map complete showing Goals percentile at each tier. Clear one-line answer: "Goals is at the Xth percentile - below/at/above median."',
    'T1.1.2': 'Curve comparison documented. Assessment of whether Goals discount progression is appropriate for new launch. Specific recommendation if any tier discount should change.',
    'T1.1.3': 'Structure audit complete. Entry point assessment with recommendation. Whale ceiling analysis documented. Structural gaps identified with recommendation.',
    'T1.1.4': 'Net revenue table complete. Margin concerns flagged. Epic revenue advantage quantified.',
    'T1.2.1': '4 scenarios modelled with revenue impact ranges. Clear recommendation on which scenario maximises launch revenue while preserving future flexibility.',
    'T1.3.1': '2-3 specific first-purchase recommendations documented with rationale and expected conversion impact. Clearly framed as complementary to core pricing.',
    'T2.1.1': 'Every country categorised (green/amber/red). Red countries have specific risk descriptions. Ranked priority list for deeper analysis.',
    'T2.1.2': 'Competitor pricing gathered for all red/amber countries. Goals position vs EA FC documented per market. Countries significantly above/below EA FC identified.',
    'T2.2.1': 'Platform compliance verified or issues flagged. Submission timeline documented. Blockers escalated to Julius.',
    'T2.3.1': 'All red/amber countries have specific, actionable recommendation cards. Each card self-contained for Julius to implement.',
    'T3.1.1': 'One clear paragraph answering "are we too low?" with specific evidence and specific recommendation. No hedging.',
    'T3.2.1': 'Minimum 7 risks documented with full detail. Ranked by severity x likelihood. Top 3 have specific, actionable mitigations.',
    'T3.3.1': '5 recommendations in full format. No vague advice. Every number specific. Every impact range stated.',
    'T4.1.1': 'One-page document Julius can follow to price any new SKU. Worked example (Celebration at $7 for Brazil) included.',
    'T4.2.1': '3-5 bullet points with specific, actionable World Cup pricing tactics. Half page max.',
    'T4.3.1': 'Half-page "What\'s Next" section. Helpful, not salesy. Positions Phase 2 naturally.',
    'T5.1.1': 'Complete document draft. All sections populated with real findings. Internally consistent. 12-18 pages.',
    'T5.2.1': 'Final document approved by Glen and Tom. Sent to client 24h before review call. Call scheduled.',
    'T5.3.1': 'Call completed. Follow-up email sent within 24h. Phase 2 interest level documented.',
}

# Timeline dates (from plan Day allocation)
timeline = {
    'T1.0.1': ('2026-04-22', '2026-04-22'),  # Day 1
    'T1.0.2': ('2026-04-22', '2026-04-22'),  # Day 1
    'T1.1.1': ('2026-04-22', '2026-04-22'),  # Day 1
    'T1.1.2': ('2026-04-22', '2026-04-22'),  # Day 1
    'T1.1.3': ('2026-04-23', '2026-04-23'),  # Day 2
    'T1.1.4': ('2026-04-23', '2026-04-23'),  # Day 2
    'T1.2.1': ('2026-04-23', '2026-04-23'),  # Day 2
    'T1.3.1': ('2026-04-23', '2026-04-23'),  # Day 2
    'T2.1.1': ('2026-04-23', '2026-04-24'),  # Day 2-3
    'T2.1.2': ('2026-04-24', '2026-04-24'),  # Day 3
    'T2.2.1': ('2026-04-24', '2026-04-24'),  # Day 3
    'T2.3.1': ('2026-04-24', '2026-04-25'),  # Day 3-4
    'T3.1.1': ('2026-04-25', '2026-04-25'),  # Day 4
    'T3.2.1': ('2026-04-25', '2026-04-25'),  # Day 4
    'T3.3.1': ('2026-04-25', '2026-04-25'),  # Day 4
    'T4.1.1': ('2026-04-25', '2026-04-25'),  # Day 4 (conditional)
    'T4.2.1': ('2026-04-25', '2026-04-25'),  # Day 4 (conditional)
    'T4.3.1': ('2026-04-25', '2026-04-25'),  # Day 4 (conditional)
    'T5.1.1': ('2026-04-28', '2026-04-28'),  # Day 5
    'T5.2.1': ('2026-04-28', '2026-04-28'),  # Day 5
    'T5.3.1': ('2026-04-28', '2026-04-28'),  # Day 5
}

# Feature/Story timeline
feature_timeline = {
    'F1': ('2026-04-22', '2026-04-23'),
    'F2': ('2026-04-23', '2026-04-25'),
    'F3': ('2026-04-25', '2026-04-25'),
    'F4': ('2026-04-25', '2026-04-25'),
    'F5': ('2026-04-28', '2026-04-28'),
    'S1.0': ('2026-04-22', '2026-04-22'),
    'S1.1': ('2026-04-22', '2026-04-23'),
    'S1.2': ('2026-04-23', '2026-04-23'),
    'S1.3': ('2026-04-23', '2026-04-23'),
    'S2.1': ('2026-04-23', '2026-04-24'),
    'S2.2': ('2026-04-24', '2026-04-24'),
    'S2.3': ('2026-04-24', '2026-04-25'),
    'S3.1': ('2026-04-25', '2026-04-25'),
    'S3.2': ('2026-04-25', '2026-04-25'),
    'S3.3': ('2026-04-25', '2026-04-25'),
    'S4.1': ('2026-04-25', '2026-04-25'),
    'S4.2': ('2026-04-25', '2026-04-25'),
    'S4.3': ('2026-04-25', '2026-04-25'),
    'S5.1': ('2026-04-28', '2026-04-28'),
    'S5.2': ('2026-04-28', '2026-04-28'),
    'S5.3': ('2026-04-28', '2026-04-28'),
}

# Update all tasks via PATCH
updated = 0
for prefix, task in by_prefix.items():
    patch = {'practice_area': 'gaming'}

    if prefix in success_factors:
        patch['success_factor'] = success_factors[prefix]

    if prefix in timeline:
        patch['start_date'] = timeline[prefix][0]
        patch['end_date'] = timeline[prefix][1]
        patch['due_date'] = timeline[prefix][1]
    elif prefix in feature_timeline:
        patch['start_date'] = feature_timeline[prefix][0]
        patch['end_date'] = feature_timeline[prefix][1]
        patch['due_date'] = feature_timeline[prefix][1]

    resp = requests.patch(f'{BASE}/api/tasks/{task["id"]}', headers=HEADERS, json=patch)
    if resp.status_code == 200:
        updated += 1
    else:
        print(f'  WARN: Failed to update {prefix}: {resp.status_code} {resp.text}')

# Also update the project itself
for r in rows:
    if r['title'] == 'Goals Studio Pricing Engagement':
        resp = requests.patch(f'{BASE}/api/tasks/{r["id"]}', headers=HEADERS, json={
            'practice_area': 'gaming',
            'start_date': '2026-04-22',
            'end_date': '2026-04-28',
            'due_date': '2026-04-28',
        })
        if resp.status_code == 200:
            updated += 1

print(f'Updated {updated} tasks with success_factor, practice_area, and timeline dates.')

# Re-fetch all tasks for Excel export
resp = requests.get(f'{BASE}/api/tasks?limit=50', headers=HEADERS)
data = resp.json()
rows = data['rows'] if isinstance(data, dict) else data

# Build parent lookup for hierarchy display
by_id = {r['id']: r for r in rows}

def get_depth(task):
    depth = 0
    current = task
    while current.get('parent_id') and current['parent_id'] in by_id:
        depth += 1
        current = by_id[current['parent_id']]
    return depth

def get_path(task):
    path = []
    current = task
    while current.get('parent_id') and current['parent_id'] in by_id:
        current = by_id[current['parent_id']]
        path.insert(0, current['title'])
    return ' > '.join(path) if path else ''

# Filter to Goals tasks only
goals_tasks = [r for r in rows if r.get('client_id') == '6975460f-c302-42c5-a586-1d04c5fcb929']

# Sort by hierarchy: project first, then features by number, stories, tasks
def sort_key(task):
    type_order = {'project': 0, 'feature': 1, 'story': 2, 'task': 3}
    title = task['title']
    # Extract number for sorting
    import re
    nums = re.findall(r'[\d.]+', title[:20])
    num = nums[0] if nums else '99'
    return (type_order.get(task['item_type'], 4), num)

# Build tree-ordered list
def tree_sort(tasks):
    by_parent = {}
    roots = []
    for t in tasks:
        pid = t.get('parent_id')
        if pid and pid in by_id:
            by_parent.setdefault(pid, []).append(t)
        else:
            roots.append(t)

    result = []
    def add_children(parent):
        result.append(parent)
        children = by_parent.get(parent['id'], [])
        children.sort(key=sort_key)
        for c in children:
            add_children(c)

    roots.sort(key=sort_key)
    for r in roots:
        add_children(r)
    return result

ordered = tree_sort(goals_tasks)

# Create Excel workbook
wb = openpyxl.Workbook()

# === Sheet 1: Project Summary ===
ws_summary = wb.active
ws_summary.title = 'Project Summary'

header_font = Font(name='Calibri', bold=True, size=14)
section_font = Font(name='Calibri', bold=True, size=11)
body_font = Font(name='Calibri', size=10)
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
header_text = Font(name='Calibri', bold=True, size=11, color='FFFFFF')

ws_summary.column_dimensions['A'].width = 25
ws_summary.column_dimensions['B'].width = 80

summary_data = [
    ('Goals Studio Pricing Engagement', ''),
    ('', ''),
    ('Client', 'Goals Studio (Goals AB)'),
    ('Location', 'Stockholm, Sweden'),
    ('Primary Contact', 'Jonas Rundberg (CEO)'),
    ('Secondary Contact', 'Julius (Live Ops)'),
    ('NBI Team', 'Glen Pryer (domain + relationship), Tom Rieger (contract), Devin Rieger (research)'),
    ('', ''),
    ('Contract', ''),
    ('SOW Value', '100,000 SEK (~$10,000 USD)'),
    ('Budget', '~50 billable hours over 5 working days'),
    ('SOW Date', '8 April 2026'),
    ('Delivery Deadline', '27 April 2026 (written summary), review call same week'),
    ('Hard Deadline', 'Before Goals submits to PS/Xbox (~5 May for 14 May launch)'),
    ('', ''),
    ('Scope', ''),
    ('Deliverable', 'Written summary (12-18 pages) + remote review call (30-45 min)'),
    ('Area 1', 'Review competitive pricing positioning in sports/football game space'),
    ('Area 2', 'Price point selection, currency pack structure, conversion rate guidance'),
    ('Area 3', 'Regional pricing for base and regional markets (40+ countries)'),
    ('Area 4', 'Assessment of current pricing against market norms'),
    ('Area 5', 'Recommendations on adjustments and areas of risk'),
    ('', ''),
    ('The Game (GOALS)', ''),
    ('Genre', 'F2P football game'),
    ('Launch Date', '14 May 2026'),
    ('Platforms', 'Steam, Epic Games Store, Xbox, PlayStation'),
    ('Unique Mechanic', 'Generated players (not licensed), RPG progression, aging/retirement'),
    ('Monetisation', '80% skill / 20% purchased advantage (explicitly non-P2W)'),
    ('Economy Type', 'Seasonless (no battle pass, no seasonal resets, no wipes)'),
    ('Beta Stats', '400K total players, 50% D1 retention (PS), NPS +24.6, CPI $0.43'),
    ('Key Event', 'World Cup June 2026 = first major monetisation peak'),
    ('HC Economy', '6-tier pack ladder: $5.99-$99.99. 1 HC = 50 points = $0.012 USD'),
    ('', ''),
    ('Key Questions', ''),
    ('Jonas', '"Are we pricing too low?"'),
    ('Glen Principle', 'Launch high, patch down. Never raise prices post-launch.'),
    ('Conversion Risk', '10.4% payer target vs 2-5% industry norm for month 1'),
    ('', ''),
    ('Hour Budget', ''),
    ('F1: Competitive HC Pricing Benchmark', '16h (32%)'),
    ('F2: Regional Pricing Strategy', '13h (26%)'),
    ('F3: Risk Assessment & Recommendations', '10h (20%)'),
    ('F4: Value-Add (Conditional)', '4h (8%)'),
    ('F5: Deliverable Assembly & Review', '7h (14%)'),
    ('Total', '50h (100%)'),
    ('', ''),
    ('Timeline', ''),
    ('Day 1 (Tue 22 Apr)', 'Foundation + Competitive position (9h)'),
    ('Day 2 (Wed 23 Apr)', 'Competitive completion + Regional start (10h)'),
    ('Day 3 (Thu 24 Apr)', 'Regional deep-dive + Quality Gate 3 (10h)'),
    ('Day 4 (Fri 25 Apr)', 'Synthesis + Recommendations (10h)'),
    ('Day 5 (Mon 28 Apr)', 'Deliverable + Review (7h)'),
    ('', ''),
    ('Notes', ''),
    ('Buffer Strategy', 'Feature 4 is conditional - only if F1-3 complete within budget'),
    ('Quality Gate 3', 'End of Day 3 - Glen reviews draft recs before synthesis'),
    ('Old SOW Warning', 'The $62K proposal (31 March) is NOT the current SOW'),
    ('Documentation', '15 markdown files extracted from client PDFs/Excel/PPTX in Clients/Goals/'),
]

for i, (label, value) in enumerate(summary_data, 1):
    cell_a = ws_summary.cell(row=i, column=1, value=label)
    cell_b = ws_summary.cell(row=i, column=2, value=value)
    if i == 1:
        cell_a.font = header_font
    elif value == '' and label:
        cell_a.font = section_font
    else:
        cell_a.font = Font(name='Calibri', bold=True, size=10)
        cell_b.font = body_font

# === Sheet 2: Full Task Breakdown ===
ws_tasks = wb.create_sheet('Task Breakdown')

columns = [
    ('ID', 8),
    ('Type', 10),
    ('Title', 55),
    ('Status', 14),
    ('Priority', 10),
    ('Assignees', 20),
    ('Hours Est.', 10),
    ('Start Date', 12),
    ('End Date', 12),
    ('Due Date', 12),
    ('Success Criteria', 60),
    ('Description', 80),
    ('Parent Path', 40),
    ('Practice Area', 13),
]

for col_idx, (name, width) in enumerate(columns, 1):
    cell = ws_tasks.cell(row=1, column=col_idx, value=name)
    cell.font = header_text
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws_tasks.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

# Type colours
type_fills = {
    'project': PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid'),
    'feature': PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid'),
    'story': PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid'),
    'task': PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid'),
}
type_fonts = {
    'project': Font(name='Calibri', bold=True, size=11, color='FFFFFF'),
    'feature': Font(name='Calibri', bold=True, size=10, color='FFFFFF'),
    'story': Font(name='Calibri', bold=True, size=10),
    'task': Font(name='Calibri', size=10),
}

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

for row_idx, task in enumerate(ordered, 2):
    depth = get_depth(task)
    indent = '  ' * depth
    title_prefix = task['title']

    assignees = ', '.join(task.get('assignees') or [])
    parent_path = get_path(task)
    item_type = task['item_type']

    values = [
        task['title'].split(':')[0].strip() if ':' in task['title'] else item_type.upper(),
        item_type.capitalize(),
        task['title'],
        task.get('status', 'Not started'),
        task.get('priority', ''),
        assignees,
        task.get('hours_estimated', 0),
        task.get('start_date', ''),
        task.get('end_date', ''),
        task.get('due_date', ''),
        task.get('success_factor', ''),
        task.get('description', ''),
        parent_path,
        task.get('practice_area', ''),
    ]

    fill = type_fills.get(item_type, type_fills['task'])
    font = type_fonts.get(item_type, type_fonts['task'])

    for col_idx, val in enumerate(values, 1):
        cell = ws_tasks.cell(row=row_idx, column=col_idx, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        if item_type in ('project', 'feature'):
            cell.fill = fill
            cell.font = font
        elif item_type == 'story':
            cell.fill = fill
            cell.font = font
        else:
            cell.font = font

# === Sheet 3: Timeline View ===
ws_timeline = wb.create_sheet('Timeline')

timeline_cols = [('Task', 45), ('Type', 10), ('Assignee', 18), ('Hours', 8),
                 ('Day 1\n22 Apr', 12), ('Day 2\n23 Apr', 12), ('Day 3\n24 Apr', 12),
                 ('Day 4\n25 Apr', 12), ('Day 5\n28 Apr', 12)]

for col_idx, (name, width) in enumerate(timeline_cols, 1):
    cell = ws_timeline.cell(row=1, column=col_idx, value=name)
    cell.font = header_text
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws_timeline.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

day_dates = ['2026-04-22', '2026-04-23', '2026-04-24', '2026-04-25', '2026-04-28']
active_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
active_font = Font(name='Calibri', bold=True, color='FFFFFF', size=9)

task_only = [t for t in ordered if t['item_type'] == 'task']
for row_idx, task in enumerate(task_only, 2):
    ws_timeline.cell(row=row_idx, column=1, value=task['title']).font = body_font
    ws_timeline.cell(row=row_idx, column=2, value=task['item_type'].capitalize()).font = body_font
    ws_timeline.cell(row=row_idx, column=3, value=', '.join(task.get('assignees') or [])).font = body_font
    ws_timeline.cell(row=row_idx, column=4, value=task.get('hours_estimated', 0)).font = body_font

    start = task.get('start_date', '')
    end = task.get('end_date', '')
    for day_idx, day_date in enumerate(day_dates):
        col = 5 + day_idx
        if start and end and start <= day_date <= end:
            cell = ws_timeline.cell(row=row_idx, column=col, value=f"{task.get('hours_estimated', 0)}h")
            cell.fill = active_fill
            cell.font = active_font
            cell.alignment = Alignment(horizontal='center')
        else:
            ws_timeline.cell(row=row_idx, column=col, value='')

    for col_idx in range(1, len(timeline_cols) + 1):
        ws_timeline.cell(row=row_idx, column=col_idx).border = thin_border

# Freeze panes
ws_tasks.freeze_panes = 'A2'
ws_timeline.freeze_panes = 'A2'

# Save
output_path = r'd:\OneDrive\Claude_code\NBIAI_TEAM\Clients\Goals\goals_pricing_engagement_plan.xlsx'
wb.save(output_path)
print(f'Excel saved to: {output_path}')
print(f'Sheets: {wb.sheetnames}')
print(f'Task rows: {len(ordered)}')
