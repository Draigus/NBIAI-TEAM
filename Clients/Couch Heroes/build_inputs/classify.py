"""Classify every Couch Heroes feature into type/layer with confidence.

Reads Couch_Heroes_Features_v2.xlsx, applies the rules from the brief,
writes a pipe-delimited markdown file at game_classification.md.
"""

import openpyxl
import re
from pathlib import Path

SRC = Path(r'D:/OneDrive/Claude_code/NBIAI_TEAM/Clients/Couch Heroes/Couch_Heroes_Features_v2.xlsx')
OUT = Path(r'D:/OneDrive/Claude_code/NBIAI_TEAM/Clients/Couch Heroes/build_inputs/game_classification.md')

# ---------- keyword sets for layer classification ----------

FOUNDATION_KW = [
    'engine', 'save', 'load', 'scene', 'event bus', 'network', 'online service',
    'server', 'replication', 'persistence', 'authentication', 'auth ', 'login',
    'database', 'backend', 'matchmak', 'sharding', 'instance management',
    'session', 'account', 'session ', 'patch', 'build pipeline', 'devops',
    'telemetry', 'analytics infra', 'live ops infrastructure', 'deployment',
    'cloud', 'cdn', 'data layer', 'platform', 'sdk', 'framework',
]

CORE_KW = [
    'physics', 'input', 'controller', 'movement', 'locomotion', 'jump',
    'collision', 'character controller', 'camera', 'animation system',
    'state machine', 'targeting', 'hit detection', 'damage calculation',
    'combat system', 'inventory grid', 'inventory system',
]

FEATURE_KW = [
    'quest', 'crafting', 'profession', 'magic', 'spell system', 'economy',
    'ai ', ' ai', 'enemy', 'creature', 'mount', 'companion', 'progression',
    'level up', 'achievement', 'faction', 'corruption', 'puzzle', 'platform',
    'loot', 'drop', 'shop', 'trade', 'auction', 'marketplace', 'guild',
    'party', 'group', 'social', 'friend', 'chat', 'mail', 'mailbox',
    'housing', 'user space', 'wardrobe', 'cosmetic', 'crystal', 'enchant',
    'potion', 'fishing', 'cooking', 'alchemy', 'gathering', 'recipe',
    'reputation', 'dialogue', 'lore', 'narrative', 'event', 'season',
    'battle pass', 'monetisation', 'monetization', 'tutorial',
    'skill tree', 'ability', 'class', 'archetype', 'batsuit', 'gadget',
]

PRESENTATION_KW = [
    'ui', 'hud', 'menu', 'vfx', 'audio', 'sound', 'sfx', 'music', 'icon',
    'tooltip', 'splash', 'cinematic', 'cutscene', 'screen', 'widget',
    'minimap', 'compass', 'notification', 'banner', 'overlay', 'dialog box',
    'animation', 'particle', 'lighting', 'post-process', 'shader',
]

POLISH_KW = [
    'accessibility', 'optimisation', 'optimization', 'performance',
    'qa pass', 'polish', 'localis', 'localiz', 'subtitles',
    'colourblind', 'colorblind', 'accessib',
]

CONTENT_ASSET_KW = [
    'icon', 'animation', 'model', 'mesh', 'texture', 'vfx', 'sfx', 'sound',
    'concept art', 'asset', 'rig', 'particle', 'shader', 'material',
    'environment', 'prop', 'building', 'character model', 'npc model',
    'foliage', 'skybox', 'audio cue', 'music track', 'voiceover',
]

SYSTEM_HINT_KW = [
    'system', 'epic', 'framework', 'service', 'pipeline', 'engine',
    'architecture', 'manager', 'controller',
]


def is_blank(v):
    return v is None or (isinstance(v, str) and v.strip() == '')


def clean(v):
    if v is None:
        return ''
    s = str(v).strip()
    s = s.replace('\t', ' ').replace('\n', ' ').replace('|', '/')
    s = re.sub(r'\s+', ' ', s)
    return s


def title_for_classification(epic, feature, story, task):
    """Pick the most specific non-blank field to classify on."""
    for v in [task, story, feature, epic]:
        if not is_blank(v):
            return clean(v)
    return ''


def detect_layer(text, default='Feature'):
    t = text.lower()
    # Check most-specific first (Polish, Presentation, Foundation, Core, Feature)
    for kw in POLISH_KW:
        if kw in t:
            return 'Polish'
    for kw in PRESENTATION_KW:
        if re.search(r'\b' + re.escape(kw) + r'\b', t):
            return 'Presentation'
    for kw in FOUNDATION_KW:
        if kw in t:
            return 'Foundation'
    for kw in CORE_KW:
        if kw in t:
            return 'Core'
    for kw in FEATURE_KW:
        if kw in t:
            return 'Feature'
    return default


def is_system_title(text, epic, feature, story, task):
    """Heuristic for whether row is a System (architectural).

    A row is a System when:
    - It is an epic-level row (only epic populated, others blank), OR
    - It is the feature-level row whose feature value names a system (e.g. "Player Progression System"), OR
    - The most specific populated field is the title and it ends with/equals a system word
      (NOT just contains the word in a [bracketed prefix]).
    """
    # Epic-level row: only epic populated
    if not is_blank(epic) and is_blank(feature) and is_blank(story) and is_blank(task):
        return True
    # Feature-level row where the feature column itself is populated and contains a system word
    if not is_blank(feature) and is_blank(story) and is_blank(task):
        f = str(feature).lower()
        if re.search(r'\b(system|framework|service|engine|pipeline|architecture)\b', f):
            return True
    # Strip any [bracketed prefix] before checking the residual title
    residual = re.sub(r'^\s*\[[^\]]*\]\s*', '', text).strip().lower()
    # Now if residual itself ends with "system"/"framework"/"service"/"engine" it is a system
    if re.search(r'\b(system|framework|service|engine|pipeline|architecture)\s*$', residual):
        return True
    # Or residual is empty and bracket-only contained a system word -> still a system
    if residual == '':
        full = text.lower()
        if re.search(r'\b(system|framework|service|engine|pipeline|architecture)\b', full):
            return True
    return False


def is_content_asset(text):
    t = text.lower()
    for kw in CONTENT_ASSET_KW:
        if re.search(r'\b' + re.escape(kw) + r'\b', t):
            return True
    return False


def classify_game_feature_row(epic, feature, story, task):
    """Return (type, layer, confidence, reason)."""
    title = title_for_classification(epic, feature, story, task)
    if not title:
        return ('Feature', 'Feature', 'low', 'all classification fields blank')

    t = title.lower()

    # Asset-style title -> Content
    if is_content_asset(title) and not is_system_title(title, epic, feature, story, task):
        layer = detect_layer(title, default='Presentation')
        return ('Content', layer, 'medium', 'asset-style title in Game Feature List (verify it is not a gameplay feature)')

    # System detection
    if is_system_title(title, epic, feature, story, task):
        layer = detect_layer(title, default='Feature')
        # Confidence: high if has explicit "System"/"Framework"/"Service" word
        if re.search(r'\b(system|framework|service|engine|pipeline|architecture)\b', t):
            return ('System', layer, 'high', '')
        # Epic-only row, no explicit system word
        return ('System', layer, 'medium', 'epic-level row inferred as System; no explicit System keyword')

    # Default to Feature
    layer = detect_layer(title, default='Feature')
    # Confidence
    if layer == 'Feature' and not any(kw in t for kw in FEATURE_KW):
        return ('Feature', layer, 'low', 'no clear layer keyword matched; defaulted to Feature/Feature')
    return ('Feature', layer, 'high', '')


def classify_tdd_row(epic, feature, story, task):
    title = title_for_classification(epic, feature, story, task)
    if not title:
        return ('System', 'Feature', 'low', 'all classification fields blank')
    layer = detect_layer(title, default='Feature')
    return ('System', layer, 'high', '')


def classify_platform_row(epic, feature, story, task):
    title = title_for_classification(epic, feature, story, task)
    if not title:
        return ('Platform', 'Foundation', 'low', 'all classification fields blank')
    return ('Platform', 'Foundation', 'high', '')


def classify_content_row(epic, feature, content_task, art_team):
    """Content List has different schema: epic, feature, content_task, art_team."""
    title = title_for_classification(epic, feature, '', content_task)
    if not title:
        # Maybe art_team has it
        if not is_blank(art_team):
            title = clean(art_team)
        else:
            return ('Content', 'Presentation', 'low', 'all classification fields blank')

    t = title.lower()
    # Layer
    layer = 'Presentation'
    if any(kw in t for kw in ['accessibility', 'optimisation', 'optimization', 'performance', 'localis', 'localiz']):
        layer = 'Polish'
    elif any(kw in t for kw in ['save', 'load', 'pipeline', 'tooling', 'asset pipeline']):
        layer = 'Foundation'

    # Confidence
    art_team_str = clean(art_team).lower() if not is_blank(art_team) else ''
    if art_team_str or any(kw in t for kw in ['vfx', 'sfx', 'audio', 'icon', 'animation', 'model', 'concept', 'ux', 'ui', 'art', 'texture', 'rig', 'mesh', 'shader', 'material', 'particle']):
        return ('Content', layer, 'high', '')
    return ('Content', layer, 'medium', 'no explicit art team or asset keyword; verify it is content not feature')


def classify_liveservice_row(epic, feature, story, task):
    title = title_for_classification(epic, feature, story, task)
    if not title:
        return ('LiveService', 'Feature', 'low', 'all classification fields blank')
    return ('LiveService', 'Feature', 'high', '')


def classify_feature_versions_row(name):
    """Feature Versions sheet — these are maturity overlays."""
    if is_blank(name):
        return None
    n = clean(name)
    nl = n.lower()
    # Skip header-ish / definition rows
    if nl in {'feature name', 'not started', 'r&d', 'prototype', 'mvp', 'launch', 'optimize', 'expand'}:
        return None
    # Platform-prefixed -> type=Platform, layer=Foundation
    if n.startswith('[Platform]') or '[platform]' in nl:
        return ('Platform', 'Foundation', 'high', '')
    return ('Maturity', 'Feature', 'high', '')


# ---------- main ----------

def row_get(row, idx):
    return row[idx] if idx < len(row) else None


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    out_lines = []
    out_lines.append('| source_sheet | epic | feature | story | task | type | layer | confidence | needs_review_reason |')
    out_lines.append('|---|---|---|---|---|---|---|---|---|')

    counts_by_type = {}
    low_count = 0
    total = 0

    def add_row(source, epic, feature, story, task, typ, layer, conf, reason):
        nonlocal low_count, total
        out_lines.append(f'| {source} | {clean(epic)} | {clean(feature)} | {clean(story)} | {clean(task)} | {typ} | {layer} | {conf} | {reason} |')
        counts_by_type[typ] = counts_by_type.get(typ, 0) + 1
        if conf == 'low':
            low_count += 1
        total += 1

    # Schema for sheets with Tag/Epic/Feature/Story/Task layout:
    #   col 0 = Tag (category label, repeats on every row)
    #   col 1 = Epic
    #   col 2 = Feature
    #   col 3 = Story
    #   col 4 = Task
    # We drop the Tag column from output (per the brief's output schema:
    # epic, feature, story, task) and use cols 1..4 as the four hierarchy fields.

    def fold_tag(row):
        """Use col 0 (Tag) as the epic when col 1 (Epic) is blank.

        This captures section-header rows where only the Tag column is populated.
        Returns (epic, feature, story, task) for sheets with the
        Tag/Epic/Feature/Story/Task layout.
        """
        c0 = row_get(row, 0)
        c1 = row_get(row, 1)
        c2 = row_get(row, 2)
        c3 = row_get(row, 3)
        c4 = row_get(row, 4)
        # Promote Tag into Epic slot when Epic is blank AND there is no other content.
        # When Epic is populated, ignore Tag (it duplicates the parent epic on every row).
        if is_blank(c1) and is_blank(c2) and is_blank(c3) and is_blank(c4):
            return (c0, '', '', '')
        return (c1, c2, c3, c4)

    # ---- Game Feature List ----
    ws = wb['Game Feature List']
    for row in ws.iter_rows(min_row=2, values_only=True):
        epic, feature, story, task = fold_tag(row)
        if all(is_blank(v) for v in [epic, feature, story, task]):
            continue
        typ, layer, conf, reason = classify_game_feature_row(epic, feature, story, task)
        add_row('Game Feature List', epic, feature, story, task, typ, layer, conf, reason)

    # ---- TDD ----
    ws = wb['TDD']
    for row in ws.iter_rows(min_row=2, values_only=True):
        epic, feature, story, task = fold_tag(row)
        if all(is_blank(v) for v in [epic, feature, story, task]):
            continue
        typ, layer, conf, reason = classify_tdd_row(epic, feature, story, task)
        add_row('TDD', epic, feature, story, task, typ, layer, conf, reason)

    # ---- Platform Feature List ----
    ws = wb['Platform Feature List']
    for row in ws.iter_rows(min_row=2, values_only=True):
        epic, feature, story, task = fold_tag(row)
        if all(is_blank(v) for v in [epic, feature, story, task]):
            continue
        typ, layer, conf, reason = classify_platform_row(epic, feature, story, task)
        add_row('Platform Feature List', epic, feature, story, task, typ, layer, conf, reason)

    # ---- Content List ----
    # Content List schema: col0=Tag, col1=Epic, col2=Feature, col3=Content Task, col4=Art Team
    # No Story column. Content Task occupies the task slot in the output.
    ws = wb['Content List']
    for row in ws.iter_rows(min_row=2, values_only=True):
        c0 = row_get(row, 0)
        c1 = row_get(row, 1)
        c2 = row_get(row, 2)
        c3 = row_get(row, 3)
        c4 = row_get(row, 4)
        # Promote Tag to Epic when Epic blank and no other content
        if is_blank(c1) and is_blank(c2) and is_blank(c3):
            epic, feature, content_task, art_team = (c0, '', '', c4)
        else:
            epic, feature, content_task, art_team = (c1, c2, c3, c4)
        if all(is_blank(v) for v in [epic, feature, content_task, art_team]):
            continue
        typ, layer, conf, reason = classify_content_row(epic, feature, content_task, art_team)
        add_row('Content List', epic, feature, '', content_task, typ, layer, conf, reason)

    # ---- Live Service Features ----
    ws = wb['Live Service Features']
    for row in ws.iter_rows(min_row=2, values_only=True):
        epic, feature, story, task = fold_tag(row)
        if all(is_blank(v) for v in [epic, feature, story, task]):
            continue
        typ, layer, conf, reason = classify_liveservice_row(epic, feature, story, task)
        add_row('Live Service Features', epic, feature, story, task, typ, layer, conf, reason)

    # ---- Feature Versions ----
    ws = wb['Feature Versions']
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row_get(row, 0)
        result = classify_feature_versions_row(name)
        if result is None:
            continue
        typ, layer, conf, reason = result
        add_row('Feature Versions', '', name, '', '', typ, layer, conf, reason)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text('\n'.join(out_lines), encoding='utf-8')

    print(f'Total rows classified: {total}')
    print('Breakdown by type:')
    for k, v in sorted(counts_by_type.items()):
        print(f'  {k}: {v}')
    print(f'Low confidence rows: {low_count}')
    print(f'Output: {OUT}')


if __name__ == '__main__':
    main()
