import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb_src = openpyxl.load_workbook('Clients/Couch Heroes/CH_AMA_Questions_Broken_Out2.xlsx')
ws_src = wb_src.active

# Read all source questions
source = []
for r in range(2, ws_src.max_row + 1):
    q = ws_src.cell(row=r, column=3).value
    if not q or not str(q).strip():
        continue
    source.append({
        'qnum': ws_src.cell(row=r, column=1).value,
        'topic': ws_src.cell(row=r, column=2).value,
        'question': str(q).strip(),
        'orig_row': ws_src.cell(row=r, column=4).value,
        'crit_p': ws_src.cell(row=r, column=5).value,
        'crit_s': ws_src.cell(row=r, column=6).value,
    })

# Build merge map: old_qnum -> keep_qnum
# These are the 14 merges identified and approved by Glen
merges = {
    5: 3,    # "USP + experience" into "What exactly is our USP"
    8: 3,    # "What makes different" into USP
    21: 11,  # "Game pillars" into "Core features/pillars"
    31: 11,  # "Non-negotiable pillars" into core pillars
    49: 48,  # "Monetisation system" into "How monetise"
    53: 52,  # "Processes for content lasting" into "Long-term continuity"
    63: 62,  # "Is culture defined" into "Culture beyond nice people"
    66: 64,  # "Why outside hires" into "Promote from within"
    69: 68,  # "Office plans" into "Always remote"
    41: 40,  # "Brutally honest not working" into "Truth avoiding"
    58: 57,  # "Decisions delayed" into "Ownership unclear"
    60: 59,  # "Better communication" into "Communication breakdown"
    82: 79,  # "Mistakes punished/learned" into "Leaders handle being wrong"
    83: 80,  # "Accountable regardless of role" into "High performers accountable"
}

# Build enriched questions for merged items
merge_additions = {}
for old_q, keep_q in merges.items():
    old_item = next((s for s in source if s['qnum'] == old_q), None)
    if old_item:
        if keep_q not in merge_additions:
            merge_additions[keep_q] = []
        merge_additions[keep_q].append(old_item['question'])

# Update kept questions to incorporate merged angle
question_updates = {
    3: "What exactly is our USP? What experience are we focusing on for players, and what makes this game different from other games?",
    11: "What are the core features, foundational pillars, and non-negotiable pillars of the game (excluding platform, partnerships, lore)? How do they shape feature and priority decisions?",
    48: "How is our game going to monetise? What will the monetisation system be and how will it work?",
    52: "How are we ensuring long-term continuity - reusing/building on previous work instead of reworking? What processes ensure content built now lasts?",
    62: "What is our culture as a studio, beyond just nice people? Is it formally defined?",
    64: "What happened to the Promote from Within culture? Why are outside hires filling positions instead?",
    68: "Are we going to always be remote? Are there real plans for offices in Greece/UK?",
    40: "What is the truth we are avoiding about the game, platform, or team? If we were brutally honest, what is not working and how can we solve it?",
    57: "Where is ownership unclear? What decisions are delayed because no one wants to take responsibility?",
    59: "Where does communication break down between teams? How do we ensure better cross-discipline communication moving forward?",
    79: "How do leaders handle being wrong? How are mistakes handled - punished or learned from?",
    80: "Are we holding high performers accountable if they hurt the team? Are people held accountable regardless of role?",
}

# Glen's answers from conversation (Q1-Q4)
glen_answers = {}
glen_answers[1] = """The demand we're chasing is absolutely MMO players. Some of the most engaged, loyal, community-driven players in the industry. We want them, and we're building for them. But we want to serve them differently than the current crop of MMOs do.

When I was at World of Warcraft, the stat that stuck with me was that the average player who'd put in more than twenty hours was then spending twenty-eight hours a week or more in the game. That's a part-time job. And those players weren't always having fun. They were stressing over reputation grinds, gear checks, raid schedules. Miss a raid night, get kicked from the guild. That's a game consuming a person's life, not being part of it.

You can see the whole industry quietly backing away from that model. Games are toning down dailies, softening progression, extending catch-up mechanics, trimming raid lockouts. They see the problem. They just can't unbuild the game they already shipped around it.

And if you look at what players are actually spending their time in right now, the signal is even clearer. Fortnite, Roblox, Animal Crossing, Palia, Sky, VRChat, the housing and hangout side of FFXIV. The numbers are enormous, and the thread running through all of them is the same. People want somewhere to be. A place to show up, meet their friends, decorate their corner of it, collect things they care about, and play around. Progression is expressive, not numerical. You don't need to clear your evening to feel like you got something out of it.

Couch Heroes is built from the ground up for that. Personalisation and a real sense of self are a core pillar, not a side feature. Your avatar, your home, your arcade, your collections. That's the progression. It's communal by design, decorated by you, shared with the friends you choose. No gear treadmill. No stats gating what you can wear. No forty-hour-a-week commitment to stay relevant. If you've got twenty minutes, you drop in, tidy up, catch up with friends, poke around a point of interest, and log off feeling good. If you've got a whole evening, there's real combat, real exploration, proper puzzles, and a world with genuine craft and story behind it. Both of those players get something worthwhile, and they're the same player on different days.

The partner layer keeps the world feeling fresh. Other studios bring their IP into Couch Heroes as arcade machines, cosmetics, events, or whole shards. For players, that means new things to discover that they already care about. For partners, it's a living audience to show their IP to without having to build the infrastructure themselves. That gives us a content engine that doesn't rely on us shipping a giant expansion every year to keep the lights on.

So when people ask why we think there's demand for this, the answer is: the demand is already there. It's visible in what players play every day, and in the way the whole industry is quietly shifting. We want to be part of our players' lives, not consume them. That's the game we're making, and that's the home we're building for them."""

glen_answers[2] = """The answer is partially, and it's always going to be partially when you're trying to do something new. When you orient a product of this size at the market, there's always going to be that question, and anybody who tells you "yes, we're a hundred per cent confident" is probably not being truthful. You can't be, at this stage, on anything that's genuinely new.

What we do know is that we have clear unique selling points. We focus heavily on integrating with people's lives and putting the player first in a new way, which for us means the game is built around the player's life rather than demanding they redesign their life around it. We combine that in the recipe with the rest of the game's depth and its theme, between grit core and myth core, and the way the lore and the story lay out in the activities we build into the game. You can feel it in the corruption mechanic threading through combat and exploration, in the way the arcade ties a player's collection back to partner content they actually care about, in the way the home becomes a real reflection of who they are in the world. That recipe, we believe, is genuinely rather unique.

We know we have a unique story in a unique setting, with a mixture of future and past that you don't see directly in other products. We know that the style of play we're building is unique, based on what we've structured internally, and we'll continue to test it against the market constantly. That's the whole point of iterating as you go. Whether it's RuneScape or World of Warcraft or any other MMO out there, the real danger isn't getting it wrong on day one. It's the inability to iterate to the market as you understand more.

And the honest truth is that the validation isn't mine to make. It's the work everyone in this call does every day that will prove it. The concept is a starting point. What turns it into a game people actually want to live in is what you're all building.

Do we believe we have an incredibly strong premise and a strong core? Absolutely. Have we validated that this is differentiated in the market? Absolutely. Will we continue to validate and drive that hypothesis as we build? Absolutely. That's not a one-time exercise. That's the whole discipline of how we're making this game, and it's on all of us."""

glen_answers[3] = """The USPs will iterate to some extent as we learn more, but what won't change is this.

We are very clearly an MMO. A massively multiplayer online game with a rich and robust role-playing experience, real combat, real exploration, and a robust social system. That sets us firmly in what category of game we are. MMO players are who we're building for.

What makes us unique in that category is we're heavily focused on the identity, friends, and collections side of the game. We'll be able to partner with other games to allow a broad reach of experiences that no individual game can solve on its own. Partner games and partner IP come into our world as arcade machines, cosmetics, events, whole shards. You play, you collect, you show off, you host. The rest of the industry becomes part of your world instead of being scattered across twenty launchers and storefronts you never open.

Everything else is how we build around that.

We're building around the player's life, not the other way round. That means the game isn't structured to require grind treadmills or forty-hour weeks to stay competitive, or chasing a gear score to feel like you belong, or showing up on a Tuesday night so your guild doesn't kick you. Those pressures are what chew people up and spit them out of the genre, and we've designed past them.

Identity is a pillar, not an afterthought. You build who you are in our world. Your avatar, your home, your arcade, your collections. How you dress, how you decorate, the people you hang out with. Most games treat that as a skin on top of the real game. For us it is the game.

Skill over stats. What you do matters more than what you're wearing. Combat, progression, and mastery come from how you play, not from gear inflation. Keeps it competitive without turning into an arms race only the hardcore can survive.

And the world itself is built to move with the players in it. The economy, the story, the environment, all of it responds to what players actually do. Our content pipeline is slower and more considered than the industry's annual-expansion treadmill, and the partner ecosystem means the world keeps getting new stuff we don't have to ship entirely on our own.

The pillars we run the studio by fall straight out of all that. Player identity as the product. Play has real value. Social first, game second. Skill over stats. A living, player-influenced world.

So if you want the one-liner, the thing you can tell your friends when they ask what you're working on: Couch Heroes is an MMO you can actually live with, where your identity, your friends, and your collections carry across the whole gaming world, and where the rest of the industry plugs in to keep it alive.

[Note: Q5 asked about experience focus - your Q3 answer covers the USP comprehensively. The experience focus specifically: we're building for explorers and socialisers first, achievers second. The moment-to-moment feel is cosy exploration with real depth. 20 minutes or 4 hours both feel worthwhile. Q8 asked what makes us different - fully covered above.]"""

glen_answers[4] = """Yes, our pitch will work in the current market. And I won't kid you, the shenanigans in the MMO space lately don't make that any easier. Part of the pitch is establishing the differentiation we have from what's caused other games to fail.

I'm sure if I opened up the table right now, you could give me a long critique on what affected New World. Some of you in this room may have worked on that game. But what has to be said in our pitch, whether it's investors or players or partners, is how we differentiate from what's jaded people looking at the MMO market in the first place.

And the honest answer starts with the audience. There's a massive number of players who loved MMOs but walked away because the genre asked too much of their lives. They didn't stop wanting the experience. They stopped being able to afford the time. Look at where those players went. Fortnite social, FFXIV housing, Animal Crossing, Palia. The behaviour is still there. We're building for those players, and we're building for the MMO players who are still in the genre and want something that respects them more. We're not asking people to choose us over their existing game. We're asking them to come home between sessions of whatever else they play.

That changes the pitch fundamentally, because we're not a WoW killer. I've protected against that deliberately. I watched it happen at Blizzard and after I left. Many games try to become WoW killers. That is not our objective. That is not our purpose.

The other thing that changes the pitch is how we stay alive. Most MMOs that fail don't actually fail at launch. They fail twelve months later when the content pipeline can't keep up. Our partner ecosystem changes that equation. Other studios bring their IP into our world as arcade machines, cosmetics, events, whole shards. The world keeps getting fresh content we don't have to ship entirely on our own. That's not hope, that's a production strategy.

Now underneath all of that is something people don't always think about, but it matters as much as any feature. The health of the studio comes out directly in the game. The direction of leadership manifests in the product. Anybody here who's worked on New World, or any game for that matter, can feel the influences that affect how a game gets built and why. A healthy studio shows up in the game. An unhealthy one shows up too, and usually louder.

We understand that. The studio is as important as the game from a pitch perspective. What matters most about the product itself is that we have a very clear discipline on what we're building and why, when it goes in, how it integrates, and what it creates. We know the themes we're fitting all the content to and we know why those themes.

One of the contexts we keep coming back to is balance. The game can't be too small, or it loses players because there's nothing to do. It can't aim too big, or the purpose becomes convoluted and we can't execute against it. The tone and goals inside the game need to be oriented with absolute clarity, and the studio behind it has to be of one mind, building in one direction.

That is a big part of our pitch. And it's a big reason why we're here talking to you today."""

# Filter out merged questions, keep the rest
keep_qnums = set()
merged_into = set(merges.keys())
for s in source:
    if s['qnum'] not in merged_into:
        keep_qnums.add(s['qnum'])

# Build output
wb_out = openpyxl.Workbook()
ws = wb_out.active
ws.title = "Glen Answers"

header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
body_font = Font(name="Calibri", size=10, color="000000")
wrap_align = Alignment(wrap_text=True, vertical="top")
thin_border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)
answered_fill = PatternFill(start_color="CCE5CC", end_color="CCE5CC", fill_type="solid")
unanswered_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

headers = ["Q#", "Topic", "Individual Question", "Orig Row", "Crit (Personal)", "Crit (Studio)", "Glen", "Status"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = wrap_align
    cell.border = thin_border

new_qnum = 0
for s in source:
    if s['qnum'] in merged_into:
        continue
    new_qnum += 1
    old_qnum = s['qnum']

    question_text = question_updates.get(old_qnum, s['question'])
    glen_text = glen_answers.get(old_qnum, '')
    status = "ANSWERED" if glen_text else "PENDING"

    row = new_qnum + 1
    values = [
        new_qnum,
        s['topic'],
        question_text,
        s['orig_row'],
        s['crit_p'],
        s['crit_s'],
        glen_text,
        status,
    ]
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = body_font
        cell.alignment = wrap_align
        cell.border = thin_border

    status_cell = ws.cell(row=row, column=8)
    if status == "ANSWERED":
        status_cell.fill = answered_fill
    else:
        status_cell.fill = unanswered_fill

col_widths = [5, 28, 65, 8, 10, 10, 80, 12]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.row_dimensions[1].height = 30
for r in range(2, new_qnum + 2):
    ws.row_dimensions[r].height = 100

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:H{new_qnum+1}"

output_path = 'Clients/Couch Heroes/CH_AMA_Questions_Broken_Out2.xlsx'
wb_out.save(output_path)
print(f"Saved: {output_path}")
print(f"Total questions: {new_qnum} (was 87, merged 14)")
answered = sum(1 for q in range(1, new_qnum+1) if q <= 4)
print(f"Answered: {answered}, Pending: {new_qnum - answered}")

# Print the new numbering for reference
print("\n=== NEW QUESTION LIST ===")
new_q = 0
for s in source:
    if s['qnum'] in merged_into:
        continue
    new_q += 1
    old_qnum = s['qnum']
    qt = question_updates.get(old_qnum, s['question'])
    has_a = "DONE" if old_qnum in glen_answers else "    "
    print(f"Q{new_q:2d} [{has_a}] [{s['topic']}] {qt[:110]}")
