"""
Build the Couch Heroes WorkSage import workbook + true-up inventory.

Reads:
- Glen_work_list_v10.csv (150 items)
- CH_Artifacts_Project_Plan_FINAL_v3.xlsx (UK Company artifacts)
- Merged_Hiring_Plan_final_v2.xlsx (30 hiring positions)
- Couch Heroes Features.xlsx (421 game features)

Writes:
- _trueup_inventory.md (audit trail)
- CH_WorkSage_import_v1.xlsx (deliverable)

Applies Glen's Q1-Q13 decisions from spec.
"""
import csv
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment

CH_DIR = '.'

# ============================================================
# Q1-Q13 DECISIONS
# ============================================================
DECISIONS = {
    'Q1_CTO_dead': True,                 # Cam, Neil, Nia all cancelled
    'Q2_coaching_to_HR': True,           # Exec coaching → HR not EP Hiring
    'Q3_lead_animator_done': True,       # Alon hired
    'Q4_anim_questions_done': True,      # Question set marked Done; locate-and-document task added
    'Q5_lili_separate': True,            # New COS hire
    'Q7_dino_review_done': True,         # Performance review delivered; exit timeline tentative
    'Q9_jira_walk_postponed': True,      # Target ~mid-May post-offsite
    'Q10_offsite_4days': True,           # Apr 27-30
    'Q11_use_features_xlsx': True,       # Couch Heroes Features.xlsx IS the Miro export
    'Q12_1on1_programme_plus_tasks': True, # 1 story + 8 tasks
    'Q13_sow_removed': True,             # SOW belongs in NBI
}

# ============================================================
# CSV PARENT TASK → (PROJECT, FEATURE) MAPPING
# ============================================================
# Project list (locked structure 7+AI):
PROJECTS = [
    'Studio Operations',
    'HR',
    'UK Company Setup',
    'Production Strategy & Implementation',
    'Game',
    'Studio Business',
    'AI',
]

# Direct CSV parent → (project, feature) mappings
PARENT_MAP = {
    'Coaching Programme':              ('HR', 'Coaching Programme'),
    'JIRA Implementation & Rollout':   ('Production Strategy & Implementation', 'JIRA Implementation & Rollout'),
    'Gate System Design & Rollout':    ('Production Strategy & Implementation', 'Gate System Design & Rollout'),
    'Agilefall Methodology Rollout':   ('Production Strategy & Implementation', 'Agilefall Methodology Rollout'),
    'Roadmap & Cadence Rollout':       ('Production Strategy & Implementation', 'Roadmap & Cadence Rollout'),
    'CTO Hiring':                      ('HR', 'Hiring Pipeline — Engineering'),
    'Executive Producer Hiring':       ('HR', 'Hiring Pipeline — Production & Design'),
    'Hiring Process & Infrastructure': ('HR', 'Hiring Pipeline — Process'),
    'Pay & Performance Governance':    ('HR', 'Pay & Performance'),
    'Performance Review Process':      ('HR', 'Pay & Performance'),
    'Salary Review Delivery':          ('HR', 'Salary Review Delivery (CLOSED)'),
    'Publishing and BD':               ('Studio Business', 'BD'),
    'Gamescom Preparations':           ('__KILLED__', 'Gamescom (Cancelled)'),
    'UK Transition Planning':          ('UK Company Setup', 'UK Transition Operational'),
    'Roles & Responsibilities Framework': ('Studio Operations', 'Org Design & Structure'),
    'Org Design Presentation & Strategy': ('Studio Operations', 'Org Design & Structure'),
    'Game Vision and Design':          ('Game', 'Game Vision & Design'),
}

# Items where "Org Design Presentation & Strategy" parent is wrong — disambiguate by title
TITLE_OVERRIDES = {
    # Production items mis-parented under Org Design
    'Gate System Design & Rollout':       ('Production Strategy & Implementation', 'Gate System Design & Rollout'),
    'Agilefall Methodology Rollout':      ('Production Strategy & Implementation', 'Agilefall Methodology Rollout'),
    'Roadmap & Cadence Rollout':          ('Production Strategy & Implementation', 'Roadmap & Cadence Rollout'),
    # QA / UX Research items mis-parented under Org Design — these go to HR Hiring + Studio Ops
    'QA Structure & UX Research':         ('Studio Operations', 'Org Design & Structure'),
    'Meet with Hannah: cover QA plan, headcount needs, internal vs external split':
                                          ('HR', 'Hiring Pipeline — QA'),
    'Review Hannah Pickard for combined QA / UX Research role':
                                          ('HR', 'Hiring Pipeline — QA'),
    'Set up one with Hannah':             ('HR', 'Hiring Pipeline — QA'),
    'Review and assess primary tool set, specifically test rail and purchase tool set':
                                          ('Studio Operations', 'IT & Systems'),
    'Align with Hannah on overall team size at phases of development':
                                          ('HR', 'Hiring Pipeline — QA'),
    'Get bids for external UX research':  ('HR', 'Hiring Pipeline — Production & Design'),
    'Send UX agency details':             ('HR', 'Hiring Pipeline — Production & Design'),
    # Hiring items mis-parented
    'Lead Animator Hiring':               ('HR', 'Hiring Pipeline — DONE'),
    'Build question set for Lead Animator interviews':
                                          ('HR', 'Hiring Pipeline — DONE'),
    'create lead level designer JD':      ('HR', 'Hiring Pipeline — Production & Design'),
    'CTO Hiring':                         ('HR', 'Hiring Pipeline — Engineering'),
    'Hire Lead Animator':                 ('HR', 'Hiring Pipeline — DONE'),
    'Talk with David about Lead Animator (current candidate may not be best choice)':
                                          ('HR', 'Hiring Pipeline — DONE'),
    # Non-CH item per Q13
    'Statement of Work administration':   ('__REMOVED__', 'SOW (moved to NBI)'),
    # Coaching items wrongly under EP Hiring (Q2)
    'Executive Coaching Programme':       ('HR', 'Coaching Programme'),
    # 1:1 items → consolidate per Q12
    'Establish and maintain one-on-ones with Dino':     ('HR', '1:1s — Weekly Programme'),
    'Establish and maintain one-on-ones with Lorenza':  ('HR', '1:1s — Weekly Programme'),
    'Establish and maintain one-on-ones with Vardis':   ('HR', '1:1s — Weekly Programme'),
    'Establish and maintain one-on-ones with Aris':     ('HR', '1:1s — Weekly Programme'),
    'Establish and maintain one-on-ones with Mustafa':  ('HR', '1:1s — Weekly Programme'),
    'Establish and maintain one-on-ones with Robin':    ('HR', '1:1s — Weekly Programme'),
    'Establish and maintain one-on-ones with David':    ('HR', '1:1s — Weekly Programme'),
    'Establish and maintain one-on-ones with Valeria':  ('HR', '1:1s — Weekly Programme'),
    'Weekly coaching: Dino 1:1':                        ('HR', '1:1s — Weekly Programme'),
    'Weekly coaching: Department Leads + bi-weekly 1:1s': ('HR', '1:1s — Weekly Programme'),
    # Department Leads Development
    'Department Leads Development':                     ('HR', 'Coaching Programme'),
    'AI Policy: verify final presentation item':        ('AI', 'Usage Policy & Governance'),
}

# Status overrides per Q&A
STATUS_OVERRIDES = {
    # Q1: CTO restart — kill Cam-related items
    'Help close the CTO role': 'Cancelled',  # Cam-specific
    'Identify key CTO candidates': 'Cancelled',  # Cam was the candidate
    'CTO Hiring': 'In progress',  # restart
    # Q3 + Q4: Lead Animator
    'Lead Animator Hiring': 'Done',
    'Build question set for Lead Animator interviews': 'Done',
    # Q7: Dino performance review delivered
    "Decide on Dino's performance review (on hold until Aris)": 'Done',
    # Q9: JIRA walk postponed
    'Walk through with leads on February 18th': 'Not started',
    # Q13: SOW removed - status not relevant
}

# Description appendices (notes, references)
DESC_APPEND = {
    'CTO Hiring': '\n\nQ1 (2026-04-26): Cam, Neil, Nia ALL DEAD. Restarting search. Question set staged at Clients/Couch Heroes/CTO_Interview_Questions_Couch_Heroes.docx. Scorecard at CTO_Scorecard.xlsx. Recruiting brief at CTO_Recruiting_Brief_v6.pdf. Need to refresh and resource new candidates.',
    'Help close the CTO role': '\n\nQ1: Cam dead. Cancelled.',
    'Identify key CTO candidates': '\n\nQ1: Cam was the candidate. Cancelled. Need fresh candidate identification.',
    'Lead Animator Hiring': '\n\nQ3 (2026-04-26): Same role as Alon. Marked Done.',
    'Build question set for Lead Animator interviews': '\n\nQ4 (2026-04-26): Done. ACTION: locate the questions actually used (no separate doc on disk), document them, attach during WorkSage upload.',
    "Decide on Dino's performance review (on hold until Aris)": '\n\nQ7 (2026-04-26): Performance review was already DELIVERED. Exit timeline (1mo transition + 2mo garden leave) is TENTATIVE — see Dino exit story for tracking.',
    'Walk through with leads on February 18th': '\n\nQ9 (2026-04-26): Postponed. Target ~2 weeks after offsite (mid-May), conditional on roadmap + process flow agreement first.',
    'Statement of Work administration': '\n\nQ13 (2026-04-26): NOT a Couch Heroes item. Belongs in NBI. Removed from CH backlog.',
    'Executive Coaching Programme': '\n\nQ2 (2026-04-26): Moved from EP Hiring to HR → Coaching Programme. The coaching is separate work from hiring Graham (who has already accepted, ~May 18 start).',
}

# ============================================================
# READ CSV
# ============================================================
def read_csv():
    rows = []
    with open(f'{CH_DIR}/Glen_work_list_v10.csv','r',encoding='utf-8') as f:
        reader = csv.reader(f)
        header = next(reader)
        for r in reader:
            if not any(r):
                continue
            # pad
            while len(r) < 9:
                r.append('')
            rows.append({
                'title': r[0].strip(),
                'parent': r[1].strip(),
                'status': r[2].strip(),
                'priority': r[3].strip(),
                'description': r[4].strip(),
                'assignee': r[5].strip(),
                'files': r[6].strip(),
                'health': r[7].strip(),
                'attachment': r[8].strip(),
            })
    return rows

# ============================================================
# MAP A CSV ROW → (project, feature, story_or_task)
# ============================================================
def map_row(row):
    title = row['title']
    parent = row['parent']
    # Title-level override first
    if title in TITLE_OVERRIDES:
        proj, feat = TITLE_OVERRIDES[title]
    elif parent in PARENT_MAP:
        proj, feat = PARENT_MAP[parent]
    elif title in PARENT_MAP:
        # Sometimes the title IS a parent name (parent of itself)
        proj, feat = PARENT_MAP[title]
    elif title == '' and parent in PARENT_MAP:
        proj, feat = PARENT_MAP[parent]
    elif parent == '':
        # Top-level orphan
        proj, feat = ('__UNMAPPED__', '__UNMAPPED__')
    else:
        proj, feat = ('__UNMAPPED__', f'parent={parent}')
    return proj, feat

# ============================================================
# READ UK ARTIFACTS
# ============================================================
def read_uk_artifacts():
    """Returns list of UK artifact dicts."""
    try:
        wb = load_workbook(f'{CH_DIR}/CH_Artifacts_Project_Plan_FINAL_v3.xlsx', data_only=True)
    except Exception as e:
        print('UK xlsx read error:', e)
        return []
    artifacts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # find header row
        header_row = None
        for r in range(1, min(15, ws.max_row+1)):
            row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column+1)]
            if any(isinstance(v, str) and ('Artifact' in v or 'Ref' in v or 'Reference' in v or 'Title' in v) for v in row if v):
                header_row = r
                headers = [str(v) if v else f'col{i}' for i, v in enumerate(row)]
                break
        if not header_row:
            continue
        for r in range(header_row+1, ws.max_row+1):
            row = [ws.cell(row=r, column=c).value for c in range(1, len(headers)+1)]
            if not any(row):
                continue
            artifacts.append({
                'sheet': sheet_name,
                'data': dict(zip(headers, row)),
            })
    return artifacts

# ============================================================
# READ HIRING PLAN
# ============================================================
def read_hiring_plan():
    """Roles in 'C Level approved' sheet: col D = role name, col E = annual salary, col H = start month."""
    try:
        wb = load_workbook(f'{CH_DIR}/Merged_Hiring_Plan_final_v2.xlsx', data_only=True)
    except Exception as e:
        print('Hiring xlsx read error:', e)
        return []
    rows = []
    ws = wb['C Level approved']
    seen = set()
    for r in range(5, ws.max_row+1):
        role = ws.cell(row=r, column=4).value
        if not role or not isinstance(role, str):
            continue
        role = role.strip()
        # Skip empty / non-role rows
        if not role or role in seen or len(role) < 3:
            continue
        # Filter obvious headers / sub-totals
        if role.lower() in ('total','subtotal','approved','hiring plan','headcount & salaries'):
            continue
        seen.add(role)
        salary = ws.cell(row=r, column=5).value
        bench = ws.cell(row=r, column=6).value
        currency = ws.cell(row=r, column=7).value
        start_mo = ws.cell(row=r, column=8).value
        cost_type = ws.cell(row=r, column=9).value
        rows.append({'sheet': 'C Level approved', 'data': {
            'Role': role,
            'Annual Salary (000s)': salary,
            'Benchmark Salary': bench,
            'Original Currency': currency,
            'Start Month': start_mo,
            'Cost Type': cost_type,
        }})
    return rows

# ============================================================
# READ FEATURES V2 (Couch_Heroes_Features_v2.xlsx) - GAME
# ============================================================
# Glen's structure (confirmed 2026-04-27):
#   Game (project)
#   - Game Features         <- Game Feature List sheet
#   - Game Systems          <- TDD sheet
#   - Platform Features     <- Platform Feature List sheet
#   - Live Service Features <- Live Service Features sheet
#   - Content               <- Content List sheet
#   - Feature Versions      <- Feature Versions sheet (V0-V7 maturity per feature)

def _read_hier_sheet(ws, sub_feature, level_cols=(2,3,4,5),
                     progress_col=9, priority_col=13, dod_col=15,
                     comments_col=10, owner_col=11, integration_col=7,
                     start_row=2):
    """Generic reader for V2 sheets that have Tag/Epic/Feature/Story/Task hierarchy.

    level_cols indices: (epic, feature, story, task) - 1-indexed Excel columns.
    Returns list of items with epic/feature/story/task tracking.
    """
    epic_col, feature_col, story_col, task_col = level_cols
    rows = []
    current_epic = None
    current_feature = None
    current_story = None
    for r in range(start_row, ws.max_row+1):
        epic = ws.cell(row=r, column=epic_col).value
        feature = ws.cell(row=r, column=feature_col).value
        story = ws.cell(row=r, column=story_col).value
        task = ws.cell(row=r, column=task_col).value
        progress = ws.cell(row=r, column=progress_col).value if progress_col else None
        priority = ws.cell(row=r, column=priority_col).value if priority_col else None
        dod = ws.cell(row=r, column=dod_col).value if dod_col else None
        comments = ws.cell(row=r, column=comments_col).value if comments_col else None
        owner = ws.cell(row=r, column=owner_col).value if owner_col else None
        integration = ws.cell(row=r, column=integration_col).value if integration_col else None

        if epic:
            current_epic = str(epic).strip()
            current_feature = None
            current_story = None
        if feature:
            current_feature = str(feature).strip()
            current_story = None
        if story:
            current_story = str(story).strip()

        # lowest-level item determines this row's level
        if task:
            level, title = 'task', str(task).strip()
        elif story:
            level, title = 'story', str(story).strip()
        elif feature:
            level, title = 'feature', str(feature).strip()
        elif epic:
            level, title = 'epic', str(epic).strip()
        else:
            continue

        # Build description from comments + owner + integration + dod
        desc_parts = []
        if dod:
            desc_parts.append(f'DoD: {dod}')
        if comments:
            desc_parts.append(f'Comments: {comments}')
        if owner:
            desc_parts.append(f'Owner: {owner}')
        if integration:
            desc_parts.append(f'Integration: {integration}')
        desc = ' | '.join(str(x) for x in desc_parts)

        rows.append({
            'sub_feature': sub_feature,
            'epic': current_epic or 'Unspecified',
            'feature': current_feature or '',
            'story': current_story or '',
            'level': level,
            'title': title,
            'progress': progress,
            'priority': priority,
            'description': desc,
        })
    return rows


def read_game_v2():
    """Read all 6 game-relevant sheets from V2 file. Returns list of items."""
    try:
        wb = load_workbook(f'{CH_DIR}/Couch_Heroes_Features_v2.xlsx', data_only=True)
    except Exception as e:
        print('Features V2 read error:', e)
        return []

    items = []

    # 1. Game Features - Game Feature List sheet
    ws = wb['Game Feature List']
    # Headers row 1: Tag/Fability co, Epic, Feature, Story, Task, System Required, Integration, GDD, Progress, Comments, Owner, Player Promise, Player Need, Success Factor, Definition of Done
    items.extend(_read_hier_sheet(ws, 'Game Features',
        level_cols=(2,3,4,5), progress_col=9, priority_col=None, dod_col=15,
        comments_col=10, owner_col=11, integration_col=7))

    # 2. Game Systems - TDD sheet (Tag, Epic, Feature, Story, Task, System Required, Integration, TDD, Progress, Comments, Owner, Player Promise, Priority, ETA, Actual Time)
    ws = wb['TDD']
    items.extend(_read_hier_sheet(ws, 'Game Systems',
        level_cols=(2,3,4,5), progress_col=9, priority_col=13, dod_col=None,
        comments_col=10, owner_col=11, integration_col=7))

    # 3. Platform Features - Platform Feature List sheet (same structure as Game Feature List)
    ws = wb['Platform Feature List']
    items.extend(_read_hier_sheet(ws, 'Platform Features',
        level_cols=(2,3,4,5), progress_col=9, priority_col=None, dod_col=15,
        comments_col=10, owner_col=11, integration_col=7))

    # 4. Live Service Features - same structure
    ws = wb['Live Service Features']
    items.extend(_read_hier_sheet(ws, 'Live Service Features',
        level_cols=(2,3,4,5), progress_col=9, priority_col=None, dod_col=15,
        comments_col=10, owner_col=11, integration_col=7))

    # 5. Content - Content List sheet (Tag, Epic, Feature, Content Task, Art Team, System Required, Integration, GDD, Progress, Comments, Owner, Definition of Done, Priority, ETA, Actual Time)
    # No 'Story' col here - col 4 is 'Content Task'. Treat content tasks as tasks.
    # NB: source has occasional multi-task cells (newline-separated). Split those.
    ws = wb['Content List']
    # Use level_cols=(2,3,4,4) so the 'Content Task' col fills both story/task slots harmlessly
    # Actually we want: epic=2, feature=3, task=4 (no story level)
    # The reader handles missing levels - pass story_col same as task_col? Better: pass story_col as a non-existent col then handle:
    # Simpler: read manually
    current_epic = None
    current_feature = None
    for r in range(2, ws.max_row+1):
        epic = ws.cell(row=r, column=2).value
        feature = ws.cell(row=r, column=3).value
        ctask = ws.cell(row=r, column=4).value
        art_team = ws.cell(row=r, column=5).value
        progress = ws.cell(row=r, column=9).value
        comments = ws.cell(row=r, column=10).value
        owner = ws.cell(row=r, column=11).value
        dod = ws.cell(row=r, column=12).value
        priority = ws.cell(row=r, column=13).value
        if epic:
            current_epic = str(epic).strip()
            current_feature = None
        if feature:
            current_feature = str(feature).strip()
        # Build a list of titles to emit. Split on newlines if a single cell contains multiple tasks.
        emit_titles = []
        if ctask:
            ct_str = str(ctask).strip()
            # Split on newline if multi-line cell
            for sub in ct_str.split('\n'):
                sub = sub.strip()
                if sub:
                    emit_titles.append(('task', sub))
        elif feature:
            emit_titles.append(('feature', str(feature).strip()))
        elif epic:
            emit_titles.append(('epic', str(epic).strip()))
        else:
            continue
        for level, title in emit_titles:
            desc_parts = []
            if art_team:
                desc_parts.append(f'Art team: {art_team}')
            if dod:
                desc_parts.append(f'DoD: {dod}')
            if comments:
                desc_parts.append(f'Comments: {comments}')
            if owner:
                desc_parts.append(f'Owner: {owner}')
            items.append({
                'sub_feature': 'Content',
                'epic': current_epic or 'Unspecified',
                'feature': current_feature or '',
                'story': '',
                'level': level,
                'title': title,
                'progress': progress,
                'priority': priority,
                'description': ' | '.join(str(x) for x in desc_parts),
            })

    # 6. Feature Versions - Feature Versions sheet
    # Header row 1: Feature Name, V0, V1, V2, V3, V4, V5, V6, V7
    # Each row = one feature with progression definitions per version
    ws = wb['Feature Versions']
    for r in range(2, ws.max_row+1):
        feature_name = ws.cell(row=r, column=1).value
        if not feature_name or not isinstance(feature_name, str):
            continue
        v_descs = []
        for vc in range(2, 10):  # cols B..I = V0..V7
            v_val = ws.cell(row=r, column=vc).value
            if v_val and str(v_val).strip() not in ('-',''):
                v_descs.append(f'V{vc-2}: {str(v_val).strip()}')
        if not v_descs:
            continue
        items.append({
            'sub_feature': 'Feature Versions',
            'epic': 'Feature Versions Maturity Map',
            'feature': '',
            'story': '',
            'level': 'story',
            'title': str(feature_name).strip(),
            'progress': None,
            'priority': None,
            'description': ' | '.join(v_descs),
        })

    return items

# ============================================================
# STATUS / PRIORITY NORMALISATION
# ============================================================
# ============================================================
# CLASSIFICATION FRAMEWORK (per Q14, derived from create-architecture.md
# layer model + create-epics.md epic-per-architectural-module + team-live-ops.md)
# ============================================================
# Two dimensions:
#   type  = System | Feature | Content | Platform | LiveService
#   layer = Foundation | Core | Feature | Presentation | Polish
#
# Layer definitions (from create-architecture.md):
#   Foundation = engine integration, save/load, scene management, event bus, networking, online services
#   Core       = physics, input, combat, movement, base inventory mechanics
#   Feature    = gameplay systems built on Core (quests, AI, crafting, magic, economy)
#   Presentation = UI, HUD, menus, VFX, audio, social-facing surfaces
#   Polish     = accessibility, tutorial, optimisations
#
# Type definitions:
#   System  = architectural module backbone — built once, foundation for everything else
#   Feature = player-facing capability built on top of a System
#   Content = data/asset instance populating Systems and Features
#   Platform = web/mobile companion product (separate stack/team/cadence)
#   LiveService = post-launch operational (events, seasons, battle pass)

# Heuristic keywords per layer
LAYER_KEYWORDS = {
    'Foundation': [
        'online services', 'backend', 'save', 'load', 'persistence', 'serialization',
        'serialisation', 'scene management', 'event bus', 'networking', 'multiplayer',
        'server', 'shard', 'replication', 'database', 'analytics system',
        'telemetry', 'metric', 'instrumentation', 'cdn', 'patch', 'manifest',
        'engine integration', 'devops', 'build pipeline', 'tooling',
    ],
    'Core': [
        'combat', 'movement', 'navigation', 'physics', 'input', 'controls',
        'item core', 'inventory system', 'stat system', 'attribute', 'character controller',
        'camera', 'targeting', 'hit detection', 'damage', 'core combat',
    ],
    'Feature': [
        'crafting', 'gathering', 'quest', 'magic', 'corruption', 'fishing',
        'profession', 'achievement', 'faction', 'guild', 'pets', 'forge',
        'minigame', 'trading', 'auction', 'currency', 'economy', 'loot',
        'equipment', 'weapon', 'ability', 'skill', 'talent', 'puzzle',
        'platforming', 'world cycle', 'day/night', 'weather', 'corruption',
        'tutorial', 'progression', 'pet', 'mount', 'companion',
    ],
    'Presentation': [
        'ui', 'hud', 'menu', 'screen', 'interface', 'tooltip', 'notification',
        'icon', 'vfx', 'visual', 'presentation', 'animation', 'audio', 'sound',
        'music', 'sfx', 'chat', 'mail', 'friend', 'social', 'voice', 'subtitle',
        'localization', 'localisation', 'minimap',
    ],
    'Polish': [
        'accessibility', 'colorblind', 'screen reader', 'remap', 'optimization',
        'optimisation', 'qa', 'release engineering', 'support system', 'help codex',
        'modern ux',
    ],
}

# Indicators a row is a SYSTEM (architectural module)
SYSTEM_INDICATORS = [
    ' system', ' epic', 'framework', 'architecture', 'pipeline', 'service',
    'engine', 'controller', 'backend', 'platform certification',
]

# Indicators a row is CONTENT (asset/data, not mechanic)
CONTENT_INDICATORS = [
    'icon', 'animations', 'animation', 'model', 'models', 'vfx', 'concept',
    'character art', 'texture', 'asset', 'sound effect', 'music track',
    'cinematic', 'cutscene', 'lore text', 'dialogue line', 'flavor text',
    'flavour text', 'splash', 'wallpaper', 'logo', 'partner skin',
]

def classify_game_item(title, sub_feature_source, epic='', feature='', story=''):
    """Return (type, layer, confidence) for a Game project item.

    Auto-classify by source sheet for unambiguous cases; heuristic for Game Feature List.
    """
    title_l = (title or '').lower()
    epic_l = (epic or '').lower()
    feature_l = (feature or '').lower()

    # Hard rules by source sheet
    if sub_feature_source == 'Content':
        return 'Content', 'Presentation', 'high'
    if sub_feature_source == 'Platform Features':
        return 'Platform', 'Foundation', 'high'
    if sub_feature_source == 'Live Service Features':
        return 'LiveService', 'Feature', 'high'
    if sub_feature_source == 'Game Systems':
        # TDD sheet — these are the technical systems
        layer = _infer_layer(title_l, epic_l, feature_l)
        return 'System', layer, 'high'
    if sub_feature_source == 'Feature Versions':
        # Maturity overlay — type/layer N/A
        return 'Maturity', 'Polish', 'high'

    # Game Feature List — heuristic classify
    # First, content check
    if any(c in title_l for c in CONTENT_INDICATORS):
        return 'Content', 'Presentation', 'medium'
    # System check
    if any(s in title_l for s in SYSTEM_INDICATORS):
        layer = _infer_layer(title_l, epic_l, feature_l)
        return 'System', layer, 'medium'
    # Default: Feature
    layer = _infer_layer(title_l, epic_l, feature_l)
    # Confidence: high if any layer keyword strongly matches; else medium; else low
    confidence = 'medium'
    if not any(kw in (title_l + epic_l + feature_l) for kws in LAYER_KEYWORDS.values() for kw in kws):
        confidence = 'low'
    return 'Feature', layer, confidence


def _infer_layer(title_l, epic_l, feature_l):
    text = ' '.join([title_l, epic_l, feature_l])
    # Score each layer by keyword hits
    best_layer = 'Feature'  # default
    best_score = 0
    for layer, keywords in LAYER_KEYWORDS.items():
        score = sum(1 for kw in keywords if kw in text)
        if score > best_score:
            best_score = score
            best_layer = layer
    return best_layer


# ============================================================
# AGENT-FILE INGEST (5 markdown tables produced by parallel agents)
# ============================================================

def read_md_table(filepath):
    """Read a pipe-delimited markdown table. Skip headers, separators, narrative.

    Returns list of dicts keyed by header column names.
    """
    import os
    if not os.path.exists(filepath):
        return []
    with open(filepath, 'r', encoding='utf-8') as f:
        lines = f.read().splitlines()
    headers = None
    rows = []
    for line in lines:
        line = line.strip()
        if not line.startswith('|'):
            continue
        # split, strip outer pipes, then cells
        cells = [c.strip() for c in line.strip('|').split('|')]
        # Separator row (---)
        if all(set(c) <= set('-: ') for c in cells if c):
            continue
        if headers is None:
            headers = cells
            continue
        if len(cells) < len(headers):
            cells += [''] * (len(headers) - len(cells))
        rows.append(dict(zip(headers, cells[:len(headers)])))
    return rows


def read_game_classifications():
    """Build lookup map: key=(source_sheet, title) -> classification dict.

    Agent's format: only one of (epic, feature, story, task) is filled per row,
    and that's the title at that level. So we collapse to (sheet, title) for lookup.
    Collisions (same title in same sheet) are rare; later wins.
    """
    rows = read_md_table('./build_inputs/game_classification.md')
    lookup = {}
    for r in rows:
        sheet = (r.get('source_sheet') or '').strip()
        # The title is whichever of epic/feature/story/task is filled (lowest priority)
        title = (
            (r.get('task') or '').strip() or
            (r.get('story') or '').strip() or
            (r.get('feature') or '').strip() or
            (r.get('epic') or '').strip()
        )
        if not (sheet and title):
            continue
        lookup[(sheet, title)] = {
            'type': (r.get('type') or '').strip(),
            'layer': (r.get('layer') or '').strip(),
            'confidence': (r.get('confidence') or '').strip(),
            'reason': (r.get('needs_review_reason') or '').strip(),
        }
    return lookup


# Map V2 sub_feature names to source_sheet labels used in classification file
SUB_FEATURE_TO_SHEET = {
    'Game Features': 'Game Feature List',
    'Game Systems': 'TDD',
    'Platform Features': 'Platform Feature List',
    'Live Service Features': 'Live Service Features',
    'Content': 'Content List',
    'Feature Versions': 'Feature Versions',
}


def normalise_priority_extended(p):
    """Extended priority normaliser to handle all agent-output deviations."""
    if not p or not isinstance(p, str):
        return 'Medium'
    pl = p.strip().lower()
    if pl in ('critical', 'critical act', 'crit'):
        return 'Critical ACT'
    if pl == 'urgent':
        return 'Urgent'
    if pl == 'high':
        return 'High'
    if pl == 'medium' or pl == 'med':
        return 'Medium'
    if pl == 'low':
        return 'Low'
    return normalise_priority(p)


def normalise_status_extended(s):
    """Extended status normaliser to handle agent-output deviations."""
    if not s or not isinstance(s, str):
        return 'Not started'
    sl = s.strip().lower()
    # Agent invented vocabulary: open / in_progress / blocked / please_review / resolved
    mapping = {
        'open': 'Not started',
        'in_progress': 'In progress', 'in progress': 'In progress',
        'blocked': 'Blocked',
        'please_review': 'In Review', 'please review': 'In Review', 'in review': 'In Review',
        'resolved': 'Done', 'done': 'Done',
        'not started': 'Not started', 'not_started': 'Not started',
        'planning': 'Planning',
        'drafted': 'Drafted', 'draft': 'Drafted',
        'cancelled': 'Cancelled', 'canceled': 'Cancelled',
    }
    if sl in mapping:
        return mapping[sl]
    return normalise_status(s)


def normalise_health(h):
    """Normalise health values - Green/Yellow/Red/Waiting on Client/blank."""
    if not h or not isinstance(h, str):
        return ''
    hl = h.strip().lower()
    mapping = {
        'green': 'Green',
        'yellow': 'Yellow', 'amber': 'Yellow',
        'red': 'Red',
        'waiting on client': 'Waiting on Client', 'waiting': 'Waiting on Client',
        'unknown': '', 'tbd': '', '': '',
    }
    return mapping.get(hl, '')


def normalise_work_type(w):
    """Normalise work_type values to spec set: Research/Strategy/Implementation/Assessment/Ongoing Mgmt."""
    if not w or not isinstance(w, str):
        return 'Implementation'
    valid = {'Research', 'Strategy', 'Implementation', 'Assessment', 'Ongoing Mgmt'}
    w = w.strip()
    if w in valid:
        return w
    wl = w.lower()
    # Research
    if any(k in wl for k in ('research', 'discovery', 'analysis', 'audit', 'investigation')):
        return 'Research'
    # Strategy
    if any(k in wl for k in ('strateg', 'plan', 'design', 'org', 'decision', 'workshop')):
        return 'Strategy'
    # Assessment
    if any(k in wl for k in ('assess', 'review', 'verification', 'measurement', 'reporting')):
        return 'Assessment'
    # Ongoing Mgmt
    if any(k in wl for k in ('ongoing', 'mgmt', 'manage', 'governance', 'process', 'compliance', 'change')):
        return 'Ongoing Mgmt'
    # Implementation (default for: tooling, config, policy, comms, asset, build, integration, migration,
    # pilot, training, enablement, onboarding, automation, documentation, communications, knowledge, technical)
    return 'Implementation'


# UK Company Setup duplicate detection: keywords that indicate an item duplicates
# a UK artifact already tracked in the UK Company Setup project.
UK_DUPLICATE_PATTERNS = [
    ('director identity verification', 'UK 5.3 Director Identity Verification'),
    ('register for corporation tax', 'UK 5.4 Corporation Tax Registration'),
    ('vat registration', 'UK 5.5 VAT Registration'),
    ('auto-enrolment pension', 'UK 1.6 Pension Auto-Enrolment Letter'),
    ('right to work check', 'UK 1.4 Right to Work Checks'),
    ('health and safety policy', 'UK 1.5 Health and Safety Policy'),
    ('bribery act', 'UK 4.9 Bribery Act Policy'),
    ('modern slavery', 'UK 4.10 Modern Slavery Statement'),
    ('employee handbook', 'UK 2.1 Employee Handbook'),
    ('dpia', 'UK 2.5 DPIA Template'),
    ('cross-border data transfer', 'UK 2.6 Cross-Border Data Transfer Agreement'),
    ('contract template - employment', 'UK 1.1 Employment Contract'),
    ('contract template - mutual nda', 'UK 2.3 Confidentiality / NDA'),
    ('record retention', 'UK 2.7 Record Retention Policy'),
    ('sponsor licence application', 'UK 3.x Sponsor Licence'),
    ('sponsor licence hr compliance', 'UK 3.1 HR Compliance Pack'),
    ('sponsored worker file', 'UK 3.2 Sponsored Worker File Template'),
    ('confirmation statement', 'UK 5.x Companies House Filings'),
    ('employers\' liability insurance', 'UK 1.7 EL Insurance Certificate'),
    ('emi option scheme', 'UK 4.12 EMI Share Scheme'),
]


def detect_uk_duplicate(title):
    """If title likely duplicates a UK Company Setup artifact, return the UK ref. Else None."""
    if not title or not isinstance(title, str):
        return None
    tl = title.lower()
    for pattern, ref in UK_DUPLICATE_PATTERNS:
        if pattern in tl:
            return ref
    return None


def normalise_status(s):
    if s is None or s == '':
        return 'Not started'
    if not isinstance(s, str):
        # Numeric progress (e.g., 0.4 = 40% complete) -> map to status
        try:
            f = float(s)
            if f >= 1.0:
                return 'Done'
            if f > 0:
                return 'In progress'
            return 'Not started'
        except (TypeError, ValueError):
            return 'Not started'
    s = s.strip()
    valid = ['Not started','In progress','Planning','Drafted','In Review','Blocked','Done','Cancelled']
    if s in valid:
        return s
    # heuristics
    sl = s.lower()
    if 'done' in sl and 'partial' in sl:
        return 'In progress'
    if 'done' in sl:
        return 'Done'
    if 'progress' in sl:
        return 'In progress'
    if 'paused' in sl or 'block' in sl:
        return 'Blocked'
    if 'not started' in sl:
        return 'Not started'
    if 'plan' in sl:
        return 'Planning'
    if 'draft' in sl:
        return 'Drafted'
    if 'review' in sl:
        return 'In Review'
    return 'Not started'

def normalise_priority(p):
    if p is None or p == '':
        return 'Medium'
    if not isinstance(p, str):
        return 'Medium'
    p = p.strip()
    valid = ['Low','Medium','High','Urgent','Critical ACT']
    if p in valid:
        return p
    pl = p.lower()
    if pl == 'must':
        return 'Critical ACT'
    if pl == 'should':
        return 'High'
    if pl == 'could':
        return 'Medium'
    return 'Medium'

# ============================================================
# BUILD INVENTORY + WORKBOOK
# ============================================================
def main():
    csv_rows = read_csv()
    uk_artifacts = read_uk_artifacts()
    hiring_rows = read_hiring_plan()
    feature_rows = read_game_v2()

    print(f'CSV rows: {len(csv_rows)}')
    print(f'UK artifacts: {len(uk_artifacts)}')
    print(f'Hiring rows: {len(hiring_rows)}')
    print(f'Feature rows: {len(feature_rows)}')

    # ---- INVENTORY MARKDOWN ----
    inv = ['# Couch Heroes True-Up Inventory', '',
           f'Generated 2026-04-26 from CSV ({len(csv_rows)} rows), UK artifacts ({len(uk_artifacts)}), Hiring plan ({len(hiring_rows)}), Features ({len(feature_rows)} rows).', '',
           '## CSV Items — Mapped', '',
           '| # | Title | Original parent | Original status | Project | Feature | New status | Notes |',
           '|---|---|---|---|---|---|---|---|']
    needs_glen = []
    unmapped = []
    final_items = []  # for Excel build
    for i, row in enumerate(csv_rows, 1):
        title = row['title'] or '(blank)'
        proj, feat = map_row(row)
        new_status = STATUS_OVERRIDES.get(row['title'], normalise_status(row['status']))
        notes = ''
        if proj == '__REMOVED__':
            notes = 'REMOVED per Q13'
        elif proj == '__KILLED__':
            notes = 'Killed (Gamescom)'
            new_status = 'Cancelled'
        elif proj == '__UNMAPPED__':
            unmapped.append((i, row['title'], row['parent']))
            notes = 'NEEDS GLEN'
        # Build final item for Excel
        if proj not in ('__REMOVED__','__UNMAPPED__'):
            desc = row['description'] + DESC_APPEND.get(row['title'], '')
            if row['attachment']:
                desc += f"\n\nAttachment: {row['attachment'][:200]}"
            target_proj = 'Closed/Killed' if proj == '__KILLED__' else proj
            final_items.append({
                'project': target_proj,
                'feature': feat,
                'parent_path': f'Couch Heroes / {target_proj} / {feat}',
                'title': row['title'] or feat,
                'description': desc,
                'status': new_status,
                'priority': normalise_priority(row['priority']),
                'health_state': row['health'] or '',
                'work_type': '',
                'assignees': row['assignee'] or '',
                'due_date': '',
                'dependencies': '',
                'practice_area': 'Game Studio Operations',
                'source': f'CSV row {i}',
                'notes': notes,
            })
        inv.append(f"| {i} | {title[:60]} | {row['parent'][:40]} | {row['status']} | {proj} | {feat[:40]} | {new_status} | {notes} |")

    # ---- ADD UK ARTIFACTS ----
    inv.extend(['', f'## UK Artifacts ({len(uk_artifacts)})', '',
                '| Sheet | Ref | Title | Status | Mapped to |',
                '|---|---|---|---|---|'])
    for art in uk_artifacts:
        d = art['data']
        ref = next((str(v) for k,v in d.items() if k and 'Ref' in str(k) and v), '')
        title_a = next((str(v) for k,v in d.items() if k and ('Artifact' in str(k) or 'Title' in str(k)) and v), '')
        status_a = next((str(v) for k,v in d.items() if k and 'Status' in str(k) and v), 'Not started')
        if not title_a:
            continue
        inv.append(f"| {art['sheet']} | {ref} | {title_a[:60]} | {status_a} | UK Company Setup |")
        final_items.append({
            'project': 'UK Company Setup',
            'feature': art['sheet'],
            'parent_path': f'Couch Heroes / UK Company Setup / {art["sheet"]}',
            'title': f'{ref} {title_a}'.strip(),
            'description': str({k:v for k,v in d.items() if v})[:1000],
            'status': normalise_status(status_a),
            'priority': 'High',
            'health_state': '',
            'work_type': 'Implementation',
            'assignees': 'Lorenza',
            'due_date': '',
            'dependencies': '',
            'practice_area': 'Legal & Compliance',
            'source': f'UK xlsx / {art["sheet"]} / ref {ref}',
            'notes': '',
        })

    # ---- ADD HIRING POSITIONS ----
    inv.extend(['', f'## Hiring Positions ({len(hiring_rows)})', '',
                'All become stories under HR → Hiring Pipeline (sub-feature by discipline).', ''])
    for h in hiring_rows[:60]:
        d = h['data']
        role = next((str(v) for k,v in d.items() if k and 'Role' in str(k) and v), '')
        if not role:
            continue
        # Categorise role into a hiring sub-feature
        rl = role.lower()
        if any(t in rl for t in ['cto','engineer','engineering','full stack','gameplay','network','tools','td','technical director','backend','client']):
            sub = 'Hiring Pipeline — Engineering'
        elif any(t in rl for t in ['producer','design','jira','systems designer','ux','balance','economy']):
            sub = 'Hiring Pipeline — Production & Design'
        elif any(t in rl for t in ['art','animator','vfx','lighting','environment','character','concept']):
            sub = 'Hiring Pipeline — Art'
        elif 'audio' in rl:
            sub = 'Hiring Pipeline — Audio'
        elif 'qa' in rl:
            sub = 'Hiring Pipeline — QA'
        else:
            sub = 'Hiring Pipeline — Operations'
        final_items.append({
            'project': 'HR',
            'feature': sub,
            'parent_path': f'Couch Heroes / HR / {sub}',
            'title': f'Hire: {role}',
            'description': str({k:v for k,v in d.items() if v})[:500],
            'status': 'In progress',
            'priority': 'High',
            'health_state': '',
            'work_type': 'Implementation',
            'assignees': '',
            'due_date': '',
            'dependencies': '',
            'practice_area': 'Talent Acquisition',
            'source': f'Hiring xlsx / {h["sheet"]}',
            'notes': '',
        })

    # ---- ADD GAME PROJECT (V2) with CLASSIFICATIONS ----
    # Per Glen 2026-04-27: Game project structured as 6 sub-features, each fed by a V2 sheet.
    # Classification (type/layer/confidence) loaded from build_inputs/game_classification.md
    classifications = read_game_classifications()
    print(f'Loaded {len(classifications)} game classifications')
    classified_hits = 0
    classified_misses = 0
    inv.extend(['', f'## Game Project (V2) ({len(feature_rows)} items)', '',
                'From Couch_Heroes_Features_v2.xlsx. 6 sub-features:',
                '- Game Features (Game Feature List sheet)',
                '- Game Systems (TDD sheet)',
                '- Platform Features (Platform Feature List sheet)',
                '- Live Service Features (Live Service Features sheet)',
                '- Content (Content List sheet)',
                '- Feature Versions (Feature Versions sheet, V0-V7 maturity per feature)',
                '',
                'Each item carries type / layer / confidence / needs_review_reason from classification pass.', ''])
    for f in feature_rows:
        sub = f['sub_feature']
        epic = f['epic'] or 'Unspecified'
        feature_name = f.get('feature') or ''
        story_name = f.get('story') or ''
        # Build path: Couch Heroes / Game / SubFeature / Epic / Feature [/ Story]
        path_parts = ['Couch Heroes', 'Game', sub]
        if epic and epic != 'Unspecified':
            path_parts.append(epic)
        if feature_name and feature_name != epic:
            path_parts.append(feature_name)
        if story_name and f['level'] == 'task':
            path_parts.append(story_name)
        parent = ' / '.join(path_parts)

        # Classification lookup by (sheet, title) - agent records title in only
        # one of (epic, feature, story, task) per row (the level it represents).
        sheet = SUB_FEATURE_TO_SHEET.get(sub, sub)
        title = f['title']
        cls = classifications.get((sheet, title))
        # Fallback: when title not in classifications (e.g. items split from a multi-line cell),
        # apply sheet-based default since the source sheet alone determines type unambiguously.
        if not cls:
            sheet_defaults = {
                'Content List': ('Content', 'Presentation', 'medium'),
                'Platform Feature List': ('Platform', 'Foundation', 'medium'),
                'Live Service Features': ('LiveService', 'Feature', 'medium'),
                'TDD': ('System', 'Feature', 'medium'),
                'Feature Versions': ('Maturity', 'Polish', 'medium'),
            }
            if sheet in sheet_defaults:
                t, l, c = sheet_defaults[sheet]
                cls = {'type': t, 'layer': l, 'confidence': c,
                       'reason': 'Title not in classification file - sheet-based default'}
        if cls:
            classified_hits += 1
            type_ = cls['type']; layer = cls['layer']
            conf = cls['confidence']; reason = cls['reason']
        else:
            classified_misses += 1
            type_ = ''; layer = ''; conf = 'unknown'; reason = 'not in classification file'

        notes_parts = []
        if conf == 'low' and reason:
            notes_parts.append(f'NEEDS OFFSITE REVIEW: {reason}')
        elif conf == 'medium' and reason:
            notes_parts.append(f'Medium confidence: {reason}')

        final_items.append({
            'project': 'Game',
            'feature': sub,
            'parent_path': parent,
            'title': title,
            'description': f.get('description') or '',
            'status': normalise_status(f.get('progress')),
            'priority': normalise_priority(f.get('priority')),
            'health_state': '',
            'work_type': 'Implementation',
            'assignees': '',
            'due_date': '',
            'dependencies': '',
            'practice_area': 'Game Development',
            'source': f'Features V2 / {sub}',
            'notes': ' | '.join(notes_parts),
            'type': type_,
            'layer': layer,
            'confidence': conf,
        })
    print(f'Classification: {classified_hits} hits, {classified_misses} misses')

    # ---- ADD STUDIO OPERATIONS (from agent file, with UK duplicate detection) ----
    so_rows = read_md_table('./build_inputs/studio_operations_items.md')
    so_dup_count = 0
    inv.extend(['', f'## Studio Operations ({len(so_rows)} items from build_inputs/studio_operations_items.md)', ''])
    for r in so_rows:
        feat = r.get('feature','').strip()
        title = r.get('story','').strip()
        if not feat or not title:
            continue
        # Flag UK duplicates
        notes = r.get('notes','')
        uk_dup = detect_uk_duplicate(title)
        if uk_dup:
            so_dup_count += 1
            notes = f'POSSIBLE DUPLICATE OF {uk_dup} - confirm at offsite. ' + notes
        final_items.append({
            'project': 'Studio Operations',
            'feature': feat,
            'parent_path': f'Couch Heroes / Studio Operations / {feat}',
            'title': title,
            'description': r.get('description',''),
            'status': normalise_status_extended(r.get('status','')),
            'priority': normalise_priority_extended(r.get('priority','')),
            'health_state': normalise_health(r.get('health_state','')),
            'work_type': normalise_work_type(r.get('work_type','')),
            'assignees': r.get('assignees',''),
            'due_date': '',
            'dependencies': r.get('dependencies',''),
            'practice_area': 'Studio Operations',
            'source': 'Agent / Studio Operations',
            'notes': notes,
            'type': '', 'layer': '', 'confidence': '',
        })
    print(f'Studio Ops: {so_dup_count} items flagged as possible UK Company Setup duplicates')

    # ---- ADD STUDIO BUSINESS (from agent file, with assignee heuristics) ----
    # Glen is fractional CPO covering Marketing/Brand/PR/Community/BD/Web at Couch Heroes
    # per NBI Brain. Default assignees by feature when agent gave TBD.
    SB_DEFAULT_ASSIGNEES = {
        'GTM and Forecasting': 'Glen, Vardis',
        'BD': 'Glen, Vardis',
        'PR': 'Glen',
        'Community': 'Glen, Barb',
        'Website': 'Mustafa, Glen',
        'Studio Naming': 'Vardis, Glen',
    }
    sb_rows = read_md_table('./build_inputs/studio_business_items.md')
    sb_assigned_by_default = 0
    inv.extend(['', f'## Studio Business ({len(sb_rows)} items from build_inputs/studio_business_items.md)', ''])
    for r in sb_rows:
        feat = r.get('feature','').strip()
        title = r.get('story','').strip()
        if not feat or not title:
            continue
        # Apply assignee heuristic if agent left TBD or empty
        assignees = r.get('assignees','').strip()
        notes = r.get('notes','')
        if not assignees or assignees.upper() == 'TBD':
            default = SB_DEFAULT_ASSIGNEES.get(feat, 'Glen')
            assignees = default
            sb_assigned_by_default += 1
            notes = (notes + ' | Assignee defaulted from feature heuristic - confirm at offsite').strip(' |')
        final_items.append({
            'project': 'Studio Business',
            'feature': feat,
            'parent_path': f'Couch Heroes / Studio Business / {feat}',
            'title': title,
            'description': r.get('description',''),
            'status': normalise_status_extended(r.get('status','')),
            'priority': normalise_priority_extended(r.get('priority','')),
            'health_state': normalise_health(r.get('health_state','')),
            'work_type': normalise_work_type(r.get('work_type','')),
            'assignees': assignees,
            'due_date': '',
            'dependencies': r.get('dependencies',''),
            'practice_area': 'Studio Business',
            'source': 'Agent / Studio Business',
            'notes': notes,
            'type': '', 'layer': '', 'confidence': '',
        })
    print(f'Studio Business: {sb_assigned_by_default} items got default assignees (TBD replaced)')

    # ---- ADD AI PROJECT (from agent file) ----
    ai_rows = read_md_table('./build_inputs/ai_project_items.md')
    inv.extend(['', f'## AI Project ({len(ai_rows)} items from build_inputs/ai_project_items.md)', ''])
    for r in ai_rows:
        feat = r.get('feature','').strip()
        title = r.get('story','').strip()
        if not feat or not title:
            continue
        final_items.append({
            'project': 'AI',
            'feature': feat,
            'parent_path': f'Couch Heroes / AI / {feat}',
            'title': title,
            'description': r.get('description',''),
            'status': normalise_status_extended(r.get('status','')),
            'priority': normalise_priority_extended(r.get('priority','')),
            'health_state': normalise_health(r.get('health_state','')),
            'work_type': normalise_work_type(r.get('work_type','')),
            'assignees': r.get('assignees',''),
            'due_date': '',
            'dependencies': r.get('dependencies',''),
            'practice_area': 'AI Governance',
            'source': 'Agent / AI Project',
            'notes': r.get('notes',''),
            'type': '', 'layer': '', 'confidence': '',
        })

    # ---- ADD PRODUCTION STRATEGY ENRICHMENTS (NEW items only - existing CSV items already added) ----
    ps_rows = read_md_table('./build_inputs/production_strategy_items.md')
    ps_added = 0
    inv.extend(['', f'## Production Strategy enrichments (NEW items from build_inputs/production_strategy_items.md)', ''])
    for r in ps_rows:
        notes = r.get('notes','').strip()
        # Only add NEW items - existing CSV items already in workbook
        if 'Existing CSV' in notes or 'existing' in notes.lower() and 'csv' in notes.lower():
            continue
        feat = r.get('feature','').strip()
        title = r.get('story','').strip()
        if not feat or not title:
            continue
        final_items.append({
            'project': 'Production Strategy & Implementation',
            'feature': feat,
            'parent_path': f'Couch Heroes / Production Strategy & Implementation / {feat}',
            'title': title,
            'description': r.get('description',''),
            'status': normalise_status_extended(r.get('status','')),
            'priority': normalise_priority_extended(r.get('priority','')),
            'health_state': normalise_health(r.get('health_state','')),
            'work_type': normalise_work_type(r.get('work_type','')),
            'assignees': r.get('assignees',''),
            'due_date': '',
            'dependencies': r.get('dependencies',''),
            'practice_area': 'Production Strategy',
            'source': 'Agent / Production Strategy',
            'notes': notes,
            'type': '', 'layer': '', 'confidence': '',
        })
        ps_added += 1
    print(f'Production Strategy: added {ps_added} NEW items (skipped existing-CSV duplicates)')

    # ---- ADD NEW STORIES FROM Q&A ----
    new_from_qa = [
        {
            'project': 'HR', 'feature': 'Hiring Pipeline — Engineering',
            'parent_path': 'Couch Heroes / HR / Hiring Pipeline — Engineering',
            'title': 'CTO restart: source new candidates',
            'description': 'Q1 (2026-04-26): Cam, Neil, Nia ALL DEAD. Restart sourcing from scratch with refreshed brief.',
            'status': 'Not started', 'priority': 'Critical ACT',
            'health_state': 'Red', 'work_type': 'Implementation',
            'assignees': 'Glen', 'due_date':'','dependencies':'',
            'practice_area':'Talent Acquisition', 'source':'Q1 (2026-04-26)','notes':'NEW from Q&A',
        },
        {
            'project': 'HR', 'feature': 'Hiring Pipeline — Engineering',
            'parent_path': 'Couch Heroes / HR / Hiring Pipeline — Engineering',
            'title': 'CTO scorecard refresh',
            'description': 'Q1 (2026-04-26): Existing scorecard at Clients/Couch Heroes/CTO_Scorecard.xlsx. Review, refresh against new candidate criteria post-Cam/Neil/Nia.',
            'status': 'Not started','priority':'High','health_state':'',
            'work_type':'Implementation','assignees':'Glen','due_date':'','dependencies':'CTO_Scorecard.xlsx',
            'practice_area':'Talent Acquisition','source':'Q1 (2026-04-26)','notes':'NEW from Q&A',
        },
        {
            'project': 'HR', 'feature': 'Hiring Pipeline — Engineering',
            'parent_path': 'Couch Heroes / HR / Hiring Pipeline — Engineering',
            'title': 'CTO question set refresh',
            'description': 'Q1 (2026-04-26): Existing question set at Clients/Couch Heroes/CTO_Interview_Questions_Couch_Heroes.docx. Review and refresh.',
            'status': 'Not started','priority':'High','health_state':'',
            'work_type':'Implementation','assignees':'Glen','due_date':'','dependencies':'CTO_Interview_Questions_Couch_Heroes.docx',
            'practice_area':'Talent Acquisition','source':'Q1 (2026-04-26)','notes':'NEW from Q&A',
        },
        {
            'project': 'HR', 'feature': 'Hiring Pipeline — DONE',
            'parent_path': 'Couch Heroes / HR / Hiring Pipeline — DONE',
            'title': 'Document Lead Animator question set artifact',
            'description': 'Q4 (2026-04-26): No separate question-set doc on disk for Lead Animator interview. ACTION: locate the questions actually used (likely informal / verbal / Slack), document them as a reusable artifact, attach to Lead Animator hiring story when uploading to WorkSage.',
            'status': 'Not started','priority':'Medium','health_state':'',
            'work_type':'Implementation','assignees':'Glen','due_date':'','dependencies':'',
            'practice_area':'Talent Acquisition','source':'Q4 (2026-04-26)','notes':'NEW from Q&A',
        },
        {
            'project': 'HR', 'feature': 'Hiring Pipeline — Operations',
            'parent_path': 'Couch Heroes / HR / Hiring Pipeline — Operations',
            'title': 'Hire: Lili Zhao (Chief of Staff, ~£135k)',
            'description': 'Q5 (2026-04-26): NEW HIRE separate from Lorenza (who is existing staff handling Dino transition). Per Granola SayBrook discussion ~£135k for COS role.',
            'status': 'In progress','priority':'High','health_state':'',
            'work_type':'Implementation','assignees':'Glen','due_date':'','dependencies':'',
            'practice_area':'Talent Acquisition','source':'Q5 (2026-04-26)','notes':'NEW from Q&A',
        },
        {
            'project': 'HR', 'feature': 'Hiring Pipeline — Operations',
            'parent_path': 'Couch Heroes / HR / Hiring Pipeline — Operations',
            'title': 'HR Operations role — ongoing interviews',
            'description': 'Q6 (2026-04-26): Two weak candidates interviewed; sourcing continues. Reference: Slack engineering candidates channel https://couchheroes.slack.com/archives/C0AJBRJTJTA + Google Sheet of all candidates https://docs.google.com/spreadsheets/d/1NaZABJP_8HQCVdrQR9Uoq9g3NYVYdyKHLB8qc3nSLRo/edit',
            'status': 'In progress','priority':'High','health_state':'',
            'work_type':'Implementation','assignees':'Glen','due_date':'','dependencies':'Slack channel C0AJBRJTJTA / Google Sheet',
            'practice_area':'Talent Acquisition','source':'Q6 (2026-04-26)','notes':'NEW from Q&A',
        },
        {
            'project': 'HR', 'feature': 'Staff Changes',
            'parent_path': 'Couch Heroes / HR / Staff Changes',
            'title': 'Dino exit — transition + garden leave',
            'description': 'Q7 (2026-04-26): Performance review DELIVERED. Exit timeline 1mo transition + 2mo garden leave is TENTATIVE. Decided Apr 9. Settlement details with Riley/legal. Lorenza extracting documentation.',
            'status': 'In progress','priority':'High','health_state':'Yellow',
            'work_type':'Implementation','assignees':'Glen,Lorenza,Riley','due_date':'','dependencies':'',
            'practice_area':'HR Operations','source':'Granola Apr 9 + Q7 (2026-04-26)','notes':'NEW from Q&A — dates tentative',
        },
        {
            'project': 'Studio Operations', 'feature': 'Org Design & Structure',
            'parent_path': 'Couch Heroes / Studio Operations / Org Design & Structure',
            'title': 'Attach Miro org-structure artifact',
            'description': 'Q8 (2026-04-26): Glen provided Miro screenshot of likely structure. Captured at Clients/Couch Heroes/org_structure_likely_2026-04-26.md. Replace with finalised version after Apr 27-30 offsite.',
            'status': 'In progress','priority':'High','health_state':'',
            'work_type':'Strategy','assignees':'Glen','due_date':'','dependencies':'org_structure_likely_2026-04-26.md',
            'practice_area':'Org Design','source':'Q8 (2026-04-26)','notes':'NEW from Q&A',
        },
        {
            'project': 'HR', 'feature': '1:1s — Weekly Programme',
            'parent_path': 'Couch Heroes / HR / 1:1s — Weekly Programme',
            'title': 'Weekly 1:1 Programme (8 leads)',
            'description': 'Q12 (2026-04-26): One programme story + 8 weekly sub-tasks per cycle. Leads: Dino, Lorenza, Vardis, Aris, Mustafa, Robin, David, Valeria. Bi-weekly 1:1s + weekly coaching cadence.',
            'status': 'In progress','priority':'High','health_state':'Green',
            'work_type':'Ongoing Mgmt','assignees':'Glen','due_date':'','dependencies':'',
            'practice_area':'HR Operations','source':'Q12 (2026-04-26)','notes':'NEW from Q&A — programme story',
        },
    ]
    final_items.extend(new_from_qa)

    inv.extend(['', f'## NEW Items From Q&A ({len(new_from_qa)})',''])
    for it in new_from_qa:
        inv.append(f"- {it['project']} / {it['feature']} / **{it['title']}** ({it['status']}, {it['priority']}) — {it['notes']}")

    if unmapped:
        inv.extend(['', f'## UNMAPPED ({len(unmapped)}) — NEEDS GLEN',''])
        for i, t, p in unmapped:
            inv.append(f"- Row {i}: '{t}' (parent: '{p}')")

    # Write inventory
    with open(f'{CH_DIR}/_trueup_inventory.md','w',encoding='utf-8') as f:
        f.write('\n'.join(inv))
    print(f'Inventory written ({len(inv)} lines)')

    # ---- BUILD WORKBOOK ----
    wb = Workbook()
    wb.remove(wb.active)
    cols = ['parent_path','title','description','status','priority','health_state','work_type','assignees','due_date','dependencies','practice_area','type','layer','confidence','source','notes']
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2D3748', end_color='2D3748', fill_type='solid')
    # Column widths (16 columns)
    widths = [55, 55, 65, 14, 14, 16, 16, 24, 12, 30, 22, 12, 14, 12, 28, 50]

    def col_letter(i):
        # 1-indexed column to letter
        s = ''
        while i:
            i, r = divmod(i-1, 26)
            s = chr(65+r) + s
        return s

    def setup_sheet(ws, items_for_sheet):
        """Apply headers, append rows, freeze pane, autofilter, widths."""
        ws.append(cols)
        for c in range(1, len(cols)+1):
            ws.cell(row=1, column=c).font = header_font
            ws.cell(row=1, column=c).fill = header_fill
        for it in items_for_sheet:
            ws.append([it.get(c,'') for c in cols])
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[col_letter(i)].width = w
        # Freeze header row, enable autofilter across all columns
        ws.freeze_panes = 'A2'
        last_col = col_letter(len(cols))
        ws.auto_filter.ref = f'A1:{last_col}{ws.max_row}'

    # Master sheet
    ws_m = wb.create_sheet('MASTER')
    setup_sheet(ws_m, final_items)

    # Per-project sheets
    by_project = {}
    for it in final_items:
        by_project.setdefault(it['project'], []).append(it)
    for proj, items in by_project.items():
        sheet_name = proj.replace('/','-').replace('\\','-').replace('?','').replace('*','').replace('[','(').replace(']',')').replace(':','-')[:31]
        ws = wb.create_sheet(sheet_name)
        setup_sheet(ws, items)

    # Validation sheet - flag everything that needs offsite review
    ws_v = wb.create_sheet('_VALIDATION')
    ws_v.append(['Severity', 'Issue', 'Detail'])
    for c in range(1, 4):
        ws_v.cell(row=1, column=c).font = header_font
        ws_v.cell(row=1, column=c).fill = header_fill
    flag_count = 0
    for i, it in enumerate(final_items, 1):
        if not it.get('title'):
            ws_v.append(['ERROR', 'Missing title', f'Row {i}'])
            flag_count += 1
        if not it.get('parent_path'):
            ws_v.append(['ERROR', 'Missing parent_path', f"Row {i}: {it.get('title')}"])
            flag_count += 1
        # Game items: low confidence
        if it.get('confidence') == 'low':
            ws_v.append(['REVIEW', 'Game item: low classification confidence',
                         f"{it.get('parent_path','')} | {it.get('title','')[:80]} | reason: {it.get('notes','')[:100]}"])
            flag_count += 1
        # Unknown classification
        if it.get('confidence') == 'unknown':
            ws_v.append(['REVIEW', 'Game item: classification not found',
                         f"{it.get('parent_path','')} | {it.get('title','')[:80]}"])
            flag_count += 1
        # UK duplicates flagged
        if it.get('notes') and 'POSSIBLE DUPLICATE' in it.get('notes',''):
            ws_v.append(['REVIEW', 'Studio Operations possible UK duplicate',
                         f"{it.get('title','')[:80]} | {it.get('notes','')[:120]}"])
            flag_count += 1
        # Missing assignee on non-Game projects
        if it.get('project') not in ('Game', 'UK Company Setup') and (not it.get('assignees') or 'TBD' in (it.get('assignees') or '').upper()):
            ws_v.append(['REVIEW', 'Missing or TBD assignee',
                         f"{it.get('parent_path','')} | {it.get('title','')[:80]}"])
            flag_count += 1
    if unmapped:
        for u in unmapped:
            ws_v.append(['ERROR', 'Unmapped CSV item', f'Row {u[0]}: {u[1]} (parent: {u[2]})'])
            flag_count += 1
    # Format validation sheet
    ws_v.freeze_panes = 'A2'
    ws_v.auto_filter.ref = f'A1:C{ws_v.max_row}'
    ws_v.column_dimensions['A'].width = 12
    ws_v.column_dimensions['B'].width = 50
    ws_v.column_dimensions['C'].width = 120
    print(f'Validation sheet: {flag_count} flags')

    # Stats sheet
    ws_s = wb.create_sheet('_STATS', 0)
    ws_s.append(['Project', 'Item count'])
    ws_s.cell(row=1, column=1).font = header_font
    ws_s.cell(row=1, column=2).font = header_font
    ws_s.cell(row=1, column=1).fill = header_fill
    ws_s.cell(row=1, column=2).fill = header_fill
    for proj, items in by_project.items():
        ws_s.append([proj, len(items)])
    ws_s.append(['TOTAL', len(final_items)])
    ws_s.column_dimensions['A'].width = 40
    ws_s.column_dimensions['B'].width = 14

    # ---- AUTOMATIC EM-DASH SWEEP (Glen's hard rule) ----
    EM = '—'
    em_count = 0
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and EM in cell.value:
                    cell.value = cell.value.replace(EM, '-')
                    em_count += 1
    print(f'Em-dash sweep on workbook: {em_count} cells cleaned')

    out = f'{CH_DIR}/CH_WorkSage_import_v1.xlsx'
    wb.save(out)
    print(f'Workbook written: {out}')
    print(f'Total items: {len(final_items)}')
    print(f'Projects: {list(by_project.keys())}')
    print(f'Unmapped: {len(unmapped)}')

    # ---- AUTOMATIC EM-DASH SWEEP on markdown deliverables ----
    import os
    for fn in ['_trueup_inventory.md', 'features_v1_vs_v2_comparison.md',
               'org_structure_likely_2026-04-26.md',
               '../../docs/HANDOFF.md',
               '../../docs/superpowers/specs/2026-04-24-couch-heroes-backlog-consolidation.md']:
        if not os.path.exists(fn):
            continue
        with open(fn,'r',encoding='utf-8') as f:
            t = f.read()
        if EM in t:
            n = t.count(EM)
            with open(fn,'w',encoding='utf-8') as f:
                f.write(t.replace(EM, '-'))
            print(f'Em-dash sweep on {fn}: {n} removed')

if __name__ == '__main__':
    main()
