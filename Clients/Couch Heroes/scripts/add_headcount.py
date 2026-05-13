"""
Add headcount summary to Glen's cleaned v10.
- Total man-days per role from Gate 2 (EARLY PROD, cols 85-122)
- Convert to hours (x8), adjust for 6hrs/day productivity
- Calculate headcount needed for the build period
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import sys
sys.stdout.reconfigure(encoding='utf-8')

src = r'C:\Users\gpbea\Downloads\CouchHeroes_Man_Day_Work_Plan_v10.xlsx'
dst = r'C:\Users\gpbea\Downloads\CouchHeroes_Man_Day_Work_Plan_v10.xlsx'  # overwrite Glen's copy

wb = openpyxl.load_workbook(src)

# === Collect totals from Plan sheet ===
ws = wb['Plan']

ROLES = [
    ('Game Director', 95),
    ('Game Design Lead', 96),
    ('Level Design Lead', 97),
    ('Game Designer', 98),
    ('Combat Designer', 99),
    ('Level Designer', 100),
    ('Technical Designer', 101),
    ('Writer', 102),
    ('Art Director', 103),
    ('Concept Artist', 104),
    ('UI/UX Designer', 105),
    ('Graphic Designer', 106),
    ('Character Artist', 107),
    ('Environment Artist', 108),
    ('Prop Artist', 109),
    ('Animator', 110),
    ('Character Rigger', 111),
    ('VFX Artist', 112),
    ('Technical Artist', 113),
    ('CTO', 114),
    ('Tech Lead', 115),
    ('Gameplay Developer', 116),
    ('Full Stack Developer', 117),
    ('Backend Developer', 118),
    ('Network Engineer', 119),
    ('QA', 120),
    ('Sound Designer', 121),
    ('Platform Team', 122),
]

# Sum each role column across feature rows (5-159)
role_totals = {}
for role_name, col in ROLES:
    total = 0
    for row in range(5, 160):
        val = ws.cell(row=row, column=col).value
        if val and isinstance(val, (int, float)):
            total += val
    if total > 0:
        role_totals[role_name] = total

# === Also update Plan sheet row 161+ with proper SUM formulas for Gate 2 ===
# Row 161: Total PW (man-days) — add SUM formulas for Gate 2 columns
for role_name, col in ROLES:
    col_letter = get_column_letter(col)
    ws.cell(row=161, column=col).value = f'=SUM({col_letter}5:{col_letter}159)'
    ws.cell(row=161, column=col).font = Font(bold=True)

# === Create Headcount Summary sheet ===
if 'Headcount Summary' in wb.sheetnames:
    del wb['Headcount Summary']

hc = wb.create_sheet('Headcount Summary', 1)  # Insert after Plan

# Styles
header_font = Font(bold=True, size=12, color='FFFFFF')
header_fill = PatternFill(start_color='404040', end_color='404040', fill_type='solid')
subheader_font = Font(bold=True, size=10)
subheader_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
number_font = Font(size=10)
total_font = Font(bold=True, size=11)
total_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Column widths
hc.column_dimensions['A'].width = 28
hc.column_dimensions['B'].width = 18
hc.column_dimensions['C'].width = 18
hc.column_dimensions['D'].width = 22
hc.column_dimensions['E'].width = 18
hc.column_dimensions['F'].width = 20

# Title
hc.merge_cells('A1:F1')
hc.cell(row=1, column=1).value = 'COUCH HEROES — VERTICAL SLICE HEADCOUNT PLAN'
hc.cell(row=1, column=1).font = Font(bold=True, size=14, color='FFFFFF')
hc.cell(row=1, column=1).fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
hc.cell(row=1, column=1).alignment = Alignment(horizontal='center')
for col in range(1, 7):
    hc.cell(row=1, column=col).fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')

# Assumptions block
hc.cell(row=3, column=1).value = 'Assumptions'
hc.cell(row=3, column=1).font = Font(bold=True, size=11)
hc.cell(row=4, column=1).value = 'Man-days budgeted at'
hc.cell(row=4, column=2).value = 8
hc.cell(row=4, column=3).value = 'hours/day'
hc.cell(row=5, column=1).value = 'Productive delivery rate'
hc.cell(row=5, column=2).value = 6
hc.cell(row=5, column=3).value = 'hours/day (meetings, overhead, etc.)'
hc.cell(row=6, column=1).value = 'Build period (Early Prod)'
hc.cell(row=6, column=2).value = 4
hc.cell(row=6, column=3).value = 'months'
hc.cell(row=7, column=1).value = 'Working days/month'
hc.cell(row=7, column=2).value = 20
hc.cell(row=8, column=1).value = 'Total working days in period'
hc.cell(row=8, column=2).value = '=B6*B7'
hc.cell(row=8, column=2).font = Font(bold=True)

for r in range(4, 9):
    hc.cell(row=r, column=2).number_format = '0'

# Headers
row = 10
headers = ['Role', 'Man-Days\n(at 8hrs/day)', 'Total Hours', 'Adjusted Days\n(at 6hrs/day)', 'Headcount\n(for build period)', 'Notes']
for i, h in enumerate(headers, 1):
    cell = hc.cell(row=row, column=i)
    cell.value = h
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', wrap_text=True, vertical='center')
    cell.border = thin_border
hc.row_dimensions[row].height = 36

# Group roles by department
DEPARTMENTS = [
    ('LEADERSHIP & PRODUCTION', [
        'Game Director', 'Game Design Lead', 'Level Design Lead',
    ]),
    ('DESIGN', [
        'Game Designer', 'Combat Designer', 'Level Designer',
        'Technical Designer', 'Writer',
    ]),
    ('ART', [
        'Art Director', 'Concept Artist', 'UI/UX Designer', 'Graphic Designer',
        'Character Artist', 'Environment Artist', 'Prop Artist',
        'Animator', 'Character Rigger', 'VFX Artist', 'Technical Artist',
    ]),
    ('ENGINEERING', [
        'CTO', 'Tech Lead', 'Gameplay Developer', 'Full Stack Developer',
        'Backend Developer', 'Network Engineer', 'Platform Team',
    ]),
    ('QA & AUDIO', [
        'QA', 'Sound Designer',
    ]),
]

current_row = 11
data_start_row = None

for dept_name, dept_roles in DEPARTMENTS:
    # Department header
    hc.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    cell = hc.cell(row=current_row, column=1)
    cell.value = dept_name
    cell.font = subheader_font
    cell.fill = subheader_fill
    for c in range(1, 7):
        hc.cell(row=current_row, column=c).fill = subheader_fill
        hc.cell(row=current_row, column=c).border = thin_border
    current_row += 1

    for role in dept_roles:
        days = role_totals.get(role, 0)
        if data_start_row is None:
            data_start_row = current_row

        cell_a = hc.cell(row=current_row, column=1)
        cell_a.value = role
        cell_a.font = number_font
        cell_a.border = thin_border

        cell_b = hc.cell(row=current_row, column=2)
        cell_b.value = days
        cell_b.number_format = '#,##0'
        cell_b.font = number_font
        cell_b.border = thin_border
        cell_b.alignment = Alignment(horizontal='center')

        # Total Hours = man-days * 8
        cell_c = hc.cell(row=current_row, column=3)
        cell_c.value = f'=B{current_row}*$B$4'
        cell_c.number_format = '#,##0'
        cell_c.font = number_font
        cell_c.border = thin_border
        cell_c.alignment = Alignment(horizontal='center')

        # Adjusted Days = total_hours / 6
        cell_d = hc.cell(row=current_row, column=4)
        cell_d.value = f'=C{current_row}/$B$5'
        cell_d.number_format = '#,##0.0'
        cell_d.font = number_font
        cell_d.border = thin_border
        cell_d.alignment = Alignment(horizontal='center')

        # Headcount = adjusted_days / total_working_days
        cell_e = hc.cell(row=current_row, column=5)
        cell_e.value = f'=IF(D{current_row}=0,"",ROUND(D{current_row}/$B$8,1))'
        cell_e.number_format = '0.0'
        cell_e.font = Font(bold=True, size=10)
        cell_e.border = thin_border
        cell_e.alignment = Alignment(horizontal='center')

        # Notes
        cell_f = hc.cell(row=current_row, column=6)
        cell_f.border = thin_border
        if days == 0:
            cell_f.value = 'No VS hours assigned'
            cell_f.font = Font(italic=True, size=9, color='999999')

        current_row += 1

data_end_row = current_row - 1

# TOTALS row
current_row += 1
total_row = current_row
hc.cell(row=total_row, column=1).value = 'TOTAL'
hc.cell(row=total_row, column=1).font = total_font
hc.cell(row=total_row, column=1).fill = total_fill
hc.cell(row=total_row, column=1).border = thin_border

for col in range(2, 6):
    cell = hc.cell(row=total_row, column=col)
    col_letter = get_column_letter(col)
    cell.value = f'=SUM({col_letter}{data_start_row}:{col_letter}{data_end_row})'
    cell.font = total_font
    cell.fill = total_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')
    if col == 2:
        cell.number_format = '#,##0'
    elif col == 3:
        cell.number_format = '#,##0'
    elif col == 4:
        cell.number_format = '#,##0.0'
    elif col == 5:
        cell.number_format = '0.0'

hc.cell(row=total_row, column=6).fill = total_fill
hc.cell(row=total_row, column=6).border = thin_border

# Only-with-hours total
current_row += 1
hc.cell(row=current_row, column=1).value = 'Roles with hours'
hc.cell(row=current_row, column=1).font = Font(italic=True, size=10)
roles_with_hours = len([r for r in role_totals.values() if r > 0])
hc.cell(row=current_row, column=2).value = f'{roles_with_hours} of {len(ROLES)}'
hc.cell(row=current_row, column=2).font = Font(italic=True, size=10)

# Key insight
current_row += 2
hc.cell(row=current_row, column=1).value = 'KEY'
hc.cell(row=current_row, column=1).font = Font(bold=True, size=11)
current_row += 1
hc.cell(row=current_row, column=1).value = 'Headcount shows how many full-time people per role you need'
current_row += 1
hc.cell(row=current_row, column=1).value = 'for the 4-month Early Prod build period, at 6 productive hours/day.'
current_row += 1
hc.cell(row=current_row, column=1).value = 'Change B5 (productive hours) or B6 (months) to remodel.'

wb.save(dst)
print(f"Saved to: {dst}")

# Print summary
print(f"\n=== HEADCOUNT SUMMARY (6hrs/day, 4 months, 80 working days) ===\n")
grand_days = 0
grand_hours = 0
grand_headcount = 0
for dept_name, dept_roles in DEPARTMENTS:
    print(f"\n{dept_name}")
    for role in dept_roles:
        days = role_totals.get(role, 0)
        if days == 0:
            continue
        hours = days * 8
        adj_days = hours / 6
        headcount = adj_days / 80
        grand_days += days
        grand_hours += hours
        grand_headcount += headcount
        print(f"  {role:28s} {days:5d} man-days  {hours:6d} hrs  {adj_days:7.1f} adj-days  {headcount:5.1f} headcount")

print(f"\n{'TOTAL':28s} {grand_days:5d} man-days  {grand_hours:6d} hrs  {grand_hours/6:7.1f} adj-days  {grand_headcount:5.1f} headcount")
