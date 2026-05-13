import openpyxl

wb = openpyxl.load_workbook('Clients/Couch Heroes/CH_AMA_Questions_Broken_Out2.xlsx')
ws = wb.active

answers = {}

answers[1] = """The demand we're chasing is absolutely MMO players. Some of the most engaged, loyal, community-driven players in the industry. We want them, and we're building for them. But we want to serve them differently than the current crop of MMOs do.

When I was at World of Warcraft, the stat that stuck with me was that the average player who'd put in more than twenty hours was then spending twenty-eight hours a week or more in the game. That's a part-time job. And those players weren't always having fun. They were stressing over reputation grinds, gear checks, raid schedules. Miss a raid night, get kicked from the guild. That's a game consuming a person's life, not being part of it.

You can see the whole industry quietly backing away from that model. Games are toning down dailies, softening progression, extending catch-up mechanics, trimming raid lockouts. They see the problem. They just can't unbuild the game they already shipped around it.

And if you look at what players are actually spending their time in right now, the signal is even clearer. Fortnite, Roblox, Animal Crossing, Palia, Sky, VRChat, the housing and hangout side of FFXIV. The numbers are enormous, and the thread running through all of them is the same. People want somewhere to be. A place to show up, meet their friends, decorate their corner of it, collect things they care about, and play around. Progression is expressive, not numerical. You don't need to clear your evening to feel like you got something out of it.

Couch Heroes is built from the ground up for that. Personalisation and a real sense of self are a core pillar, not a side feature. Your avatar, your home, your arcade, your collections. That's the progression. It's communal by design, decorated by you, shared with the friends you choose. No gear treadmill. No stats gating what you can wear. No forty-hour-a-week commitment to stay relevant. If you've got twenty minutes, you drop in, tidy up, catch up with friends, poke around a point of interest, and log off feeling good. If you've got a whole evening, there's real combat, real exploration, proper puzzles, and a world with genuine craft and story behind it. Both of those players get something worthwhile, and they're the same player on different days.

The partner layer keeps the world feeling fresh. Other studios bring their IP into Couch Heroes as arcade machines, cosmetics, events, or whole shards. For players, that means new things to discover that they already care about. For partners, it's a living audience to show their IP to without having to build the infrastructure themselves. That gives us a content engine that doesn't rely on us shipping a giant expansion every year to keep the lights on.

So when people ask why we think there's demand for this, the answer is: the demand is already there. It's visible in what players play every day, and in the way the whole industry is quietly shifting. We want to be part of our players' lives, not consume them. That's the game we're making, and that's the home we're building for them."""

answers[2] = """The answer is partially, and it's always going to be partially when you're trying to do something new. When you orient a product of this size at the market, there's always going to be that question, and anybody who tells you "yes, we're a hundred per cent confident" is probably not being truthful. You can't be, at this stage, on anything that's genuinely new.

What we do know is that we have clear unique selling points. We focus heavily on integrating with people's lives and putting the player first in a new way, which for us means the game is built around the player's life rather than demanding they redesign their life around it. We combine that in the recipe with the rest of the game's depth and its theme, between grit core and myth core, and the way the lore and the story lay out in the activities we build into the game. You can feel it in the corruption mechanic threading through combat and exploration, in the way the arcade ties a player's collection back to partner content they actually care about, in the way the home becomes a real reflection of who they are in the world. That recipe, we believe, is genuinely rather unique.

We know we have a unique story in a unique setting, with a mixture of future and past that you don't see directly in other products. We know that the style of play we're building is unique, based on what we've structured internally, and we'll continue to test it against the market constantly. That's the whole point of iterating as you go. Whether it's RuneScape or World of Warcraft or any other MMO out there, the real danger isn't getting it wrong on day one. It's the inability to iterate to the market as you understand more.

And the honest truth is that the validation isn't mine to make. It's the work everyone in this call does every day that will prove it. The concept is a starting point. What turns it into a game people actually want to live in is what you're all building.

Do we believe we have an incredibly strong premise and a strong core? Absolutely. Have we validated that this is differentiated in the market? Absolutely. Will we continue to validate and drive that hypothesis as we build? Absolutely. That's not a one-time exercise. That's the whole discipline of how we're making this game, and it's on all of us."""

answers[3] = """The USPs will iterate to some extent as we learn more, but what won't change is this.

We are very clearly an MMO. A massively multiplayer online game with a rich and robust role-playing experience, real combat, real exploration, and a robust social system. That sets us firmly in what category of game we are. MMO players are who we're building for.

What makes us unique in that category is we're heavily focused on the identity, friends, and collections side of the game. We'll be able to partner with other games to allow a broad reach of experiences that no individual game can solve on its own. Partner games and partner IP come into our world as arcade machines, cosmetics, events, whole shards. You play, you collect, you show off, you host. The rest of the industry becomes part of your world instead of being scattered across twenty launchers and storefronts you never open.

Everything else is how we build around that.

We're building around the player's life, not the other way round. That means the game isn't structured to require grind treadmills or forty-hour weeks to stay competitive, or chasing a gear score to feel like you belong, or showing up on a Tuesday night so your guild doesn't kick you. Those pressures are what chew people up and spit them out of the genre, and we've designed past them.

Identity is a pillar, not an afterthought. You build who you are in our world. Your avatar, your home, your arcade, your collections. How you dress, how you decorate, the people you hang out with. Most games treat that as a skin on top of the real game. For us it is the game.

Skill over stats. What you do matters more than what you're wearing. Combat, progression, and mastery come from how you play, not from gear inflation. Keeps it competitive without turning into an arms race only the hardcore can survive.

And the world itself is built to move with the players in it. The economy, the story, the environment, all of it responds to what players actually do. Our content pipeline is slower and more considered than the industry's annual-expansion treadmill, and the partner ecosystem means the world keeps getting new stuff we don't have to ship entirely on our own.

The pillars we run the studio by fall straight out of all that. Player identity as the product. Play has real value. Social first, game second. Skill over stats. A living, player-influenced world.

So if you want the one-liner, the thing you can tell your friends when they ask what you're working on: Couch Heroes is an MMO you can actually live with, where your identity, your friends, and your collections carry across the whole gaming world, and where the rest of the industry plugs in to keep it alive."""

answers[4] = """Yes, our pitch will work in the current market. And I won't kid you, the shenanigans in the MMO space lately don't make that any easier. Part of the pitch is establishing the differentiation we have from what's caused other games to fail.

I'm sure if I opened up the table right now, you could give me a long critique on what affected New World. Some of you in this room may have worked on that game. But what has to be said in our pitch, whether it's investors or players or partners, is how we differentiate from what's jaded people looking at the MMO market in the first place.

And the honest answer starts with the audience. There's a massive number of players who loved MMOs but walked away because the genre asked too much of their lives. They didn't stop wanting the experience. They stopped being able to afford the time. Look at where those players went. Fortnite social, FFXIV housing, Animal Crossing, Palia. The behaviour is still there. We're building for those players, and we're building for the MMO players who are still in the genre and want something that respects them more. We're not asking people to choose us over their existing game. We're asking them to come home between sessions of whatever else they play.

That changes the pitch fundamentally, because we're not a WoW killer. I've protected against that deliberately. I watched it happen at Blizzard and after I left. Many games try to become WoW killers. That is not our objective. That is not our purpose.

The other thing that changes the pitch is how we stay alive. Most MMOs that fail don't actually fail at launch. They fail twelve months later when the content pipeline can't keep up. Our partner ecosystem changes that equation. Other studios bring their IP into our world as arcade machines, cosmetics, events, whole shards. The world keeps getting fresh content we don't have to ship entirely on our own. That's not hope, that's a production strategy.

Now underneath all of that is something people don't always think about, but it matters as much as any feature. The health of the studio comes out directly in the game. The direction of leadership manifests in the product. Anybody here who's worked on New World, or any game for that matter, can feel the influences that affect how a game gets built and why. A healthy studio shows up in the game. An unhealthy one shows up too, and usually louder.

We understand that. The studio is as important as the game from a pitch perspective. What matters most about the product itself is that we have a very clear discipline on what we're building and why, when it goes in, how it integrates, and what it creates. We know the themes we're fitting all the content to and we know why those themes.

One of the contexts we keep coming back to is balance. The game can't be too small, or it loses players because there's nothing to do. It can't aim too big, or the purpose becomes convoluted and we can't execute against it. The tone and goals inside the game need to be oriented with absolute clarity, and the studio behind it has to be of one mind, building in one direction.

That is a big part of our pitch. And it's a big reason why we're here talking to you today."""

answers[5] = """This is a really good question, and I think it's one of the reasons we're actually on this call today.

Historically, I don't think we have been aligned, at least not to the level we want to be. And I want to be clear about what I mean by that, because I'm not saying the work hasn't been good. It has. I think all of you have done great work and delivered great things. I think the vision is amazing. I think the market is ready for this kind of game. I think we'll find a great audience. All of that is real.

What I'm saying is that the consistency hasn't been there. Some of you have absolute clarity on what you're building and why. Some areas have had amazing delivery. But it hasn't been consistent across the board, and until it is, we're leaving quality on the table. Not because anyone's doing bad work, but because good work pointed in slightly different directions doesn't add up the way it should.

What we need is to function as a tribe. One purpose, one set of goals, lots of different skills and perspectives, all pointed at one thing. Every department, every discipline, every person in this room bringing their best to the same vision, not their own version of what they think the vision might be. A tribe doesn't mean everyone does the same thing. It means everyone knows why they're doing their thing and how it connects to what the person next to them is doing.

So here's what we're doing about it. This AMA is the start. After this, we've got the offsite. After the offsite, the leads sessions. After the leads, the rollout to the teams. Each step gets more specific. We're going to create clear documentation so the vision and the pillars and the direction aren't just things people heard in a meeting once, they're things you can point at and say "this is what we agreed". We're going to create spaces where you can ask questions when you're not sure, because not being sure is fine. Not asking isn't.

And going forward, every feature decision gets tested against the pillars. Does this serve player identity? Does this reward curiosity? Does this respect the player's time? If it doesn't pass, we question it. That's not a one-off exercise. That's the operating discipline we're building into how this studio runs, every day, in every team.

I'm not pretending we'll have perfect clarity in every space overnight. But we're setting the road, marking the boundaries, and making sure we're all travelling together. This AMA, right now, is the flag in the ground. This is where that starts."""

answers[6] = "Your game world home. An MMO built around your life, your identity, and the games you love, not the other way round. Filled with an ever-evolving, ever-alive story that you can explore."

answers[7] = """This is always a trap question because it has multiple dimensions.

At the feature and system level, we have loads of competitors. We're building inventory systems, quest chains, combat with world bosses. Plenty of games do those things and many do them well. At that level we're competing with every MMO and a good chunk of action RPGs, and we should be studying them and building our versions better.

At the macro level, who competes with us is broader than people think. Netflix. World of Warcraft. Table tennis. Steam. Anything competing for a player's evening is a competitor. That's exactly why "built around your life, not the other way round" matters. We're not asking for someone's whole evening, we're asking to be worth part of it.

At the direct-competitor level, it gets interesting. FFXIV does social and housing brilliantly. Roblox proves the platform model works. Fortnite proves the social hub pivot has legs. Palia proves the cosy MMO audience exists. Animal Crossing proves low-pressure expressive gaming sells in the hundreds of millions. Every one of those games validates a piece of our thesis, and every one of them is an adjacent competitor we should be watching.

But none of them combines all of it. Persistent identity, partner ecosystem, MMO-lite depth, a world that evolves with the players in it. That specific combination doesn't have a direct competitor. So at the feature level, we go broad. At the product level, we're building something those games haven't built.

And whether that combination works is on us to deliver."""

answers[8] = """On a technical level, we'll be scaling for the low-digit millions. The architecture is being built for that. On the marketing and brand level, as we get closer to launch, we'll scale as far as we can based on what finances we have to market the game to a broad audience.

But here's the thing. The key isn't just how many people we bring in the front door. For a game that's going to launch in a couple of years, it's about how we're building systems, features, and a cohesive gameplay experience that draws players back to play again and again.

We'd celebrate a million players on day one. But we could have a tragedy if we see only a hundred thousand by day three or day seven. What we need is retention. The game's depth has to be clear, and it has to draw players forward socially, personally, and gameplay-wise. I'd rather have fifty thousand players who come back every week than a million who try it and leave.

So if you're looking for me to give you a forecast right now, I can't give you an honest one. We're two years out on a game we haven't built yet, and we just talked about getting aligned as a tribe. Any number I give you right now is a guess at best.

What I will tell you is what actually matters, and what you can use in whatever your craft is towards building this game together. We want extremely high retention numbers. Day three. Day seven. Day thirty. Day one-eighty. That's the target. Not a vanity number on launch day, but players who stay, players who come back, players who bring their friends.

Every single one of you affects that number. The quest designer affects day-three retention. The systems designer affects day-thirty. The person building social features affects day-one-eighty. If you're ever wondering whether the thing you're working on matters, ask yourself: does this make someone come back tomorrow? If the answer is yes, you're building the right thing.

If we get the retention right, then we know we have a strong central recipe and we can grow that."""

answers[9] = """So this question has two layers to it and I want to be clear about both because they do different jobs.

The first layer is the vision pillars. These are the values the game is built on. They tell you why we make the decisions we make. Player identity as the product. Play has real value. Social first, game second. Skill over stats. A living, player-influenced world. Those five don't bend. Every decision in the studio runs through them.

The second layer is the foundational pillars of the game itself. The non-negotiable things the game actually has. The stuff you're building. And these need to be concrete, because you can't go to your desk and build "social first, game second". You can build a friends system. So here's what the game doesn't ship without.

Class system and customisation. No locked classes. Your weapons, your off-hand, your build, your approach. You choose it, you change it, you're never stuck with a decision you made at character creation. The whole system is designed around freedom and choice.

Discovery is persistent through everything. Quests, hidden lore, hidden pieces of the world you stumble on when you weren't looking for them. The act of discovering things isn't a side activity. It's threaded through the entire game. You should be able to turn a corner and find something worth finding, every session, whether you were on a quest or just wandering.

Meaningful player choice. And I mean broadly. Not just how your character looks, although that matters. How you approach combat. How you approach a puzzle. How you set up your space. Your home, your arcade, your collections. The game should feel like it's responding to your choices, not funnelling you down one path.

Social is easy and your actions matter to the people around you. It should be effortless to play with friends, to form groups, to be part of a faction. And what you do should have a beneficial impact on the people you're connected to. Not just "I helped my friend kill a boss". Your progress, your discoveries, your contributions should ripple out to the people who matter to you in the game. That's what makes social sticky rather than just a chat window bolted on.

Those are the foundational pillars. The vision pillars tell you why. These tell you what. And when you put a feature on the table, it should serve at least one from each layer, or we need to ask why we're building it.

As far as how they shape decisions, these are razors. You use them to shape and recut proposed features or systems to make sure they're in line with the objectives. If a feature is in opposition to these, it should be questioned. Prioritisation combines with the golden path, which is the core player journey, the experience we want every player to have. We get that established first, and then we flesh out the branches from there."""

answers[10] = "About 70% of the game should feel familiar to any MMO player. Quests, combat, inventory, crafting, guilds, housing - we're not reinventing wheels that work. Players shouldn't have to relearn what an MMO is. The 30% that's new is where we differentiate: the corruption mechanic (no MMO has done this systemically), stat-free cosmetics as THE progression, the soft-class weapon system with no locks, and the partner ecosystem bringing other games into our world. The familiar stuff earns trust. The new stuff is why they stay."

answers[11] = "Absolutely. The partner layer is one of our biggest differentiators. Partners bring their IP into our world through our pipeline, in our style, meeting our quality bar. They can have instanced environments, shop items, arcade cabinets. But the core single-player experience, the main story, the art direction - that's ours. Partners are the seasoning, not the recipe."

answers[12] = """This is one of my favourite questions because it gets at something we're doing deliberately differently.

In most MMOs, the player fantasy is "you're the chosen one". You're the hero of prophecy, the saviour of the world, the one person who can stop the ancient evil. And that's great until you're standing next to 850,000 other chosen ones all doing the same quest. The fantasy breaks. Everyone knows it breaks. Studios just kind of hope you don't think about it too hard.

We don't want that. You're not the solo protagonist. You're an explorer and a discoverer at the core. You showed up in a world that was torn apart long before you got here, a world that's been rebuilding itself for centuries, and there's a struggle going on in the background that's deeper and stranger than it first appears. The lore is layered. Some of it's on the surface. Some of it's buried. Some of it you'll stumble on by accident and it'll change how you see everything else.

But here's the thing. We don't want to hand you one fantasy. We want you to build your own.

That's what the open class system is for. That's what the customisation is for. You want to be the ultra shadowy assassin who moves through the world like smoke? Build that. You want to be the huge heavy warrior who walks through the front door of every fight? Build that. You want to be the collector who's more interested in what's on the shelves of their home than what's in their weapon hand? Build that. The game doesn't tell you who to be. You tell the game.

Your identity is about the choices you make. How you look. How you fight. How you move through the world. What you explore. What you collect. Who you run with. And the world notices. It responds to what you do. The corruption you push back stays pushed back. The people you help remember. The space you build reflects who you've become. This isn't a theme park where you ride the rides and nothing changes. Your choices land.

Some players will dig deep into the lore and start to understand the hidden purpose of the world and what Darwin really is and what the corruption actually means. Some players just want to run around and kill stuff. And both of those players are playing the game exactly right, because the fantasy is theirs, not ours.

So the player fantasy isn't about being the unkillable hero. It's about being the self the player wants to be, in a world with a deep and rich story that shows up in the architecture and the lore throughout the world. And that depth is there whenever you're ready for it."""

answers[13] = """That's going to be unique to every player, and it should be. But the thing I want us to inspire is long-lasting memories. The kind that stick with you years after you've stopped playing.

For the deep story player, that's the moment they finally uncover a major reveal in the lore. Something that reframes everything they thought they knew about the world. That becomes a memory they carry around and talk about.

For the combat player, it's "I just pulled that off". The battle was close, it could have gone either way, and they won. That rush. That's a memory.

But the one that really sits at the core of what we're building is the social one. You came together with a bunch of friends and you laughed until you cried. Maybe it was a conversation about something you earned in a partner game and brought back and how cool that was. Maybe it was a battle you all fought together that you didn't think you could win. Maybe it was a piece of content you just did and someone did something stupid and it became the thing you all reference for the next six months.

I'll tell you from personal experience. I ran one of the top-rated MMO guilds for a long time, across a very large MMO. And I don't remember the achievements as much as I remember the moments. The unique, delightful moments where we achieved something I didn't think we could pull off, even if we had to try it twelve times to get there. The twelfth attempt where it finally clicked and everyone lost their minds. That level of achievement and that level of camaraderie is what I want people to remember.

So that's the answer. We want three kinds of memories. The story reveal that changes how you see the world. The personal victory that felt earned. And the shared moment with the people you care about that you'll still be talking about years from now.

If we're building systems and content that create those moments, we're doing our job. If a player logs off and thinks about where they want to go next time, we've won."""

answers[14] = """Beneath the surface, the game is really about connection. And it starts with the world itself.

Everything you see in the game is telling you that two worlds have merged into one. Mythcore and gritcore. Literally the visuals are saying it. The ancient, which is actually the high-tech, and the new, which is the lower-tech, living side by side. There's a friction there. You can feel it in the architecture, in the landscape, in the way things look and sit next to each other. That friction is baked into the lore. It's the setting. It's why the world looks the way it does and feels the way it does.

The people within that world are the second layer. The NPCs have found a harmony of sorts. They've adapted. They live between those two juxtaposed elements and they've made it work. There's something in that, the idea of two opposing things learning to coexist, that runs through everything we're building.

And underneath all of that is a hidden and ancient conflict that set up an entire ecosystem. It created a world where those polarised elements exist in balance, and playing the game should let you feel both the tension and the harmony. That's the setting doing its job.

From the player's perspective, what the game is really about is discovery. But discovery means something different to every player. For one person it's the lore that leads them forward through the story arc. For another it's achievements they want to complete. For another it's making friends who love the same things you do, or even completely different things, and building lasting relationships while you play something you both really enjoy.

That's what sits beneath the surface. A world built on the tension and harmony of two things that shouldn't fit together but do. And players finding their own version of connection within it, whether that's with the world, with the story, or with each other."""

# Q15 removed - covered by Q14

answers[16] = """This is an MMO, so the honest answer is it's going to feel different depending on what part of the world you're in and what you're doing. And that's deliberate. We don't want one mood. We want the right mood for the right moment.

When you're moving through an area you haven't explored before, it should feel like wonder. A sense of curiosity pulling you forward. A light in the distance, a path you haven't taken, something you want to go and look at just because it's there. And agency. You can go and look at whatever you want. Nobody's telling you where to walk. If you're in that exploration mood, the world should reward it.

When you're sleuthing out a piece of the lore and following a quest line, you should feel engaged. Properly engaged. You should be telling yourself what you think the next answer is going to be, building your own theory, only to feel that bit of surprise when it doesn't work out exactly like you planned. Think of it like an M. Night Shyamalan film. You thought you knew. You didn't. And that's the fun of it.

When you're in combat, it should feel fast-paced and fun. Not blood-borne level tension where your hands are shaking on the controller. But impactful. Grandiose. When you swing, you should feel it connect. When you pull off something clever with your build, it should feel like you earned it. The interaction should make you feel powerful, not stressed.

And when you're just hanging out, decorating your space, browsing your collections, sitting in your home with friends around, it should feel unhurried. No timer ticking. No daily quest guilt. No anxiety about falling behind. Just a place you want to be.

The thread running through all of those is that the game should always feel like you chose to be here, not like you have to be. Curious when you're exploring. Engaged when you're chasing the story. Powerful when you're fighting. At ease when you're home. Every part of the game should earn the player's attention, not demand it."""

answers[17] = """So I'm going to try to give a specific answer here, but there's context that matters.

For some players, the central conflict is going to be other players. They're going to go straight to whatever PvP we put in the game and that's their opposition. That's valid and we're building for it.

There will be factions, because we want players to feel like they're part of something. Their friend group, their guild, their faction. Those factions will come into conflict at different points, and that tension is part of the experience.

But for the core game, the central conflict is corruption. It's a looming presence and an active threat as you explore through the world. Corrupted terrain, corrupted creatures, corrupted zones. You can see it. You can feel it. And you push it back. That's the opposition that every player shares regardless of playstyle.

Here's where it gets interesting. Players may not learn in the beginning why that corruption is there. It's just there. It's dangerous and it needs fighting. But as the player progresses, they'll begin to learn who the power behind the corruption is and why it exists. They'll begin to understand where it comes from. And the somewhat bittersweet idea is that it's not just a straight-up villain. There's something more complicated underneath it. That's the intent behind the central conflict.

And I'm being intentionally vague about that, because I don't think we're ready to roll the whole backchannel storyline out just yet. You'll have to stay tuned until we're ready to put that out at a detail level that makes sense for everybody. But trust me when I say the depth is there, and when we do reveal it, it's going to reframe things."""

answers[18] = """There is a deep narrative behind the game, and it deserves its own walkthrough. What I can tell you now is what I said a moment ago: the core story is the spine that every player walks. You arrive in a world that's been broken and is rebuilding. You encounter corruption. You learn about the factions. And you gradually uncover why this world exists and what's really going on underneath it. That arc is non-negotiable.

Everything that branches off that, faction-specific storylines, side quests, partner content, hidden lore you find by exploring, that's the extra story. And both are written with the same care.

But the full detail of how those layers work together, the narrative structure, the reveals, the pacing, that deserves its own session rather than a five-minute answer in an AMA. So we'll schedule a follow-up to go through it in depth. Robin and the narrative team will be part of that."""

answers[19] = """This is an honest and good question. Right now, we are building what we believe they want, and we're fitting to a game that we know we want to build, but understanding what players actually want is a process. As a part of that process, we'll be doing playtests and market studies and talking to players. The core game of what we're building today has a lot of key indicators across close competitors in the market that would say, "Yes, we are building things players want." The question will be, "Can all of us, everybody hearing this, everybody at Couch Heroes, put that into a full environment where they enjoy doing what they want?\""""

answers[20] = """We are very early in production. I don't even know how to answer this question right now. I don't know that I have a good feel of what gameplay and narrative is yet in a firm enough fashion to be able to put that together. If you're feeling disconnected, yes, I agree, and that's a focus point of what we do going forward, but it's going to feel disconnected for some time, as it did in every other MMO I worked on at this phase of development. So I wouldn't worry too much about this yet. If it stays that way, then that'll be of concern, but that alignment is coming as we look at creative direction and design aligning with the actual output of the gameplay. If we have to true up gameplay back to the design, we will."""

answers[21] = """I don't know how to answer this question, to be honest with you, because there's a lot of pieces that are hard to fix.

We are bringing on staff like network engineers so that we get the multiplayer to scale. We are going to iterate across combat until we get it right. It's got to have the right feel for the game we're making. Part of doing this AMA and the follow-up sessions is about staging and prioritizing all pieces of the game in a way that is coherent, that any of you can see. This is a question you'll never have to ask again."""

answers[22] = """I have about a million answers to this question, and I think it is not only in the general themes but in the details: butterflies on bushes and narrative hooks that lead you to a moment of surprise. A world that looks and feels lived in, persistent events that go on around the player instead of for the player. There's a lot of pieces in here that need to be in motion. The world needs to feel like it's in motion, but this is a great concept and something that we want to dig into deeply as we think about all of the different pieces of the world coming together. But when the world is growing and moving outside of the player's purview, and it's evolving with or without the player, it truly makes the world feel alive."""

answers[23] = """There are many ways to do this, and we will use a multitude of them. Many games have a progression bar, and you watch your progression bar go up, and that unlocks skills, and that's how you define the basics of progression. I know all of you know that. In our game, specifically, we want to broaden that horizon a little bit. It will be signaled by:
- your accumulation
- your house
- your cosmetics
- the number of games you visited
If anybody doesn't think that your steam catalog isn't life progression, then we got to talk.
But you've also got reputation and faction standing and guild progression. There's a lot of ways to think about progression for our game. We want to have multiple vectors, not just "I have to kill 400 wolves to get to level six.\""""

answers[24] = "You start in the Tutorial Cave with guided combat, your first corruption encounter, and your first traversal moments. Then you hit Downtime, the main city, and it opens up: your first home via House of Houses, your first faction interaction, your first proper quest chain. Within the first hour you should have fought, explored, customised, and met people. No twenty-minute cutscene. No reading a manual. Play first, understand later."

answers[25] = """That's very dependent on the player, but we want the world, when you first come into it, to have a sense of how huge the environment is and the mixture of old and new rolling together. Then later, it's about the ability to heal the land if that's something that appeals to you. Every player is going to have an individual wow moment, as we talked about in earlier questions. What we need to do is to be able to catalog the most common of those and make sure they are represented through the design of the game."""

answers[26] = "Early game is guided: tutorial, first quests, first home, first faction choice. Mid-game opens up: you choose your own path, deeper systems unlock, crafting and professions become available, the world gets bigger. Late game becomes social and expressive: world events, collections, partner content, guild progression, your home as a destination. The systems grow with you rather than just scaling numbers up."

answers[27] = """So, not finding the fun is not an option, and what will the studio do to make the game fun? I am talking to the studio. The studio is not Vardy or Robin; the studio is the studio. We need to aim collectively in one direction, as best we can, for what we believe that fun to be. We need to talk to players, we need to play tests, and we need to prove the hypothesis of fun as we go through it. What we know is:
- We know what each and every one of us finds is fun and we can go through that.
- We know exploring new worlds is fun.
- We know making friends for some is fun.
- We know winning in combat is fun.
It's how do we combine those collectively in a coherent design and delivery that is fun for everyone. If you wanted me to tell you how do we make the game fun, the answer is I'm asking you to be part of that."""

answers[28] = """The games platform provides architecture and features to support our game. It's a mechanical piece of the value we deliver as our world. It enables connection to partners. It'll set up the vacation into the game and ID and store, and it's augmentative to the game itself. So the best way to think about it is a codebase and interface that is, in many ways, just another feature or set of features for the game."""

answers[29] = "Partners show up in multiple ways. Arcade cabinets with playable mini-games from their IP. Cosmetic items in shops themed to their brand. Quest lines that introduce their world within ours. Dedicated shards or instanced environments for deeper integration. And items/rewards that cross between their game and ours. All of it goes through our art pipeline so it fits our visual language. Partners add flavour and variety. They don't change the recipe."

answers[30] = """No, we're not planning a smaller fallback version or an expand version. We're going to create the core of the game, and then we will grow out from there. I know there's lots of ideation that has been mixed up with production, and that's what the off-site and the layout of the roadmap and the core plan and the vertical slice and the walk down through the team are all about. We'll create what's necessary for our launch client, and then we will plan out a roadmap beyond that.

We need to plan both the studio's growth, which is certainly something we're doing right now, and the scope of the game in unison. They have to be in parallel, because the size of the studio and the size of the features and the criticality of the features against the pillars we need to deliver are all part of how this comes together with one direction."""

answers[31] = "So why is there no dedicated commercial sales person for partnerships? There will be. That's a role that we have to build out. Matter of fact, that may be a team we have to build out going forward, depending on how the early days go."

answers[32] = "Investors give studios money based on the value they think that the product the team can pull together and deliver to the market. Then, from that, they will exit or cash out their value for a higher worth than what they put in. That's as simple as it can be. Each investor has different timing based on the negotiations."

answers[33] = """There's a huge list of things we're doing right now that we won't want to do later. As we harden the production process, making iterations mid-sprint is something we're going to try to avoid, for example. Once we're live with players in the game, we can't make the kind of decisions we're making today about build quality and integrity, for example.

This is such a giant box of things; it's hard to narrow it down, but I can tell you our team size will change, our process will change and mature. Our bottleneck behind one person is dangerous as we move forward, so we need to have some redundancy. Many of these things are going to evolve a great deal."""

answers[34] = "We are optimizing for clarity, fun, quality, and then speed. That's why we're here today. That's why we're talking through all these pieces, because clarity and the criticality of fun lead to quality. Once we've achieved and maintained those, then we start working on our speed. But a company can't pick one thing to work on. It just doesn't work that way. We have to move all of those forward as best we can."

answers[35] = """Hindsight questions are always interesting. We would have locked the vision a little cleaner. We would have hardened the pillars on the front end. We would have matured and structured the production structure more. But to be honest with you, hindsight is also illusory. Part of the iteration that we've gone through makes us who we are today. Part of that hardening and understanding that we've earned through that journey is what puts us in a position to now harden the production cycle and now drive a clean vision with pillars and direction."""

answers[36] = """So there are a number of questions that have this theme to it. It's a bit of an aggregate, to be honest. We're not avoiding anything about the game platform or team, and if there's a sense that something is being avoided, you guys should speak up. The Slack channel that opens after this AMA, I would love to hear that.

I will tell you that, as far as I'm aware (and I've been here only since October of last year), there's no sense of avoidance. There's more about cognitive load and what can we fix first and when? If you feel like something is being avoided, I would tell you you're completely protected; call it out in the Slack channel and we'll talk through it. There is no intent to avoid; there is most certainly an intent to prioritize."""

answers[37] = """So this also came up as a theme: I am an outsider. I started in October; I've shipped 27 games before now, so I came in. The very first thing I did was put together a very long document for Vardis and Aris to review on what I critiqued. Yes, the team size needs to grow to fit the ambition. Yes, we need to harden production. Yes, there are a lot of changes that we're going to make and evolve on, but what we're doing right now is actually doing that.

I feel incredibly heartened that a lot of the pain points that you guys have called out while I've been here have come up in the critique. This AMA and all the steps after it are meant to mitigate those critiques and put us on a path to deliver the scope, the quality, and the fun that we want to."""

answers[38] = "Organization process and scope, and again that's why we're doing what we're doing."

answers[39] = "Once we get the production flow nailed down and we have really good clarity, then we'll start looking at marketing and business dev heads to come into the company to start working on their pieces of the puzzle. I think realistically it's later this year."

answers[40] = "Yes, the MMO genre is definitely viable. If you go look at the number of users playing MMOs today, I'd say there's still a lot of appeal to them. Our design is intended to approach the MMO game genre from a slightly different angle and something that's more attuned to what players want that is our belief."

answers[41] = "We are doing the AMA and walking through all the pieces I've answered in previous questions."

answers[42] = "Once the game is built, then the real work begins. We move into live and content expansion."

answers[43] = "We'll be looking at a live service model that, at least in the high levels of design, will be subscription and cosmetic-based. We're not going to be pay-to-win, that I can tell you, and we're not doing loot boxes or bitcoin or anything crazy. We're going to keep the monetization very clean and very healthy, but the full spec of that has not been laid out yet."

answers[44] = """I hate it when people ask when the release is going to be.

The reality is we need to get the production flow nailed down, the roadmap, the estimated time to completion by features, and the play testing done. There's a lot of variability between here and when we put this in front of players as a full launch, so I can't tell you a date. I can tell you it's not this year and it's not probably next year. What I can say is probably a little infuriatingly is we'll release when it's ready."""

answers[45] = "The roadmap and pipelines, and the seniors who are adding to the team, and the gating systems that we're putting in place are all about helping to prevent this going forward. So I hope you've heard an answer to this question as we've gone through this. If there are still follow-up questions, the Slack channel is available."

answers[46] = "This is near and dear to my heart. I despise rework. Getting to clean designs and technical design docs up front, with clear briefs on what kind of art and content we want, in good communication pipelines, is a cornerstone of the three-day off-site. The leads follow up right after it so that work is actually going on, starting next week."

answers[47] = "Right now, we're about 55 staff. We're already set up to hire about 16 or 17 more, and we'll continue to grow that until we meet our scope requirements. There's no way to provide an exact number; there isn't such a thing. It depends on the team and their ability to work together, the tool sets they use, and the scope of the game. We're roughly half the size we'll be, probably, by the time we launch."

answers[48] = "Because we're missing leads in multiple departments and it's overtaxing the senior folks that we have, so they can't do many of the things they need to on a daily basis. Also means that we have coherence and direction."

answers[49] = "This will probably be a whole topic that we'll discuss in detail when the time is right, but AI is an incredibly powerful tool set. It is also a polarizing one, so we will be very selective about AI usage. We don't want to alienate our player base, and at the same time we want to accelerate so that we can make more game. There are many tasks, like network node monitoring and validation, that AI can do for us, so we're definitely going to be looking at topics like that."

answers[50] = "From an ownership perspective, where there is no lead, you go to your department director. Where you have a personnel problem that is personal or friction with the department director, you go talk to Lorenza. As we grow out the HR team, you'll have more people to be able to reach out to. Part of the pipelining process is addressing these concerns directly."

# Write all answers to Excel
for r in range(2, ws.max_row + 1):
    qnum = ws.cell(row=r, column=1).value
    if qnum is not None and int(qnum) in answers:
        ws.cell(row=r, column=7, value=answers[int(qnum)])
        ws.cell(row=r, column=8, value="ANSWERED")

# Mark Q15 as removed
for r in range(2, ws.max_row + 1):
    qnum = ws.cell(row=r, column=1).value
    if qnum is not None and int(qnum) == 15:
        ws.cell(row=r, column=7, value="[REMOVED - Covered by Q14]")
        ws.cell(row=r, column=8, value="REMOVED")

wb.save('Clients/Couch Heroes/CH_AMA_Questions_Broken_Out2.xlsx')

answered = 0
pending = 0
removed = 0
for r in range(2, ws.max_row + 1):
    s = ws.cell(row=r, column=8).value
    if s == "ANSWERED": answered += 1
    elif s == "REMOVED": removed += 1
    elif ws.cell(row=r, column=3).value: pending += 1
print(f"Saved. Answered: {answered}, Removed: {removed}, Pending: {pending}")
