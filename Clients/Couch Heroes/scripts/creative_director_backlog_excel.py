"""Generate the Creative Director backlog as a formatted Excel workbook."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "CD Backlog"

# Styles
title_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
title_fill = PatternFill(start_color='1B2A4A', end_color='1B2A4A', fill_type='solid')
feature_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
feature_fill = PatternFill(start_color='2E5090', end_color='2E5090', fill_type='solid')
story_font = Font(name='Calibri', size=11, bold=True, color='1B2A4A')
story_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
task_font = Font(name='Calibri', size=10)
task_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
header_font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='404040', end_color='404040', fill_type='solid')
wrap = Alignment(wrap_text=True, vertical='top')
wrap_center = Alignment(wrap_text=True, vertical='top', horizontal='center')
thin_border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

# Column config
cols = {
    'A': ('Type', 8),
    'B': ('ID', 8),
    'C': ('Title', 40),
    'D': ('Description', 60),
    'E': ('Player Story', 50),
    'F': ('Success Factors', 55),
    'G': ('Tasks', 55),
    'H': ('Status', 12),
}

for col_letter, (name, width) in cols.items():
    ws.column_dimensions[col_letter].width = width

# Title row
ws.merge_cells('A1:H1')
cell = ws['A1']
cell.value = 'Couch Heroes MMO - Creative Director Backlog'
cell.font = title_font
cell.fill = title_fill
cell.alignment = Alignment(vertical='center')
ws.row_dimensions[1].height = 36

# Subtitle
ws.merge_cells('A2:H2')
ws['A2'].value = 'Owner: Creative Director (Aris)  |  Generated: 2026-05-11  |  Hierarchy: Feature > Story > Task'
ws['A2'].font = Font(name='Calibri', size=9, italic=True, color='666666')
ws.row_dimensions[2].height = 20

# Headers
row = 3
for col_letter, (name, _) in cols.items():
    cell = ws[f'{col_letter}{row}']
    cell.value = name
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = wrap_center
    cell.border = thin_border
ws.row_dimensions[row].height = 22

ws.freeze_panes = 'A4'

# --- DATA ---

backlog = [
    # Feature 1
    {
        'id': 'F1',
        'title': 'Creative Vision & Pillars',
        'description': 'Establish and document the creative north star for Couch Heroes MMO. Define the core creative pillars that every discipline uses to make judgment calls - what makes this game feel like this game and not a generic MMO.',
        'stories': [
            {
                'id': 'S1.1',
                'title': 'Define Core Creative Pillars',
                'description': 'Identify and articulate 3-5 creative pillars that define the Couch Heroes experience. These are the non-negotiable truths about the game\'s identity - the lens through which every feature, asset, and design decision is evaluated. Pillars must reinforce the existing design pillars (Discovery & Exploration, Corruption Mechanic, Player Expression, Soft-Class System, Partner Games Integration) while extending them into a creative framework all disciplines can use.',
                'player_story': 'As a player, when I watch a trailer or play for 5 minutes, I can immediately tell this game apart from every other MMO on the market. It has a clear identity I can describe to a friend in one sentence.',
                'success_factors': '- Pillars are distinct, memorable, and not interchangeable with competitor MMOs\n- Each pillar has a "this, not that" contrast\n- Every discipline lead can recite the pillars and give an example of how they apply to their work\n- Pillars are validated against the V2 feature list - nothing contradicts them',
                'tasks': '- Audit competitor MMO creative identities (WoW, FFXIV, GW2, New World, Palia)\n- Workshop with discipline leads to surface what they believe makes CH unique\n- Draft pillar candidates with "this, not that" definitions\n- Validate pillars against V2 feature list\n- Write Creative Pillars document with rationale and examples\n- Present to studio leadership for sign-off',
            },
            {
                'id': 'S1.2',
                'title': 'Creative Vision Document (CVD)',
                'description': 'Write the master Creative Vision Document that translates the pillars into actionable guidance for every department. Covers visual tone, narrative tone, gameplay tone, audio tone, and UX tone. This is the single reference document that answers "what does Couch Heroes feel like?" The byte-punk art style, the corruption/restoration dynamic, the young adult edgy-but-not-GTA narrative tone, and the cosy-but-deep gameplay philosophy must all be codified here.',
                'player_story': 'As a player, every part of the game - the combat, the world, the menus, the music, the story - feels like it was made by people who agreed on what they were making. Nothing feels out of place or bolted on.',
                'success_factors': '- Document covers visual, narrative, gameplay, audio, and UX tone\n- Each section includes concrete reference material (moodboards, audio references, gameplay clips)\n- Document is usable by a new hire on day one\n- Reviewed and endorsed by Art Director, Audio Director, Lead Designer, Narrative Lead',
                'tasks': '- Write visual tone section with moodboards (byte-punk reference library)\n- Write narrative tone section with sample dialogue and world voice\n- Write gameplay tone section with feel targets and pacing references\n- Write audio tone section with sonic palette references\n- Write UX tone section with interaction philosophy\n- Compile reference media library\n- Review cycle with all discipline leads\n- Final sign-off and distribution',
            },
            {
                'id': 'S1.3',
                'title': 'Competitive Differentiation Map',
                'description': 'Map Couch Heroes\' creative positioning against the MMO market. Identify where CH occupies unique creative space (corruption mechanic, partner portals, no-stat cosmetics, soft-class system) and where it deliberately overlaps with proven approaches. Feeds both creative direction and BD/marketing messaging.',
                'player_story': 'As a player choosing my next MMO, I can see exactly what Couch Heroes offers that I can\'t get elsewhere. It\'s not trying to be WoW or FFXIV - it\'s doing something I haven\'t seen before.',
                'success_factors': '- Map covers at least 8 competitor/adjacent titles\n- Identifies at least 3 areas of genuine creative differentiation\n- Identifies proven patterns to adopt (not reinvent)\n- Validated by BD/marketing as usable for investor and partner conversations',
                'tasks': '- Research and catalogue creative identities of 8-10 MMO competitors\n- Plot CH features against competitor feature sets on differentiation matrix\n- Identify creative white space opportunities\n- Identify proven patterns to adopt\n- Write competitive differentiation summary\n- Review with BD and marketing teams',
            },
        ]
    },
    # Feature 2
    {
        'id': 'F2',
        'title': 'World Identity & Lore',
        'description': 'Define the world of Couch Heroes - its history (including the AI proving ground / Darwin/Digit1 sci-fi-beneath-fantasy foundation), mythology, geography, cultures, factions (Allegiant, Athenaeum, Adeticus, Mendarium, Purple), and the central corruption/restoration dynamic driven by the antagonist Drystan\'s virus.',
        'stories': [
            {
                'id': 'S2.1',
                'title': 'World Bible - History & Mythology',
                'description': 'Write the foundational world bible covering the creation mythology (the world as an AI proving ground - Darwin/Digit1), major historical events, and the origin of Drystan\'s corruption virus. Establish why corruption exists, what restoration means, and what the player\'s role is. The sci-fi-beneath-fantasy lore foundation is unique to CH and must be handled with care - players discover it gradually, not dumped upfront.',
                'player_story': 'As a player, I feel like this world existed before I arrived. The ruins, the NPCs\' stories, the corruption spreading across the land - it all connects to something bigger. When I eventually discover the truth beneath the fantasy, it recontextualises everything.',
                'success_factors': '- History is internally consistent with no timeline contradictions\n- Corruption origin (Drystan\'s virus) is baked into world history\n- Sci-fi foundation is revealed gradually, not front-loaded\n- Lore is layered - surface-level for casual players, deep enough for lore hunters\n- Structured so quest writers and level designers can reference without reading the whole thing',
                'tasks': '- Draft creation mythology and the Darwin/Digit1 AI proving ground origin\n- Write Drystan\'s history and the origin of the corruption virus\n- Define major historical eras (3-5 ages/epochs)\n- Define the restoration force and its relationship to corruption\n- Map historical events to geographic locations\n- Create timeline with key dates and events\n- Write lore cheat sheet (1-page summary)\n- Review with narrative team for consistency',
            },
            {
                'id': 'S2.2',
                'title': 'Zone Identity Guide',
                'description': 'Define the creative identity for each major zone/region. Each zone needs a distinct visual theme, narrative purpose, corruption state, and emotional tone. Known zones include Tutorial Cave, Downtime (main city with five faction districts), Portal Peak, Hidden Hills, Farmlands, Mistrun, and Whispering zones. Corruption manifests differently per zone - not copy-paste corruption overlays.',
                'player_story': 'As a player, every time I enter a new zone, I feel like I\'ve travelled somewhere genuinely different. The colours change, the music shifts, the NPCs talk differently, and the corruption looks and feels different here.',
                'success_factors': '- Every zone in the V2 World Systems feature set has a defined identity\n- Zones have clear visual, audio, and narrative differentiation\n- Corruption manifests differently per zone\n- Zone identities support the progression curve\n- Each zone identity doc is detailed enough for environment artists and level designers',
                'tasks': '- Define zone count and geographic layout\n- Write identity briefs per zone: visual theme, narrative hook, corruption state, emotional tone, ambient audio\n- Define corruption visual language per zone (forest vs desert vs urban vs Portal Peak)\n- Create zone mood boards with reference imagery\n- Map zone identities to the player progression curve\n- Define biome-specific fauna (Portal Peak: animals, slimes, goblins, wolves)\n- Review with environment art lead and level design lead',
            },
            {
                'id': 'S2.3',
                'title': 'Faction & Culture Design',
                'description': 'Design the five factions with full creative identities. Allegiant (Red) - military, conservative, under-resourced. Athenaeum (Blue) - bureaucrats, knowledge hoarders, rigidly hierarchical. Adeticus (Yellow) - mercenaries, traders, scavengers, decentralised. Mendarium (Green) - caretakers, artisans, nature-focused, under-powered. Purple - ruling class / Mayoress faction. Each faction occupies a district in Downtime and drives player choice, quest content, and reputation systems.',
                'player_story': 'As a player, I feel genuine loyalty to my faction. They have a clear identity, a cause I believe in, and enemies I understand. Choosing sides matters because the factions feel real and different.',
                'success_factors': '- All five factions have distinct visual identity, motivation, and gameplay niche\n- No faction is objectively "the good one" - each has strengths, flaws, and moral grey areas\n- Faction designs support the reputation/faction system from V2\n- Cultural details are deep enough for quest writers to build on\n- Faction districts in Downtime have distinct architectural and environmental identity',
                'tasks': '- Write faction identity briefs: origin story, visual identity, values, relationship to corruption\n- Design inter-faction relationship matrix (allies, rivals, neutral, hostile)\n- Define faction visual markers (armour styles, colour palettes, architecture, banners)\n- Create faction leadership characters (named NPCs, including the Mayoress)\n- Map factions to zones (Downtime districts, territorial influence)\n- Define faction-specific quest tone and narrative hooks\n- Review with narrative team and game design lead',
            },
            {
                'id': 'S2.4',
                'title': 'NPC Archetype Library',
                'description': 'Define recurring NPC archetypes across the world - quest givers, merchants, faction officers, the Barbersmith, ambient townsfolk, named characters (Mayor, Goblins, Skeletons, Frogs, Pirates), and hostile creatures. Each archetype needs visual, behavioural, and narrative guidelines that adapt to faction and corruption state.',
                'player_story': 'As a player, NPCs feel alive. The blacksmith in a corruption-free village acts differently from the one in a besieged outpost. I can tell a character\'s faction and role at a glance.',
                'success_factors': '- Archetypes cover all NPC types in V2 (quest givers, vendors, Barbersmith, faction NPCs, combat NPCs)\n- Each archetype has visual, behavioural, and dialogue guidelines\n- Archetypes adapt to zone and corruption state\n- Named characters (Barbersmith, Mayor, faction leaders) have distinct personalities\n- Library is structured as a toolkit for quest designers and character artists',
                'tasks': '- Catalogue all NPC types from V2 feature list and GDD\n- Define archetype categories (quest giver, merchant, faction officer, Barbersmith, ambient, hostile, boss)\n- Write archetype guides: visual template, behaviour patterns, dialogue tone, corruption variants\n- Create naming convention guidelines per faction/culture\n- Design ambient life guidelines (townsfolk behaviours, patrol routes, idle actions)\n- Define named character personality briefs (Mayor, Barbersmith, key NPCs)\n- Review with character art lead and quest design team',
            },
        ]
    },
    # Feature 3
    {
        'id': 'F3',
        'title': 'Art Direction',
        'description': 'Establish the visual identity of Couch Heroes across all art disciplines. The byte-punk art style (high visual fidelity, stylised fantasy with digital/tech undertones) must be codified into actionable direction for character, environment, UI, VFX, and technical art.',
        'stories': [
            {
                'id': 'S3.1',
                'title': 'Art Style Guide (Byte-Punk Definition)',
                'description': 'Define and document the byte-punk art style - the overarching visual identity of Couch Heroes. Byte-punk is the intersection of high-fidelity fantasy and digital/technological undertones that reflects the world\'s AI proving ground origin. This guide sets proportions, colour philosophy, rendering approach, level of detail, and stylistic boundaries for all art production.',
                'player_story': 'As a player, the game\'s art style is distinctive and appealing. I can recognise a Couch Heroes screenshot instantly. The byte-punk aesthetic makes this world feel like no other MMO.',
                'success_factors': '- "Byte-punk" is clearly defined with visual references, not just a label\n- Style guide defines proportions, silhouette philosophy, colour palette approach, rendering\n- Includes "do" and "don\'t" examples for common art decisions\n- Style supports corruption/restoration dynamic visually\n- Art team can self-review work before CD review\n- High visual fidelity target is maintained without sacrificing readability',
                'tasks': '- Define byte-punk visual language with reference library\n- Establish character proportion standard\n- Create colour philosophy (palette structure, saturation rules, corruption colour language)\n- Define rendering approach (PBR baseline, stylised elements)\n- Create silhouette standards (class identification at distance)\n- Produce "do/don\'t" reference sheets per art discipline\n- Write material and texture guidelines\n- Review with David (Director of Art) and art team',
            },
            {
                'id': 'S3.2',
                'title': 'Character Design Direction',
                'description': 'Set the creative direction for all player and NPC character designs. Must support the no-stat cosmetics pillar ("if you can buy or find it, you can wear it"), the soft-class system (weapon + off-hand visual identity), character creation at the Barbersmith, and the garment system. Batsuits (stored end-game loadouts) need visual distinction.',
                'player_story': 'As a player, my character looks amazing and unique. I can see my class fantasy in my weapon loadout. The character creator gave me real choices. Other players\' characters all look different because cosmetics carry no stats.',
                'success_factors': '- Direction supports character creation, garment system, and Barbersmith customisation\n- Weapon + off-hand combinations are visually distinct (torch, shield, gauntlets, crossbow etc.)\n- Batsuit visual identity is defined (how stored loadouts look when equipped)\n- No-stat cosmetics means every outfit must look good at every level\n- Corruption effects on character models are defined',
                'tasks': '- Define body type range and proportions for playable races\n- Establish weapon + off-hand visual identity (each off-hand archetype visually distinct)\n- Design Batsuit visual language (how loadouts look, switching animations)\n- Set garment system visual rules (layering, dye system, no stat gating)\n- Define corruption effects on character models\n- Create character design review checklist\n- Produce reference sheets for character artists',
            },
            {
                'id': 'S3.3',
                'title': 'Environment Art Direction',
                'description': 'Define the visual approach to world environments within the byte-punk style. Covers terrain, modular environment sets, five faction district architecture in Downtime, vegetation, lighting, weather, and the corruption/restoration visual system. Portal Peak zone has specific requirements as the partner games hub.',
                'player_story': 'As a player, the world is breathtaking. Every zone has its own colour palette and architecture. Downtime\'s five faction districts each feel like a different neighbourhood. The corruption is unsettling and restoration is satisfying.',
                'success_factors': '- Direction covers all zone types from V2 (biomes, dungeons, Downtime districts, User Spaces, Portal Peak)\n- Corruption visual system defined with multiple stages\n- Modular environment sets support efficient art production\n- Five faction district architectures in Downtime are distinct\n- Day/night cycle and dynamic weather interact with art direction intentionally',
                'tasks': '- Define biome visual language per zone type\n- Design corruption visual progression (4-5 stages per biome)\n- Design restoration visual system\n- Establish architecture style guides per faction district in Downtime\n- Define modular environment set guidelines\n- Define day/night and weather visual impact\n- Set Portal Peak zone-specific art direction (partner portal aesthetic)\n- Create environment art review checklist',
            },
            {
                'id': 'S3.4',
                'title': 'VFX Direction',
                'description': 'Define visual effects philosophy for combat abilities (weapon crystals, custom crystal effects - trails, glows, sounds), corruption/restoration magic, environmental effects, and UI feedback. Crystal system VFX (combat crystals, custom crystals, shatter crystals, cursed crystals, dice crystals) are a major differentiator.',
                'player_story': 'As a player, I can read combat clearly even in group content. My weapon crystal effects feel powerful and unique. Corruption effects are menacing. Custom crystal cosmetic effects let me personalise my combat style.',
                'success_factors': '- VFX direction supports combat readability in 30-40 player instances\n- Crystal system VFX are visually categorised (combat, custom, shatter, cursed, dice)\n- Corruption and restoration VFX are distinct families\n- Performance budgets per VFX type are defined\n- Custom crystal cosmetic effects have clear creative boundaries',
                'tasks': '- Define VFX style within byte-punk aesthetic\n- Create combat VFX readability hierarchy\n- Design crystal system VFX families (combat, custom, shatter, cursed, dice)\n- Design corruption VFX language (spreading, pulsing, decay)\n- Design restoration VFX language (cleansing, growing, light)\n- Define status effect VFX library (buffs, debuffs, crowd control)\n- Set VFX performance budgets per content type\n- Create VFX review checklist',
            },
            {
                'id': 'S3.5',
                'title': 'UI Art Direction',
                'description': 'Define the visual identity of all user interfaces including the Hero Gadget (personal portable computer for environment interaction, data gathering, and UI). The Hero Gadget is the in-world justification for UI - it should feel like a device the character carries, not a traditional MMO menu system.',
                'player_story': 'As a player, the UI is beautiful and feels like it belongs in this world. The Hero Gadget makes menus feel diegetic - I\'m using a device, not navigating a spreadsheet.',
                'success_factors': '- UI art direction covers all UI surfaces from V2 (character sheet, inventory, crafting, map, combat HUD, housing, social)\n- Hero Gadget aesthetic is defined (how the device looks, opens, navigates)\n- Direction prioritises readability and usability\n- Consistent iconography system across all features\n- Accessible to colourblind players and supports text scaling',
                'tasks': '- Define Hero Gadget visual language and interaction model\n- Create UI colour system for states (default, active, disabled, corrupted)\n- Design iconography standards and icon grid\n- Set typography hierarchy\n- Define HUD layout philosophy (minimal, contextual, always-on)\n- Create accessibility guidelines (colourblind modes, scaling)\n- Produce UI art direction reference sheet',
            },
        ]
    },
    # Feature 4
    {
        'id': 'F4',
        'title': 'Audio Direction',
        'description': 'Establish the sonic identity of Couch Heroes - music direction, sound design philosophy, ambient soundscapes, and voice direction. Audio must reinforce the byte-punk aesthetic, the corruption/restoration dynamic, and the young adult edgy-but-fun tone.',
        'stories': [
            {
                'id': 'S4.1',
                'title': 'Music Direction & Sonic Identity',
                'description': 'Define the musical identity - genre, instrumentation palette, emotional range, and how music adapts to gameplay states (exploration, combat, corruption zones, restoration moments, social spaces in Downtime, User Space housing, Portal Peak, minigames). The music must work for the "cosy-but-deep" experience - 20 minutes or 4 hours both feel worthwhile.',
                'player_story': 'As a player, the music makes me feel things. Exploration makes me want to keep walking. Combat raises my heart rate. Corrupted zones sound wrong. The soundtrack is something I\'d listen to outside the game.',
                'success_factors': '- Music direction defines genre, instrumentation, emotional palette for byte-punk world\n- Adaptive music system specified (gameplay state transitions)\n- Corruption and restoration have distinct musical identities\n- Zone themes are musically distinct but share common sonic DNA\n- Music supports solo meditative play and group combat',
                'tasks': '- Define musical genre and instrumentation palette\n- Create emotional arc mapping per gameplay state\n- Design adaptive music system specifications\n- Define corruption sonic signature\n- Define restoration sonic signature\n- Create zone music briefs with reference tracks\n- Set music technical requirements (loops, transitions, stems)',
            },
            {
                'id': 'S4.2',
                'title': 'Sound Design Direction',
                'description': 'Define sound design philosophy for all game sounds - weapon combat (melee, ranged, spells, off-hand abilities), crystal effects, potion effects, UI (Hero Gadget sounds), environment, creatures, and corruption/restoration. The soft-class system means weapon sounds are identity-defining.',
                'player_story': 'As a player, every action feels weighty and real. My sword hits have impact. Spells crackle. The corrupted forest whispers. I can tell where I am by sound alone. My weapon sounds match my class fantasy.',
                'success_factors': '- Sound design covers all V2 feature areas (combat, movement, UI/Hero Gadget, world, creatures, abilities)\n- Weapon-type sounds support class identity (each off-hand combo sounds distinct)\n- Crystal system sounds are categorised and distinct\n- Environmental soundscapes per zone are specified\n- Mix priority hierarchy defined',
                'tasks': '- Define sound design philosophy (realistic, hyper-real, stylised)\n- Create weapon + off-hand sound identity per class archetype\n- Design crystal system sound families\n- Create combat sound hierarchy (impact types, ability categories)\n- Design environmental soundscape briefs per biome\n- Define corruption/restoration sound languages\n- Set UI/Hero Gadget sound design guidelines\n- Create creature sound archetypes\n- Define mix priority hierarchy',
            },
            {
                'id': 'S4.3',
                'title': 'Voice Direction',
                'description': 'Define voice scope and direction. The narrative tone is young adult, edgy but not GTA, accessible fantasy, witty about the meta of MMOs, never cynical about its own world. Five factions need distinct vocal identities. Key characters (Mayor, Barbersmith, faction leaders) need casting direction.',
                'player_story': 'As a player, the characters sound like real people in this world. Each faction sounds different. The tone is witty and fun without breaking immersion.',
                'success_factors': '- Clear decision on voice scope (full VO, partial VO, barks only)\n- If VO: casting direction per faction defined\n- Voice direction matches young adult, edgy-but-fun narrative tone\n- Faction vocal identities are distinct\n- Player character voice approach defined (silent, selectable, full)',
                'tasks': '- Decide voice scope and create VO scope document\n- Define vocal identity per faction (accent, tone, vocabulary)\n- Write voice casting direction briefs for key characters\n- Define combat bark style and intensity\n- Set narration style (if using narrator)\n- Create voice recording direction guidelines\n- Define localisation voice strategy',
            },
        ]
    },
    # Feature 5
    {
        'id': 'F5',
        'title': 'Player Fantasy & Experience Design',
        'description': 'Define what fantasy the player lives in Couch Heroes and how the experience evolves from Tutorial Cave through Downtime to endgame. The game targets explorers and socialisers first, achievers second. The moment-to-moment feel is "cosy exploration with real depth."',
        'stories': [
            {
                'id': 'S5.1',
                'title': 'Core Player Fantasy Definition',
                'description': 'Articulate the core player fantasy specific to Couch Heroes. The player is not the protagonist of the universe - they are an explorer and restorer in a world threatened by Drystan\'s corruption virus. The fantasy scales from "I\'m learning to fight corruption in Tutorial Cave" through "I\'m a veteran restorer with a fully built User Space and Batsuits." The soft-class system means the fantasy is "I am whatever I choose to be today."',
                'player_story': 'As a player, I feel like a hero whose actions matter. I\'m not just levelling up - I\'m pushing back corruption and restoring the world. I can switch my entire playstyle by changing my weapon loadout. My User Space is my trophy room.',
                'success_factors': '- Fantasy is specific to CH, not generic "be a hero"\n- Fantasy connects to corruption/restoration as player-driven narrative\n- Fantasy scales from Tutorial Cave to endgame\n- Fantasy supports multiple playstyles (combat, crafting, social, exploration)\n- Fantasy is testable - playtesters can articulate it unprompted',
                'tasks': '- Write core fantasy statement (one paragraph, evocative)\n- Define fantasy arc per game phase (Tutorial Cave, Downtime, levelling, endgame)\n- Map fantasy delivery to specific game systems\n- Define how soft-class system creates personal fantasy expression\n- Define how User Space / Batsuits create a personal identity narrative\n- Write fantasy validation questions for playtest surveys\n- Review with game design and narrative teams',
            },
            {
                'id': 'S5.2',
                'title': 'Class/Role Fantasy Design (Weapon + Off-Hand System)',
                'description': 'Define the power fantasy for each weapon + off-hand combination. Torch = Adventurer, healing spells = Priest, thrown potions = Support, second melee = Rogue, gauntlets = Monk, shield = Tank, crossbow/pistol = Ranged, spells = Mage, second ranged = Gunslinger. Each combo needs a fantasy that\'s creatively distinct and supported by visuals, audio, and abilities.',
                'player_story': 'As a Tank with shield, I feel like an immovable wall. As an Adventurer with torch, I feel like a pioneer exploring corruption. Each weapon combo has its own identity and I love switching between them.',
                'success_factors': '- All 9+ weapon + off-hand combos have a defined fantasy statement\n- Fantasies are distinct - no two combos feel like reskins\n- Fantasies inform art direction (visual), audio (sounds), and design (abilities)\n- Batsuit switching between fantasies feels like a meaningful identity choice\n- Fantasies connect to the corruption/restoration theme where possible',
                'tasks': '- Define all weapon + off-hand archetypes and their fantasy statements\n- Map fantasy delivery to specific abilities, visuals, and audio per combo\n- Define visual identity hooks per archetype\n- Define audio identity hooks per archetype\n- Design Batsuit switching as a creative moment (animation, sound, visual shift)\n- Create class fantasy validation criteria for playtesting\n- Review with combat design, art direction, audio direction',
            },
            {
                'id': 'S5.3',
                'title': 'Player Journey Map (Tutorial Cave to Endgame)',
                'description': 'Map the complete player experience. Key phases: Tutorial Cave (first contact, learn movement + combat + corruption), Downtime arrival (discover city, factions, User Space via House of Houses, Barbersmith), zone exploration (Portal Peak as stretch), midgame (professions, guild joining, deeper combat), endgame (Batsuits, User Space decoration, partner content, social status). Time targets: 20 minutes should feel worthwhile, 4 hours should feel like a natural session.',
                'player_story': 'As a player, I never feel lost or bored. Tutorial Cave taught me the basics without boring me. Discovering Downtime was exciting. Key moments - first boss kill, discovering corruption, restoring my first region, building my User Space - felt earned.',
                'success_factors': '- Journey map covers Tutorial Cave, Downtime, early zones, midgame, endgame\n- Each phase has defined emotional targets\n- Key "hero moments" are identified and designed for impact\n- Pacing targets defined (time to first combat, first group content, endgame)\n- Churn risk points identified with creative interventions\n- Supports the 20-minute casual and 4-hour deep session targets',
                'tasks': '- Define journey phases and duration targets\n- Map emotional arc per phase\n- Identify 8-12 "hero moments" and design their creative impact\n- Define pacing targets for key milestones\n- Identify churn risk points and creative interventions\n- Map journey to V2 features\n- Write creative briefs for each hero moment\n- Review with game design, UX, and production',
            },
            {
                'id': 'S5.4',
                'title': 'Social Experience Design',
                'description': 'Define how social interaction feels. CH targets explorers and socialisers first - the social experience is a pillar, not a bolt-on. Housing (User Space via House of Houses) is identity-first, not sandbox UGC. Visitors can portal into User Spaces. Minigames (Hot Potato, Archway Racetrack) provide casual social play. Guilds, factions, and community features drive long-term social bonds.',
                'player_story': 'As a player, I made friends in this game. It was easy to group up. My User Space became a social hub friends visited. The minigames gave us things to do together that weren\'t just combat.',
                'success_factors': '- Social design supports User Space visiting, guild systems, party play, minigames\n- Social onboarding is low-friction (easy to group, no punishment for casual play)\n- User Space as social hub is designed (visitor features, public/private spaces)\n- Minigames serve social bonding, not just content padding\n- Social experience connects to corruption/restoration (cooperative restoration, group quests)',
                'tasks': '- Define social philosophy statement (explorer/socialiser-first positioning)\n- Design social onboarding flow (how new players find groups/guilds)\n- Define User Space as social space (visiting, public rooms, social features)\n- Design guild creative identity guidelines\n- Define minigame social design (Hot Potato, Archway Racetrack, Arcade Cabinets)\n- Set community tone guidelines (young adult, edgy but welcoming)\n- Review with social systems design and community management',
            },
        ]
    },
    # Feature 6
    {
        'id': 'F6',
        'title': 'Corruption & Restoration Creative System',
        'description': 'Define how corruption (Drystan\'s virus) and restoration look, sound, feel, and behave as a unified creative experience. Corruption is CH\'s signature mechanic - no other MMO has this at a systemic level. The "Restore the Lighthouse" quest demonstrates the core loop. This creative system spans every discipline.',
        'stories': [
            {
                'id': 'S6.1',
                'title': 'Corruption Visual & Audio Language',
                'description': 'Define the comprehensive sensory language of Drystan\'s corruption virus. What it looks like at every stage, what it sounds like, how it affects the environment, characters, and UI (Hero Gadget). Corruption should be immediately recognisable and viscerally unsettling without being horror-genre - maintaining the young adult tone. Must look different per biome (Portal Peak corruption vs Farmlands corruption vs Downtime corruption).',
                'player_story': 'As a player, corruption terrifies me in a compelling way. I can see it spreading across the landscape. The sounds change. Even the Hero Gadget UI starts to feel wrong. I want to fight it.',
                'success_factors': '- Corruption has a clear visual progression (5 stages from pristine to consumed)\n- Corruption looks and sounds different per biome\n- Corruption affects all sensory channels (visual, audio, UI/Hero Gadget, gameplay)\n- Corruption is unsettling but not horror-genre (young adult appropriate)\n- All disciplines have clear guidelines for implementing corruption',
                'tasks': '- Design 5-stage corruption visual progression with reference art per biome\n- Design corruption audio progression (ambient changes, music distortion)\n- Define corruption effects on NPC behaviour and dialogue\n- Define corruption effects on Hero Gadget UI elements\n- Create corruption VFX library specifications\n- Set corruption intensity standards per content type\n- Produce corruption reference bible (single cross-discipline document)',
            },
            {
                'id': 'S6.2',
                'title': 'Restoration Visual & Audio Language',
                'description': 'Define the sensory language of restoration - the emotional payoff for CH\'s core loop. The "Restore the Lighthouse" quest is the template. Restoration should feel heroic, satisfying, and earned. This is the moment players share clips of. Player actions must visibly drive restoration - not automatic timers.',
                'player_story': 'As a player, restoring a corrupted zone is one of the most satisfying things I\'ve ever done in a game. Colour returns, creatures come back, music swells. I feel like I made a real difference.',
                'success_factors': '- Restoration has a clear visual progression that reverses corruption\n- Restoration is emotionally uplifting and satisfying\n- Restored zones feel alive, not just "default"\n- Restoration creates memorable shareable moments\n- Player actions visibly drive restoration',
                'tasks': '- Design restoration visual progression (reversal of corruption, plus "thriving" state)\n- Design restoration audio progression (music restoration, ambient life returning)\n- Define restoration climax moments (what happens at full restoration)\n- Create restoration VFX specifications\n- Define player feedback for restoration actions (visual and audio rewards)\n- Design persistent restoration markers (how players see their contribution)\n- Produce restoration reference bible',
            },
            {
                'id': 'S6.3',
                'title': 'Corruption/Restoration Integration with Game Systems',
                'description': 'Define how the creative system integrates with every major game system - combat (corrupted weapons, abilities), crafting (corrupted materials, purification recipes), progression (corruption/restoration milestones), crystal system (cursed crystals), housing (can corruption threaten User Spaces?), economy, questing (Restore the Lighthouse template), potions (corruption antidotes?), and partner portals.',
                'player_story': 'As a player, corruption affects everything, not just the scenery. My weapon crystals can become corrupted. My crafting materials change. My User Space can be threatened. It\'s not a gimmick - it\'s the world.',
                'success_factors': '- Integration points defined for every major V2 feature area\n- Integration adds depth without adding friction\n- Creates meaningful player choices (embrace corruption for power vs resist for purity)\n- Creative direction for each integration documented\n- Cursed crystals and corruption-themed content have creative guidelines',
                'tasks': '- Audit V2 feature list for all corruption/restoration integration points\n- Define corruption effects on combat (corrupted weapons, abilities, enemy behaviours)\n- Define corruption effects on crystal system (cursed crystals creative direction)\n- Define corruption effects on crafting and potions\n- Define corruption effects on User Space/housing\n- Define corruption effects on economy\n- Write creative integration briefs per feature area\n- Review with game design leads per system',
            },
        ]
    },
    # Feature 7
    {
        'id': 'F7',
        'title': 'Partner Portals Creative System',
        'description': 'Define the creative direction for Partner Games Integration - CH\'s most unique pillar. Partner studios\' games integrate as portals within the CH world. Partner content includes in-world quests, arcade cabinets with playable mini-games, partner item shops, and quest rewards. Portal Peak is the primary zone. The game serves as a "home" players return to between adventures.',
        'stories': [
            {
                'id': 'S7.1',
                'title': 'Partner Portal Player Experience Design',
                'description': 'Define exactly how players discover, approach, enter, experience, and return from partner portals. The portal experience must feel magical and seamless - not a loading screen or a launcher. Players should feel like they\'re stepping through a doorway into another world, and coming home to CH afterwards. Partner content spans roguelikes, mini RPGs, puzzle games, racing, narrative shorts, retro arcade. The byte-punk aesthetic provides a natural frame for digital portals.',
                'player_story': 'As a player, discovering a new partner portal is exciting. I see it glowing in Portal Peak, walk up to it, and step through into a completely different game experience. When I come back, I have new loot, new stories, and my friends ask "where did you just go?" The portal system makes CH feel like a living hub of adventures.',
                'success_factors': '- Portal discovery is a moment of wonder, not a menu click\n- Portal entry feels like crossing a threshold between worlds\n- Partner game aesthetic coexists with byte-punk without breaking CH\'s identity\n- Return to CH feels like "coming home" with rewards to show\n- Experience works for both first-time discovery and repeat visits\n- 30+ potential partner game types (roguelikes to narrative shorts) all fit the framework',
                'tasks': '- Design portal discovery flow (how players find new portals in the world)\n- Design portal approach experience (visual and audio escalation as player gets closer)\n- Design portal entry sequence (the transition from CH world to partner content)\n- Design in-portal experience framing (how partner content is presented within CH context)\n- Design portal exit and return sequence (coming back to CH with rewards)\n- Define portal visual taxonomy (do different partner game types look different?)\n- Design "portal memory" system (how returning to a visited portal differs from first visit)\n- Create portal experience wireframes/storyboards\n- Review with game design, UX, and BD teams',
            },
            {
                'id': 'S7.2',
                'title': 'Partner Content Art Constraints & Style Bridge',
                'description': 'Define the art constraints for partner content integration. Partner games will have wildly different art styles - a pixel-art roguelike, a watercolour puzzle game, a retro arcade shooter. CH\'s byte-punk frame must contextualise these without diluting either CH\'s identity or the partner\'s. Define the "style bridge" - how CH wraps partner content so it feels intentional.',
                'player_story': 'As a player, partner games look different from CH\'s world, and that\'s the point - I\'m visiting another dimension. But the portal frame, the arcade cabinet, the partner shop all feel like CH. It never feels like a broken mod or a cheap embed.',
                'success_factors': '- Art constraints are clear enough for partner studios to self-assess fit\n- Style bridge framework handles diverse art styles without breaking CH identity\n- Portal frame / arcade cabinet frame is defined as a visual "container"\n- Partner items in CH world (shop items, quest rewards) adapt to byte-punk style\n- Technical brief for partners (included in the partner integration tech spec) covers art requirements',
                'tasks': '- Define the style bridge concept (how CH frames partner content visually)\n- Create portal frame art direction (the visual "doorway" wrapping partner content)\n- Create arcade cabinet art direction (the physical in-world object housing partner games)\n- Define partner item adaptation rules (how partner IP items look when in CH inventory/world)\n- Define partner shop visual integration (how partner stores fit in Downtime/zones)\n- Write art constraint spec for partner tech brief (resolution, palette, animation, asset limits)\n- Create 3-4 example mock-ups showing diverse partner styles within CH frames\n- Review with David (Director of Art) and BD team',
            },
            {
                'id': 'S7.3',
                'title': 'Partner Portal World Integration',
                'description': 'Define how partner portals exist within the CH game world narratively and spatially. Portal Peak is the primary zone for portals, but portals also appear in User Spaces (housing), Guild spaces, and potentially other zones. The world lore must explain why portals exist. The partner portal system ties into the "home" concept - CH is where players live, portals are where they adventure.',
                'player_story': 'As a player, portals feel like a natural part of this world, not a monetisation bolt-on. The lore explains why other dimensions exist. Portal Peak is my favourite zone because every visit there\'s something new to discover. I even have a portal in my User Space to my favourite partner game.',
                'success_factors': '- Portals have a lore justification within the world bible (connected to the AI proving ground / corruption origin?)\n- Portal Peak zone has a defined creative identity that celebrates the portal concept\n- User Space portals feel personal and earned\n- Guild portals serve social function\n- New partner portals appearing over time feels like world evolution, not patching\n- Portal placement follows world design logic, not arbitrary billboard placement',
                'tasks': '- Write portal lore (why other dimensions/worlds exist in CH\'s fiction)\n- Define Portal Peak zone creative identity and layout\n- Define User Space portal placement rules (how players earn/choose their home portal)\n- Define Guild portal integration (how guilds access partner content together)\n- Design "new portal arrival" world event (how the world reacts when a new partner portal opens)\n- Define portal density and distribution rules across the world map\n- Create portal world integration guidelines for level designers\n- Review with narrative team, level design, and BD',
            },
            {
                'id': 'S7.4',
                'title': 'Partner Content Quality & Tone Standards',
                'description': 'Define quality and tone standards for partner content. Not every game fits CH\'s world - the CD must set creative guardrails. Partner content must match CH\'s young adult, edgy-but-fun tone. Excessively violent, adult, or thematically clashing content needs clear rejection criteria. Quality floor must be defined so partner content doesn\'t undermine CH\'s polish.',
                'player_story': 'As a player, every partner game I try through a portal is genuinely fun and fits the vibe of CH. I\'ve never stepped through a portal and thought "this doesn\'t belong here." The quality is always at least good.',
                'success_factors': '- Tone guardrails are defined (what themes/content are acceptable, what\'s rejected)\n- Quality floor is specified per partner content type (arcade game, quest, item, full portal experience)\n- Review process for partner content is defined (who reviews, at what stage, turnaround)\n- Partner content certification process includes creative sign-off, not just technical\n- Guardrails are included in the partner-facing tech spec and briefing materials',
                'tasks': '- Define tone guardrails (acceptable themes, content rating alignment, rejection criteria)\n- Set quality floor per partner content type\n- Design creative review process for partner submissions\n- Write creative section of partner certification checklist\n- Define "partner content showcase" standards (how CH promotes partner games in-world)\n- Create partner creative brief template (what partners receive to understand CH\'s standards)\n- Review with BD, legal, and studio leadership',
            },
        ]
    },
    # Feature 8
    {
        'id': 'F8',
        'title': 'Content Direction & Quality Standards',
        'description': 'Establish what content gets made, to what quality standard, and in what order. The CD sets the quality bar and content strategy. With ~1,193 items in the V2 feature list across 6 sheets, creative prioritisation is essential to avoid scope paralysis.',
        'stories': [
            {
                'id': 'S8.1',
                'title': 'Content Quality Standards',
                'description': 'Define the minimum quality bar for all content types across the V2 feature list. CH targets high visual fidelity - quality can\'t slip just because the feature list is enormous. Standards must distinguish between "minimum shippable" and "target quality" so production can make informed trade-offs.',
                'player_story': 'As a player, nothing feels unfinished. Every quest has genuine writing. Every zone has detail. The quality is consistent throughout the world.',
                'success_factors': '- Quality standards defined per content type (art, animation, audio, narrative, UI)\n- Standards are objective enough for reviewers to apply consistently\n- "Minimum shippable" vs "target quality" tiers exist\n- A review checklist exists per content type\n- Standards are achievable within production constraints',
                'tasks': '- Define quality standards per content type with exemplar references\n- Create review checklists per content type\n- Set "minimum shippable" vs "target quality" vs "stretch quality" tiers\n- Define quality gate process (who reviews, when, escalation path)\n- Create quality exemplar library (approved examples for reference)\n- Review with production and discipline leads for feasibility',
            },
            {
                'id': 'S8.2',
                'title': 'Content Prioritisation Framework',
                'description': 'Define the creative framework for deciding which content gets made first. With 1,193 V2 items, creative prioritisation answers "when we have limited time, what matters most for the player experience?" The Minimum Viable Creative Experience (MVCE) for launch must be defined separately from the full vision.',
                'player_story': 'As a player, the launch content feels complete and polished. The core experience is fully realised. There aren\'t obvious gaps or placeholder zones.',
                'success_factors': '- Framework provides clear creative criteria for prioritisation\n- MVCE for launch is defined\n- Framework has been applied to V2 features to produce priority tiers\n- Post-launch content roadmap is outlined\n- Framework is usable by production to resolve "what do we cut?" decisions',
                'tasks': '- Define creative prioritisation criteria (fantasy impact, differentiation, replayability)\n- Apply framework to V2 feature list to produce priority tiers\n- Define MVCE for launch\n- Identify "must-have at launch" vs "strong post-launch" content\n- Write creative content roadmap (launch, month 1, month 3, month 6, year 1)\n- Review with production, game design, and studio leadership',
            },
            {
                'id': 'S8.3',
                'title': 'Creative Review Process',
                'description': 'Establish how creative work is reviewed. The CD (Aris as CEO/CD) must balance creative oversight with not becoming a bottleneck. Define tiers of review, delegation, cadence, and feedback format. Integrate with the Agilefall gate system the studio uses.',
                'player_story': 'As a player, the game\'s creative quality is consistently high across all zones and features. Nothing feels rushed next to something amazing.',
                'success_factors': '- Review tiers defined (CD review, discipline lead review, peer review)\n- Content types mapped to review tiers\n- Review cadence integrates with Agilefall gates\n- Standardised feedback format\n- Process doesn\'t create bottlenecks (CD reviews critical items, delegates routine)',
                'tasks': '- Define review tiers and delegation rules\n- Map content types to review tiers\n- Establish review cadence per Agilefall gate\n- Create standardised feedback template\n- Integrate with gate system from production strategy\n- Define "creative hold" process\n- Review with production team for schedule impact',
            },
        ]
    },
    # Feature 9
    {
        'id': 'F9',
        'title': 'IP & Brand Creative Cohesion',
        'description': 'Ensure the Couch Heroes IP is creatively cohesive across all touchpoints - in-game, marketing, social media, partner communications, The Platform (web + mobile app), and any future media extensions. The CD guards IP consistency.',
        'stories': [
            {
                'id': 'S9.1',
                'title': 'Brand Visual Identity Guidelines',
                'description': 'Define how CH looks outside the game. Covers logo, marketing art (how byte-punk translates to social tiles, trailer cards, Steam capsules, banner ads), streaming overlays, merchandise direction. Must work across The Platform (web + mobile app) as well as traditional marketing channels. The brand voice is young adult, edgy but not GTA, witty about MMO meta.',
                'player_story': 'As a player, I see a CH ad on social media and immediately recognise it. The marketing looks like the game. The merch looks like the game. Everything is connected.',
                'success_factors': '- Brand guidelines cover logo, colour palette, typography, imagery\n- Marketing art direction consistent with in-game byte-punk style\n- The Platform (web + mobile) visual identity defined\n- Guidelines usable by external agencies and partners\n- Brand adapts to contexts (trailer, tweet, merch, investor deck) while staying cohesive',
                'tasks': '- Define brand visual identity package\n- Create marketing art direction guidelines (byte-punk to marketing translation)\n- Design The Platform (web + mobile) visual identity\n- Design social media visual templates\n- Set merchandise art direction guidelines\n- Write streaming/content creator asset guidelines\n- Create brand guideline document for external partners\n- Review with marketing and BD teams',
            },
            {
                'id': 'S9.2',
                'title': 'Marketing Creative Alignment',
                'description': 'Ensure all marketing output accurately represents the game and reinforces the creative pillars. CH\'s corruption visuals, partner portal concept, and byte-punk aesthetic are the marketing hooks. The CD approves or delegates approval of all marketing creative. Marketing must never misrepresent the game.',
                'player_story': 'As a player, the game I\'m playing is the game that was advertised. The trailers showed real gameplay. The corruption clips looked as good in-game as they did in the trailer.',
                'success_factors': '- Marketing creative approval process defined\n- Marketing accurately represents actual gameplay\n- Key art, trailers, screenshots reinforce creative pillars\n- Marketing tone matches young adult, edgy-but-fun voice\n- Shared approved asset library exists',
                'tasks': '- Define marketing creative approval workflow\n- Create marketing tone guidelines (do/don\'t phrases, sample posts)\n- Establish screenshot and trailer capture guidelines\n- Build shared approved asset library\n- Set key art direction per campaign type\n- Define "red lines" for marketing (never misrepresent)\n- Review with marketing team',
            },
        ]
    },
]

# Write data
row = 4
for feature in backlog:
    # Feature row
    ws.merge_cells(f'C{row}:H{row}')
    ws[f'A{row}'].value = 'FEATURE'
    ws[f'A{row}'].font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    ws[f'A{row}'].fill = feature_fill
    ws[f'A{row}'].alignment = wrap_center
    ws[f'A{row}'].border = thin_border
    ws[f'B{row}'].value = feature['id']
    ws[f'B{row}'].font = feature_font
    ws[f'B{row}'].fill = feature_fill
    ws[f'B{row}'].alignment = wrap_center
    ws[f'B{row}'].border = thin_border
    c = ws[f'C{row}']
    c.value = f"{feature['title']}\n\n{feature['description']}"
    c.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    c.fill = feature_fill
    c.alignment = wrap
    c.border = thin_border
    for col in ['D','E','F','G','H']:
        ws[f'{col}{row}'].fill = feature_fill
        ws[f'{col}{row}'].border = thin_border
    ws.row_dimensions[row].height = 60
    row += 1

    for story in feature['stories']:
        # Story row
        ws[f'A{row}'].value = 'STORY'
        ws[f'A{row}'].font = Font(name='Calibri', size=10, bold=True, color='2E5090')
        ws[f'A{row}'].fill = story_fill
        ws[f'A{row}'].alignment = wrap_center
        ws[f'A{row}'].border = thin_border
        ws[f'B{row}'].value = story['id']
        ws[f'B{row}'].font = story_font
        ws[f'B{row}'].fill = story_fill
        ws[f'B{row}'].alignment = wrap_center
        ws[f'B{row}'].border = thin_border
        ws[f'C{row}'].value = story['title']
        ws[f'C{row}'].font = story_font
        ws[f'C{row}'].fill = story_fill
        ws[f'C{row}'].alignment = wrap
        ws[f'C{row}'].border = thin_border
        ws[f'D{row}'].value = story['description']
        ws[f'D{row}'].font = task_font
        ws[f'D{row}'].fill = story_fill
        ws[f'D{row}'].alignment = wrap
        ws[f'D{row}'].border = thin_border
        ws[f'E{row}'].value = story['player_story']
        ws[f'E{row}'].font = Font(name='Calibri', size=10, italic=True)
        ws[f'E{row}'].fill = story_fill
        ws[f'E{row}'].alignment = wrap
        ws[f'E{row}'].border = thin_border
        ws[f'F{row}'].value = story['success_factors']
        ws[f'F{row}'].font = task_font
        ws[f'F{row}'].fill = story_fill
        ws[f'F{row}'].alignment = wrap
        ws[f'F{row}'].border = thin_border
        ws[f'G{row}'].value = story['tasks']
        ws[f'G{row}'].font = task_font
        ws[f'G{row}'].fill = story_fill
        ws[f'G{row}'].alignment = wrap
        ws[f'G{row}'].border = thin_border
        ws[f'H{row}'].value = 'Not Started'
        ws[f'H{row}'].font = task_font
        ws[f'H{row}'].fill = story_fill
        ws[f'H{row}'].alignment = wrap_center
        ws[f'H{row}'].border = thin_border
        ws.row_dimensions[row].height = 140
        row += 1

    # spacer
    row += 1

# Summary sheet
ws2 = wb.create_sheet('Summary')
ws2.column_dimensions['A'].width = 45
ws2.column_dimensions['B'].width = 12
ws2.column_dimensions['C'].width = 12

ws2.merge_cells('A1:C1')
ws2['A1'].value = 'Creative Director Backlog Summary'
ws2['A1'].font = title_font
ws2['A1'].fill = title_fill
ws2.row_dimensions[1].height = 30

for i, (hdr, w) in enumerate([('Feature', 45), ('Stories', 12), ('Tasks', 12)], 1):
    cell = ws2.cell(row=2, column=i, value=hdr)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = wrap_center
    cell.border = thin_border

for i, feature in enumerate(backlog, 3):
    total_tasks = sum(len(s['tasks'].split('\n')) for s in feature['stories'])
    ws2.cell(row=i, column=1, value=f"{feature['id']}: {feature['title']}").font = Font(name='Calibri', size=10, bold=True)
    ws2.cell(row=i, column=1).border = thin_border
    ws2.cell(row=i, column=2, value=len(feature['stories'])).alignment = wrap_center
    ws2.cell(row=i, column=2).border = thin_border
    ws2.cell(row=i, column=3, value=total_tasks).alignment = wrap_center
    ws2.cell(row=i, column=3).border = thin_border

total_row = 3 + len(backlog)
ws2.cell(row=total_row, column=1, value='TOTAL').font = Font(name='Calibri', size=10, bold=True)
ws2.cell(row=total_row, column=1).border = thin_border
ws2.cell(row=total_row, column=2, value=sum(len(f['stories']) for f in backlog)).font = Font(bold=True)
ws2.cell(row=total_row, column=2).alignment = wrap_center
ws2.cell(row=total_row, column=2).border = thin_border
ws2.cell(row=total_row, column=3, value=sum(sum(len(s['tasks'].split('\n')) for s in f['stories']) for f in backlog)).font = Font(bold=True)
ws2.cell(row=total_row, column=3).alignment = wrap_center
ws2.cell(row=total_row, column=3).border = thin_border

output_path = r'd:\OneDrive\Claude_code\NBIAI_TEAM\Clients\Couch Heroes\Couch_Heroes_Creative_Director_Backlog.xlsx'
wb.save(output_path)
print(f'Saved to {output_path}')
