import json, re, html, os
import openpyxl

# === LOAD MIRO DATA AND BUILD FEATURE HOURS (same as build_hours_map.py) ===
base = r'C:\Users\gpbea\.claude\projects\d--OneDrive-Claude-code-NBIAI-TEAM\d41e7002-7785-4c2a-ba67-02f379183b49\tool-results'

shapes = []
for f in ['mcp-claude_ai_Miro-board_list_items-1777998367328.txt',
          'mcp-claude_ai_Miro-board_list_items-1777998983493.txt',
          'mcp-claude_ai_Miro-board_list_items-1777998996501.txt',
          'mcp-claude_ai_Miro-board_list_items-1777999013112.txt']:
    shapes.extend(json.load(open(os.path.join(base, f), encoding='utf-8'))['data'])
stickies = json.load(open(os.path.join(base, 'mcp-claude_ai_Miro-board_list_items-1777998158578.txt'), encoding='utf-8'))['data']
texts = json.load(open(os.path.join(base, 'mcp-claude_ai_Miro-board_list_items-1777998375179.txt'), encoding='utf-8'))['data']

def strip_html(s):
    if not s: return ''
    s = re.sub(r'<br\s*/?>', ' ', s)
    s = re.sub(r'<[^>]+>', '', s)
    return html.unescape(s).strip()

def get_color(item):
    style = item.get('style') or {}
    return (style.get('fillColor') or style.get('backgroundColor') or '').lower()

def get_pos(item):
    pos = item.get('position') or {}
    geo = item.get('geometry') or {}
    return pos.get('x', 0), pos.get('y', 0), geo.get('width', 0), geo.get('height', 0)

def get_content(item):
    c = item.get('content') or ''
    if not c:
        d = item.get('data') or {}
        c = d.get('content') or ''
    return strip_html(c)

feat_colors = {'light_yellow','orange','dark_blue','light_blue','green','light_green','yellow','gray','grey'}
features = []
for item in stickies:
    color = get_color(item)
    content = get_content(item)
    x, y, w, h = get_pos(item)
    if color in feat_colors and 12000 < y < 28000 and w > 500 and len(content) > 2:
        features.append({'content': content, 'color': color, 'x': x, 'y': y, 'w': w, 'h': h})

grey_set = {'#e7e7e7','#b0b0b0','#d3d3d3','#c0c0c0','#a9a9a9','#e6e6e6','#cccccc','#f5f5f5','gray','grey','#808080','#ebebeb','#d9d9d9'}
grey_boxes = []
for item in shapes + stickies + texts:
    color = get_color(item)
    content = get_content(item)
    x, y, w, h = get_pos(item)
    is_grey = color in grey_set
    has_days = bool(re.search(r'\d+\s*[dD](?:\b|[)\s,\-])', content))
    has_estimate = bool(re.search(r'designer|engineer|concept|props|layout|anim|vfx|audio|rigging|UI|tech art|blockout|narrative|enviro', content, re.I))
    if is_grey and (has_days or has_estimate) and 12000 < y < 30000:
        grey_boxes.append({'content': content, 'color': color, 'x': x, 'y': y, 'w': w, 'h': h})

# Also include white text items that are biome art estimates
for item in texts:
    color = get_color(item)
    content = get_content(item)
    x, y, w, h = get_pos(item)
    if color in ('#ffffff', 'white', '') and 13500 < y < 15500:
        has_art = bool(re.search(r'concept.*\d+|props.*\d+|layout.*\d+|foliage.*\d+', content, re.I))
        if has_art and x > 236000 and x < 250000:
            grey_boxes.append({'content': content, 'color': 'white_art_est', 'x': x, 'y': y, 'w': w, 'h': h})

def match_grey_to_feature(gb, features):
    best = None
    best_score = float('inf')
    gx, gy = gb['x'], gb['y']
    for feat in features:
        fx, fy = feat['x'], feat['y']
        x_dist = abs(fx - gx)
        y_dist = fy - gy
        if x_dist < 800 and 0 < y_dist < 1500:
            score = x_dist * 2 + y_dist
            if score < best_score:
                best_score = score
                best = feat
    return best

matches = []
for gb in grey_boxes:
    feat = match_grey_to_feature(gb, features)
    if feat:
        matches.append({'feature': feat['content'], 'estimate': gb['content'],
                       'feat_x': feat['x'], 'feat_y': feat['y']})

def parse_days(content):
    roles = {}
    patterns = [
        (r'designer\s*\((\d+)[dD]\)', 'Game Designer'),
        (r'engineer\s*\(?(\d+)[dD]\)?', 'Gameplay Developer'),
        (r'UI\s*\(?(\d+)[dD]\)?', 'UI/UX Designer'),
        (r'UI\s+(\d+)d\b', 'UI/UX Designer'),
        (r'VFX\s*\(?(\d+)[dD]\)?', 'VFX Artist'),
        (r'VFX\s+(\d+)d\b', 'VFX Artist'),
        (r'[Aa]udio\s*\(?(\d+)[dD]\)?', 'Sound Designer'),
        (r'[Pp]rops?\s*\(?(\d+)[dD]\)?', 'Prop Artist'),
        (r'[Pp]rops?\s+(\d+)d\b', 'Prop Artist'),
        (r'[Aa]nim\s*\(?(\d+)[dD]\)?', 'Animator'),
        (r'[Aa]nim\s+(\d+)d\b', 'Animator'),
        (r'[Tt]ech\s*[Aa]rt\s*\(?(\d+)[dD]\)?', 'Technical Artist'),
        (r'[Tt]ech\s*[Aa]rt\s+(\d+)d\b', 'Technical Artist'),
        (r'[Tt]ech[Aa]rt\s*\(?(\d+)[dD]\)?', 'Technical Artist'),
        (r'[Cc]oncept\s*\(?(\d+)[dD]?\)?', 'Concept Artist'),
        (r'[Cc]oncept\s+(\d+)\b', 'Concept Artist'),
        (r'[Ee]nviro?\s*\(?(\d+)[dD]\)?', 'Environment Artist'),
        (r'[Ee]nv\s*(?:art)?\s*\(?(\d+)[dD]\)?', 'Environment Artist'),
        (r'[Ee]nv\s+(\d+)d\b', 'Environment Artist'),
        (r'[Ll]vl\s+design\s*\(?(\d+)[dD]\)?', 'Level Designer'),
        (r'[Bb]lockout\s*\(?(\d+)[dD]\)?', 'Level Designer'),
        (r'[Bb]lockout\s+(\d+)d\b', 'Level Designer'),
        (r'[Ll]ayout\s*\(?(\d+)[dD]\)?', 'Level Designer'),
        (r'[Ll]ayout\s+(\d+)d\b', 'Level Designer'),
        (r'[Nn]arrative\s*\(?(\d+)[dD]\)?', 'Writer'),
        (r'[Nn]arrative\((\d+)[dD]\)', 'Writer'),
        (r'[Rr]igging\s*\(?(\d+)[dD]?\)?', 'Character Rigger'),
        (r'[Rr]igging\s+(\d+)\b', 'Character Rigger'),
        (r'[Tt]ech\s+[Aa]nim\s*\(?(\d+)[dD]\)?', 'Character Rigger'),
        (r'[Tt]ech\s+[Aa]nim\s+(\d+)d\b', 'Character Rigger'),
        (r'[Cc]ombat\s+[Dd]esign(?:er)?\s*\(?(\d+)[dD]\)?', 'Combat Designer'),
        (r'[Cc]har\s*(?:\s+art)?\s*\(?(\d+)[dD]?\)?', 'Character Artist'),
        (r'[Cc]har\s+(\d+)d?\b', 'Character Artist'),
        (r'[Ff]oliage\s*\(?(\d+)[dD]\)?', 'Environment Artist'),
        (r'[Ff]oliage\s+(\d+)d\b', 'Environment Artist'),
        (r'GD\s*\(?(\d+)[dD]\)?', 'Game Director'),
        (r'GD\+CD\s*\(?(\d+)[dD]\)?', 'Game Director'),
        (r'CD\s*\(?(\d+)[dD]\)?', 'Combat Designer'),
        (r'economy\s+designer\s*\(?(\d+)[dD]\)?', 'Game Designer'),
        (r'commerce\s+designer[/\w]*\s*\(?(\d+)[dD]\)?', 'Game Designer'),
        (r'rod\s+(\d+)[dD]', 'Prop Artist'),
        (r'sign\s+(\d+)[dD]', 'Prop Artist'),
        (r'fish\s+(\d+)[dD]', 'Prop Artist'),
        (r'table\s+(\d+)[dD]', 'Prop Artist'),
        (r'modeling\s+\w+\s+(\d+)[dD]', 'Prop Artist'),
        (r'(?<!\w)[Dd]esign\s*\((\d+)[dD]\)', 'Game Designer'),
        (r'(?<!\w)[Dd]esign\s+(\d+)d\b', 'Game Designer'),
    ]
    for pattern, role in patterns:
        for m in re.finditer(pattern, content):
            try:
                days = int(m.group(1))
                roles[role] = roles.get(role, 0) + days
            except (ValueError, IndexError):
                pass
    return roles

feature_hours = {}
for m in matches:
    feat_name = m['feature'].split('\n')[0].strip()
    roles = parse_days(m['estimate'])
    if feat_name not in feature_hours:
        feature_hours[feat_name] = {}
    for role, days in roles.items():
        feature_hours[feat_name][role] = feature_hours[feat_name].get(role, 0) + days

# Exclusions
excluded_features = {'In Game Cinematic (IGC)', 'Meta quest system', 'Quest Generator system'}

# === MIRO NAME -> SPREADSHEET NAME MAPPING ===
NAME_MAP = {
    'FTUE (tutorial cave segment)': 'FTUE',
    'CombatBalance & Telemetry': 'Combat Balance & Telemetry',
    'Combat framework': 'Combat Framework',
    'Ranged combat': 'Ranged Combat',
    'Combat AI behavior': 'Combat AI Behavior',
    'PvP combat': 'PvP Combat',
    'User Space Server & Instance': 'User Space Server & Instance',
    'User space Decoration': 'User Space Decoration',
    'P2P/P2G User Space Connection': 'P2P / P2G User Space Connection',
    'Anywhere Door': 'Anywhere Door Placement',
    'Mounts/ Pets': 'Mounts / Pets',
    'Virtual Currency (VC) System': 'Currency System',
    'In game Economy': 'In-game Economy',
    'RMT game Economy': 'RMT Game Economy',
    'Quests Backend/Tooling': 'Quests Backend / Tooling',
    'Quests : Courier, Escort, Kill, Collect, research, time obj, Partner Quest, corruption': 'Quest Types - Courier / Escort / Kill / Collect / Mystery / Time / Partner / Lockpicking',
    'Quest reward system': 'Quest Reward System',
    'Quest log': 'Quest Log',
    'Quest dialog GUI': 'Quest Dialog GUI',
    'Quest Narrative Content': 'Quest Narrative Content',
    'Guilds System': 'Guilds',
    'Co-op/Party': 'Co-op / Party',
    'Co-op/Raids': 'Co-op / Raids',
    'In-game mail': 'In-game Mail',
    'LFG / Pick up groups': 'LFG / Pick-up Groups',
    'Day/Night cycle': 'Day/Night Cycle',
    'Dynamic weather': 'Dynamic Weather',
    'Player Account Creation/Deletion': 'Player Account Creation / Deletion',
    'Player Account Authentification': 'Player Account Authentication',
    '[Player Dashboard] Player Account Features': 'Player Dashboard - Account Features',
    '[Player Dashboard] Game Dashboard': 'Player Dashboard - Game Dashboard',
    '[Player Dashboard] Friends': 'Player Dashboard - Friends',
    '[Player Dashboard] Search & Notifications': 'Player Dashboard - Search & Notifications',
    '[Player Dashboard] Marketplace': 'Player Dashboard - Marketplace',
    '[Developer Dashboard] Developer Account': 'Developer Dashboard - Developer Account',
    '[Developer Dashboard] Game Dashboard': 'Developer Dashboard - Game Dashboard',
    'Red Books (emergency guidelines for Live Service)': 'Red Books (Live Service emergency guidelines)',
    'Monetization Bible': 'Monetization Bible',
    'Targeted VO': 'Targeted VO',
    'IG / RMT Monitoring System': 'IG / RMT Monitoring System',
    'Player Account Privacy Settings & Legal': 'Player Account Privacy Settings & Legal',
    'Partner Portal': 'Partner Portal',
    'Garment Shop': None,  # Not in spreadsheet
    'Partner Shop': None,  # Not in spreadsheet
    'Guild House': None,  # Not in spreadsheet
    'NPCs': None,  # Not in spreadsheet
    'Quest Item': None,  # Not in spreadsheet - might be under Quest types
    'For VS we need LG planned out and prototyped by the end of Early prod (tech, strategy etc)': None,
}

def normalize(s):
    return re.sub(r'[^a-z0-9]', '', s.lower())

# === ROLE -> COLUMN MAPPING (Gate 2 = EARLY PROD, starts at col 85/CG) ===
GATE2_START = 85
ROLE_TO_OFFSET = {
    'Game Director': 10,       # CQ = 95
    'Game Design Lead': 11,    # CR = 96
    'Game Designer': 13,       # CT = 98
    'Combat Designer': 14,     # CU = 99
    'Level Designer': 15,      # CV = 100
    'Technical Designer': 16,  # CW = 101
    'Writer': 17,              # CX = 102
    'Concept Artist': 19,      # CZ = 104
    'UI/UX Designer': 20,      # DA = 105
    'Character Artist': 22,    # DC = 107
    'Environment Artist': 23,  # DD = 108
    'Prop Artist': 24,         # DE = 109
    'Animator': 25,            # DF = 110
    'Character Rigger': 26,    # DG = 111
    'VFX Artist': 27,          # DH = 112
    'Technical Artist': 28,    # DI = 113
    'Gameplay Developer': 31,  # DL = 116
    'Sound Designer': 36,      # DQ = 121
}

# === LOAD AND UPDATE EXCEL ===
src = r'D:\OneDrive\Claude_code\NBIAI_TEAM\Clients\Couch Heroes\CouchHeroes_Man_Day_Work_Plan_v9.xlsx'
dst = r'D:\OneDrive\Claude_code\NBIAI_TEAM\Clients\Couch Heroes\CouchHeroes_Man_Day_Work_Plan_v10.xlsx'

wb = openpyxl.load_workbook(src)
ws = wb['Plan']

# Build spreadsheet feature name -> row mapping
sheet_features = {}
for row in range(5, 160):
    val = ws.cell(row=row, column=2).value
    if val:
        sheet_features[val.strip()] = row
        sheet_features[normalize(val.strip())] = row

# Step 1: WIPE all existing hours (columns G onwards for data rows 5-159)
# Gates span from col 7 (G) to col 434 (PR)
wiped_count = 0
for row in range(5, 160):
    for col in range(7, 435):
        cell = ws.cell(row=row, column=col)
        if cell.value is not None and isinstance(cell.value, (int, float)):
            cell.value = None
            wiped_count += 1

print(f"Wiped {wiped_count} hour cells")

# Step 2: FILL hours from grey boxes
filled = 0
not_found = []
skipped_excluded = []

for miro_name, roles in feature_hours.items():
    # Check exclusions
    if miro_name in excluded_features:
        skipped_excluded.append(miro_name)
        continue

    # Skip non-spreadsheet features
    if miro_name in NAME_MAP and NAME_MAP[miro_name] is None:
        continue

    # Map name
    sheet_name = NAME_MAP.get(miro_name, miro_name)

    # Find row
    row = sheet_features.get(sheet_name) or sheet_features.get(normalize(sheet_name))
    if not row:
        # Try fuzzy: check if normalized miro name is substring of any sheet feature
        norm_name = normalize(sheet_name)
        for sf, sr in sheet_features.items():
            if isinstance(sf, str) and len(sf) > 5:
                if norm_name in normalize(sf) or normalize(sf) in norm_name:
                    row = sr
                    break

    if not row:
        not_found.append(f"{miro_name} -> {sheet_name}")
        continue

    # Fill role hours into Gate 2 columns
    for role, days in roles.items():
        if role in ROLE_TO_OFFSET and days > 0:
            col = GATE2_START + ROLE_TO_OFFSET[role]
            ws.cell(row=row, column=col).value = days
            filled += 1

print(f"Filled {filled} cells across features")
print(f"Skipped (Not for VS): {skipped_excluded}")
print(f"Not found in spreadsheet: {not_found}")

# Step 3: Save
wb.save(dst)
print(f"\nSaved to: {dst}")
