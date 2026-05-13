import json, re, html, os, sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

base = r'C:\Users\gpbea\.claude\projects\d--OneDrive-Claude-code-NBIAI-TEAM\d41e7002-7785-4c2a-ba67-02f379183b49\tool-results'

shapes = []
for f in ['mcp-claude_ai_Miro-board_list_items-1777998367328.txt',
          'mcp-claude_ai_Miro-board_list_items-1777998983493.txt',
          'mcp-claude_ai_Miro-board_list_items-1777998996501.txt',
          'mcp-claude_ai_Miro-board_list_items-1777999013112.txt']:
    shapes.extend(json.load(open(os.path.join(base, f), encoding='utf-8'))['data'])
stickies = json.load(open(os.path.join(base, 'mcp-claude_ai_Miro-board_list_items-1777998158578.txt'), encoding='utf-8'))['data']
texts = json.load(open(os.path.join(base, 'mcp-claude_ai_Miro-board_list_items-1777998375179.txt'), encoding='utf-8'))['data']

def strip_html(s):
    if not s: return ''
    return html.unescape(re.sub(r'<[^>]+>', '', re.sub(r'<br\s*/?>', ' ', s))).strip()
def get_color(item):
    return ((item.get('style') or {}).get('fillColor') or '').lower()
def get_pos(item):
    pos = item.get('position') or {}; geo = item.get('geometry') or {}
    return pos.get('x',0), pos.get('y',0), geo.get('width',0), geo.get('height',0)
def get_content(item):
    c = item.get('content') or ''
    if not c: c = (item.get('data') or {}).get('content') or ''
    return strip_html(c)

# === COLLECT ALL ITEMS ===
feat_colors = {'light_yellow','orange','dark_blue','light_blue','green','light_green','yellow','gray','grey'}
features = []
for item in stickies:
    color = get_color(item)
    content = get_content(item)
    x, y, w, h = get_pos(item)
    if color in feat_colors and 12000 < y < 28000 and w > 500 and len(content) > 2:
        features.append({'content': content, 'name': content.split('\n')[0].strip(), 'color': color, 'x': x, 'y': y, 'w': w, 'h': h})

grey_set = {'#e7e7e7','#b0b0b0','#d3d3d3','#c0c0c0','#a9a9a9','#e6e6e6','#cccccc','#f5f5f5','gray','grey','#808080','#ebebeb','#d9d9d9'}
grey_boxes = []
for item in shapes + stickies + texts:
    color = get_color(item)
    content = get_content(item)
    x, y, w, h = get_pos(item)
    is_grey = color in grey_set
    has_days = bool(re.search(r'\d+\s*[dD](?:\b|[)\s,\-])', content))
    has_est = bool(re.search(r'designer|engineer|concept|props|layout|anim|vfx|audio|rigging|UI|tech art|blockout|narrative|enviro|months', content, re.I))
    has_number = bool(re.search(r'^\d+$', content.strip()))
    if is_grey and (has_days or has_est or has_number) and 12000 < y < 30000:
        grey_boxes.append({'content': content, 'color': color, 'x': x, 'y': y, 'w': w, 'h': h})

# Also white text art estimates for biomes
for item in texts:
    color = get_color(item)
    content = get_content(item)
    x, y, w, h = get_pos(item)
    if color in ('#ffffff', 'white', '') and 13500 < y < 15500:
        has_art = bool(re.search(r'concept.*\d+|props.*\d+|layout.*\d+|foliage.*\d+', content, re.I))
        if has_art and x > 236000 and x < 250000:
            grey_boxes.append({'content': content, 'color': 'white_art', 'x': x, 'y': y, 'w': w, 'h': h})

# === IMPROVED PARSER ===
def parse_days(content):
    roles = {}
    # Normalize content
    c = content.replace('\n', ' ').replace('\r', ' ')

    # Individual discipline patterns - order matters (more specific first)
    patterns = [
        # Specific compound terms first
        (r'[Tt]ech\s*[Aa]rt\s*\(?(\d+)[dD]?\)?', 'Technical Artist'),
        (r'[Tt]ech\s*[Aa]rt\s+(\d+)\b', 'Technical Artist'),
        (r'[Tt]ech[Aa]rt\s*\(?(\d+)[dD]?\)?', 'Technical Artist'),
        (r'[Tt]ech\s+[Aa]nim\s*\(?(\d+)[dD]?\)?', 'Character Rigger'),
        (r'[Tt]ech\s+[Aa]nim\s+(\d+)\b', 'Character Rigger'),
        (r'tech\s+art\s+(\d+)d?\b', 'Technical Artist'),
        (r'tech\s+anim\s+(\d+)d?\b', 'Character Rigger'),
        (r'[Cc]ombat\s+[Dd]esign(?:er)?\s*\(?(\d+)[dD]?\)?', 'Combat Designer'),
        (r'[Cc]ombat\s+[Dd]esign\s+(\d+)\b', 'Combat Designer'),
        (r'[Cc]har(?:acter)?\s+[Aa]rt(?:ist)?\s*\(?(\d+)[dD]?\)?', 'Character Artist'),
        (r'[Cc]har(?:acter)?\s+[Aa]rt\s+(\d+)\b', 'Character Artist'),
        (r'[Ee]nv(?:iro(?:nment)?)?\s+[Aa]rt\s*\(?(\d+)[dD]?\)?', 'Environment Artist'),
        (r'[Ll]vl\s+design(?:er)?\s*\(?(\d+)[dD]?\)?', 'Level Designer'),
        (r'[Ll]evel\s+[Dd]esign(?:er)?\s*\(?(\d+)[dD]?\)?', 'Level Designer'),
        (r'[Ll]ive\s+service\s+designer[/\s\w]*\s+(\d+)[dD]', 'Game Designer'),
        (r'economy\s+designer\s*\(?(\d+)[dD]?\)?', 'Game Designer'),
        (r'commerce\s+designer[/\w]*\s*\(?(\d+)[dD]?\)?', 'Game Designer'),
        (r'[Gg]ame\s+[Aa]rt\s*\(?concept\s*\(?(\d+)[dD]?\)?\)?', 'Concept Artist'),

        # GD and CD
        (r'GD\+CD\s*\(?(\d+)[dD]?\)?', 'Game Director'),
        (r'GD\s*\(?(\d+)[dD]?\)?', 'Game Director'),
        (r'(?<![a-zA-Z])CD\s*\(?(\d+)[dD]?\)?', 'Game Director'),  # CD usually = Creative Director in this context

        # Single-word disciplines
        (r'(?<![a-zA-Z])designer\s*\(?(\d+)[dD]?\)?', 'Game Designer'),
        (r'(?<![a-zA-Z])engineer\s*\(?(\d+)[dD]?\)?', 'Gameplay Developer'),
        (r'(?<![a-zA-Z])UI\s*\(?(\d+)[dD]?\)?', 'UI/UX Designer'),
        (r'(?<![a-zA-Z])UI\s+(\d+)d?\b', 'UI/UX Designer'),
        (r'(?<![a-zA-Z])VFX\s*\(?(\d+)[dD]?\)?', 'VFX Artist'),
        (r'(?<![a-zA-Z])VFX\s+(\d+)d?\b', 'VFX Artist'),
        (r'(?<![a-zA-Z])[Aa]udio\s*\(?(\d+)[dD]?\)?', 'Sound Designer'),
        (r'(?<![a-zA-Z])[Aa]udio\s+(\d+)\b', 'Sound Designer'),
        (r'(?<![a-zA-Z])[Pp]rops?\s*\(?(\d+)[dD]?\)?', 'Prop Artist'),
        (r'(?<![a-zA-Z])[Pp]rops?\s+(\d+)d?\b', 'Prop Artist'),
        (r'(?<![a-zA-Z])[Aa]nim\s*\(?(\d+)[dD]?\)?', 'Animator'),
        (r'(?<![a-zA-Z])[Aa]nim\s+(\d+)d?\b', 'Animator'),
        (r'(?<![a-zA-Z])[Cc]oncept\s*\(?(\d+)[dD]?\)?', 'Concept Artist'),
        (r'(?<![a-zA-Z])[Cc]oncept\s+(\d+)\b', 'Concept Artist'),
        (r'(?<![a-zA-Z])[Ee]nviro?\s*\(?(\d+)[dD]?\)?', 'Environment Artist'),
        (r'(?<![a-zA-Z])[Ee]nv\s*\(?(\d+)[dD]?\)?', 'Environment Artist'),
        (r'(?<![a-zA-Z])[Ee]nv\s+(\d+)d?\b', 'Environment Artist'),
        (r'(?<![a-zA-Z])[Bb]lockout\s*\(?(\d+)[dD]?\)?', 'Level Designer'),
        (r'(?<![a-zA-Z])[Bb]lockout\s+(\d+)\b', 'Level Designer'),
        (r'(?<![a-zA-Z])[Ll]ayout\s*\(?(\d+)[dD]?\)?', 'Level Designer'),
        (r'(?<![a-zA-Z])[Ll]ayout\s+(\d+)d?\b', 'Level Designer'),
        (r'(?<![a-zA-Z])[Nn]arrative\s*\(?(\d+)[dD]?\)?', 'Writer'),
        (r'(?<![a-zA-Z])[Nn]arrative\((\d+)[dD]?\)', 'Writer'),
        (r'(?<![a-zA-Z])[Rr]igging\s*\(?(\d+)[dD]?\)?', 'Character Rigger'),
        (r'(?<![a-zA-Z])[Rr]igging\s+(\d+)\b', 'Character Rigger'),
        (r'(?<![a-zA-Z])[Ff]oliage\s*\(?(\d+)[dD]?\)?', 'Environment Artist'),
        (r'(?<![a-zA-Z])[Ff]oliage\s+(\d+)d?\b', 'Environment Artist'),
        (r'(?<![a-zA-Z])[Cc]har\s*\(?(\d+)[dD]?\)?', 'Character Artist'),
        (r'(?<![a-zA-Z])[Cc]har\s+(\d+)d?\b', 'Character Artist'),

        # Specific prop items
        (r'rod\s+(\d+)[dD]', 'Prop Artist'),
        (r'sign\s+(\d+)[dD]', 'Prop Artist'),
        (r'fish\s+(\d+)[dD]', 'Prop Artist'),
        (r'table\s+(\d+)[dD]', 'Prop Artist'),
        (r'banner\s+(\d+)[dD]', 'Prop Artist'),
        (r'modeling\s+\w+\s+(\d+)[dD]', 'Prop Artist'),
        (r'unique\s+props?\s*\(?(\d+)[dD]?\)?', 'Prop Artist'),

        # Generic "Design" (careful - only match if not part of compound term)
        (r'(?<![a-zA-Z])[Dd]esign\s*\((\d+)[dD]?\)', 'Game Designer'),
        (r'(?<![a-zA-Z])[Dd]esign\s+(\d+)d\b', 'Game Designer'),

        # "env world building"
        (r'env\s+world\s+building\s+(\d+)d?\b', 'Environment Artist'),

        # TV/video tech art
        (r'TV\s+\w+\s+\w+\s+\w+\s+\(tech\s+art\s+(\d+)[dD]\)', 'Technical Artist'),

        # Lockpicking specifics
        (r'[Ll]ockpicking\s+game\s+art\s+\(concept\s+\((\d+)[dD]\)', 'Concept Artist'),
    ]

    for pattern, role in patterns:
        for m in re.finditer(pattern, c):
            try:
                days = int(m.group(1))
                roles[role] = roles.get(role, 0) + days
            except (ValueError, IndexError):
                pass

    return roles

def extract_all_numbers(content):
    """Extract all numbers that look like day values from content"""
    nums = []
    for m in re.finditer(r'(\d+)\s*[dD](?:\b|[)\s,\-])', content):
        nums.append(int(m.group(1)))
    # Also catch standalone numbers after discipline names
    for m in re.finditer(r'(?:designer|engineer|UI|VFX|audio|props?|anim|concept|enviro?|env|blockout|layout|narrative|rigging|foliage|char|tech art|techart|tech anim|GD|CD|combat design)\s*\(?(\d+)\)?', content, re.I):
        v = int(m.group(1))
        if v not in nums:
            nums.append(v)
    return nums

# === MATCH GREY BOXES TO FEATURES ===
def match_grey_to_feature(gb, features):
    best = None; best_score = float('inf')
    gx, gy = gb['x'], gb['y']
    for feat in features:
        x_dist = abs(feat['x'] - gx)
        y_dist = feat['y'] - gy
        if x_dist < 800 and 0 < y_dist < 1500:
            score = x_dist * 2 + y_dist
            if score < best_score:
                best_score = score; best = feat
    return best

# Build matches
matches = []
for gb in grey_boxes:
    feat = match_grey_to_feature(gb, features)
    if feat:
        matches.append({'feature_name': feat['name'], 'estimate_text': gb['content'],
                       'feat_x': feat['x'], 'feat_y': feat['y'],
                       'est_x': gb['x'], 'est_y': gb['y']})

# === FULL AUDIT ===
feature_hours = {}
audit_log = []
issues = []

for m in matches:
    feat_name = m['feature_name']
    raw_text = m['estimate_text']
    roles = parse_days(raw_text)
    parsed_total = sum(roles.values())

    # Extract all numbers from raw text for comparison
    all_nums = extract_all_numbers(raw_text)
    expected_total = sum(all_nums)

    if feat_name not in feature_hours:
        feature_hours[feat_name] = {}
    for role, days in roles.items():
        feature_hours[feat_name][role] = feature_hours[feat_name].get(role, 0) + days

    # Check for parse gaps
    gap = expected_total - parsed_total
    status = "OK" if gap == 0 else f"GAP: {gap}D unparsed"
    if gap > 0:
        issues.append(f"  {feat_name}: parsed {parsed_total}D of {expected_total}D ({gap}D gap)")
        issues.append(f"    Raw: {raw_text[:200]}")
        issues.append(f"    Parsed: {roles}")

    audit_log.append(f"{feat_name}: {parsed_total}D (raw nums sum={expected_total}D) [{status}]")
    audit_log.append(f"  Raw: {raw_text[:200]}")
    audit_log.append(f"  Parsed: {roles}")

# Handle Professions special case (no day numbers, says "3 months")
# Estimate: 1 designer + 1 engineer + UI + VFX = 3 months ~ 60 working days
# But we can't guess - flag it
if 'Professions' in feature_hours and sum(feature_hours['Professions'].values()) == 0:
    issues.append("  Professions: Grey box says '1 designer + 1 engineer + UI + VFX = 3 months' but no day numbers")

print("=== FULL AUDIT LOG ===\n")
for line in audit_log:
    print(line)

print(f"\n=== PARSE ISSUES ({len(issues)} lines) ===\n")
for issue in issues:
    print(issue)

# === UNMATCHED GREY BOXES ===
matched_texts = set(m['estimate_text'] for m in matches)
unmatched = [gb for gb in grey_boxes if gb['content'] not in matched_texts]

# Try to match unmatched to biomes by X position
# Get all biome-like features
biome_features = [f for f in features if 'biome' in f['name'].lower() or 'portal peak' in f['name'].lower()
                  or 'hidden hills' in f['name'].lower() or 'farmlands' in f['name'].lower()
                  or 'mistrun' in f['name'].lower() or 'whispering' in f['name'].lower()
                  or 'tranquil' in f['name'].lower() or 'torrential' in f['name'].lower()
                  or 'clockwork' in f['name'].lower() or 'downtime' in f['name'].lower()]

print(f"\n=== UNMATCHED GREY BOXES: {len(unmatched)} ===\n")
for gb in sorted(unmatched, key=lambda g: (g['y'], g['x'])):
    # Skip non-estimate items
    if not re.search(r'\d', gb['content']):
        continue
    if any(skip in gb['content'].lower() for skip in ['required for rti', 'year-end', 'tranquil sands', 'enviromental elemental', 'scale of the world']):
        print(f"  SKIP (label/note): x={gb['x']:.0f} y={gb['y']:.0f} | {gb['content'][:100]}")
        continue

    # Try biome matching by X proximity
    matched_biome = None
    for bf in biome_features:
        if abs(bf['x'] - gb['x']) < 200:
            matched_biome = bf['name']
            break

    if matched_biome:
        roles = parse_days(gb['content'])
        if roles:
            print(f"  BIOME MATCH: {matched_biome} <- {gb['content'][:100]}")
            print(f"    Parsed: {roles}")
            if matched_biome not in feature_hours:
                feature_hours[matched_biome] = {}
            for role, days in roles.items():
                feature_hours[matched_biome][role] = feature_hours[matched_biome].get(role, 0) + days
        else:
            print(f"  BIOME MATCH but 0 parsed: {matched_biome} <- {gb['content'][:100]}")
    else:
        print(f"  UNMATCHED: x={gb['x']:.0f} y={gb['y']:.0f} | {gb['content'][:120]}")

# === PRINT FINAL FEATURE HOURS TABLE ===
print(f"\n=== FINAL FEATURE HOURS ({len(feature_hours)} features) ===\n")
grand_total = 0
for feat in sorted(feature_hours.keys()):
    roles = feature_hours[feat]
    total = sum(roles.values())
    grand_total += total
    role_str = ", ".join(f"{r}={d}" for r, d in sorted(roles.items()))
    print(f"{feat}: {total}D | {role_str}")
print(f"\nGrand total: {grand_total}D")

# === WRITE EXCEL ===
NAME_MAP = {
    'FTUE (tutorial cave segment)': 'FTUE',
    'CombatBalance & Telemetry': 'Combat Balance & Telemetry',
    'Combat framework': 'Combat Framework',
    'Ranged combat': 'Ranged Combat',
    'Combat AI behavior': 'Combat AI Behavior',
    'PvP combat': 'PvP Combat',
    'User Space Server & Instance': 'User Space Server & Instance',
    'User space Decoration': 'User Space Decoration',
    'P2P/P2G User Space Connection': 'P2P / P2G User Space Connection',
    'Anywhere Door': 'Anywhere Door Placement',
    'Mounts/ Pets': 'Mounts / Pets',
    'Virtual Currency (VC) System': 'Currency System',
    'In game Economy': 'In-game Economy',
    'RMT game Economy': 'RMT Game Economy',
    'Quests Backend/Tooling': 'Quests Backend / Tooling',
    'Quests : Courier, Escort, Kill, Collect, research, time obj, Partner Quest, corruption': 'Quest Types - Courier / Escort / Kill / Collect / Mystery / Time / Partner / Lockpicking',
    'Quest reward system': 'Quest Reward System',
    'Quest log': 'Quest Log',
    'Quest dialog GUI': 'Quest Dialog GUI',
    'Quest Narrative Content': 'Quest Narrative Content',
    'Guilds System': 'Guilds',
    'Co-op/Party': 'Co-op / Party',
    'Co-op/Raids': 'Co-op / Raids',
    'In-game mail': 'In-game Mail',
    'LFG / Pick up groups': 'LFG / Pick-up Groups',
    'Day/Night cycle': 'Day/Night Cycle',
    'Dynamic weather': 'Dynamic Weather',
    'Player Account Creation/Deletion': 'Player Account Creation / Deletion',
    'Player Account Authentification': 'Player Account Authentication',
    '[Player Dashboard] Player Account Features': 'Player Dashboard - Account Features',
    '[Player Dashboard] Game Dashboard': 'Player Dashboard - Game Dashboard',
    '[Player Dashboard] Friends': 'Player Dashboard - Friends',
    '[Player Dashboard] Search & Notifications': 'Player Dashboard - Search & Notifications',
    '[Player Dashboard] Marketplace': 'Player Dashboard - Marketplace',
    '[Developer Dashboard] Developer Account': 'Developer Dashboard - Developer Account',
    '[Developer Dashboard] Game Dashboard': 'Developer Dashboard - Game Dashboard',
    'Red Books (emergency guidelines for Live Service)': 'Red Books (Live Service emergency guidelines)',
    'Monetization Bible': 'Monetization Bible',
    'Targeted VO': 'Targeted VO',
    'Partner Portal': 'Partner Portal',
    'Garment Shop': None,
    'Partner Shop': None,
    'Guild House': None,
    'NPCs': None,
    'Quest Item': None,
    'For VS we need LG planned out and prototyped by the end of Early prod (tech, strategy etc)': None,
    'Biome - Portal Peak': 'Biome - Portal Peak',
    'Biome - Hidden Hills': 'Biome - Hidden Hills',
    'Biome - Farmlands': 'Biome - Farmlands',
    'Biome - Mistrun Shore': 'Biome - Mistrun Shore',
    'Biome - Whispering Woods': 'Biome - Whispering Woods',
    'Biome - Tranquil Sands': 'Biome - Tranquil Sands',
    'Biome - Torrential Vale': 'Biome - Torrential Vale',
    'Biome - Clockwork': 'Biome - Clockwork',
    'Biome - Downtime': 'Biome - Downtime',
}

excluded = {'In Game Cinematic (IGC)', 'Meta quest system', 'Quest Generator system'}

GATE2_START = 85
ROLE_TO_OFFSET = {
    'Game Director': 10, 'Game Design Lead': 11, 'Game Designer': 13,
    'Combat Designer': 14, 'Level Designer': 15, 'Technical Designer': 16,
    'Writer': 17, 'Concept Artist': 19, 'UI/UX Designer': 20,
    'Character Artist': 22, 'Environment Artist': 23, 'Prop Artist': 24,
    'Animator': 25, 'Character Rigger': 26, 'VFX Artist': 27,
    'Technical Artist': 28, 'Gameplay Developer': 31, 'Sound Designer': 36,
}

src = r'D:\OneDrive\Claude_code\NBIAI_TEAM\Clients\Couch Heroes\CouchHeroes_Man_Day_Work_Plan_v9.xlsx'
dst = r'D:\OneDrive\Claude_code\NBIAI_TEAM\Clients\Couch Heroes\CouchHeroes_Man_Day_Work_Plan_v10.xlsx'

wb = openpyxl.load_workbook(src)
ws = wb['Plan']

sheet_features = {}
for row in range(5, 160):
    val = ws.cell(row=row, column=2).value
    if val:
        sheet_features[val.strip()] = row
        sheet_features[re.sub(r'[^a-z0-9]', '', val.strip().lower())] = row

# Wipe hours
wiped = 0
for row in range(5, 160):
    for col in range(7, 435):
        cell = ws.cell(row=row, column=col)
        if cell.value is not None and isinstance(cell.value, (int, float)):
            cell.value = None
            wiped += 1

# Fill
filled = 0
not_found = []
for miro_name, roles in feature_hours.items():
    if miro_name in excluded: continue
    if miro_name in NAME_MAP and NAME_MAP[miro_name] is None: continue
    sheet_name = NAME_MAP.get(miro_name, miro_name)
    norm = re.sub(r'[^a-z0-9]', '', sheet_name.lower())
    row = sheet_features.get(sheet_name) or sheet_features.get(norm)
    if not row:
        for sf, sr in sheet_features.items():
            if isinstance(sf, str) and len(sf) > 5:
                if norm in re.sub(r'[^a-z0-9]', '', sf.lower()) or re.sub(r'[^a-z0-9]', '', sf.lower()) in norm:
                    row = sr; break
    if not row:
        not_found.append(f"{miro_name} -> {sheet_name}")
        continue
    for role, days in roles.items():
        if role in ROLE_TO_OFFSET and days > 0:
            col = GATE2_START + ROLE_TO_OFFSET[role]
            ws.cell(row=row, column=col).value = days
            filled += 1

wb.save(dst)
print(f"\n=== EXCEL WRITTEN ===")
print(f"Wiped: {wiped} cells")
print(f"Filled: {filled} cells")
print(f"Not found: {not_found}")
print(f"Saved to: {dst}")
