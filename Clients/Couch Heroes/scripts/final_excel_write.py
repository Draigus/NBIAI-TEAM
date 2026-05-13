"""
Final Excel writer with manually verified corrections for every parse issue.
Every value below was hand-read from the raw Miro grey box text.
"""
import openpyxl, re

# === MANUALLY VERIFIED FEATURE HOURS ===
# Features marked [OK] had parser output matching raw text sums exactly.
# Features marked [FIXED] had parse errors — values below are hand-computed from raw text.

FEATURE_HOURS = {
    # --- Player Build ---
    'Player Progression': {'Game Designer':7,'Gameplay Developer':20,'UI/UX Designer':3,'VFX Artist':2,'Sound Designer':2},  # 34D [OK]
    'Skill System': {'Game Designer':20,'Gameplay Developer':35,'UI/UX Designer':15,'VFX Artist':5,'Sound Designer':7},  # 82D [OK]
    'Achievements & Trophies': {'Game Designer':7,'Gameplay Developer':35,'UI/UX Designer':5,'VFX Artist':3,'Prop Artist':5,'Sound Designer':2},  # 57D [OK]
    # Factions/reputation: no grey box on Miro
    'FTUE': {'Game Designer':8,'Level Designer':8,'Gameplay Developer':15,'UI/UX Designer':5,'VFX Artist':12,'Environment Artist':15,'Sound Designer':2,'Concept Artist':5,'Animator':5,'Technical Artist':5},  # 80D [FIXED: was 103D from double-count on lvl designer→designer, concept overlap]
    # Professions: grey box says "1 designer + 1 engineer + UI + VFX = 3 months" — no day numbers, SKIP
    # Character Management: no grey box
    'Character Creation & Customization': {'Game Designer':5,'Gameplay Developer':5,'UI/UX Designer':20,'Technical Artist':5},  # 35D [FIXED: was 40D, TechArt double-counted]
    'Garment System': {'Game Designer':20,'Gameplay Developer':60,'UI/UX Designer':20,'Technical Artist':5},  # 105D [FIXED: was 110D, TechArt double-counted]
    # Persona System: no grey box
    # Emote System: no grey box

    # --- World Systems ---
    'Day/Night Cycle': {'Game Designer':10,'UI/UX Designer':3,'Prop Artist':11,'Animator':15},  # 39D [OK] — note: "tech designer" mapped to Game Designer (closest role)
    'Dynamic Weather': {'Game Designer':17,'UI/UX Designer':20,'VFX Artist':5,'Prop Artist':10,'Gameplay Developer':60,'Technical Artist':45},  # 157D [FIXED: was 202D, TechArt double-counted]
    'Instancing': {'Gameplay Developer':120,'Game Designer':5,'Technical Artist':5},  # 130D [FIXED: was 135D, TechArt double-counted]
    'Dungeon Mechanics': {'Gameplay Developer':20},  # 20D [OK]
    'The Dungeon': {'Game Designer':35,'Combat Designer':10,'Level Designer':15,'Environment Artist':30,'Gameplay Developer':10},  # 100D [OK]
    'Partner Portal': {'Environment Artist':5,'VFX Artist':17,'Gameplay Developer':20},  # 42D [OK]
    # Biomes — manually matched from unmatched grey boxes by X position:
    'Biome - Portal Peak': {'Concept Artist':15,'Prop Artist':40,'Level Designer':70,'Technical Artist':15,'VFX Artist':30,'Character Artist':20,'Character Rigger':30,'Animator':20,'Game Designer':20},  # 260D
    'Biome - Hidden Hills': {'Concept Artist':10,'Prop Artist':30,'Level Designer':32,'Technical Artist':10,'VFX Artist':20,'Sound Designer':5},  # 107D
    'Biome - Farmlands': {'Concept Artist':10,'Prop Artist':30,'Level Designer':45,'Technical Artist':10,'VFX Artist':20,'Game Designer':5,'Sound Designer':3},  # 123D
    'Biome - Mistrun Shore': {'Concept Artist':10,'Environment Artist':20,'Level Designer':30,'Technical Artist':10,'VFX Artist':10,'Combat Designer':15,'Sound Designer':20,'Writer':4},  # 119D
    'Biome - Whispering Woods': {'Concept Artist':10,'Environment Artist':15,'Prop Artist':20,'Level Designer':45,'Character Artist':20,'Character Rigger':35,'Animator':30,'Technical Artist':15,'VFX Artist':20,'Game Designer':8,'Sound Designer':20,'Writer':5},  # 243D
    'Biome - Tranquil Sands': {'Concept Artist':10,'Environment Artist':20,'Prop Artist':25,'Level Designer':60,'VFX Artist':20,'Character Artist':20,'Character Rigger':35,'Animator':40,'Game Designer':8,'Sound Designer':10},  # 248D
    'Biome - Torrential Vale': {'Concept Artist':20,'Environment Artist':40,'Prop Artist':40,'Level Designer':65,'Character Artist':20,'Character Rigger':35,'Animator':30,'Technical Artist':15,'VFX Artist':25,'Game Designer':10,'Writer':7},  # 307D
    'Biome - Clockwork': {'Game Designer':20,'Level Designer':20,'Writer':7},  # 47D (art estimate not found)
    # Biome - Downtime: no grey box

    # --- Combat ---
    'Combat Framework': {'Game Designer':24,'Gameplay Developer':36,'Sound Designer':6,'Technical Artist':60,'UI/UX Designer':9,'VFX Artist':30},  # 165D [OK]
    'Ranged Combat': {'Gameplay Developer':20,'Animator':50},  # 70D [FIXED: was 20D, "Animation (50d)" missed — "Animation" not "Anim"]
    'Combat AI Behavior': {'Game Designer':14,'UI/UX Designer':6,'Animator':15,'Character Rigger':5,'VFX Artist':25,'Technical Artist':20,'Gameplay Developer':6,'Sound Designer':10},  # 101D [OK]
    'Enemies Art': {'Game Designer':10,'Concept Artist':30,'Gameplay Developer':20,'UI/UX Designer':30,'Sound Designer':2},  # 92D [FIXED: was 175D from spurious multi-match; Concept (maps) (10D+10D+10D) = 30D not 10D]
    'PvP Combat': {'Game Designer':16,'UI/UX Designer':3,'Gameplay Developer':3,'Sound Designer':2},  # 24D [OK]
    'Combat Balance & Telemetry': {'Gameplay Developer':40},  # 40D [OK]
    'Combat Abilities': {'Game Designer':20,'Concept Artist':2,'UI/UX Designer':10,'Animator':25,'Character Rigger':10,'VFX Artist':20,'Gameplay Developer':30,'Sound Designer':20},  # 137D [FIXED: was 147D, overcounted by 10]
    'Magic': {'Game Director':5,'Game Designer':30,'Combat Designer':5,'UI/UX Designer':10,'Animator':30,'Character Rigger':10,'VFX Artist':20,'Gameplay Developer':30},  # 140D [OK]
    'Full Motion Video (FMV)': {'Animator':60,'Character Rigger':20},  # 80D — from "NPCs" grey box which is char anim work

    # --- User Space ---
    'User Space Server & Instance': {'Gameplay Developer':60},  # 60D [OK]
    'User Space Decoration': {'Gameplay Developer':30,'Game Designer':7,'Sound Designer':2,'Level Designer':7,'Concept Artist':20,'Environment Artist':35,'Prop Artist':15,'VFX Artist':10},  # 126D [FIXED: was 88D, env modeling/set dressing/skybox missed]
    'P2P / P2G User Space Connection': {'Game Designer':1,'Gameplay Developer':30},  # 31D [OK]
    'Building Houses': {'Game Designer':15,'Sound Designer':2,'Gameplay Developer':60,'Environment Artist':25,'UI/UX Designer':3},  # 105D [FIXED: was 65D, "designer = (15D)" and "Env modeling 15d + world building (10d)" missed]
    'Saving House Layouts': {'Gameplay Developer':30},  # 30D [OK]
    'Anywhere Door Placement': {'Game Designer':3,'Sound Designer':3,'Environment Artist':15,'VFX Artist':15,'Concept Artist':5,'Gameplay Developer':5},  # 46D [FIXED: was 31D, "Env Modeling (15d)" and "(3 days)" format missed]

    # --- Items & Inventory ---
    'Item Core System': {'Gameplay Developer':30},  # 30D [OK]
    'Inventory System': {'Gameplay Developer':20,'Game Designer':15,'UI/UX Designer':30,'Concept Artist':10,'VFX Artist':20,'Sound Designer':2},  # 97D [FIXED: was 82D, "design 15 D" (space before D) missed]
    'Equipment System': {'Game Designer':10,'Concept Artist':10,'UI/UX Designer':40,'Technical Artist':10,'Gameplay Developer':20},  # 90D [OK — check: 10+10+40+10+20=90]
    'Loot Distribution': {'Gameplay Developer':40,'VFX Artist':15,'Game Designer':3,'UI/UX Designer':5},  # 63D [FIXED: was 73D; "1 engineer(30D) 10d" = 30+10 engineering, VFX 15d, designer 3D, UI 5d]
    'Mounts / Pets': {'Gameplay Developer':20,'Game Designer':5,'Combat Designer':2,'Character Artist':5,'Animator':10,'VFX Artist':5,'Technical Artist':2,'Character Rigger':3},  # 52D [FIXED: was 50D; CD=Combat Designer not Game Director; reuse gritcore=Character Artist; anim=10 not 13]
    'Banks': {'Gameplay Developer':20,'Environment Artist':20,'Concept Artist':5,'Sound Designer':1,'Game Designer':1},  # 47D [FIXED: was 27D, "env model (20d)" missed]
    'P2P Item Transfer': {'Gameplay Developer':30},  # 30D [OK]
    'Consumables': {'UI/UX Designer':2,'Animator':5,'Sound Designer':2},  # 9D [OK]
    # Durable Items: no grey box

    # --- Player Economy ---
    'Trading System': {'Game Designer':10,'UI/UX Designer':30,'Gameplay Developer':40,'Sound Designer':3},  # 83D [OK]
    'Currency System': {'Game Designer':10,'Combat Designer':2,'Concept Artist':2,'UI/UX Designer':30,'Gameplay Developer':20},  # 64D [OK]
    'Auction House': {'Gameplay Developer':120},  # 120D [OK]
    'In-game Store': {'Gameplay Developer':60},  # 60D [OK]
    'In-game Economy': {'Game Designer':90,'UI/UX Designer':30},  # 120D [OK]
    'RMT Game Economy': {'Gameplay Developer':60},  # 60D [OK]

    # --- Quest System ---
    'Quests Backend / Tooling': {'Gameplay Developer':60},  # 60D [OK]
    'Quests as Events': {'Gameplay Developer':30,'UI/UX Designer':30},  # 60D [OK]
    'Quest Multiplayer': {'Game Designer':5,'Gameplay Developer':30,'UI/UX Designer':30},  # 65D [OK]
    'Quest Types - Courier / Escort / Kill / Collect / Mystery / Time / Partner / Lockpicking': {'Game Director':2,'Game Designer':3,'Gameplay Developer':30},  # 35D [OK]
    'Quest Reward System': {'Gameplay Developer':20,'UI/UX Designer':10},  # 30D [OK]
    'Quest Log': {'Gameplay Developer':20,'UI/UX Designer':10},  # 30D [OK]
    'Quest Dialog GUI': {'Game Designer':4,'Combat Designer':2,'Writer':4,'UI/UX Designer':30},  # 40D [OK]
    'Quest Narrative Content': {'Game Director':3,'Game Designer':4,'Writer':7,'UI/UX Designer':20},  # 34D [OK]
    # In Game Cinematic: EXCLUDED (Not for VS)
    # Meta quest system: EXCLUDED (Not for VS)
    # Quest Generator system: EXCLUDED (Not for VS)
    'Targeted VO': {'Combat Designer':2,'Writer':20,'Sound Designer':10},  # 32D [OK]

    # --- Social & Multiplayer ---
    'Game Chat System': {'Game Director':1,'Game Designer':12,'UI/UX Designer':20,'Gameplay Developer':60},  # 93D [OK]
    'Friends': {'UI/UX Designer':30,'Gameplay Developer':20},  # 50D [OK]
    'Guilds': {'Game Director':1,'Game Designer':20,'UI/UX Designer':10,'Gameplay Developer':40},  # 71D [OK]
    'Co-op / Party': {'Gameplay Developer':75,'Game Designer':5,'Game Director':2,'UI/UX Designer':3},  # 85D [FIXED: was 75D; "1 engineer(60D) 15d" = 75D total engineering]
    'Co-op / Raids': {'Gameplay Developer':40},  # 40D [OK]
    'In-game Mail': {'Game Director':1,'Game Designer':3,'Writer':2,'Concept Artist':1,'UI/UX Designer':10,'Gameplay Developer':20},  # 37D [OK]
    'Notifications': {'Gameplay Developer':30},  # 30D [OK]
    'LFG / Pick-up Groups': {'Gameplay Developer':40},  # 40D [OK]

    # --- Game Bibles ---
    'Monetization Bible': {'Game Designer':40},  # 40D [OK]
    'Red Books (Live Service emergency guidelines)': {'Gameplay Developer':30},  # 30D [OK]

    # --- Platform ---
    'Player Account Creation / Deletion': {'Gameplay Developer':10},  # 10D [OK]
    'Player Account Authentication': {'Gameplay Developer':15},  # 15D [OK]
    'Player Account Privacy Settings & Legal': {'Gameplay Developer':10},  # 10D [OK]
    'Player Dashboard - Account Features': {'Gameplay Developer':30},  # 30D [OK]
    'Player Dashboard - Game Dashboard': {'Gameplay Developer':30},  # 30D [OK]
    'Player Dashboard - Friends': {'Gameplay Developer':30},  # 30D [OK]
    'Player Dashboard - Search & Notifications': {'Gameplay Developer':30},  # 30D [OK]
    'Player Dashboard - Marketplace': {'Gameplay Developer':30},  # 30D [OK]
    'Quest Tracking': {'Gameplay Developer':20},  # 20D [OK]
    'Social Media Support': {'Gameplay Developer':20},  # 20D [OK]
}

# === ROLE -> GATE 2 COLUMN MAPPING ===
GATE2_START = 85  # CG
ROLE_COL = {
    'Game Director': 95, 'Game Design Lead': 96, 'Game Designer': 98,
    'Combat Designer': 99, 'Level Designer': 100, 'Technical Designer': 101,
    'Writer': 102, 'Concept Artist': 104, 'UI/UX Designer': 105,
    'Character Artist': 107, 'Environment Artist': 108, 'Prop Artist': 109,
    'Animator': 110, 'Character Rigger': 111, 'VFX Artist': 112,
    'Technical Artist': 113, 'Gameplay Developer': 116, 'Sound Designer': 121,
}

# === WRITE EXCEL ===
src = r'D:\OneDrive\Claude_code\NBIAI_TEAM\Clients\Couch Heroes\CouchHeroes_Man_Day_Work_Plan_v9.xlsx'
dst = r'D:\OneDrive\Claude_code\NBIAI_TEAM\Clients\Couch Heroes\CouchHeroes_Man_Day_Work_Plan_v10.xlsx'

wb = openpyxl.load_workbook(src)
ws = wb['Plan']

# Build feature name -> row mapping
sheet_rows = {}
for row in range(5, 160):
    val = ws.cell(row=row, column=2).value
    if val:
        sheet_rows[val.strip()] = row

# Wipe ALL hours
wiped = 0
for row in range(5, 160):
    for col in range(7, 435):
        cell = ws.cell(row=row, column=col)
        if cell.value is not None and isinstance(cell.value, (int, float)):
            cell.value = None
            wiped += 1

# Fill from manually verified data
filled = 0
not_found = []
for feat_name, roles in FEATURE_HOURS.items():
    row = sheet_rows.get(feat_name)
    if not row:
        not_found.append(feat_name)
        continue
    for role, days in roles.items():
        col = ROLE_COL.get(role)
        if col and days > 0:
            ws.cell(row=row, column=col).value = days
            filled += 1

wb.save(dst)

# === VERIFICATION ===
print(f"Wiped: {wiped} cells")
print(f"Filled: {filled} cells")
if not_found:
    print(f"NOT FOUND in spreadsheet: {not_found}")

# Verify totals
print(f"\n=== VERIFICATION ===\n")
grand_total = 0
feat_count = 0
role_names = ['CEO','Creative Director','Head of Legal','COO','HR Manager','Executive Producer',
              'Producer','Game Producer','Tech Producer','Art Producer','Game Director',
              'Game Design Lead','Level Design Lead','Game Designer','Combat Designer',
              'Level Designer','Technical Designer','Writer','Art Director','Concept Artist',
              'UI/UX Designer','Graphic Designer','Character Artist','Environment Artist',
              'Prop Artist','Animator','Character Rigger','VFX Artist','Technical Artist',
              'CTO','Tech Lead','Gameplay Developer','Full Stack Developer','Backend Developer',
              'Network Engineer','QA','Sound Designer','Platform Team']

for row in range(5, 160):
    feat = ws.cell(row=row, column=2).value
    if not feat: continue
    hours = []
    row_total = 0
    for i, rn in enumerate(role_names):
        val = ws.cell(row=row, column=85+i).value
        if val and isinstance(val, (int, float)) and val > 0:
            hours.append(f"{rn}={int(val)}")
            row_total += int(val)
    if hours:
        feat_count += 1
        grand_total += row_total
        # Verify against our input
        expected = FEATURE_HOURS.get(feat.strip(), {})
        expected_total = sum(expected.values())
        match = "OK" if row_total == expected_total else f"MISMATCH ({expected_total})"
        print(f"Row {row:3d} | {feat.strip()}: {row_total}D [{match}] | {', '.join(hours)}")

print(f"\nFeatures: {feat_count}")
print(f"Grand total: {grand_total}D")
print(f"\nSaved to: {dst}")
