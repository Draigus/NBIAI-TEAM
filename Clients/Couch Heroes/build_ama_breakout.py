import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb_src = openpyxl.load_workbook('Clients/Couch Heroes/CH AMA (Responses) (1).xlsx')
ws_src = wb_src['Form Responses 1']

source_rows = []
for i, row in enumerate(ws_src.iter_rows(min_row=2, max_row=30, values_only=True)):
    if not row[1]:
        continue
    source_rows.append({
        'orig_row': i+2,
        'question_raw': str(row[1]) if row[1] else '',
        'topic': str(row[2]) if row[2] else '',
        'crit_personal': row[3],
        'crit_studio': row[6],
        'clarification': str(row[7]) if row[7] else '',
        'robin': str(row[8]) if row[8] else '',
        'vardis': str(row[9]) if row[9] else '',
        'aris': str(row[10]) if row[10] else '',
        'david': str(row[11]) if row[11] else '',
        'mustafa': str(row[12]) if row[12] else '',
        'valeria': str(row[13]) if row[13] else '',
        'glen': str(row[14]) if row[14] and len(row) > 14 else '',
    })

broken_out = []

def add_q(orig_idx, question, topic, robin='', vardis='', aris='', david='', mustafa='', valeria='', glen='', crit_p=None, crit_s=None, clarification=''):
    s = source_rows[orig_idx]
    broken_out.append({
        'orig_row': s['orig_row'],
        'question': question.strip(),
        'topic': topic if topic else s['topic'],
        'crit_personal': crit_p if crit_p is not None else s['crit_personal'],
        'crit_studio': crit_s if crit_s is not None else s['crit_studio'],
        'clarification': clarification if clarification else s['clarification'],
        'robin': robin if robin else '',
        'vardis': vardis if vardis else '',
        'aris': aris if aris else '',
        'david': david if david else '',
        'mustafa': mustafa if mustafa else '',
        'valeria': valeria if valeria else '',
        'glen': glen if glen else '',
    })

# --- ROW 2 (idx 0): Two questions ---
s = source_rows[0]
add_q(0, "What is our culture as a studio, beyond just nice people?", "Culture & Values",
    robin=s['robin'], vardis=s['vardis'], aris=s['aris'], david=s['david'], valeria=s['valeria'])
add_q(0, "Is our studio culture formally defined?", "Culture & Values",
    robin=s['robin'], vardis=s['vardis'], aris=s['aris'], david=s['david'], valeria=s['valeria'])

# --- ROW 3 (idx 1): Single ---
s = source_rows[1]
add_q(1, "What happened to the Promote from Within culture we used to have?", "Culture & Values",
    robin=s['robin'], vardis=s['vardis'], aris=s['aris'])

# --- ROW 4 (idx 2): Two distinct questions ---
s = source_rows[2]
add_q(2, "Are we aware of the scope and personnel requirements an MMO needs, and how understaffed we are?", "Team / Organisation",
    robin=s['robin'], vardis=s['vardis'], aris=s['aris'], david=s['david'], valeria=s['valeria'])
add_q(2, "Why are we prioritising leadership/management hires over execution-level personnel in key areas?", "Team / Organisation",
    robin=s['robin'], vardis=s['vardis'], aris=s['aris'], david=s['david'], valeria=s['valeria'])

# --- ROW 5 (idx 3): Four distinct questions ---
s = source_rows[3]
add_q(3, "Why do we believe there is demand for our game?", "Game Vision & USP",
    robin=s['robin'], vardis=s['vardis'], aris=s['aris'], david=s['david'], valeria=s['valeria'])
add_q(3, "Has the USP been validated?", "Game Vision & USP",
    robin=s['robin'], vardis=s['vardis'], aris=s['aris'], david=s['david'], valeria=s['valeria'])
add_q(3, "What exactly is our USP?", "Game Vision & USP",
    robin=s['robin'], vardis=s['vardis'], aris=s['aris'], david=s['david'], valeria=s['valeria'])
add_q(3, "Will our pitch work in the current market, given how other MMOs have failed?", "Game Vision & USP",
    robin=s['robin'], vardis=s['vardis'], aris=s['aris'], david=s['david'], valeria=s['valeria'])

# --- ROW 6 (idx 4): Single ---
s = source_rows[4]
add_q(4, "Are we planning a smaller/fallback version of the game if we cannot execute the full vision?", "Company Strategy",
    robin=s['robin'], vardis=s['vardis'], aris=s['aris'], valeria=s['valeria'])

# --- ROW 7 (idx 5): Single ---
s = source_rows[5]
add_q(5, "With our current trajectory, how long until v1.0 release?", "Priorities & Roadmap",
    robin=s['robin'], vardis=s['vardis'], aris=s['aris'], david=s['david'], valeria=s['valeria'])

# --- ROW 8 (idx 6): Two questions ---
s = source_rows[6]
add_q(6, "What are the core features and foundational pillars of the game itself (excluding platform, partnerships, lore)?", "Core Game Pillars",
    robin=s['robin'], vardis=s['vardis'], aris=s['aris'], clarification=s['clarification'])
add_q(6, "Which core features are comparable to existing titles, and what unique innovations are we introducing?", "Core Game Pillars",
    robin=s['robin'], vardis=s['vardis'], aris=s['aris'], clarification=s['clarification'])

# --- ROW 9 (idx 7): Single ---
s = source_rows[7]
add_q(7, "Why don't we have a dedicated commercial/sales person, and shouldn't that be a priority given we're building a partnership platform?", "Company Strategy",
    robin=s['robin'], vardis=s['vardis'], aris=s['aris'], valeria=s['valeria'])

# --- ROW 10 (idx 8): Three embedded questions ---
s = source_rows[8]
add_q(8, "What does growth (skill, responsibility, financial) look like for existing team members?", "Culture & Values",
    robin=s['robin'], vardis=s['vardis'], clarification=s['clarification'])
add_q(8, "Why are outside hires filling positions instead of promoting internally?", "Culture & Values",
    robin=s['robin'], vardis=s['vardis'], clarification=s['clarification'])
add_q(8, "If there are financial caps based on country, how does one grow? Does the cap cancel out performance reviews?", "Culture & Values",
    robin=s['robin'], vardis=s['vardis'], clarification=s['clarification'])

# --- ROW 11 (idx 9): Two questions ---
s = source_rows[9]
add_q(9, "What is our game's USP and what experience are we focusing on for players?", "Game Vision & USP",
    robin=s['robin'], valeria=s['valeria'], clarification=s['clarification'])
add_q(9, "How are all departments working towards the same USP/vision?", "Game Vision & USP",
    robin=s['robin'], valeria=s['valeria'], clarification=s['clarification'])

# --- ROW 12 (idx 10): Two questions ---
s = source_rows[10]
add_q(10, "How are investors expecting their money back and when?", "Company Strategy",
    robin=s['robin'], vardis=s['vardis'], valeria=s['valeria'])
add_q(10, "How is our game going to monetise?", "Monetisation",
    robin=s['robin'], vardis=s['vardis'], valeria=s['valeria'])

# --- ROW 13 (idx 11): Single ---
s = source_rows[11]
add_q(11, "Are we still having partnerships that define minigames or visual style inside our game?", "Core Game Pillars",
    robin=s['robin'], vardis=s['vardis'], david=s['david'], valeria=s['valeria'])

# --- ROW 14 (idx 12): Single ---
s = source_rows[12]
add_q(12, "How are we going to use AI as a tool in development or inside the game?", "Team / Organisation",
    robin=s['robin'], david=s['david'])

# --- ROW 15 (idx 13): Two questions ---
s = source_rows[13]
add_q(13, "Are we going to always be remote?", "Culture & Values",
    robin=s['robin'], vardis=s['vardis'])
add_q(13, "Are there real plans for offices and relocation in Greece/UK?", "Culture & Values",
    robin=s['robin'], vardis=s['vardis'])

# --- ROW 16 (idx 14): MARIA'S MEGA SUBMISSION ---
# Glen's comment was a meta-instruction ("we just break them down"), not an answer to any question
s = source_rows[14]
g = ''
# Game Vision block
add_q(14, "What is the one-sentence pitch of the game?", "Game Vision & USP", glen=g)
add_q(14, "What makes this game different from other games?", "Game Vision & USP", glen=g)
add_q(14, "What is the player fantasy? (Who do I get to be?)", "Core Game Pillars", glen=g)
add_q(14, "What should players remember after playing?", "Core Game Pillars", glen=g)
add_q(14, "What is the game really about beneath the surface?", "Core Game Pillars", glen=g)
add_q(14, "What themes are being explored by the player?", "Core Game Pillars", glen=g)
add_q(14, "What should the player feel moment-to-moment?", "Core Game Pillars", glen=g)
add_q(14, "What is the central conflict in the game? Who are our bad guys / what opposes the player?", "Core Game Pillars", glen=g)
add_q(14, "What is core story vs extra story? What is the main story all about?", "Core Game Pillars", glen=g)
add_q(14, "What are our game pillars?", "Core Game Pillars", glen=g)
# Platform/partnerships
add_q(14, "What is the game's platform and how do partnerships work?", "Partnerships & Platform", glen=g)
add_q(14, "How will partnerships work inside of our game world?", "Partnerships & Platform", glen=g)
# Culture
add_q(14, "What behaviours get rewarded at Couch Heroes?", "Culture & Values", glen=g)
add_q(14, "What behaviours are not tolerated at Couch Heroes?", "Culture & Values", glen=g)
add_q(14, "What behaviours are we unintentionally rewarding?", "Culture & Values", glen=g)
add_q(14, "Where do people feel unsafe speaking up?", "Culture & Values", glen=g)
add_q(14, "Are we hiring for culture fit or culture growth?", "Culture & Values", glen=g)
add_q(14, "How do people give feedback to each other?", "Culture & Values", glen=g)
# Accountability & Leadership
add_q(14, "Where is ownership unclear?", "Team / Organisation", glen=g)
add_q(14, "What decisions are delayed because no one wants to take responsibility?", "Team / Organisation", glen=g)
add_q(14, "How do leaders handle being wrong?", "Leadership & Accountability", glen=g)
add_q(14, "Are we holding high performers accountable if they hurt the team?", "Leadership & Accountability", glen=g)
add_q(14, "Is it safe to disagree with leadership?", "Leadership & Accountability", glen=g)
add_q(14, "How are mistakes handled - punished or learned from?", "Leadership & Accountability", glen=g)
# Communication & Transparency
add_q(14, "Are employees informed about changes early or late?", "Communication & Transparency", glen=g)
add_q(14, "What information is typically withheld from the team and why?", "Communication & Transparency", glen=g)
add_q(14, "Are people held accountable regardless of role?", "Leadership & Accountability", glen=g)
add_q(14, "What are the plans for how to inform everyone about changes going forward?", "Communication & Transparency", glen=g)
# Competitors
add_q(14, "Who are our competitors?", "Game Vision & USP", glen=g)
# Strategy
add_q(14, "What are we doing today that will not work in 2-3 years? Are we solving the right problems?", "Company Strategy", glen=g)
add_q(14, "Are we optimising for speed or quality, and is that the right choice?", "Company Strategy", glen=g)
add_q(14, "Where does communication break down between teams?", "Team / Organisation", glen=g)
add_q(14, "What benefits does Couch Heroes offer its contractors?", "Culture & Values", glen=g)
# Player/game quality
add_q(14, "Are we building what players actually want, or what we assume they want?", "Core Game Pillars", glen=g)
add_q(14, "Where does gameplay and narrative feel disconnected?", "Core Game Pillars", glen=g)
add_q(14, "What part of the game are we avoiding because it is hard to fix?", "Core Game Pillars", glen=g)
# Candour
add_q(14, "What is something employees complain about that leadership disagrees with?", "Leadership & Accountability", glen=g)
add_q(14, "What would we do differently if we started Couch Heroes today?", "Company Strategy", glen=g)
add_q(14, "What is the truth we are avoiding about the game, platform, or team?", "Company Strategy", glen=g)
add_q(14, "If we were brutally honest, what is not working and how can we solve it?", "Company Strategy", glen=g)
add_q(14, "What would an outsider immediately criticise about Couch Heroes?", "Company Strategy", glen=g)

# --- ROW 17 (idx 15): Single ---
s = source_rows[15]
add_q(15, "How do we plan to maintain team morale and clarity when direction changes happen?", "Culture & Values",
    robin=s['robin'], vardis=s['vardis'], valeria=s['valeria'])

# --- ROW 18 (idx 16): Single ---
s = source_rows[16]
add_q(16, "What does success look like for the team internally, not just for investors?", "Culture & Values",
    robin=s['robin'], vardis=s['vardis'], david=s['david'])

# --- ROW 19 (idx 17): Single ---
s = source_rows[17]
add_q(17, "How do we plan to avoid previous issues with unclear direction and rework cycles?", "Priorities & Roadmap",
    robin=s['robin'], vardis=s['vardis'], david=s['david'])

# --- ROW 20 (idx 18): Single ---
s = source_rows[18]
add_q(18, "How do we ensure better cross-discipline communication moving forward?", "Team / Organisation",
    robin=s['robin'], vardis=s['vardis'], david=s['david'], valeria=s['valeria'])

# --- ROW 21 (idx 19): Single ---
s = source_rows[19]
add_q(19, "What are the biggest risks for the project right now, and how are we addressing them?", "Company Strategy",
    robin=s['robin'], vardis=s['vardis'])

# --- ROW 22 (idx 20): Two parts ---
s = source_rows[20]
add_q(20, "How are we ensuring long-term continuity - reusing/building on previous work instead of reworking core areas?", "Priorities & Roadmap",
    robin=s['robin'], vardis=s['vardis'])
add_q(20, "What processes are we putting in place to ensure content built now lasts and contributes long-term?", "Priorities & Roadmap",
    robin=s['robin'], vardis=s['vardis'])

# --- ROW 23 (idx 21): Single ---
s = source_rows[21]
add_q(21, "When will there be Marketing and Commercial hires at Couch Heroes?", "Company Strategy",
    robin=s['robin'], vardis=s['vardis'], valeria=s['valeria'])

# --- ROW 24 (idx 22): Single ---
s = source_rows[22]
add_q(22, "What is our HR escalation process? Who do employees reach out to for next-level review?", "Team / Organisation",
    robin=s['robin'], vardis=s['vardis'], valeria=s['valeria'])

# --- ROW 25 (idx 23): Two numbered ---
s = source_rows[23]
add_q(23, "How do we make a world that feels truly alive and makes players want to live there?", "Core Game Pillars",
    robin=s['robin'], vardis=s['vardis'])
add_q(23, "How do we make character growth feel constantly rewarding from hour 1 to month 12?", "Core Game Pillars",
    robin=s['robin'], vardis=s['vardis'])

# --- ROW 26 (idx 24): Three embedded ---
s = source_rows[24]
add_q(24, "What specific gameplay activities engage a new player entering the game world?", "Core Game Pillars",
    robin=s['robin'], vardis=s['vardis'])
add_q(24, "What is the intended wow moment early in the player experience?", "Core Game Pillars",
    robin=s['robin'], vardis=s['vardis'])
add_q(24, "How do early gameplay activities evolve over time - do they scale with playtime or transition into different systems?", "Core Game Pillars",
    robin=s['robin'], vardis=s['vardis'])

# --- ROW 27 (idx 25): Single (overlaps monetisation from row 12) ---
s = source_rows[25]
add_q(25, "What will be our monetisation system and how will it work?", "Monetisation",
    robin=s['robin'], valeria=s['valeria'])

# --- ROW 28 (idx 26): Three embedded ---
s = source_rows[26]
add_q(26, "Is the MMORPG genre still viable - will the game be recognised by the market at launch?", "Company Strategy",
    robin=s['robin'])
add_q(26, "How does the company plan to steer development to increase chances of success?", "Company Strategy",
    robin=s['robin'])
add_q(26, "What will the company do to make the game feel fun and appealing? What if even we do not find it fun?", "Core Game Pillars",
    robin=s['robin'])

# --- ROW 29 (idx 27): Two ---
s = source_rows[27]
add_q(27, "What are the non-negotiable pillars of the game and how do they shape feature/priority decisions?", "Core Game Pillars",
    robin=s['robin'])
add_q(27, "After the game is built, what is the next step for Couch Heroes as a studio?", "Company Strategy",
    robin=s['robin'])

# --- ROW 30 (idx 28): Single ---
s = source_rows[28]
add_q(28, "What is the target player base number in the next 2-3 years?", "Game Vision & USP")


# ============================================================
# Build the output Excel
# ============================================================
wb_out = openpyxl.Workbook()
ws = wb_out.active
ws.title = "AMA Questions Broken Out"

header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
no_answer_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
has_answer_fill = PatternFill(start_color="CCE5CC", end_color="CCE5CC", fill_type="solid")
body_font = Font(name="Calibri", size=10, color="000000")
wrap_align = Alignment(wrap_text=True, vertical="top")
thin_border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)

headers = [
    "Q#", "Topic", "Individual Question", "Orig Row",
    "Crit (Personal)", "Crit (Studio)",
    "Robin", "Vardis", "Aris", "David", "Mustafa", "Valeria", "Glen",
    "Has Direct Answer?", "Clarification / Notes"
]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = wrap_align
    cell.border = thin_border

topic_order = [
    "Game Vision & USP", "Core Game Pillars", "Partnerships & Platform",
    "Company Strategy", "Monetisation", "Priorities & Roadmap",
    "Team / Organisation", "Culture & Values",
    "Leadership & Accountability", "Communication & Transparency"
]

def topic_sort_key(q):
    t = q['topic']
    if t in topic_order:
        return topic_order.index(t)
    return 99

broken_out.sort(key=topic_sort_key)

respondents = ['robin', 'vardis', 'aris', 'david', 'mustafa', 'valeria', 'glen']
for i, q in enumerate(broken_out):
    row_num = i + 2
    answers = [q[r] for r in respondents if q[r] and q[r] not in ['None', '[NONE]', '']]
    has_answer = "YES" if answers else "NO"

    values = [
        i + 1,
        q['topic'],
        q['question'],
        q['orig_row'],
        q['crit_personal'],
        q['crit_studio'],
        q['robin'] if q['robin'] and q['robin'] not in ['None', '[NONE]'] else '',
        q['vardis'] if q['vardis'] and q['vardis'] not in ['None', '[NONE]'] else '',
        q['aris'] if q['aris'] and q['aris'] not in ['None', '[NONE]'] else '',
        q['david'] if q['david'] and q['david'] not in ['None', '[NONE]'] else '',
        q['mustafa'] if q['mustafa'] and q['mustafa'] not in ['None', '[NONE]'] else '',
        q['valeria'] if q['valeria'] and q['valeria'] not in ['None', '[NONE]'] else '',
        q['glen'] if q['glen'] and q['glen'] not in ['None', '[NONE]'] else '',
        has_answer,
        q['clarification'] if q['clarification'] and q['clarification'] != 'None' else '',
    ]

    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.font = body_font
        cell.alignment = wrap_align
        cell.border = thin_border

    answer_cell = ws.cell(row=row_num, column=14)
    if has_answer == "NO":
        answer_cell.fill = no_answer_fill
    else:
        answer_cell.fill = has_answer_fill

col_widths = [5, 28, 60, 8, 10, 10, 50, 50, 50, 50, 30, 50, 50, 14, 45]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.row_dimensions[1].height = 30
for r in range(2, len(broken_out) + 2):
    ws.row_dimensions[r].height = 80

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:O{len(broken_out)+1}"

# ============================================================
# Summary sheet
# ============================================================
ws2 = wb_out.create_sheet("Topic Summary")
for col, h in enumerate(["Topic", "Total Questions", "Answered", "Unanswered"], 1):
    c = ws2.cell(row=1, column=col, value=h)
    c.font = header_font
    c.fill = header_fill

topic_counts = {}
for q in broken_out:
    t = q['topic']
    if t not in topic_counts:
        topic_counts[t] = {'total': 0, 'answered': 0}
    topic_counts[t]['total'] += 1
    answers = [q[r] for r in respondents if q[r] and q[r] not in ['None', '[NONE]', '']]
    if answers:
        topic_counts[t]['answered'] += 1

r = 2
for t in topic_order:
    if t in topic_counts:
        ws2.cell(row=r, column=1, value=t)
        ws2.cell(row=r, column=2, value=topic_counts[t]['total'])
        ws2.cell(row=r, column=3, value=topic_counts[t]['answered'])
        ws2.cell(row=r, column=4, value=topic_counts[t]['total'] - topic_counts[t]['answered'])
        r += 1

ws2.column_dimensions['A'].width = 30
ws2.column_dimensions['B'].width = 16
ws2.column_dimensions['C'].width = 12
ws2.column_dimensions['D'].width = 14

output_path = 'Clients/Couch Heroes/CH_AMA_Questions_Broken_Out.xlsx'
wb_out.save(output_path)
print(f"Created: {output_path}")
print(f"Total individual questions: {len(broken_out)}")
print(f"\nTopic breakdown:")
for t in topic_order:
    if t in topic_counts:
        c = topic_counts[t]
        print(f"  {t}: {c['total']} questions ({c['answered']} answered, {c['total']-c['answered']} unanswered)")
