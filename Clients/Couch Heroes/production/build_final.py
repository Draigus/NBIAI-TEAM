"""
CouchHeroes v12 — Final build.
Takes v11 template, modifies Early Prod to Min/Max per role,
writes ALL epics/features/stories/tasks with proper hierarchy formatting.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os, re

BASE = r"D:\OneDrive\Claude_code\NBIAI_TEAM\Clients\Couch Heroes\production"
TEMPLATE = os.path.join(BASE, "CouchHeroes_Man_Day_Work_Plan_v11.xlsx")
OUTPUT = os.path.join(BASE, "CouchHeroes_v12.xlsx")
DESIGN_FILE = os.path.join(BASE, "VS Estimation Design - General Sheet.xlsx")
ENG_FILE = os.path.join(BASE, "VS Estimation Engineering - General Sheet (1).xlsx")
HOURS_FILE = os.path.join(BASE, "hours_mapping.txt")

# ================================================================
# CONSTANTS
# ================================================================

ROLE_NAMES = [
    "CEO","Creative Director","Head of Legal","COO","HR Manager",
    "Executive Producer","Producer","Game Producer","Tech Producer","Art Producer",
    "Game Director","Game Design Lead","Level Design Lead","Game Designer",
    "Combat Designer","Level Designer","Technical Designer","Writer",
    "Art Director","Concept Artist","UI/UX Designer","Graphic Designer",
    "Character Artist","Environment Artist","Prop Artist","Animator",
    "Character Rigger","VFX Artist","Technical Artist","CTO","Tech Lead",
    "Gameplay Developer","Full Stack Developer","Backend Developer",
    "Network Engineer","QA","Sound Designer","Platform Team",
]
ROLE_IDX = {r: i for i, r in enumerate(ROLE_NAMES)}
ROLES_PER_PHASE = len(ROLE_NAMES)  # 38

TEAM_TO_ROLE = {
    "Game design": "Game Designer", "Level design": "Level Designer",
    "Gameplay Engineering": "Gameplay Developer", "Backend Engineering": "Backend Developer",
    "UX/UI": "UI/UX Designer", "Concept Art": "Concept Artist",
    "Environment Art": "Environment Artist", "Character Art": "Character Artist",
    "Animation & Rigging": "Animator", "VFX": "VFX Artist",
    "TechArt": "Technical Artist", "Audio": "Sound Designer",
    "QA": "QA", "Narrative": "Writer",
}
TEAM_TO_DEPT = {
    "Game design": "Design", "Level design": "Design", "Narrative": "Design",
    "UX/UI": "Art", "Concept Art": "Art", "Environment Art": "Art",
    "Character Art": "Art", "Animation & Rigging": "Art", "VFX": "Art", "TechArt": "Art",
    "Gameplay Engineering": "Engineering", "Backend Engineering": "Engineering",
    "QA": "QA", "Audio": "Audio",
}
EPIC_NORM = {
    "Combat System": "Combat", "Social": "Social & Multiplayer",
    "Rename to in-game camera squancer during dialog": "World Systems",
    "Phasing system": "World Systems", "Live Service Support": "Live Game",
}

# Phase definitions: (super_group, gate_label, sub_label)
PHASES = [
    ("CONCEPT",        "GATE\n0",  "CONCEPT\n~2 months\nTarget: T0 Ideation"),
    ("PRE-PRODUCTION", "GATE\n1",  "PRE-PRODUCTION\nTarget: T1-T2 R&D / GDD-TDD-Brief"),
    ("PRODUCTION",     "GATE\n2",  "EARLY PROD\nTarget: T3 Prototype"),
    ("PRODUCTION",     "GATE\n3",  "MID PRO\nTarget: T4 MVP"),
    ("PRODUCTION",     "GATE\n4",  "END PROD\nTarget: T5 Feature Complete"),
    ("RELEASE",        "GATE\n5",  "ALPHA\nTarget: T6 Player Tested"),
    ("RELEASE",        "GATE\n6",  "BETA\nTarget: T6 Player Tested"),
    ("RELEASE",        "GATE\n7",  "LAUNCH\nTarget: T6 Player Ready"),
    ("RELEASE",        "GATE\n8",  "FIRST RELEASE\n~1.5 months\nTarget: T6-T7"),
    ("LIVE SERVICE",   "GATE\n9",  "LS 1\nLive\nTarget: T7 Expand"),
    ("LIVE SERVICE",   "GATE\n10", "LS 2\nLive\nTarget: T8 Scale"),
]

# Phase fills
PHASE_FILLS = {
    "CONCEPT": PatternFill(start_color="FFF2CC", fill_type="solid"),
    "PRE-PRODUCTION": PatternFill(start_color="D9E2F3", fill_type="solid"),
    "PRODUCTION": PatternFill(start_color="E2EFDA", fill_type="solid"),
    "RELEASE": PatternFill(start_color="FCE4D6", fill_type="solid"),
    "LIVE SERVICE": PatternFill(start_color="D6E4F0", fill_type="solid"),
}

# Styles
DARK = PatternFill(start_color="404040", fill_type="solid")
YELLOW = PatternFill(start_color="FFD966", fill_type="solid")
GATE_FILL = PatternFill(start_color="FFD966", fill_type="solid")
EP_MIN_FILL = PatternFill(start_color="C6EFCE", fill_type="solid")
EP_MAX_FILL = PatternFill(start_color="FFC7CE", fill_type="solid")
EPIC_FILL = PatternFill(start_color="FFD966", fill_type="solid")
STORY_FILL = PatternFill(start_color="659DF2", fill_type="solid")
TASK_FILL = PatternFill(start_color="DEDAFF", fill_type="solid")
THIN = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
WRAP = Alignment(wrap_text=True, vertical="top")
ROLE_ALIGN = Alignment(wrap_text=True, text_rotation=90, horizontal="center", vertical="bottom")
WHITE_BOLD = Font(bold=True, color="FFFFFF")
WHITE_SM = Font(bold=True, size=8, color="FFFFFF")

# ================================================================
# DATA LOADING
# ================================================================

def clean(v):
    if v is None: return None
    s = str(v).strip()
    return s if s else None

def load_miro():
    bids = {}
    with open(HOURS_FILE, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            m = re.match(r"^(.+?):\s*(\d+)D\s+total\s*\|\s*(.*)$", line)
            if m:
                roles = {}
                for p in m.group(3).split(","):
                    rm = re.match(r"\s*(.+?)=(\d+)D\s*$", p.strip())
                    if rm: roles[rm.group(1).strip()] = int(rm.group(2))
                bids[m.group(1).strip()] = {"total": int(m.group(2)), "roles": roles}
        return bids

def load_sheet(path, sheet, is_eng=False):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    rows = []
    cur_e = cur_f = cur_s = ""
    for r in range(3, ws.max_row + 1):
        epic = clean(ws.cell(r,1).value)
        feat = clean(ws.cell(r,2).value)
        story = clean(ws.cell(r,3).value)
        task = clean(ws.cell(r,4).value)
        team = clean(ws.cell(r,5).value)
        desc = clean(ws.cell(r,6).value)
        dos = clean(ws.cell(r,7).value)
        if epic: cur_e = EPIC_NORM.get(epic, epic); cur_s = ""
        if feat: cur_f = feat; cur_s = ""
        if story: cur_s = story
        if not feat and not story and not task: continue
        eff_story = story if story else (cur_s if task else None)

        # Estimator values
        est = [ws.cell(r,c).value for c in range(8, 26)]  # 9 pairs = 18 vals
        if is_eng:
            adj_min = ws.cell(r,26).value
            adj_max = ws.cell(r,27).value
            notes = clean(ws.cell(r,28).value)
        else:
            final = ws.cell(r,26).value
            adj_min = adj_max = notes = None

        rows.append({
            "epic": cur_e, "feature": cur_f,
            "is_feat_hdr": bool(feat and not story and not task),
            "story": eff_story, "task": task,
            "team": team, "dept": TEAM_TO_DEPT.get(team,"") if team else "",
            "desc": desc, "dos": dos,
            "src": "Eng" if is_eng else "Des",
            "est": est,
            "final": None if is_eng else final,
            "adj_min": adj_min if is_eng else None,
            "adj_max": adj_max if is_eng else None,
            "notes": notes,
        })
    wb.close()
    return rows

def merge_data(des, eng):
    ek = {}
    for r in eng:
        k = (r["epic"],r["feature"],r["story"] or "",r["task"] or "",r["team"] or "")
        ek[k] = r
    out = []; seen = set()
    for r in des:
        k = (r["epic"],r["feature"],r["story"] or "",r["task"] or "",r["team"] or "")
        if k in ek:
            r["eng"] = ek[k]; r["src"] = "Both"; seen.add(k)
        else:
            r["eng"] = None
        out.append(r)
    for r in eng:
        k = (r["epic"],r["feature"],r["story"] or "",r["task"] or "",r["team"] or "")
        if k not in seen: r["eng"] = None; out.append(r)
    return out

def get_template_order():
    """Read exact epic and feature order from template."""
    wb = openpyxl.load_workbook(TEMPLATE, data_only=True)
    ws = wb["Plan"]
    epic_order = []
    feat_order = {}
    ce = ""
    fi = 0
    for r in range(5, 166):
        e = ws.cell(r, 1).value
        f = ws.cell(r, 2).value
        if e:
            ce = str(e).strip()
            if ce not in epic_order:
                epic_order.append(ce)
        if f:
            key = f"{ce}::{str(f).strip()}"
            feat_order[key] = fi
            fi += 1
    wb.close()
    epic_rank = {e: i for i, e in enumerate(epic_order)}
    return epic_rank, feat_order

EPIC_RANK = None
FEAT_RANK = None

def sort_data(rows):
    global EPIC_RANK, FEAT_RANK
    if EPIC_RANK is None:
        EPIC_RANK, FEAT_RANK = get_template_order()

    def sort_key(r):
        er = EPIC_RANK.get(r["epic"], 999)
        key = f"{r['epic']}::{r['feature']}"
        fr = FEAT_RANK.get(key, 9999)
        return (er, fr, 0 if r["is_feat_hdr"] else 1, r["story"] or "", r["task"] or "")

    return sorted(rows, key=sort_key)

MIRO_MAP = {
    "In-game Tutorial":"FTUE (tutorial cave segment)", "Melee Combat Framework":"Combat framework",
    "Combat abilities":"Combat Abilities", "Basic Magic system":"Magic",
    "Ranged (Corruption Gun)":"Ranged combat", "PVP Combat":"PvP combat",
    "Guilds (S)":"Guilds System", "Co-op - Party":"Co-op/Party",
    "In-game Mail":"In-game mail", "Skill System (1 skill tree active)":"Skill System",
    "Anywhere door placement":"Anywhere Door", "User Space Decoration":"User space Decoration",
    "Pets (mechanical parrot reuse)":"Mounts/ Pets", "Auction House economy":"Auction House",
    "In-game economy":"In game Economy", "Dungeons mechanics":"Dungeon Mechanics",
    "Various Quest Types Support":"Quests : Courier, Escort, Kill, Collect, research, time obj, Partner Quest, corruption",
    "Quest Tools":"Quests Backend/Tooling", "Quest Rewards System":"Quest reward system",
    "Bugged Dungeon Integration":"The Dungeon", "In game RMT Store":"In-game Store",
    "Chat system":"Game Chat System", "Dungeon finder":"LFG / Pick up groups",
    "Party system":"Co-op/Party", "Leaderboard system":"CombatBalance & Telemetry",
    "Combat component":"Combat framework", "Combat director":"Combat framework",
    "Subscription page on the website":"RMT game Economy",
}

def match_miro(feat, bids):
    if not feat: return None
    if feat in MIRO_MAP and MIRO_MAP[feat] in bids: return bids[MIRO_MAP[feat]]
    if feat in bids: return bids[feat]
    lo = {k.lower():v for k,v in bids.items()}
    if feat.lower() in lo: return lo[feat.lower()]
    for mk,mv in bids.items():
        if feat.lower() in mk.lower() or mk.lower() in feat.lower(): return mv
    return None

# ================================================================
# BUILD WORKBOOK
# ================================================================

def build():
    print("Loading data...")
    miro = load_miro()
    des = load_sheet(DESIGN_FILE, "Sorted by Epic", False)
    eng = load_sheet(ENG_FILE, "Sheet1", True)
    merged = merge_data(des, eng)
    rows = sort_data(merged)
    print(f"  {len(rows)} rows from {len(des)} design + {len(eng)} engineering")

    # Template metadata
    wb_t = openpyxl.load_workbook(TEMPLATE, data_only=True)
    ws_t = wb_t["Plan"]
    tmeta = {}; ce=""
    for r in range(5,166):
        e=ws_t.cell(r,1).value; f=ws_t.cell(r,2).value
        if e: ce=str(e).strip()
        if f: tmeta[f"{ce}::{str(f).strip()}"]={"size":ws_t.cell(r,3).value,"tier":ws_t.cell(r,4).value}
    wb_t.close()

    # Copy template, then copy headers from its Plan sheet cell-by-cell
    wb_src = openpyxl.load_workbook(TEMPLATE)
    ws_src = wb_src["Plan"]

    # Collect header data (rows 1-4) and merged cells before deleting
    from copy import copy as ccopy
    hdr_data = {}
    for r in range(1, 5):
        for c in range(1, ws_src.max_column + 1):
            cell = ws_src.cell(r, c)
            hdr_data[(r, c)] = {
                "val": cell.value,
                "font": ccopy(cell.font),
                "fill": ccopy(cell.fill),
                "align": ccopy(cell.alignment),
                "border": ccopy(cell.border),
                "nf": cell.number_format,
            }
    hdr_merges = [str(m) for m in ws_src.merged_cells.ranges if m.min_row <= 4]
    src_max_col = ws_src.max_column

    # Delete Plan, create fresh
    del wb_src["Plan"]
    ws = wb_src.create_sheet("Plan", 0)

    # Paste headers exactly
    for (r, c), d in hdr_data.items():
        cell = ws.cell(r, c)
        cell.value = d["val"]
        cell.font = d["font"]
        cell.fill = d["fill"]
        cell.alignment = d["align"]
        cell.border = d["border"]
        cell.number_format = d["nf"]
    for m in hdr_merges:
        try:
            ws.merge_cells(m)
        except Exception:
            pass

    # The template's column structure is PRESERVED exactly.
    # Early Prod roles: cols 85-122 (38 single cols).
    # We ADD 38 new cols for Max AFTER col 122 (shifting everything right).
    # Actually — to avoid shifting, we add Min/Max sections AFTER the last phase col (434).

    META = 5
    EP_START = 85   # Early Prod role start (template col)
    EP_END = 122    # Early Prod role end (template col)
    TOTAL_COLS = src_max_col  # 434

    # Add EP Min and EP Max sections after col 434
    MIN_START = TOTAL_COLS + 2  # col 436
    MAX_START = MIN_START + ROLES_PER_PHASE + 1  # col 475 (38 + gap)

    min_fill = PatternFill(start_color="C6EFCE", fill_type="solid")
    max_fill = PatternFill(start_color="FFC7CE", fill_type="solid")

    ws.cell(2, MIN_START, "EARLY PROD — MIN").font = Font(bold=True, size=10, color="FFFFFF")
    ws.cell(2, MIN_START).fill = DARK
    ws.cell(2, MAX_START, "EARLY PROD — MAX").font = Font(bold=True, size=10, color="FFFFFF")
    ws.cell(2, MAX_START).fill = DARK

    for ri, rname in enumerate(ROLE_NAMES):
        c_min = MIN_START + ri
        c_max = MAX_START + ri
        ws.cell(4, c_min, rname).font = Font(bold=True, size=8)
        ws.cell(4, c_min).fill = min_fill; ws.cell(4, c_min).alignment = ROLE_ALIGN
        ws.cell(4, c_max, rname).font = Font(bold=True, size=8)
        ws.cell(4, c_max).fill = max_fill; ws.cell(4, c_max).alignment = ROLE_ALIGN
        ws.column_dimensions[get_column_letter(c_min)].width = 5
        ws.column_dimensions[get_column_letter(c_max)].width = 5

    # Copy column widths from template
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 6
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 22
    for c in range(6, TOTAL_COLS + 1):
        ws.column_dimensions[get_column_letter(c)].width = 5

    ws.row_dimensions[3].height = 60
    ws.row_dimensions[4].height = 110
    ws.freeze_panes = "F5"

    wb = wb_src  # use this workbook going forward

    # ================================================================
    # DATA ROWS
    # ================================================================
    print("Writing data rows...")
    dr = 5
    prev_e = prev_f = ""
    stats = {"epics":0, "feats":0, "stories":0, "tasks":0, "est_cells":0}

    for row in rows:
        epic = row["epic"]; feat = row["feature"]
        story = row["story"]; task = row["task"]

        # ---- EPIC ROW ----
        if epic != prev_e:
            end_col = MAX_START + ROLES_PER_PHASE - 1
            ws.cell(dr, 1, epic).font = Font(bold=True, size=11)
            for c in range(1, end_col + 1):
                ws.cell(dr, c).fill = EPIC_FILL
            ws.merge_cells(start_row=dr, start_column=1, end_row=dr, end_column=end_col)
            prev_e = epic; prev_f = ""
            stats["epics"] += 1
            dr += 1

        # ---- FEATURE ROW ----
        if feat != prev_f:
            ws.cell(dr, 2, feat).font = Font(size=10)
            mk = f"{epic}::{feat}"
            if mk in tmeta:
                ws.cell(dr, 3, tmeta[mk]["size"])
                ws.cell(dr, 4, tmeta[mk]["tier"])

            # Miro bid -> EP role columns (agreed totals)
            mb = match_miro(feat, miro)
            if mb:
                for rn, days in mb["roles"].items():
                    if rn in ROLE_IDX:
                        ws.cell(dr, EP_START + ROLE_IDX[rn], days)
                        ws.cell(dr, MIN_START + ROLE_IDX[rn], days)
                        ws.cell(dr, MAX_START + ROLE_IDX[rn], days)
                        stats["est_cells"] += 3

            stats["feats"] += 1
            prev_f = feat

            if row["is_feat_hdr"]:
                write_est(ws, dr, row, EP_START, MIN_START, MAX_START, stats)
                if row["team"]:
                    ws.cell(dr, 5, f"{row['dept']} / {row['team']}")
                dr += 1
                continue

        # ---- STORY ROW (plain text, indented, no fill) ----
        if story and not task:
            ws.cell(dr, 2, f"  {story}").font = Font(bold=True, size=9)
            stats["stories"] += 1

        # ---- TASK ROW (plain text, further indented, no fill) ----
        elif task:
            ws.cell(dr, 2, f"    {task}").font = Font(size=9)
            stats["tasks"] += 1

        # Notes: dept/team
        if row["team"]:
            ws.cell(dr, 5, f"{row['dept']} / {row['team']}")

        # Estimates
        write_est(ws, dr, row, EP_START, MIN_START, MAX_START, stats)

        dr += 1

    # ---- SUMMARY ----
    print(f"\n  Epics: {stats['epics']}")
    print(f"  Features: {stats['feats']}")
    print(f"  Stories: {stats['stories']}")
    print(f"  Tasks: {stats['tasks']}")
    print(f"  Estimate cells: {stats['est_cells']}")
    print(f"  Total rows: {dr - 5}")

    # ---- ROW HEIGHT for data rows ----
    for r in range(5, dr):
        ws.row_dimensions[r].height = 18

    print(f"\nSaving {OUTPUT}...")
    wb.save(OUTPUT)
    wb.close()
    print("Done!")


def write_est(ws, dr, row, ep_start, min_start, max_start, stats):
    """Write estimates: agreed value in EP cols, min/max in separate sections."""
    team = row["team"]
    if not team or team not in TEAM_TO_ROLE:
        return
    role = TEAM_TO_ROLE[team]
    if role not in ROLE_IDX:
        return
    ri = ROLE_IDX[role]
    col_ep = ep_start + ri
    col_min = min_start + ri
    col_max = max_start + ri

    def num(v):
        return v is not None and not isinstance(v, str)

    # Design: agreed = Final, min = lowest estimator min, max = highest estimator max
    if row["src"] in ("Des", "Both") and row["est"]:
        est = row["est"]
        mins = [float(est[i]) for i in range(0, len(est), 2) if num(est[i])]
        maxs = [float(est[i]) for i in range(1, len(est), 2) if num(est[i])]
        if mins:
            ws.cell(dr, col_min, round(min(mins), 1))
        if maxs:
            ws.cell(dr, col_max, round(max(maxs), 1))
        if num(row["final"]):
            ws.cell(dr, col_ep, row["final"])
            stats["est_cells"] += 1

    # Engineering: agreed = Mustafa max, min/max from Mustafa adjusted
    if row["src"] in ("Eng", "Both"):
        if num(row["adj_min"]):
            ws.cell(dr, col_min, row["adj_min"])
        if num(row["adj_max"]):
            ws.cell(dr, col_max, row["adj_max"])
            ws.cell(dr, col_ep, row["adj_max"])
            stats["est_cells"] += 1

    # Merged row: engineering data overrides
    if row.get("eng"):
        er = row["eng"]
        if num(er["adj_min"]):
            ws.cell(dr, col_min, er["adj_min"])
        if num(er["adj_max"]):
            ws.cell(dr, col_max, er["adj_max"])
            ws.cell(dr, col_ep, er["adj_max"])


if __name__ == "__main__":
    build()
