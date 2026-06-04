import openpyxl

wb = openpyxl.load_workbook('VS Estimation Design - General Sheet.xlsx', data_only=True)
ws = wb['Sorted by Epic']

design_teams = {'Game design', 'Level design', 'Narrative'}
non_design = {}

for r in range(3, ws.max_row + 1):
    e = ws.cell(row=r, column=5).value
    c = ws.cell(row=r, column=3).value
    d = ws.cell(row=r, column=4).value
    b = ws.cell(row=r, column=2).value
    if not e:
        continue
    team = str(e).strip()
    if team in design_teams:
        continue
    name = str(d or c or b or '?')[:50]
    has_est = any(
        isinstance(ws.cell(row=r, column=ci).value, (int, float))
        and ws.cell(row=r, column=ci).value > 0
        for ci in range(8, 27)
    )
    if team not in non_design:
        non_design[team] = {'total': 0, 'with_est': 0, 'examples': []}
    non_design[team]['total'] += 1
    if has_est:
        non_design[team]['with_est'] += 1
        if len(non_design[team]['examples']) < 3:
            non_design[team]['examples'].append(f"Row {r}: {name}")

print("NON-DESIGN TEAM ROWS IN THE DESIGN SHEET")
print("=" * 60)
total_non = 0
total_with = 0
for team, info in sorted(non_design.items()):
    total_non += info['total']
    total_with += info['with_est']
    with_count = info['with_est']
    print(f"  {team:25s}: {info['total']:3d} rows, {with_count} with estimates")
    for ex in info['examples']:
        print(f"    -> {ex}")

print(f"\nTotal non-design rows: {total_non}")
print(f"With estimates: {total_with}")
print(f"Blank (potential duplicates): {total_non - total_with}")
wb.close()
