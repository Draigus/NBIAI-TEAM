import openpyxl
wb = openpyxl.load_workbook('CouchHeroes_Man_Day_Work_Plan_v15.xlsx', data_only=True)
ws = wb['Plan']

total_story_rows = 0
with_data = 0
blank = 0
blank_by_team = {}

for r in range(5, ws.max_row + 1):
    c = ws.cell(row=r, column=3).value  # Story column
    if not c:
        continue
    total_story_rows += 1
    team = ws.cell(row=r, column=4).value or 'Unknown'

    has_ep = False
    for col in range(87, 163):
        v = ws.cell(row=r, column=col).value
        if v is not None:
            has_ep = True
            break

    if has_ep:
        with_data += 1
    else:
        blank += 1
        dept = str(team).split(' | ')[0] if ' | ' in str(team) else str(team)
        blank_by_team[dept] = blank_by_team.get(dept, 0) + 1

print(f"Total story rows: {total_story_rows}")
print(f"With EP data: {with_data}")
print(f"Blank (no EP data): {blank}")
print(f"Fill rate: {with_data/total_story_rows*100:.1f}%")
print(f"\nBlanks by department:")
for dept, count in sorted(blank_by_team.items(), key=lambda x: -x[1]):
    print(f"  {dept:20s}: {count}")
wb.close()
