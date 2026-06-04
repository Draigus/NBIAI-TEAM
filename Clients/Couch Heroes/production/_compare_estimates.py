import openpyxl

wb = openpyxl.load_workbook('VS Estimation Engineering - General Sheet (1).xlsx', data_only=True)
ws = wb['Sheet1']

mustafa_smaller = 0
engineers_smaller = 0
equal = 0
total_compared = 0
mustafa_total_min = 0
mustafa_total_max = 0
eng_total_min = 0
eng_total_max = 0

for r in range(3, ws.max_row + 1):
    c = ws.cell(row=r, column=3).value
    e = ws.cell(row=r, column=5).value
    if not c or not e:
        continue

    # Mustafa Adjusted (cols Z=26, AA=27)
    m_min = ws.cell(row=r, column=26).value
    m_max = ws.cell(row=r, column=27).value

    # Individual engineers (cols H-Y = 8-25, paired min/max)
    ind_mins = []
    ind_maxes = []
    for ci in range(8, 24, 2):
        mn = ws.cell(row=r, column=ci).value
        mx = ws.cell(row=r, column=ci + 1).value
        if isinstance(mn, (int, float)) and mn > 0:
            ind_mins.append(mn)
        if isinstance(mx, (int, float)) and mx > 0:
            ind_maxes.append(mx)

    has_mustafa = isinstance(m_min, (int, float)) and m_min > 0 and isinstance(m_max, (int, float)) and m_max > 0
    has_eng = len(ind_mins) > 0 and len(ind_maxes) > 0

    if has_mustafa and has_eng:
        total_compared += 1
        e_min = min(ind_mins)
        e_max = max(ind_maxes)
        mustafa_total_min += m_min
        mustafa_total_max += m_max
        eng_total_min += e_min
        eng_total_max += e_max

        m_avg = (m_min + m_max) / 2
        e_avg = (e_min + e_max) / 2

        if m_avg < e_avg:
            mustafa_smaller += 1
        elif e_avg < m_avg:
            engineers_smaller += 1
        else:
            equal += 1

print(f"Rows where both Mustafa AND engineers estimated: {total_compared}")
print(f"Mustafa smaller: {mustafa_smaller} ({mustafa_smaller/total_compared*100:.0f}%)")
print(f"Engineers smaller: {engineers_smaller} ({engineers_smaller/total_compared*100:.0f}%)")
print(f"Equal: {equal}")
print()
print(f"Mustafa totals: min={mustafa_total_min} max={mustafa_total_max} days")
print(f"Engineer totals: min={eng_total_min} max={eng_total_max} days")
print(f"Mustafa midpoint: {(mustafa_total_min + mustafa_total_max) / 2:.0f} days")
print(f"Engineer midpoint: {(eng_total_min + eng_total_max) / 2:.0f} days")
print(f"Difference: Mustafa is {((mustafa_total_min+mustafa_total_max)/2 - (eng_total_min+eng_total_max)/2) / ((eng_total_min+eng_total_max)/2) * 100:.0f}% vs engineers")
wb.close()
