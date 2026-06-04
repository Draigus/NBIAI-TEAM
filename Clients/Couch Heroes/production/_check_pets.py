import openpyxl
wb = openpyxl.load_workbook('CouchHeroes_Man_Day_Work_Plan_v15.xlsx', data_only=True)
ws = wb['Plan']
for r in range(1, ws.max_row + 1):
    b = ws.cell(row=r, column=2).value
    c = ws.cell(row=r, column=3).value
    if (b and 'pet' in str(b).lower()) or (b and 'mount' in str(b).lower()) or \
       (c and 'pet' in str(c).lower()) or (c and 'parrot' in str(c).lower()) or \
       (c and 'mount' in str(c).lower()):
        team = ws.cell(row=r, column=4).value
        vals = []
        for col in range(87, 163):
            v = ws.cell(row=r, column=col).value
            if v is not None:
                vals.append((col, v))
        marker = 'FEAT' if b and not c else 'STORY'
        ep_str = str(vals) if vals else 'BLANK'
        print(f"Row {r} [{marker}]: B={b} C={c} D={team} EP={ep_str}")

# Also check the Miro extraction for pet data
print("\n--- Miro extraction for pets ---")
with open('_miro_epics_4_10_extraction.txt', 'r') as f:
    in_pets = False
    for line in f:
        if 'Pets' in line or 'MECH PARROT' in line or 'pet' in line.lower():
            in_pets = True
        if in_pets:
            print(line.rstrip())
            if line.strip() == '' and in_pets:
                in_pets = False

# Check what MIRO_STORIES has for pets
print("\n--- MIRO_STORIES entries for pets ---")
with open('build_v12_clean.py', 'r') as f:
    for line in f:
        if 'PARROT' in line or 'Mounts / Pets' in line:
            print(line.rstrip())
wb.close()
