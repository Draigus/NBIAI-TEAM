"""Strict QA: match every source row 1:1 to output and verify actual values."""
import openpyxl
from collections import defaultdict

print('='*70)
print('STRICT QA — 1:1 ROW MATCHING WITH VALUE VERIFICATION')
print('='*70)

# Parse source rows with full detail
def scan_source(filepath, sheet_name, is_eng):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[sheet_name]
    rows = []
    current_epic = current_feature = current_story = None
    for r in range(3, ws.max_row + 1):
        a, b, c, d, e = [ws.cell(row=r, column=col).value for col in range(1, 6)]
        if a: current_epic = str(a).strip()
        if b: current_feature = str(b).strip()
        if c: current_story = str(c).strip()
        nm = str(d).strip() if d else str(c).strip() if c else None
        if not nm or not e: continue
        team = str(e).strip()

        # Get individual estimator values
        ind = {}
        for col in range(8, 26):
            v = ws.cell(row=r, column=col).value
            if v is not None:
                try:
                    fv = float(v)
                    if fv != 0:
                        ind[col] = fv
                except: pass

        # Agreed values
        agreed = {}
        if is_eng:
            for col in [26, 27]:
                v = ws.cell(row=r, column=col).value
                if v is not None:
                    try:
                        fv = float(v)
                        agreed[col] = fv
                    except: pass
        else:
            v = ws.cell(row=r, column=26).value
            if v is not None:
                try: agreed[26] = float(v)
                except: pass

        # What SHOULD appear in output
        if is_eng:
            m_min = agreed.get(26)
            m_max = agreed.get(27)
            if not m_min or not m_max:
                ind_mins = [v for c, v in ind.items() if c % 2 == 0]  # even cols = min
                ind_maxes = [v for c, v in ind.items() if c % 2 == 1]  # odd cols = max
                if not m_min and ind_mins:
                    m_min = min(ind_mins)
                if not m_max and ind_maxes:
                    m_max = max(ind_maxes)
            expected_min = m_min
            expected_max = m_max
        else:
            ind_mins = [v for c, v in ind.items() if c % 2 == 0]
            final = agreed.get(26)
            expected_min = min(ind_mins) if ind_mins else None
            expected_max = final
            if expected_max is None:
                ind_maxes = [v for c, v in ind.items() if c % 2 == 1]
                expected_max = max(ind_maxes) if ind_maxes else None

        has_data = bool(ind or any(v != 0 for v in agreed.values()))

        rows.append({
            'src_row': r, 'feature': current_feature, 'parent_story': current_story if d else None,
            'story': nm, 'team': team, 'sheet': 'Eng' if is_eng else 'Des',
            'expected_min': expected_min, 'expected_max': expected_max,
            'has_data': has_data, 'ind': ind, 'agreed': agreed,
        })
    wb.close()
    return rows

d_rows = scan_source('VS Estimation Design - General Sheet.xlsx', 'Sorted by Epic', False)
e_rows = scan_source('VS Estimation Engineering - General Sheet (1).xlsx', 'Sheet1', True)
all_source = d_rows + e_rows
print(f'Source: {len(d_rows)} design + {len(e_rows)} engineering = {len(all_source)} total')

# Load RP with min/max per row
wb = openpyxl.load_workbook('CouchHeroes_Man_Day_Work_Plan_v15.xlsx', data_only=True)
ws = wb['Resource Planner']
rp_rows = []
for row in range(2, ws.max_row + 1):
    story = ws.cell(row=row, column=3).value
    team_raw = ws.cell(row=row, column=5).value
    min_v = ws.cell(row=row, column=6).value
    max_v = ws.cell(row=row, column=7).value
    if not story: continue
    team = str(team_raw).split(' | ')[-1].strip() if team_raw else ''
    mn = mx = None
    try: mn = float(min_v) if min_v is not None else None
    except: pass
    try: mx = float(max_v) if max_v is not None else None
    except: pass
    rp_rows.append({'story': str(story).strip(), 'team': team, 'min': mn, 'max': mx})
wb.close()

# 1:1 matching: source rows -> output rows in order
# Group both by (story, team) and match by occurrence index
src_by_key = defaultdict(list)
for r in all_source:
    src_by_key[(r['story'].lower(), r['team'].lower())].append(r)

rp_by_key = defaultdict(list)
for r in rp_rows:
    rp_by_key[(r['story'].lower(), r['team'].lower())].append(r)

# Check every source row
data_rows_checked = 0
value_matches = 0
value_mismatches = []
missing_estimates = []

for key, src_list in src_by_key.items():
    rp_list = rp_by_key.get(key, [])

    for i, sr in enumerate(src_list):
        if not sr['has_data']:
            continue
        data_rows_checked += 1

        # Find matching RP row (by index within this key group)
        if i < len(rp_list):
            rp = rp_list[i]
        else:
            missing_estimates.append(sr)
            continue

        # Check if RP has ANY estimate when source does
        rp_has = (rp['min'] is not None and rp['min'] != 0) or (rp['max'] is not None and rp['max'] != 0)
        if not rp_has:
            missing_estimates.append(sr)
            continue

        # Check value accuracy
        ok = True
        if sr['expected_min'] is not None and sr['expected_min'] != 0:
            if rp['min'] != sr['expected_min']:
                ok = False
        if sr['expected_max'] is not None and sr['expected_max'] != 0:
            if rp['max'] != sr['expected_max']:
                ok = False

        if ok:
            value_matches += 1
        else:
            value_mismatches.append((sr, rp))

print(f'\nSource rows with data: {data_rows_checked}')
print(f'Value matches: {value_matches}')
print(f'Value mismatches: {len(value_mismatches)}')
print(f'Missing estimates (source has data, output blank): {len(missing_estimates)}')

if missing_estimates:
    print(f'\n--- MISSING ESTIMATES ---')
    for sr in missing_estimates[:15]:
        print(f'  {sr["sheet"]} R{sr["src_row"]}: "{sr["story"][:40]}" [{sr["team"]}] '
              f'expected min={sr["expected_min"]} max={sr["expected_max"]} '
              f'ind={sr["ind"]} agreed={sr["agreed"]}')

if value_mismatches:
    print(f'\n--- VALUE MISMATCHES ---')
    for sr, rp in value_mismatches[:15]:
        print(f'  {sr["sheet"]} R{sr["src_row"]}: "{sr["story"][:35]}" [{sr["team"]}]')
        print(f'    Expected: min={sr["expected_min"]} max={sr["expected_max"]}')
        print(f'    Got:      min={rp["min"]} max={rp["max"]}')

print(f'\n{"="*70}')
print(f'VERDICT: {"ALL CLEAR" if not missing_estimates and not value_mismatches else "ISSUES FOUND"}')
print(f'{"="*70}')
