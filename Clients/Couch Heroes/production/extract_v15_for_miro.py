"""
Extract v15 Excel Plan sheet into structured JSON for Miro board build.
Reads CouchHeroes_Man_Day_Work_Plan_v15.xlsx (Plan sheet) and outputs v15_miro_data.json.
"""

import json
import os
from openpyxl import load_workbook

# --- Mappings ---

DEPT_COLOURS = {
    'Design': '#4A90D9', 'Art': '#9B59B6', 'Engineering': '#27AE60',
    'Audio': '#F39C12', 'QA': '#E74C3C',
}

# The Excel col D contains "Dept | Team" format, e.g. "Design | Game design"
# Map the raw col D value to (short_team, dept, display_name)
TEAM_LOOKUP = {
    'Design | Game design':             ('Game design',          'Design',      'Design | Game Design'),
    'Design | Level design':            ('Level design',         'Design',      'Design | Level Design'),
    'Design | Narrative':               ('Narrative',            'Design',      'Design | Narrative'),
    'Art | UX/UI':                       ('UX/UI',               'Art',         'Art | UX/UI'),
    'Art | Concept Art':                 ('Concept Art',         'Art',         'Art | Concept Art'),
    'Art | Environment Art':             ('Environment Art',     'Art',         'Art | Environment Art'),
    'Art | Character Art':               ('Character Art',       'Art',         'Art | Character Art'),
    'Art | Animation & Rigging':         ('Animation & Rigging', 'Art',         'Art | Animation'),
    'Art | VFX':                         ('VFX',                 'Art',         'Art | VFX'),
    'Art | TechArt':                     ('TechArt',             'Art',         'Art | Tech Art'),
    'Engineering | Gameplay Engineering': ('Gameplay Engineering','Engineering', 'Eng | Gameplay'),
    'Engineering | Backend Engineering': ('Backend Engineering', 'Engineering', 'Eng | Backend'),
    'QA | QA':                           ('QA',                  'QA',          'QA'),
    'Audio | Audio':                     ('Audio',               'Audio',       'Audio | Sound'),
}


def _fmt_days(val):
    """Format a day value: whole numbers as int, fractions with 1 decimal."""
    if val == int(val):
        return str(int(val))
    return f"{val:.1f}"


def extract_estimate(row_values):
    """
    Extract estimate from a row.
    Early Prod cols 86-161 (38 roles x 2 = 76 cols, Min/Max pairs).
    Concept cols 8-45 (single values).
    Pre-prod cols 47-84 (single values).
    """
    # Early Production: Min/Max pairs at cols 86-161
    ep_mins = []
    ep_maxs = []
    for i in range(86, 162, 2):
        if i < len(row_values) and i + 1 < len(row_values):
            min_val = row_values[i]
            max_val = row_values[i + 1]
            if min_val is not None and isinstance(min_val, (int, float)) and min_val > 0:
                ep_mins.append(min_val)
            if max_val is not None and isinstance(max_val, (int, float)) and max_val > 0:
                ep_maxs.append(max_val)

    if ep_mins or ep_maxs:
        final_min = min(ep_mins) if ep_mins else 0
        final_max = max(ep_maxs) if ep_maxs else 0
        if final_min > 0 and final_max > 0:
            if final_min == final_max:
                return f"{_fmt_days(final_min)}d"
            return f"{_fmt_days(final_min)}-{_fmt_days(final_max)}d"
        elif final_min > 0:
            return f"{_fmt_days(final_min)}d"
        elif final_max > 0:
            return f"{_fmt_days(final_max)}d"

    # Fallback: Concept (8-45) and Pre-prod (47-84) single columns
    fallback_vals = []
    for i in range(8, 46):
        if i < len(row_values):
            val = row_values[i]
            if val is not None and isinstance(val, (int, float)) and val > 0:
                fallback_vals.append(val)
    for i in range(47, 85):
        if i < len(row_values):
            val = row_values[i]
            if val is not None and isinstance(val, (int, float)) and val > 0:
                fallback_vals.append(val)

    if fallback_vals:
        final_min = min(fallback_vals)
        final_max = max(fallback_vals)
        if final_min == final_max:
            return f"{_fmt_days(final_min)}d"
        return f"{_fmt_days(final_min)}-{_fmt_days(final_max)}d"

    return None


def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    xlsx_path = os.path.join(script_dir, 'CouchHeroes_Man_Day_Work_Plan_v15.xlsx')
    output_path = os.path.join(script_dir, 'v15_miro_data.json')

    print(f"Loading workbook: {xlsx_path}")
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)

    # Find the Plan sheet
    plan_sheet = None
    for name in wb.sheetnames:
        if name.lower().strip() == 'plan':
            plan_sheet = wb[name]
            break

    if plan_sheet is None:
        # Try partial match
        for name in wb.sheetnames:
            if 'plan' in name.lower():
                plan_sheet = wb[name]
                print(f"Using sheet: '{name}'")
                break

    if plan_sheet is None:
        print(f"ERROR: No 'Plan' sheet found. Available sheets: {wb.sheetnames}")
        wb.close()
        return

    print(f"Reading sheet: '{plan_sheet.title}'")

    # Read all rows into memory for speed
    rows = []
    for row in plan_sheet.iter_rows(values_only=True):
        rows.append(list(row) if row else [])

    wb.close()
    print(f"Total rows read: {len(rows)}")

    # Data starts at row 5 (0-indexed row 4)
    data_rows = rows[4:]
    print(f"Data rows (from row 5 onwards): {len(data_rows)}")

    # Build structure with sparse fill carry-forward
    epics = {}  # name -> {features: {name -> [stories]}}
    epic_order = []  # preserve insertion order
    current_epic = None
    current_feature = None
    total_stories = 0
    skipped_no_story = 0

    for row in data_rows:
        # Ensure row has enough columns
        if len(row) < 3:
            skipped_no_story += 1
            continue

        # Col A (0) = Epic, Col B (1) = Feature, Col C (2) = Story
        epic_val = row[0]
        feature_val = row[1]
        story_val = row[2]

        # Carry forward epic/feature
        if epic_val is not None and str(epic_val).strip():
            current_epic = str(epic_val).strip()
            if current_epic not in epics:
                epics[current_epic] = {}
                epic_order.append(current_epic)
            current_feature = None  # Reset feature when new epic starts

        if feature_val is not None and str(feature_val).strip():
            current_feature = str(feature_val).strip()
            # Ensure the feature exists in the epic even if no stories follow
            epic_key = current_epic or '(No Epic)'
            if epic_key not in epics:
                epics[epic_key] = {}
                epic_order.append(epic_key)
            if current_feature not in epics[epic_key]:
                epics[epic_key][current_feature] = []

        # Skip if no story name
        if story_val is None or not str(story_val).strip():
            skipped_no_story += 1
            continue

        story_name = str(story_val).strip()

        # Col D (3) = Team (format "Dept | Team"), Col E (4) = Size, Col F (5) = Tier, Col G (6) = Notes
        raw_team = str(row[3]).strip() if len(row) > 3 and row[3] else ''
        if raw_team in TEAM_LOOKUP:
            team, dept, display = TEAM_LOOKUP[raw_team]
        elif '|' in raw_team:
            # Fallback: parse "Dept | Team" format
            parts = raw_team.split('|', 1)
            dept = parts[0].strip()
            team = parts[1].strip()
            display = raw_team
        else:
            team = raw_team
            dept = ''
            display = raw_team
        colour = DEPT_COLOURS.get(dept, '#888888')

        estimate = extract_estimate(row)

        story_obj = {
            'name': story_name,
            'team': team,
            'dept': dept,
            'display': display,
            'colour': colour,
            'estimate': estimate,
        }

        # Use fallback names if epic/feature not yet encountered
        epic_key = current_epic or '(No Epic)'
        feature_key = current_feature or '(No Feature)'

        if epic_key not in epics:
            epics[epic_key] = {}
        if feature_key not in epics[epic_key]:
            epics[epic_key][feature_key] = []

        epics[epic_key][feature_key].append(story_obj)
        total_stories += 1

    # Convert to output format (preserve epic order from spreadsheet)
    output = []
    for epic_name in epic_order:
        features_dict = epics[epic_name]
        features_list = []
        epic_story_count = 0
        for feature_name, stories in features_dict.items():
            features_list.append({
                'name': feature_name,
                'stories': stories,
                'story_count': len(stories),
            })
            epic_story_count += len(stories)

        output.append({
            'name': epic_name,
            'features': features_list,
            'story_count': epic_story_count,
            'feature_count': len(features_list),
        })

    # Write JSON
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(output, f, indent=2, ensure_ascii=False)

    print(f"\nOutput written to: {output_path}")
    print(f"\n{'='*60}")
    print(f"SUMMARY")
    print(f"{'='*60}")
    print(f"Total epics:    {len(output)}")
    print(f"Total features: {sum(e['feature_count'] for e in output)}")
    print(f"Total stories:  {total_stories}")
    print(f"Rows skipped (no story name): {skipped_no_story}")
    print(f"\nPer-epic breakdown:")
    print(f"{'Epic':<40} {'Features':>8} {'Stories':>8}")
    print(f"{'-'*40} {'-'*8} {'-'*8}")
    for epic in output:
        print(f"{epic['name']:<40} {epic['feature_count']:>8} {epic['story_count']:>8}")


if __name__ == '__main__':
    main()
