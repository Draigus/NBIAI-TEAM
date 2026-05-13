"""
Build v12 — EXACT template layout preserved.
No columns added, moved, or modified. Same 434 columns.
Rows expanded for Story/Task depth in col B (Feature).
Estimates filled into Early Prod role columns (85-122).
Estimator detail + Story/Task metadata on a separate sheet.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os, re
from copy import copy

BASE = r"D:\OneDrive\Claude_code\NBIAI_TEAM\Clients\Couch Heroes\production"

# Reuse data loading functions
exec(open(os.path.join(BASE, "consolidate_template_style.py")).read().split("# ============================================================\n# WRITE OUTPUT")[0])

def build():
    template = os.path.join(BASE, "CouchHeroes_Man_Day_Work_Plan_v11.xlsx")
    output = os.path.join(BASE, "CouchHeroes_Man_Day_Work_Plan_v12_final.xlsx")

    # Load data
    print("Loading data...")
    miro_bids = load_miro_bids(HOURS_MAPPING_FILE)
    design = read_design(DESIGN_FILE)
    eng = read_engineering(ENGINEERING_FILE)
    merged = merge_rows(design, eng)
    sorted_rows = sort_rows(merged)
    print(f"  {len(sorted_rows)} rows")

    # Load template metadata (size/tier)
    wb_meta = openpyxl.load_workbook(template, data_only=True)
    ws_meta = wb_meta["Plan"]
    tmpl_meta = {}
    ce = ""
    for r in range(5, 166):
        e = ws_meta.cell(r, 1).value
        f = ws_meta.cell(r, 2).value
        if e: ce = str(e).strip()
        if f: tmpl_meta[f"{ce}::{str(f).strip()}"] = {
            "size": ws_meta.cell(r, 3).value, "tier": ws_meta.cell(r, 4).value
        }
    wb_meta.close()

    # Open template for editing
    print("Opening template...")
    wb = openpyxl.load_workbook(template)
    ws = wb["Plan"]

    # DON'T clear anything. DON'T unmerge anything.
    # The template already has the correct epic/feature structure.
    # We just need to FILL IN the role columns with Miro bid totals.

    EP_START = 85  # Early Prod role start col
    wrap = Alignment(wrap_text=True, vertical="top")

    # ========== FILL MIRO BID TOTALS INTO EXISTING FEATURE ROWS ==========
    print("Filling Miro bid totals into existing feature rows...")
    filled_features = 0

    # Walk the existing template rows and fill in estimates
    current_epic = ""
    for r in range(5, 166):
        epic_val = ws.cell(r, 1).value
        feat_val = ws.cell(r, 2).value

        if epic_val:
            current_epic = str(epic_val).strip()

        if feat_val:
            feat = str(feat_val).strip()
            mb = match_miro(feat, miro_bids)
            if mb:
                for role_name, days in mb["roles"].items():
                    if role_name in ROLE_TO_OFFSET:
                        col = EP_START + ROLE_TO_OFFSET[role_name]
                        ws.cell(r, col, days)
                filled_features += 1

    print(f"  Filled {filled_features} features with Miro bid totals")

    # ========== ADD MIN/MAX SECTIONS AFTER LAST PHASE ==========
    # After LS2 (col 434), add two 38-col sections: EP MIN and EP MAX
    MIN_START = 436  # 1-col gap after col 434
    MAX_START = MIN_START + 39  # 38 roles + 1 gap

    # Headers for Min section
    min_hdr_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    max_hdr_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    hdr_font_sm = Font(bold=True, size=8)
    role_align = Alignment(wrap_text=True, text_rotation=90, horizontal="center", vertical="bottom")
    dark_fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
    white_font = Font(bold=True, size=10, color="FFFFFF")

    # Row 2: section labels
    ws.cell(2, MIN_START, "EARLY PROD — MIN ESTIMATES").font = white_font
    ws.cell(2, MIN_START).fill = dark_fill
    ws.cell(2, MAX_START, "EARLY PROD — MAX ESTIMATES").font = white_font
    ws.cell(2, MAX_START).fill = dark_fill

    # Row 4: role names
    for ri, rname in enumerate(ROLE_NAMES):
        c_min = MIN_START + ri
        c_max = MAX_START + ri
        ws.cell(4, c_min, rname).font = hdr_font_sm
        ws.cell(4, c_min).fill = min_hdr_fill
        ws.cell(4, c_min).alignment = role_align
        ws.cell(4, c_max, rname).font = hdr_font_sm
        ws.cell(4, c_max).fill = max_hdr_fill
        ws.cell(4, c_max).alignment = role_align
        ws.column_dimensions[get_column_letter(c_min)].width = 5
        ws.column_dimensions[get_column_letter(c_max)].width = 5

    # ========== AGGREGATE MIN/MAX FROM ESTIMATION SHEETS PER FEATURE ==========
    print("Aggregating min/max estimates per feature...")

    # Build feature-level aggregates: for each feature+role, collect all min and max values
    feat_role_mins = {}  # key=(epic,feat,role) -> list of min values
    feat_role_maxs = {}

    for row in sorted_rows:
        epic = row["epic"]
        feat = row["feature"]
        team = row["team"]
        if not team or team not in TEAM_TO_ROLE:
            continue
        role = TEAM_TO_ROLE[team]
        key = (epic, feat, role)

        # Design: individual estimator mins and maxes
        if row.get("estimates") and row["source"] in ("Design", "Both"):
            est = row["estimates"]
            for i in range(0, len(est), 2):
                if est[i] is not None and not isinstance(est[i], str):
                    feat_role_mins.setdefault(key, []).append(float(est[i]))
            for i in range(1, len(est), 2):
                if est[i] is not None and not isinstance(est[i], str):
                    feat_role_maxs.setdefault(key, []).append(float(est[i]))

        # Engineering: Mustafa adjusted min/max
        if row.get("eng_finals") and row["source"] in ("Engineering", "Both"):
            ef = row["eng_finals"]
            if ef[0] is not None and not isinstance(ef[0], str):
                feat_role_mins.setdefault(key, []).append(float(ef[0]))
            if len(ef) > 1 and ef[1] is not None and not isinstance(ef[1], str):
                feat_role_maxs.setdefault(key, []).append(float(ef[1]))

    # Write aggregated min/max into the new sections, matching template feature rows
    current_epic = ""
    filled_minmax = 0
    for r in range(5, 166):
        epic_val = ws.cell(r, 1).value
        feat_val = ws.cell(r, 2).value
        if epic_val:
            current_epic = str(epic_val).strip()
        if feat_val:
            feat = str(feat_val).strip()
            for ri, rname in enumerate(ROLE_NAMES):
                key = (current_epic, feat, rname)
                # Sum mins and maxes for this feature+role
                if key in feat_role_mins:
                    ws.cell(r, MIN_START + ri, round(sum(feat_role_mins[key]), 1))
                    filled_minmax += 1
                if key in feat_role_maxs:
                    ws.cell(r, MAX_START + ri, round(sum(feat_role_maxs[key]), 1))

    print(f"  Filled min/max for {filled_minmax} feature-role combinations")
    data_row = 166  # for Miro-only features below

    total_data_rows = filled_features

    # ========== DETAIL SHEET (separate) ==========
    print("Building Detail sheet...")
    if "Estimator Detail" in wb.sheetnames:
        del wb["Estimator Detail"]
    wd = wb.create_sheet("Estimator Detail")

    # Headers
    det_headers = ["Epic", "Feature", "Story", "Task", "Department", "Team",
                   "Description", "Definition of Success"]
    # Design estimators
    for name, role in DESIGN_ESTIMATORS:
        det_headers.append(f"{name} Min")
        det_headers.append(f"{name} Max")
    det_headers.append("Design Final")
    # Engineering estimators
    for name, role in ENG_ESTIMATORS:
        det_headers.append(f"{name} Min")
        det_headers.append(f"{name} Max")
    det_headers.append("Mustafa Adj Min")
    det_headers.append("Mustafa Adj Max")
    det_headers.append("Eng Notes")

    hdr_fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
    hdr_font = Font(bold=True, size=9, color="FFFFFF")
    for i, h in enumerate(det_headers, 1):
        c = wd.cell(1, i, h)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = Alignment(wrap_text=True, horizontal="center")

    det_row = 2
    for row in sorted_rows:
        wd.cell(det_row, 1, row["epic"])
        wd.cell(det_row, 2, row["feature"])
        wd.cell(det_row, 3, row["story"])
        wd.cell(det_row, 4, row["task"])
        wd.cell(det_row, 5, row["department"])
        wd.cell(det_row, 6, row["team"])
        if row["description"]:
            wd.cell(det_row, 7, row["description"]).alignment = wrap
        if row["dos"]:
            wd.cell(det_row, 8, row["dos"]).alignment = wrap

        col = 9
        # Design estimates
        if row.get("estimates"):
            for v in row["estimates"]:
                if v is not None: wd.cell(det_row, col, v)
                col += 1
            if row.get("finals"):
                for v in row["finals"]:
                    if v is not None: wd.cell(det_row, col, v)
                    col += 1
            else:
                col += 1
        else:
            col += len(DESIGN_ESTIMATORS) * 2 + 1

        # Engineering estimates
        if row.get("eng_estimates"):
            for v in row["eng_estimates"]:
                if v is not None:
                    if isinstance(v, str):
                        wd.cell(det_row, col, v).alignment = wrap
                    else:
                        wd.cell(det_row, col, v)
                col += 1
            if row.get("eng_finals"):
                for v in row["eng_finals"]:
                    if v is not None:
                        if isinstance(v, str):
                            wd.cell(det_row, col, v).alignment = wrap
                        else:
                            wd.cell(det_row, col, v)
                    col += 1
        det_row += 1

    # Column widths for detail sheet
    det_widths = {1: 15, 2: 25, 3: 30, 4: 35, 5: 10, 6: 18, 7: 40, 8: 40}
    for cn, w in det_widths.items():
        wd.column_dimensions[get_column_letter(cn)].width = w
    for c in range(9, len(det_headers) + 1):
        wd.column_dimensions[get_column_letter(c)].width = 7
    wd.column_dimensions[get_column_letter(len(det_headers))].width = 30  # Eng Notes
    wd.freeze_panes = "A2"

    print(f"Saving {output}...")
    wb.save(output)
    wb.close()
    print(f"Done! Plan: {total_data_rows} rows. Detail: {det_row - 2} rows.")

if __name__ == "__main__":
    build()
