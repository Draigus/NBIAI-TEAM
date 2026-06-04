import openpyxl

wb = openpyxl.load_workbook('VS Estimation Design - General Sheet.xlsx', data_only=True)
ws = wb['Sorted by Epic']

# First: what are the column headers?
print("=== COLUMN HEADERS (row 1-2) ===")
for col in range(1, 30):
    h1 = ws.cell(row=1, column=col).value
    h2 = ws.cell(row=2, column=col).value
    if h1 or h2:
        print(f"  Col {col}: row1={h1} row2={h2}")

print("\n=== ALL TEAMS AND COUNTS ===")
teams = {}
for r in range(3, ws.max_row + 1):
    e = ws.cell(row=r, column=5).value
    if e:
        t = str(e).strip()
        teams[t] = teams.get(t, 0) + 1
for t, c in sorted(teams.items(), key=lambda x: -x[1]):
    print(f"  {t}: {c} rows")

# Now check design-team rows specifically
print("\n=== DESIGN TEAM ROWS (Game design + Level design) ===")
design_teams = {'Game design', 'Level design', 'game design', 'level design'}
total = 0
with_est = 0
without_est = 0
without_examples = []

for r in range(3, ws.max_row + 1):
    e = ws.cell(row=r, column=5).value
    if not e or str(e).strip() not in design_teams:
        continue
    c = ws.cell(row=r, column=3).value
    d = ws.cell(row=r, column=4).value
    b = ws.cell(row=r, column=2).value
    name = str(d or c or b or '?')[:50]
    total += 1

    # Check ALL columns from 6 onwards for any numeric value
    has_any = False
    data_cols = []
    for col in range(6, ws.max_column + 1):
        v = ws.cell(row=r, column=col).value
        if v is not None and v != '' and v != 0:
            has_any = True
            data_cols.append((col, v))

    if has_any:
        with_est += 1
    else:
        without_est += 1
        if len(without_examples) < 10:
            without_examples.append(f"  Row {r}: {name} [{e}]")

print(f"Total design-team rows: {total}")
print(f"With estimates: {with_est}")
print(f"Without estimates: {without_est}")
if without_examples:
    print(f"\nFirst 10 without estimates:")
    for ex in without_examples:
        print(ex)

# Also check ALL rows regardless of team
print("\n=== ALL ROWS WITH ANY DATA IN COLS 6+ ===")
all_total = 0
all_with = 0
for r in range(3, ws.max_row + 1):
    c = ws.cell(row=r, column=3).value
    d = ws.cell(row=r, column=4).value
    b = ws.cell(row=r, column=2).value
    if not (b or c or d):
        continue
    all_total += 1
    for col in range(6, ws.max_column + 1):
        v = ws.cell(row=r, column=col).value
        if v is not None and v != '' and v != 0:
            try:
                float(v)
                all_with += 1
                break
            except (ValueError, TypeError):
                continue

print(f"Total named rows: {all_total}")
print(f"Rows with any numeric data: {all_with}")
print(f"Rows completely blank: {all_total - all_with}")

wb.close()
