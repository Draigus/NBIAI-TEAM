"""Comprehensive QA — every line against source."""
import openpyxl, re
from collections import Counter

print('='*70)
print('COMPREHENSIVE QA — EVERY LINE AGAINST SOURCE')
print('='*70)

def scan_source(filepath, sheet_name, is_eng):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[sheet_name]
    rows = []
    current_epic = current_feature = current_story = None
    for r in range(3, ws.max_row + 1):
        a, b, c, d, e = [ws.cell(row=r, column=col).value for col in range(1, 6)]
        if a: current_epic = str(a).strip()
        if b: current_feature = str(b).strip()
        if c: current_story = str(c).strip()
        nm = str(d).strip() if d else str(c).strip() if c else None
        if not nm or not e: continue
        team = str(e).strip()
        ind_vals = {}
        for col in range(8, 26):
            v = ws.cell(row=r, column=col).value
            if v is not None:
                try: ind_vals[col] = float(v)
                except: pass
        agreed = {}
        for col in ([26] if not is_eng else [26, 27]):
            v = ws.cell(row=r, column=col).value
            if v is not None:
                try: agreed[col] = float(v)
                except: pass
        rows.append({
            'src_row': r, 'epic': current_epic, 'feature': current_feature,
            'story': nm, 'team': team, 'source': 'Eng' if is_eng else 'Des',
            'has_nonzero': any(v != 0 for v in list(ind_vals.values()) + list(agreed.values()))
        })
    wb.close()
    return rows

d_rows = scan_source('VS Estimation Design - General Sheet.xlsx', 'Sorted by Epic', False)
e_rows = scan_source('VS Estimation Engineering - General Sheet (1).xlsx', 'Sheet1', True)
all_source = d_rows + e_rows
print(f'Source: {len(d_rows)} design + {len(e_rows)} engineering = {len(all_source)} total')

# Load RP
wb_rp = openpyxl.load_workbook('CouchHeroes_Man_Day_Work_Plan_v15.xlsx', data_only=True)
ws_rp = wb_rp['Resource Planner']
rp_rows = []
for row in range(2, ws_rp.max_row + 1):
    story = ws_rp.cell(row=row, column=3).value
    team_raw = ws_rp.cell(row=row, column=5).value
    min_v = ws_rp.cell(row=row, column=6).value
    max_v = ws_rp.cell(row=row, column=7).value
    if not story: continue
    team = str(team_raw).split(' | ')[-1].strip() if team_raw else ''
    has_est = False
    try: has_est = (min_v is not None and float(min_v) != 0) or (max_v is not None and float(max_v) != 0)
    except: pass
    rp_rows.append({'story': str(story).strip(), 'team': team, 'has_est': has_est})
wb_rp.close()
print(f'Output: {len(rp_rows)} RP rows')

# CHECK 1: Row count
print(f'\n--- CHECK 1: ROW COUNT ---')
print(f'Source={len(all_source)} Output={len(rp_rows)} Match={len(all_source)==len(rp_rows)}')

# CHECK 2: Every source row in output
print(f'\n--- CHECK 2: EVERY SOURCE ROW IN OUTPUT ---')
rp_counter = Counter((r['story'].lower(), r['team'].lower()) for r in rp_rows)
src_counter = Counter((r['story'].lower(), r['team'].lower()) for r in all_source)
missing = {k: (sc, rp_counter.get(k, 0)) for k, sc in src_counter.items() if rp_counter.get(k, 0) < sc}
print(f'Source combos: {len(src_counter)}, Output combos: {len(rp_counter)}')
print(f'Missing/short: {len(missing)}')
for (s, t), (sc, rc) in list(missing.items())[:5]:
    print(f'  "{s[:40]}" [{t}] src={sc} out={rc}')

# CHECK 3: No cross-dept merging
print(f'\n--- CHECK 3: NO CROSS-DEPT MERGING ---')
src_teams = {}
for r in all_source:
    s = r['story'].lower()
    src_teams.setdefault(s, set()).add(r['team'].lower())
rp_teams = {}
for r in rp_rows:
    s = r['story'].lower()
    rp_teams.setdefault(s, set()).add(r['team'].lower())
merged = 0
for story, st in src_teams.items():
    rt = rp_teams.get(story, set())
    if st - rt:
        merged += 1
        if merged <= 5:
            print(f'  MERGED: "{story[:40]}" missing teams: {st - rt}')
print(f'Stories with missing teams: {merged}')

# CHECK 4: Estimates
print(f'\n--- CHECK 4: ESTIMATES ---')
src_with_est = [r for r in all_source if r['has_nonzero']]
rp_est_index = {}
for r in rp_rows:
    k = (r['story'].lower(), r['team'].lower())
    rp_est_index.setdefault(k, []).append(r)
est_missing = 0
for sr in src_with_est:
    k = (sr['story'].lower(), sr['team'].lower())
    if not any(r['has_est'] for r in rp_est_index.get(k, [])):
        est_missing += 1
print(f'Source rows with estimates: {len(src_with_est)}')
print(f'Source estimates missing from output: {est_missing}')

# CHECK 5: Miro
print(f'\n--- CHECK 5: MIRO ---')
wb_plan = openpyxl.load_workbook('CouchHeroes_Man_Day_Work_Plan_v15.xlsx', data_only=True)
ws_plan = wb_plan['Plan']
feat_miro = 0
for row in range(1, ws_plan.max_row + 1):
    b = ws_plan.cell(row=row, column=2).value
    c = ws_plan.cell(row=row, column=3).value
    if b and not c:
        for col in range(107, 163):
            v = ws_plan.cell(row=row, column=col).value
            if v is not None:
                try:
                    if float(v) > 0: feat_miro += 1; break
                except: pass
wb_plan.close()
with open('hours_mapping.txt') as f:
    miro_src = sum(1 for l in f if re.match(r'^.+?:\s*\d+D total\s*\|', l.strip()))
print(f'Miro source features: {miro_src}, Plan features with Miro data: {feat_miro}')

# SUMMARY
print(f'\n{"="*70}')
print('QA SUMMARY')
print('='*70)
issues = 0
for name, ok in [
    ('Row count', len(all_source) == len(rp_rows)),
    ('All source rows present', len(missing) == 0),
    ('No cross-dept merging', merged == 0),
    ('All estimates present', est_missing == 0),
]:
    status = 'PASS' if ok else 'FAIL'
    if not ok: issues += 1
    print(f'  {status}: {name}')
print(f'\nTotal issues: {issues}')
