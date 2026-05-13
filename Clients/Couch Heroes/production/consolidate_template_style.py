"""
Couch Heroes Production Plan — Template-Style Consolidation
Preserves the template's phase/role column layout, expands rows to full depth,
and adds estimator detail columns after the phase sections.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os, json, re
from copy import copy

BASE = r"D:\OneDrive\Claude_code\NBIAI_TEAM\Clients\Couch Heroes\production"
DESIGN_FILE = os.path.join(BASE, "VS Estimation Design - General Sheet.xlsx")
ENGINEERING_FILE = os.path.join(BASE, "VS Estimation Engineering - General Sheet (1).xlsx")
TEMPLATE_FILE = os.path.join(BASE, "CouchHeroes_Man_Day_Work_Plan_v11.xlsx")
OUTPUT_FILE = os.path.join(BASE, "CouchHeroes_Man_Day_Work_Plan_v12.xlsx")
HOURS_MAPPING_FILE = os.path.join(BASE, "hours_mapping.txt")

# ============================================================
# MAPPINGS
# ============================================================

TEAM_TO_DEPT = {
    "Game design": "Design", "Level design": "Design", "Narrative": "Design",
    "UX/UI": "Art", "Concept Art": "Art", "Environment Art": "Art",
    "Character Art": "Art", "Animation & Rigging": "Art", "VFX": "Art", "TechArt": "Art",
    "Gameplay Engineering": "Engineering", "Backend Engineering": "Engineering",
    "QA": "QA", "Audio": "Audio",
}

# Team name -> role column offset within a phase (0-based from phase start)
# The 38 roles in each phase: CEO=0, Creative Director=1, ..., Platform Team=37
ROLE_NAMES = [
    "CEO", "Creative Director", "Head of Legal", "COO", "HR Manager",
    "Executive Producer", "Producer", "Game Producer", "Tech Producer", "Art Producer",
    "Game Director", "Game Design Lead", "Level Design Lead", "Game Designer",
    "Combat Designer", "Level Designer", "Technical Designer", "Writer",
    "Art Director", "Concept Artist", "UI/UX Designer", "Graphic Designer",
    "Character Artist", "Environment Artist", "Prop Artist", "Animator",
    "Character Rigger", "VFX Artist", "Technical Artist", "CTO", "Tech Lead",
    "Gameplay Developer", "Full Stack Developer", "Backend Developer",
    "Network Engineer", "QA", "Sound Designer", "Platform Team",
]
ROLE_TO_OFFSET = {r: i for i, r in enumerate(ROLE_NAMES)}

TEAM_TO_ROLE = {
    "Game design": "Game Designer",
    "Level design": "Level Designer",
    "Gameplay Engineering": "Gameplay Developer",
    "Backend Engineering": "Backend Developer",
    "UX/UI": "UI/UX Designer",
    "Concept Art": "Concept Artist",
    "Environment Art": "Environment Artist",
    "Character Art": "Character Artist",
    "Animation & Rigging": "Animator",
    "VFX": "VFX Artist",
    "TechArt": "Technical Artist",
    "Audio": "Sound Designer",
    "QA": "QA",
    "Narrative": "Writer",
}

EPIC_NORMALISE = {
    "Combat System": "Combat", "Social": "Social & Multiplayer",
    "Rename to in-game camera squancer during dialog": "World Systems",
    "Phasing system": "World Systems", "Live Service Support": "Live Game",
}

DESIGN_ESTIMATORS = [
    ("Yorgos Dritsas", "Jr Game Designer"), ("Nadir Latif", "Combat Designer"),
    ("Kieron Naylor", "Technical Designer"), ("Panos Andriopoulos", "Jr Level Designer"),
    ("Nikolas Tziotis", "Jr Level Designer"), ("Seth Dahl", "Jr Level Designer"),
    ("Maria Cibej", "Jr Writer"), ("Hannah Pickard", "QA"),
    ("Aggeliki Peroutsea", "Jr Sound Designer"),
]
ENG_ESTIMATORS = [
    ("Samer Balushi", "Backend Developer"), ("Nikos Gerontakis", "Sr Full Stack Developer"),
    ("Daniel Negri", "Sr Full Stack Developer"), ("Ilya Achmetov", "Gameplay Developer"),
    ("Raynor D'Souza", "Sr Gameplay Developer"), ("Leon Huang", "Gameplay Developer"),
    ("Ignasi Ezpeleta", "Jr Gameplay Developer"), ("Roberto", "Gameplay Developer"),
    ("Matt", "Technical Animation Engineer"),
]

# Miro feature name mapping (Excel name -> Miro name)
EXCEL_TO_MIRO = {
    "In-game Tutorial": "FTUE (tutorial cave segment)",
    "Melee Combat Framework": "Combat framework",
    "Combat abilities": "Combat Abilities",
    "Basic Magic system": "Magic",
    "Ranged (Corruption Gun)": "Ranged combat",
    "PVP Combat": "PvP combat",
    "Guilds (S)": "Guilds System",
    "Co-op - Party": "Co-op/Party",
    "In-game Mail": "In-game mail",
    "Skill System (1 skill tree active)": "Skill System",
    "Anywhere door placement": "Anywhere Door",
    "User Space Decoration": "User space Decoration",
    "Pets (mechanical parrot reuse)": "Mounts/ Pets",
    "Auction House economy": "Auction House",
    "In-game economy": "In game Economy",
    "Various Quest Types Support": "Quests : Courier, Escort, Kill, Collect, research, time obj, Partner Quest, corruption",
    "Quest Tools": "Quests Backend/Tooling",
    "Quest Rewards System": "Quest reward system",
    "Bugged Dungeon Integration": "The Dungeon",
    "In game RMT Store": "In-game Store",
    "Dungeons mechanics": "Dungeon Mechanics",
    "Chat system": "Game Chat System",
    "Dungeon finder": "LFG / Pick up groups",
    "Party system": "Co-op/Party",
    "Leaderboard system": "CombatBalance & Telemetry",
    "Combat component": "Combat framework",
    "Combat director": "Combat framework",
    "Subscription page on the website": "RMT game Economy",
    "Sync hotbar to db": "Item Core System",
    "Sync last known player position": "Player Progression",
}

MIRO_ONLY_EPIC_MAP = {
    "Achievements & Trophies": "Player Build", "Banks": "Items & Inventory",
    "Character Creation & Customization": "Player Build", "Co-op/Raids": "Social & Multiplayer",
    "Combat AI behavior": "Combat", "Friends": "Social & Multiplayer",
    "Garment Shop": "Player Economy", "Guild House": "Social & Multiplayer",
    "Inventory System": "Items & Inventory", "Monetization Bible": "Game Bibles",
    "Notifications": "Social & Multiplayer", "P2P Item Transfer": "Items & Inventory",
    "P2P/P2G User Space Connection": "User Space", "Partner Shop": "Player Economy",
    "Player Account Authentification": "Platform", "Player Account Creation/Deletion": "Platform",
    "Player Account Privacy Settings & Legal": "Platform", "Professions": "Player Build",
    "Quest Item": "Quest System", "Quest Multiplayer": "Quest System",
    "Quest Tracking": "Platform", "Quest dialog GUI": "Quest System",
    "Quest log": "Quest System", "Quests as Events": "Quest System",
    "Red Books (emergency guidelines for Live Service)": "Game Bibles",
    "Saving House Layouts": "User Space", "Social Media Support": "Platform",
    "Targeted VO": "Quest System", "Virtual Currency (VC) System": "Player Economy",
    "[Developer Dashboard] Developer Account": "Platform",
    "[Developer Dashboard] Game Dashboard": "Platform",
    "[Player Dashboard] Friends": "Platform", "[Player Dashboard] Game Dashboard": "Platform",
    "[Player Dashboard] Marketplace": "Platform",
    "[Player Dashboard] Player Account Features": "Platform",
    "[Player Dashboard] Search & Notifications": "Platform",
}

MIRO_BID_ROLES = [
    "Animator", "Character Artist", "Character Rigger", "Combat Designer",
    "Concept Artist", "Environment Artist", "Game Designer", "Game Director",
    "Gameplay Developer", "Level Designer", "Prop Artist", "Sound Designer",
    "Technical Artist", "UI/UX Designer", "VFX Artist", "Writer",
]

# ============================================================
# DATA LOADING (reuse from previous script)
# ============================================================

def _clean(val):
    if val is None: return None
    s = str(val).strip()
    return s if s else None

def load_miro_bids(path):
    bids = {}
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("===") or line.startswith("x=") or line.startswith("Excluded"):
                continue
            m = re.match(r"^(.+?):\s*(\d+)D\s+total\s*\|\s*(.*)$", line)
            if m:
                roles = {}
                for part in m.group(3).strip().split(","):
                    rm = re.match(r"^\s*(.+?)=(\d+)D\s*$", part.strip())
                    if rm: roles[rm.group(1).strip()] = int(rm.group(2))
                bids[m.group(1).strip()] = {"total": int(m.group(2)), "roles": roles}
                continue
            m2 = re.match(r"^(.+?):\s*(\d+)D\s+total\s*\|?\s*$", line)
            if m2:
                bids[m2.group(1).strip()] = {"total": int(m2.group(2)), "roles": {}}
    return bids

def read_source_sheet(path, sheet_name, est_col_ranges, final_cols, is_eng=False):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name]
    rows = []
    cur_epic = cur_feat = cur_story = ""

    for r in range(3, ws.max_row + 1):
        epic = _clean(ws.cell(r, 1).value)
        feat = _clean(ws.cell(r, 2).value)
        story = _clean(ws.cell(r, 3).value)
        task = _clean(ws.cell(r, 4).value)
        team = _clean(ws.cell(r, 5).value)
        desc = _clean(ws.cell(r, 6).value)
        dos = _clean(ws.cell(r, 7).value)

        if epic:
            cur_epic = EPIC_NORMALISE.get(epic, epic)
            cur_story = ""
        if feat:
            cur_feat = feat
            cur_story = ""
        if story:
            cur_story = story
        if not feat and not story and not task:
            continue

        effective_story = story if story else (cur_story if task else None)

        # Read all individual estimator values
        estimates = []
        for col_start, col_end in est_col_ranges:
            for c in range(col_start, col_end + 1):
                estimates.append(ws.cell(r, c).value)

        # Read final/adjusted columns
        finals = []
        for c in final_cols:
            finals.append(ws.cell(r, c).value)

        # Determine the "agreed estimate" for this row's role column
        # Design: use Final (days) — last of final_cols
        # Engineering: use Mustafa Adjusted Max — final_cols[1]
        if is_eng:
            agreed = finals[1] if len(finals) > 1 and finals[1] is not None else finals[0]
        else:
            agreed = finals[0] if finals else None

        rows.append({
            "epic": cur_epic, "feature": cur_feat,
            "feature_is_header": bool(feat and not story and not task),
            "story": effective_story, "task": task,
            "team": team, "department": TEAM_TO_DEPT.get(team, "") if team else "",
            "description": desc, "dos": dos,
            "source": "Engineering" if is_eng else "Design",
            "estimates": estimates, "finals": finals, "agreed": agreed,
        })
    wb.close()
    return rows

def read_design(path):
    # Estimator cols: 8-25 (9 people x 2), Final: col 26
    return read_source_sheet(path, "Sorted by Epic", [(8, 25)], [26], is_eng=False)

def read_engineering(path):
    # Estimator cols: 8-25 (9 people x 2), Mustafa: 26-27, Notes: 28
    return read_source_sheet(path, "Sheet1", [(8, 25)], [26, 27, 28], is_eng=True)

def merge_rows(design_rows, eng_rows):
    eng_lookup = {}
    for row in eng_rows:
        key = (row["epic"], row["feature"], row["story"] or "", row["task"] or "", row["team"] or "")
        eng_lookup[key] = row
    merged = []
    merged_keys = set()
    for row in design_rows:
        key = (row["epic"], row["feature"], row["story"] or "", row["task"] or "", row["team"] or "")
        if key in eng_lookup:
            er = eng_lookup[key]
            row["eng_estimates"] = er["estimates"]
            row["eng_finals"] = er["finals"]
            row["eng_agreed"] = er["agreed"]
            row["source"] = "Both"
            if not row["description"] and er["description"]: row["description"] = er["description"]
            if not row["dos"] and er["dos"]: row["dos"] = er["dos"]
            merged_keys.add(key)
        else:
            row["eng_estimates"] = None
            row["eng_finals"] = None
            row["eng_agreed"] = None
        merged.append(row)
    for row in eng_rows:
        key = (row["epic"], row["feature"], row["story"] or "", row["task"] or "", row["team"] or "")
        if key not in merged_keys:
            row["eng_estimates"] = row["estimates"]
            row["eng_finals"] = row["finals"]
            row["eng_agreed"] = row["agreed"]
            row["estimates"] = None
            row["finals"] = None
            merged.append(row)
    return merged

def sort_rows(rows):
    order = ["Player Build","World Systems","Combat","User Space","Items & Inventory",
             "Player Economy","Quest System","Social & Multiplayer","Game Bibles",
             "Platform","Live Game","Partner Build","Product Publishing"]
    rank = {e: i for i, e in enumerate(order)}
    return sorted(rows, key=lambda r: (rank.get(r["epic"], 99), r["feature"] or "",
                                        0 if r["feature_is_header"] else 1,
                                        r["story"] or "", r["task"] or ""))

def match_miro(feat, bids):
    if not feat: return None
    if feat in EXCEL_TO_MIRO:
        m = EXCEL_TO_MIRO[feat]
        if m in bids: return bids[m]
    if feat in bids: return bids[feat]
    low = {k.lower(): v for k, v in bids.items()}
    if feat.lower() in low: return low[feat.lower()]
    for mk, mv in bids.items():
        if feat.lower() in mk.lower() or mk.lower() in feat.lower(): return mv
    return None

# ============================================================
# WRITE OUTPUT IN TEMPLATE STYLE
# ============================================================

def write_template_style(rows, template_path, output_path, miro_bids):
    # Load the template and copy it
    wb = openpyxl.load_workbook(template_path)
    ws = wb["Plan"]

    # --- Template column layout ---
    # We need to INSERT 4 new columns: Story (after Feature), Task, Department, Team
    # This shifts all phase columns right by 4
    # Original: col1=Epic, col2=Feature, col3=Size, col4=Tier, col5=Notes
    # New:      col1=Epic, col2=Feature, col3=Story, col4=Task, col5=Dept, col6=Team,
    #           col7=Size, col8=Tier, col9=Notes, col10=Description, col11=DoS
    # Phase sections start at col 12 instead of col 6

    META_SHIFT = 6  # we add 6 extra metadata columns (Story, Task, Dept, Team, Desc, DoS)

    # Unmerge ALL existing merged cells first (must be done before any cell clearing)
    merged = list(ws.merged_cells.ranges)
    for m in merged:
        ws.unmerge_cells(str(m))

    # Clear all rows
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            cell.value = None
            cell.fill = PatternFill()
            cell.font = Font()
            cell.border = Border()
            cell.alignment = Alignment()

    # Styles
    hdr_font_w = Font(bold=True, size=10, color="FFFFFF")
    hdr_fill_blue = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hdr_fill_dk = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    gate_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    concept_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    preprod_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    prod_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    release_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    ls_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    thin_border = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
    wrap = Alignment(wrap_text=True, vertical="top")
    epic_fill = PatternFill(start_color="FE9F4D", end_color="FE9F4D", fill_type="solid")
    feat_fill = PatternFill(start_color="2DC75C", end_color="2DC75C", fill_type="solid")
    story_fill = PatternFill(start_color="659DF2", end_color="659DF2", fill_type="solid")
    task_fill = PatternFill(start_color="DEDAFF", end_color="DEDAFF", fill_type="solid")

    # New metadata headers (row 4)
    new_meta = ["Epic", "Feature", "Story", "Task", "Department", "Team",
                "Description", "Definition of Success", "Size", "Current Tier", "Notes"]
    meta_count = len(new_meta)  # 11

    # Phase structure: (gate_col_offset, phase_name, sub_name, fill)
    # Each phase = 1 gate col + 38 role cols = 39 cols
    PHASE_WIDTH = 39
    phases = [
        ("CONCEPT", "GATE\n0", "CONCEPT\n~2 months\nTarget: T0 Ideation", concept_fill),
        ("PRE-PRODUCTION", "GATE\n1", "PRE-PRODUCTION\nTarget: T1-T2 R&D / GDD-TDD-Brief", preprod_fill),
        ("EARLY PROD", "GATE\n2", "EARLY PROD\nTarget: T3 Prototype", prod_fill),
        ("MID PROD", "GATE\n3", "MID PRO\nTarget: T4 MVP", prod_fill),
        ("END PROD", "GATE\n4", "END PROD\nTarget: T5 Feature Complete", prod_fill),
        ("ALPHA", "GATE\n5", "ALPHA\nTarget: T6 Player Tested", release_fill),
        ("BETA", "GATE\n6", "BETA\nTarget: T6 Player Tested", release_fill),
        ("LAUNCH", "GATE\n7", "LAUNCH\nTarget: T6 Player Ready", release_fill),
        ("FIRST RELEASE", "GATE\n8", "FIRST RELEASE\n~1.5 months\nTarget: T6-T7", release_fill),
        ("LS 1", "GATE\n9", "LS 1\nLive\nTarget: T7 Expand", ls_fill),
        ("LS 2", "GATE\n10", "LS 2\nLive\nTarget: T8 Scale", ls_fill),
    ]

    # Super-phase groups
    super_phases = {
        0: ("CONCEPT", concept_fill),          # phase 0
        1: ("PRE-PRODUCTION", preprod_fill),   # phase 1
        2: ("PRODUCTION", prod_fill),           # phases 2-4
        5: ("RELEASE", release_fill),           # phases 5-8
        9: ("LIVE SERVICE", ls_fill),           # phases 9-10
    }

    # Write metadata header row (row 4)
    for i, h in enumerate(new_meta):
        c = ws.cell(4, i + 1, h)
        c.font = hdr_font_w; c.fill = hdr_fill_blue; c.border = thin_border; c.alignment = wrap

    # Row 2: "WORK STREAMS" super header
    ws.cell(2, 1, "WORK STREAMS").font = Font(bold=True, size=11)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=meta_count)

    # Write phase headers
    phase_start_col = meta_count + 1  # col 12
    early_prod_role_start = None  # we'll record this

    for pi, (pname, gate_label, sub_label, pfill) in enumerate(phases):
        gate_col = phase_start_col + pi * PHASE_WIDTH
        role_start = gate_col + 1
        role_end = role_start + 37

        if pname == "EARLY PROD":
            early_prod_role_start = role_start

        # Row 3: gate marker
        gc = ws.cell(3, gate_col, gate_label)
        gc.font = Font(bold=True, size=9); gc.fill = gate_fill; gc.alignment = wrap; gc.border = thin_border

        # Row 3: sub-phase label
        sc = ws.cell(3, role_start, sub_label)
        sc.font = Font(bold=True, size=9); sc.fill = pfill; sc.alignment = wrap; sc.border = thin_border

        # Row 4: role names
        for ri, rname in enumerate(ROLE_NAMES):
            rc = ws.cell(4, role_start + ri, rname)
            rc.font = Font(bold=True, size=8); rc.fill = pfill
            rc.alignment = Alignment(wrap_text=True, text_rotation=90, horizontal="center")
            rc.border = thin_border

        # Row 2: super-phase headers
        if pi in super_phases:
            sp_name, sp_fill = super_phases[pi]
            ws.cell(2, gate_col, sp_name).font = Font(bold=True, size=11)
            ws.cell(2, gate_col).fill = sp_fill

    total_phase_cols = len(phases) * PHASE_WIDTH
    last_phase_col = phase_start_col + total_phase_cols - 1

    # --- Estimator detail columns AFTER phases ---
    est_start = last_phase_col + 2  # gap column

    ws.cell(2, est_start, "DESIGN ESTIMATOR DETAIL").font = Font(bold=True, size=10, color="FFFFFF")
    ws.cell(2, est_start).fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")

    col = est_start
    for name, role in DESIGN_ESTIMATORS:
        ws.cell(3, col, f"{name}\n({role})").font = Font(bold=True, size=8)
        ws.cell(3, col).fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        ws.cell(3, col).alignment = Alignment(wrap_text=True, horizontal="center")
        ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col+1)
        ws.cell(4, col, "Min").font = Font(bold=True, size=8); ws.cell(4, col).fill = hdr_fill_blue; ws.cell(4, col).font = hdr_font_w
        ws.cell(4, col+1, "Max").font = Font(bold=True, size=8); ws.cell(4, col+1).fill = hdr_fill_blue; ws.cell(4, col+1).font = hdr_font_w
        col += 2
    ws.cell(4, col, "Design Final").font = hdr_font_w; ws.cell(4, col).fill = hdr_fill_blue
    design_est_start = est_start
    design_est_end = col  # includes Final
    col += 1

    eng_est_start = col
    ws.cell(2, eng_est_start, "ENGINEERING ESTIMATOR DETAIL").font = Font(bold=True, size=10, color="FFFFFF")
    ws.cell(2, eng_est_start).fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    eng_fill = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
    eng_hdr = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")

    for name, role in ENG_ESTIMATORS:
        ws.cell(3, col, f"{name}\n({role})").font = Font(bold=True, size=8)
        ws.cell(3, col).fill = eng_fill
        ws.cell(3, col).alignment = Alignment(wrap_text=True, horizontal="center")
        ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col+1)
        ws.cell(4, col, "Min").font = hdr_font_w; ws.cell(4, col).fill = eng_hdr
        ws.cell(4, col+1, "Max").font = hdr_font_w; ws.cell(4, col+1).fill = eng_hdr
        col += 2
    ws.cell(3, col, "Mustafa Sibai\n(Head of Eng)").font = Font(bold=True, size=8)
    ws.cell(3, col).fill = eng_fill; ws.cell(3, col).alignment = Alignment(wrap_text=True, horizontal="center")
    ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col+1)
    ws.cell(4, col, "Adj Min").font = hdr_font_w; ws.cell(4, col).fill = eng_hdr
    ws.cell(4, col+1, "Adj Max").font = hdr_font_w; ws.cell(4, col+1).fill = eng_hdr
    col += 2
    ws.cell(4, col, "Eng Notes").font = hdr_font_w; ws.cell(4, col).fill = eng_hdr
    ws.cell(4, col).alignment = wrap
    eng_est_end = col
    total_cols = col

    # --- Column widths ---
    widths = {1: 18, 2: 30, 3: 35, 4: 40, 5: 12, 6: 20, 7: 40, 8: 40, 9: 6, 10: 14, 11: 14}
    for cn, w in widths.items():
        ws.column_dimensions[get_column_letter(cn)].width = w
    # Phase role columns: narrow
    for c in range(phase_start_col, last_phase_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 5
    # Estimator columns: narrow
    for c in range(est_start, total_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 6
    # Eng notes wider
    ws.column_dimensions[get_column_letter(eng_est_end)].width = 35

    # Row heights
    ws.row_dimensions[3].height = 80
    ws.row_dimensions[4].height = 100

    # ============================================================
    # DATA ROWS
    # ============================================================
    data_row = 5
    prev_epic = prev_feat = ""
    template_meta = {}
    # Load template metadata for size/tier
    wb2 = openpyxl.load_workbook(TEMPLATE_FILE, data_only=True)
    ws2 = wb2["Plan"]
    ce = ""
    for r in range(5, 166):
        e = ws2.cell(r, 1).value
        f = ws2.cell(r, 2).value
        if e: ce = str(e).strip()
        if f: template_meta[f"{ce}::{str(f).strip()}"] = {"size": ws2.cell(r, 3).value, "tier": ws2.cell(r, 4).value}
    wb2.close()

    stats = {"total": 0, "epics": set(), "features": set(), "miro_matched": 0}

    for row in rows:
        epic = row["epic"]
        feat = row["feature"]
        story = row["story"]
        task = row["task"]
        is_epic_change = (epic != prev_epic)
        is_feat_change = (feat != prev_feat)

        if is_epic_change:
            ws.cell(data_row, 1, epic).font = Font(bold=True, size=11)
            ws.cell(data_row, 1).fill = epic_fill
            stats["epics"].add(epic)
            prev_epic = epic

        if is_feat_change:
            ws.cell(data_row, 2, feat).font = Font(bold=True, size=10)
            ws.cell(data_row, 2).fill = feat_fill
            mk = f"{epic}::{feat}"
            if mk in template_meta:
                ws.cell(data_row, 9, template_meta[mk]["size"])
                ws.cell(data_row, 10, template_meta[mk]["tier"])

            # Miro bid totals -> role columns in Early Prod
            mb = match_miro(feat, miro_bids)
            if mb:
                stats["miro_matched"] += 1
                for role_name, days in mb["roles"].items():
                    if role_name in ROLE_TO_OFFSET:
                        ws.cell(data_row, early_prod_role_start + ROLE_TO_OFFSET[role_name], days)

            stats["features"].add(mk)
            prev_feat = feat

        if story:
            ws.cell(data_row, 3, story).fill = story_fill
        if task:
            ws.cell(data_row, 4, task).fill = task_fill

        ws.cell(data_row, 5, row["department"])
        ws.cell(data_row, 6, row["team"])
        if row["description"]:
            ws.cell(data_row, 7, row["description"]).alignment = wrap
        if row["dos"]:
            ws.cell(data_row, 8, row["dos"]).alignment = wrap

        # Put agreed estimate into Early Prod role column
        team = row["team"]
        if team and team in TEAM_TO_ROLE:
            role = TEAM_TO_ROLE[team]
            if role in ROLE_TO_OFFSET:
                role_col = early_prod_role_start + ROLE_TO_OFFSET[role]
                # Design agreed
                if row.get("agreed") is not None:
                    ws.cell(data_row, role_col, row["agreed"])
                # Eng agreed (for eng-only or merged rows)
                if row.get("eng_agreed") is not None:
                    ws.cell(data_row, role_col, row["eng_agreed"])

        # Write design estimator detail
        if row.get("estimates"):
            col = design_est_start
            for v in row["estimates"]:
                if v is not None: ws.cell(data_row, col, v)
                col += 1
            # Finals
            if row.get("finals"):
                for v in row["finals"]:
                    if v is not None: ws.cell(data_row, col, v)
                    col += 1

        # Write engineering estimator detail
        if row.get("eng_estimates"):
            col = eng_est_start
            for v in row["eng_estimates"]:
                if v is not None:
                    if isinstance(v, str):
                        ws.cell(data_row, col, v).alignment = wrap
                    else:
                        ws.cell(data_row, col, v)
                col += 1
            if row.get("eng_finals"):
                for v in row["eng_finals"]:
                    if v is not None:
                        if isinstance(v, str):
                            ws.cell(data_row, col, v).alignment = wrap
                        else:
                            ws.cell(data_row, col, v)
                    col += 1

        stats["total"] += 1
        data_row += 1

    # --- Add Miro-only features ---
    miro_matched_names = set()
    for row in rows:
        if row.get("feature"):
            m = match_miro(row["feature"], miro_bids)
            if m:
                if row["feature"] in EXCEL_TO_MIRO:
                    miro_matched_names.add(EXCEL_TO_MIRO[row["feature"]])
                miro_matched_names.add(row["feature"])
                for mk in miro_bids:
                    if row["feature"].lower() in mk.lower() or mk.lower() in row["feature"].lower():
                        miro_matched_names.add(mk)

    miro_only_fill = PatternFill(start_color="E8DAEF", end_color="E8DAEF", fill_type="solid")
    for mf in sorted(MIRO_ONLY_EPIC_MAP.keys()):
        if mf in miro_matched_names: continue
        bid = miro_bids.get(mf)
        if not bid or bid["total"] == 0: continue
        epic = MIRO_ONLY_EPIC_MAP[mf]

        if epic != prev_epic:
            ws.cell(data_row, 1, epic).font = Font(bold=True, size=11)
            ws.cell(data_row, 1).fill = epic_fill
            prev_epic = epic

        ws.cell(data_row, 2, mf).font = Font(bold=True, size=10)
        ws.cell(data_row, 2).fill = miro_only_fill
        ws.cell(data_row, 11, "MIRO BID ONLY")

        for role_name, days in bid["roles"].items():
            if role_name in ROLE_TO_OFFSET:
                ws.cell(data_row, early_prod_role_start + ROLE_TO_OFFSET[role_name], days)

        mk = f"{epic}::{mf}"
        if mk in template_meta:
            ws.cell(data_row, 9, template_meta[mk]["size"])
            ws.cell(data_row, 10, template_meta[mk]["tier"])

        data_row += 1
        stats["total"] += 1

    # Freeze panes
    ws.freeze_panes = ws.cell(5, meta_count + 1)

    wb.save(output_path)
    wb.close()
    return stats

# ============================================================
# MAIN
# ============================================================

def main():
    print("=" * 60)
    print("COUCH HEROES — TEMPLATE-STYLE CONSOLIDATION")
    print("=" * 60)

    print("\n1. Loading Miro bids...")
    miro_bids = load_miro_bids(HOURS_MAPPING_FILE)
    print(f"   {len(miro_bids)} features")

    print("\n2. Reading Design sheet...")
    design = read_design(DESIGN_FILE)
    print(f"   {len(design)} rows")

    print("\n3. Reading Engineering sheet...")
    eng = read_engineering(ENGINEERING_FILE)
    print(f"   {len(eng)} rows")

    print("\n4. Merging...")
    merged = merge_rows(design, eng)
    print(f"   {len(merged)} rows ({len(design)+len(eng)-len(merged)} merged)")

    print("\n5. Sorting...")
    sorted_rows = sort_rows(merged)

    print(f"\n6. Writing {os.path.basename(OUTPUT_FILE)}...")
    stats = write_template_style(sorted_rows, TEMPLATE_FILE, OUTPUT_FILE, miro_bids)

    print("\n" + "=" * 60)
    print("DONE")
    print("=" * 60)
    print(f"Output: {OUTPUT_FILE}")
    print(f"Rows: {stats['total']}, Epics: {len(stats['epics'])}, Features: {len(stats['features'])}")
    print(f"Miro matched: {stats['miro_matched']}")

if __name__ == "__main__":
    main()
