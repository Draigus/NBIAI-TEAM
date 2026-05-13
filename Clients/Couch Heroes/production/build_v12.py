"""
Build v12 from v11 template — preserves template style exactly.
Adds Story/Task/Dept/Team metadata columns.
Early Prod phase gets Min/Max sub-columns per role.
All other phases remain single-column per role.
Estimator detail appended after phase sections.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os, re
from copy import copy

BASE = r"D:\OneDrive\Claude_code\NBIAI_TEAM\Clients\Couch Heroes\production"

# ---- Reuse all mappings from consolidate_template_style.py ----
exec(open(os.path.join(BASE, "consolidate_template_style.py")).read().split("# ============================================================\n# WRITE OUTPUT")[0])

# ============================================================
# BUILD V12
# ============================================================

def build_v12():
    template = os.path.join(BASE, "CouchHeroes_Man_Day_Work_Plan_v11.xlsx")
    output = os.path.join(BASE, "CouchHeroes_Man_Day_Work_Plan_v12.xlsx")

    # Load and read the template Plan sheet to capture header styling
    wb_tmpl = openpyxl.load_workbook(template)
    ws_tmpl = wb_tmpl["Plan"]

    # Capture all formatting from header rows 2-4 for the original 5 metadata cols
    # and the full phase structure
    # We'll unmerge, read styles, then build new sheet

    # First, read all the cell styles we need
    tmpl_r2_styles = {}
    tmpl_r3_styles = {}
    tmpl_r4_styles = {}
    for c in range(1, ws_tmpl.max_column + 1):
        for r, store in [(2, tmpl_r2_styles), (3, tmpl_r3_styles), (4, tmpl_r4_styles)]:
            cell = ws_tmpl.cell(r, c)
            store[c] = {
                "value": cell.value,
                "font": copy(cell.font),
                "fill": copy(cell.fill),
                "alignment": copy(cell.alignment),
                "border": copy(cell.border),
            }

    # Read template metadata
    tmpl_meta = {}
    ce = ""
    for r in range(5, 166):
        e = ws_tmpl.cell(r, 1).value
        f = ws_tmpl.cell(r, 2).value
        if e: ce = str(e).strip()
        if f: tmpl_meta[f"{ce}::{str(f).strip()}"] = {
            "size": ws_tmpl.cell(r, 3).value,
            "tier": ws_tmpl.cell(r, 4).value
        }
    wb_tmpl.close()

    # ---- New column layout ----
    # Original metadata: cols 1-5 (Epic, Feature, Size, Tier, Notes)
    # New metadata: cols 1-9 (Epic, Feature, Story, Task, Dept, Team, Size, Tier, Notes)
    # Phase cols shift right by 4

    SHIFT = 4  # 4 extra metadata columns

    # Original phase positions (from template analysis)
    ORIG_PHASES = [
        {"gate": 6,   "roles_start": 7,   "roles_end": 44},   # Concept
        {"gate": 45,  "roles_start": 46,  "roles_end": 83},   # Pre-prod
        {"gate": 84,  "roles_start": 85,  "roles_end": 122},  # Early Prod
        {"gate": 123, "roles_start": 124, "roles_end": 161},  # Mid Prod
        {"gate": 162, "roles_start": 163, "roles_end": 200},  # End Prod
        {"gate": 201, "roles_start": 202, "roles_end": 239},  # Alpha
        {"gate": 240, "roles_start": 241, "roles_end": 278},  # Beta
        {"gate": 279, "roles_start": 280, "roles_end": 317},  # Launch
        {"gate": 318, "roles_start": 319, "roles_end": 356},  # First Release
        {"gate": 357, "roles_start": 358, "roles_end": 395},  # LS1
        {"gate": 396, "roles_start": 397, "roles_end": 434},  # LS2
    ]

    # For Early Prod (phase 2), we double the role columns for Min/Max
    # That adds 38 extra columns (76 instead of 38)
    EP_EXTRA = 38  # extra cols for Early Prod min/max

    # Calculate new positions
    # Phases 0-1: shift by SHIFT only
    # Phase 2 (Early Prod): shift by SHIFT, but now 76 role cols instead of 38
    # Phases 3+: shift by SHIFT + EP_EXTRA

    def new_gate(phase_idx):
        orig = ORIG_PHASES[phase_idx]["gate"]
        if phase_idx <= 2:
            return orig + SHIFT
        else:
            return orig + SHIFT + EP_EXTRA

    def new_roles_start(phase_idx):
        return new_gate(phase_idx) + 1

    def new_roles_end(phase_idx):
        if phase_idx == 2:
            return new_roles_start(phase_idx) + 75  # 76 cols (38 pairs)
        else:
            return new_roles_start(phase_idx) + 37  # 38 cols

    # ---- Build the new workbook ----
    wb = openpyxl.load_workbook(template)

    # Delete old Plan, create fresh
    if "Plan" in wb.sheetnames:
        idx = wb.sheetnames.index("Plan")
        del wb["Plan"]
    else:
        idx = 0
    ws = wb.create_sheet("Plan", idx)

    # ---- Styles ----
    dark_fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
    white_bold = Font(bold=True, size=12, color="FFFFFF")
    hdr_font_w = Font(bold=True, size=10, color="FFFFFF")
    hdr_font_sm = Font(bold=True, size=8, color="FFFFFF")
    role_align = Alignment(wrap_text=True, text_rotation=90, horizontal="center", vertical="bottom")
    wrap = Alignment(wrap_text=True, vertical="top")
    thin = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))

    # Phase names, sub-labels, and fills matching template
    phase_info = [
        ("CONCEPT", "GATE\n0", "CONCEPT\n~2 months\nTarget: T0 Ideation"),
        ("PRE-PRODUCTION", "GATE\n1", "PRE-PRODUCTION\nTarget: T1-T2 R&D / GDD-TDD-Brief"),
        ("PRODUCTION", "GATE\n2", "EARLY PROD\nTarget: T3 Prototype"),
        ("", "GATE\n3", "MID PRO\nTarget: T4 MVP"),
        ("", "GATE\n4", "END PROD\nTarget: T5 Feature Complete"),
        ("RELEASE", "GATE\n5", "ALPHA\nTarget: T6 Player Tested"),
        ("", "GATE\n6", "BETA\nTarget: T6 Player Tested"),
        ("", "GATE\n7", "LAUNCH\nTarget: T6 Player Ready"),
        ("", "GATE\n8", "FIRST RELEASE\n~1.5 months\nTarget: T6-T7"),
        ("LIVE SERVICE", "GATE\n9", "LS 1\nLive\nTarget: T7 Expand"),
        ("", "GATE\n10", "LS 2\nLive\nTarget: T8 Scale"),
    ]

    # Copy phase fills from template
    phase_fills = {}
    for pi, p in enumerate(ORIG_PHASES):
        cell = ws_tmpl if False else None  # template is closed, use known colours
    # Known fills from template
    concept_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    preprod_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    prod_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    release_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    ls_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    gate_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    gate_font = Font(bold=True, size=9)
    phase_fill_map = [
        concept_fill, preprod_fill, prod_fill, prod_fill, prod_fill,
        release_fill, release_fill, release_fill, release_fill,
        ls_fill, ls_fill,
    ]

    ep_min_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    ep_max_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    epic_fill = PatternFill(start_color="FE9F4D", end_color="FE9F4D", fill_type="solid")
    feat_fill = PatternFill(start_color="2DC75C", end_color="2DC75C", fill_type="solid")
    story_fill = PatternFill(start_color="659DF2", end_color="659DF2", fill_type="solid")
    task_fill = PatternFill(start_color="DEDAFF", end_color="DEDAFF", fill_type="solid")

    # ========== ROW 2: WORK STREAMS + super-phase headers ==========
    # Metadata super-header
    c = ws.cell(2, 1, "WORK STREAMS")
    c.font = white_bold; c.fill = dark_fill
    ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=9)

    # Phase super-headers (row 2)
    super_phases = {0: "CONCEPT", 1: "PRE-PRODUCTION", 2: "PRODUCTION", 5: "RELEASE", 9: "LIVE SERVICE"}
    super_ends = {0: 0, 1: 1, 2: 4, 5: 8, 9: 10}  # last phase index in this super group
    for start_pi, label in super_phases.items():
        end_pi = super_ends[start_pi]
        col_start = new_gate(start_pi)
        col_end = new_roles_end(end_pi)
        c = ws.cell(2, col_start, label)
        c.font = white_bold; c.fill = dark_fill
        if col_end > col_start:
            ws.merge_cells(start_row=2, start_column=col_start, end_row=2, end_column=col_end)

    # ========== ROW 3: Gate markers + sub-phase labels ==========
    for pi in range(11):
        gc = new_gate(pi)
        rs = new_roles_start(pi)
        re_ = new_roles_end(pi)

        # Gate column
        ws.cell(3, gc, phase_info[pi][1]).font = gate_font
        ws.cell(3, gc).fill = gate_fill
        ws.cell(3, gc).alignment = wrap
        ws.cell(3, gc).border = thin

        # Sub-phase label
        ws.cell(3, rs, phase_info[pi][2]).font = Font(bold=True, size=9)
        ws.cell(3, rs).fill = phase_fill_map[pi]
        ws.cell(3, rs).alignment = wrap
        if re_ > rs:
            ws.merge_cells(start_row=3, start_column=rs, end_row=3, end_column=re_)

    # ========== ROW 4: Metadata + role headers ==========
    meta_headers = ["Epic", "Feature", "Story", "Task", "Department", "Team",
                    "Size", "Current Tier", "Notes"]
    for i, h in enumerate(meta_headers):
        c = ws.cell(4, i + 1, h)
        c.font = hdr_font_w; c.fill = dark_fill; c.border = thin; c.alignment = wrap

    for pi in range(11):
        rs = new_roles_start(pi)
        pfill = phase_fill_map[pi]

        if pi == 2:
            # Early Prod: Min/Max pairs
            for ri, rname in enumerate(ROLE_NAMES):
                col_min = rs + ri * 2
                col_max = rs + ri * 2 + 1
                ws.cell(4, col_min, f"{rname}\nMin").font = hdr_font_sm
                ws.cell(4, col_min).fill = ep_min_fill
                ws.cell(4, col_min).alignment = role_align
                ws.cell(4, col_min).border = thin
                ws.cell(4, col_max, f"{rname}\nMax").font = hdr_font_sm
                ws.cell(4, col_max).fill = ep_max_fill
                ws.cell(4, col_max).alignment = role_align
                ws.cell(4, col_max).border = thin
        else:
            for ri, rname in enumerate(ROLE_NAMES):
                c = ws.cell(4, rs + ri, rname)
                c.font = hdr_font_sm; c.fill = pfill
                c.alignment = role_align; c.border = thin

    # ========== ESTIMATOR DETAIL SECTION (after all phases) ==========
    est_start = new_roles_end(10) + 2  # 2-col gap after LS2

    ws.cell(2, est_start, "DESIGN ESTIMATOR DETAIL").font = Font(bold=True, size=10, color="FFFFFF")
    ws.cell(2, est_start).fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    des_sub = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    col = est_start
    for name, role in DESIGN_ESTIMATORS:
        ws.cell(3, col, f"{name}\n({role})").font = Font(bold=True, size=8)
        ws.cell(3, col).fill = des_sub; ws.cell(3, col).alignment = Alignment(wrap_text=True, horizontal="center")
        ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col+1)
        ws.cell(4, col, "Min").font = hdr_font_sm; ws.cell(4, col).fill = dark_fill; ws.cell(4, col).border = thin
        ws.cell(4, col+1, "Max").font = hdr_font_sm; ws.cell(4, col+1).fill = dark_fill; ws.cell(4, col+1).border = thin
        col += 2
    ws.cell(4, col, "Design Final").font = hdr_font_sm; ws.cell(4, col).fill = dark_fill; ws.cell(4, col).border = thin
    design_detail_start = est_start
    design_detail_end = col
    col += 2

    eng_detail_start = col
    ws.cell(2, eng_detail_start, "ENGINEERING ESTIMATOR DETAIL").font = Font(bold=True, size=10, color="FFFFFF")
    ws.cell(2, eng_detail_start).fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    eng_sub = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
    eng_hdr = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    for name, role in ENG_ESTIMATORS:
        ws.cell(3, col, f"{name}\n({role})").font = Font(bold=True, size=8)
        ws.cell(3, col).fill = eng_sub; ws.cell(3, col).alignment = Alignment(wrap_text=True, horizontal="center")
        ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col+1)
        ws.cell(4, col, "Min").font = hdr_font_sm; ws.cell(4, col).fill = eng_hdr; ws.cell(4, col).border = thin
        ws.cell(4, col+1, "Max").font = hdr_font_sm; ws.cell(4, col+1).fill = eng_hdr; ws.cell(4, col+1).border = thin
        col += 2
    ws.cell(3, col, "Mustafa Sibai\n(Head of Eng)").font = Font(bold=True, size=8)
    ws.cell(3, col).fill = eng_sub; ws.cell(3, col).alignment = Alignment(wrap_text=True, horizontal="center")
    ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col+1)
    ws.cell(4, col, "Adj Min").font = hdr_font_sm; ws.cell(4, col).fill = eng_hdr; ws.cell(4, col).border = thin
    ws.cell(4, col+1, "Adj Max").font = hdr_font_sm; ws.cell(4, col+1).fill = eng_hdr; ws.cell(4, col+1).border = thin
    col += 2
    ws.cell(4, col, "Eng Notes").font = hdr_font_sm; ws.cell(4, col).fill = eng_hdr; ws.cell(4, col).border = thin
    ws.cell(4, col).alignment = wrap
    eng_detail_end = col

    # ========== COLUMN WIDTHS ==========
    widths = {1: 18, 2: 28, 3: 32, 4: 38, 5: 11, 6: 18, 7: 6, 8: 14, 9: 14}
    for cn, w in widths.items():
        ws.column_dimensions[get_column_letter(cn)].width = w
    # Phase cols: narrow
    for pi in range(11):
        rs = new_roles_start(pi)
        re_ = new_roles_end(pi)
        for c in range(new_gate(pi), re_ + 1):
            ws.column_dimensions[get_column_letter(c)].width = 4 if c == new_gate(pi) else 5
    # Estimator cols
    for c in range(design_detail_start, eng_detail_end + 1):
        ws.column_dimensions[get_column_letter(c)].width = 6
    ws.column_dimensions[get_column_letter(eng_detail_end)].width = 30  # notes

    ws.row_dimensions[3].height = 80
    ws.row_dimensions[4].height = 100

    # ========== LOAD DATA ==========
    print("Loading data...")
    miro_bids = load_miro_bids(HOURS_MAPPING_FILE)
    design = read_design(DESIGN_FILE)
    eng = read_engineering(ENGINEERING_FILE)
    merged = merge_rows(design, eng)
    sorted_rows = sort_rows(merged)
    print(f"  {len(sorted_rows)} rows to write")

    # ========== WRITE DATA ROWS ==========
    ep_rs = new_roles_start(2)  # Early Prod role start

    data_row = 5
    prev_epic = prev_feat = ""

    for row in sorted_rows:
        epic = row["epic"]
        feat = row["feature"]
        story = row["story"]
        task = row["task"]

        if epic != prev_epic:
            ws.cell(data_row, 1, epic).font = Font(bold=True, size=11)
            ws.cell(data_row, 1).fill = epic_fill
            prev_epic = epic

        if feat != prev_feat:
            ws.cell(data_row, 2, feat).font = Font(bold=True, size=10)
            ws.cell(data_row, 2).fill = feat_fill
            mk = f"{epic}::{feat}"
            if mk in tmpl_meta:
                ws.cell(data_row, 7, tmpl_meta[mk]["size"])
                ws.cell(data_row, 8, tmpl_meta[mk]["tier"])

            # Miro bid -> Early Prod role columns (as agreed totals)
            mb = match_miro(feat, miro_bids)
            if mb:
                for role_name, days in mb["roles"].items():
                    if role_name in ROLE_TO_OFFSET:
                        off = ROLE_TO_OFFSET[role_name]
                        # Put in both min and max (Miro bids are single values)
                        ws.cell(data_row, ep_rs + off * 2, days)
                        ws.cell(data_row, ep_rs + off * 2 + 1, days)
            prev_feat = feat

        if story:
            ws.cell(data_row, 3, story).fill = story_fill
        if task:
            ws.cell(data_row, 4, task).fill = task_fill

        ws.cell(data_row, 5, row["department"])
        ws.cell(data_row, 6, row["team"])
        if row["description"]:
            ws.cell(data_row, 9, row["description"]).alignment = wrap

        # Put estimate into Early Prod role Min/Max columns
        team = row["team"]
        if team and team in TEAM_TO_ROLE:
            role = TEAM_TO_ROLE[team]
            if role in ROLE_TO_OFFSET:
                off = ROLE_TO_OFFSET[role]
                col_min = ep_rs + off * 2
                col_max = ep_rs + off * 2 + 1

                # Design: estimates are individual estimator values; agreed = Final
                if row.get("estimates") and row["source"] in ("Design", "Both"):
                    est = row["estimates"]
                    mins = [float(est[i]) for i in range(0, len(est), 2) if est[i] is not None and not isinstance(est[i], str)]
                    maxes = [float(est[i]) for i in range(1, len(est), 2) if est[i] is not None and not isinstance(est[i], str)]
                    if row.get("finals") and row["finals"][0] is not None and not isinstance(row["finals"][0], str):
                        ws.cell(data_row, col_max, row["finals"][0])
                    elif maxes:
                        ws.cell(data_row, col_max, round(sum(maxes)/len(maxes), 1))
                    if mins:
                        ws.cell(data_row, col_min, round(sum(mins)/len(mins), 1))

                # Engineering: use Mustafa Adjusted min/max
                if row.get("eng_finals") and row["source"] in ("Engineering", "Both"):
                    ef = row["eng_finals"]
                    if ef[0] is not None:  # Mustafa min
                        ws.cell(data_row, col_min, ef[0])
                    if len(ef) > 1 and ef[1] is not None:  # Mustafa max
                        ws.cell(data_row, col_max, ef[1])

        # Write design estimator detail
        if row.get("estimates"):
            col = design_detail_start
            for v in row["estimates"]:
                if v is not None: ws.cell(data_row, col, v)
                col += 1
            if row.get("finals"):
                for v in row["finals"]:
                    if v is not None: ws.cell(data_row, col, v)
                    col += 1

        # Write engineering estimator detail
        if row.get("eng_estimates"):
            col = eng_detail_start
            for v in row["eng_estimates"]:
                if v is not None:
                    if isinstance(v, str):
                        ws.cell(data_row, col, v).alignment = wrap
                    else:
                        ws.cell(data_row, col, v)
                col += 1
            if row.get("eng_finals"):
                for v in row["eng_finals"]:
                    if v is not None:
                        if isinstance(v, str):
                            ws.cell(data_row, col, v).alignment = wrap
                        else:
                            ws.cell(data_row, col, v)
                    col += 1

        data_row += 1

    # Add Miro-only features
    miro_matched = set()
    for row in sorted_rows:
        if row.get("feature"):
            m = match_miro(row["feature"], miro_bids)
            if m:
                if row["feature"] in EXCEL_TO_MIRO: miro_matched.add(EXCEL_TO_MIRO[row["feature"]])
                miro_matched.add(row["feature"])
                for mk in miro_bids:
                    if row["feature"].lower() in mk.lower() or mk.lower() in row["feature"].lower():
                        miro_matched.add(mk)

    miro_only_fill = PatternFill(start_color="E8DAEF", end_color="E8DAEF", fill_type="solid")
    for mf in sorted(MIRO_ONLY_EPIC_MAP.keys()):
        if mf in miro_matched: continue
        bid = miro_bids.get(mf)
        if not bid or bid["total"] == 0: continue
        epic = MIRO_ONLY_EPIC_MAP[mf]
        if epic != prev_epic:
            ws.cell(data_row, 1, epic).font = Font(bold=True, size=11)
            ws.cell(data_row, 1).fill = epic_fill
            prev_epic = epic
        ws.cell(data_row, 2, mf).font = Font(bold=True, size=10)
        ws.cell(data_row, 2).fill = miro_only_fill
        ws.cell(data_row, 9, "MIRO BID ONLY")
        for role_name, days in bid["roles"].items():
            if role_name in ROLE_TO_OFFSET:
                off = ROLE_TO_OFFSET[role_name]
                ws.cell(data_row, ep_rs + off * 2, days)
                ws.cell(data_row, ep_rs + off * 2 + 1, days)
        mk = f"{epic}::{mf}"
        if mk in tmpl_meta:
            ws.cell(data_row, 7, tmpl_meta[mk]["size"])
            ws.cell(data_row, 8, tmpl_meta[mk]["tier"])
        data_row += 1

    # Freeze panes
    ws.freeze_panes = "A5"

    print(f"Saving to {output}...")
    wb.save(output)
    wb.close()
    print(f"Done! {data_row - 5} data rows written.")
    return output

if __name__ == "__main__":
    out = build_v12()
    print(f"\nOutput: {out}")
