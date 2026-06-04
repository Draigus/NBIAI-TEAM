import openpyxl
wb = openpyxl.load_workbook('CouchHeroes_Man_Day_Work_Plan_v15.xlsx', data_only=True)
ws = wb['Plan']

blank_by_team = {}

for r in range(5, ws.max_row + 1):
    c = ws.cell(row=r, column=3).value
    if not c:
        continue
    team = ws.cell(row=r, column=4).value or 'Unknown'

    has_ep = False
    for col in range(87, 163):
        v = ws.cell(row=r, column=col).value
        if v is not None:
            has_ep = True
            break

    if not has_ep:
        t = str(team)
        blank_by_team[t] = blank_by_team.get(t, 0) + 1

print("Blank rows by specific team:")
print("=" * 50)
for team, count in sorted(blank_by_team.items(), key=lambda x: -x[1]):
    print(f"  {team:40s}: {count}")
print(f"\nTotal blank: {sum(blank_by_team.values())}")
wb.close()
