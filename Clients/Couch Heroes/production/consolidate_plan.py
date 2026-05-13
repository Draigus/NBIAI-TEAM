"""
Couch Heroes Production Plan Consolidation Script
Extracts Design + Engineering estimation sheets + Miro bid totals into the master template.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import json
import re

# --- File paths ---
BASE = r"D:\OneDrive\Claude_code\NBIAI_TEAM\Clients\Couch Heroes\production"
DESIGN_FILE = os.path.join(BASE, "VS Estimation Design - General Sheet.xlsx")
ENGINEERING_FILE = os.path.join(BASE, "VS Estimation Engineering - General Sheet (1).xlsx")
TEMPLATE_FILE = os.path.join(BASE, "CouchHeroes_Man_Day_Work_Plan_v11.xlsx")
OUTPUT_FILE = os.path.join(BASE, "CouchHeroes_Man_Day_Work_Plan_v12_consolidated.xlsx")
HOURS_MAPPING_FILE = os.path.join(BASE, "hours_mapping.txt")

# --- Team to Department mapping ---
TEAM_TO_DEPT = {
    "Game design": "Design",
    "Level design": "Design",
    "Narrative": "Design",
    "UX/UI": "Art",
    "Concept Art": "Art",
    "Environment Art": "Art",
    "Character Art": "Art",
    "Animation & Rigging": "Art",
    "VFX": "Art",
    "TechArt": "Art",
    "Gameplay Engineering": "Engineering",
    "Backend Engineering": "Engineering",
    "QA": "QA",
    "Audio": "Audio",
}

# --- Epic name normalisation ---
EPIC_NORMALISE = {
    "Combat System": "Combat",
    "Social": "Social & Multiplayer",
    "Rename to in-game camera squancer during dialog": "World Systems",
    "Phasing system": "World Systems",
    "Live Service Support": "Live Game",
}

# --- NOT FOR VS features (from Miro analysis) ---
NOT_FOR_VS = {
    "P2P Item Transfer", "Durable Items", "Mounts",
    "Dialogue/Quest Cinematics", "Meta Quest System", "Quest Generator System",
    "In Game Cinematic (IGC)", "Meta quest system",
    "Dueling", "Notifications", "Player Config PvP Server",
    "LFG / Pick-up Groups", "LFG / Pick up groups", "Friends",
    "Co-op / Raids", "Co-op/Raids",
    "Platform", "Live Service Game", "Partner Build", "Product Publishing",
    "Red Books (Live Service emergency guidelines)", "Red Books (emergency guidelines for Live Service)",
    "Monetisation Bible", "Monetization Bible",
}

# --- Miro bid roles ---
MIRO_BID_ROLES = [
    "Animator", "Character Artist", "Character Rigger", "Combat Designer",
    "Concept Artist", "Environment Artist", "Game Designer", "Game Director",
    "Gameplay Developer", "Level Designer", "Prop Artist", "Sound Designer",
    "Technical Artist", "UI/UX Designer", "VFX Artist", "Writer",
]

# --- Design estimators (person, role) ---
DESIGN_ESTIMATORS = [
    ("Yorgos Dritsas", "Jr Game Designer"),
    ("Nadir Latif", "Combat Designer"),
    ("Kieron Naylor", "Technical Designer"),
    ("Panos Andriopoulos", "Jr Level Designer"),
    ("Nikolas Tziotis", "Jr Level Designer"),
    ("Seth Dahl", "Jr Level Designer"),
    ("Maria Cibej", "Jr Writer"),
    ("Hannah Pickard", "QA"),
    ("Aggeliki Peroutsea", "Jr Sound Designer"),
]

# --- Engineering estimators (person, role) ---
ENG_ESTIMATORS = [
    ("Samer Balushi", "Backend Developer"),
    ("Nikos Gerontakis", "Sr Full Stack Developer"),
    ("Daniel Negri", "Sr Full Stack Developer"),
    ("Ilya Achmetov", "Gameplay Developer"),
    ("Raynor D'Souza", "Sr Gameplay Developer"),
    ("Leon Huang", "Gameplay Developer"),
    ("Ignasi Ezpeleta", "Jr Gameplay Developer"),
    ("Roberto", "Gameplay Developer"),
    ("Matt", "Technical Animation Engineer"),
]

# --- Styles ---
HEADER_FONT = Font(bold=True, size=11)
HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL_BLUE = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
DESIGN_SUPER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
DESIGN_SUB_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
ENG_SUPER_FILL = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
ENG_SUB_FILL = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
MIRO_SUPER_FILL = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
MIRO_SUB_FILL = PatternFill(start_color="E2D0F0", end_color="E2D0F0", fill_type="solid")
EPIC_FILL = PatternFill(start_color="FE9F4D", end_color="FE9F4D", fill_type="solid")
EPIC_FONT = Font(bold=True, size=12)
FEAT_FILL = PatternFill(start_color="2DC75C", end_color="2DC75C", fill_type="solid")
FEAT_FONT = Font(bold=True, size=11)
STORY_FILL = PatternFill(start_color="659DF2", end_color="659DF2", fill_type="solid")
NFV_FILL = PatternFill(start_color="B0B0B0", end_color="B0B0B0", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
WRAP = Alignment(wrap_text=True, vertical="top")


# ============================================================
# DATA LOADING
# ============================================================

def load_template_metadata(template_path):
    wb = openpyxl.load_workbook(template_path, data_only=True)
    ws = wb["Plan"]
    metadata = {}
    current_epic = ""
    for r in range(5, 166):
        epic = ws.cell(r, 1).value
        feat = ws.cell(r, 2).value
        size = ws.cell(r, 3).value
        tier = ws.cell(r, 4).value
        if epic:
            current_epic = str(epic).strip()
        if feat:
            key = f"{current_epic}::{str(feat).strip()}"
            metadata[key] = {"size": size, "tier": tier}
    wb.close()
    return metadata


def load_miro_bid_totals(path):
    bid_data = {}
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("===") or line.startswith("x=") or line.startswith("Excluded"):
                continue
            match = re.match(r"^(.+?):\s*(\d+)D\s+total\s*\|\s*(.*)$", line)
            if match:
                feat_name = match.group(1).strip()
                total = int(match.group(2))
                roles_str = match.group(3).strip()
                roles = {}
                if roles_str:
                    for part in roles_str.split(","):
                        part = part.strip()
                        rmatch = re.match(r"^(.+?)=(\d+)D$", part)
                        if rmatch:
                            roles[rmatch.group(1).strip()] = int(rmatch.group(2))
                bid_data[feat_name] = {"total": total, "roles": roles}
                continue
            match2 = re.match(r"^(.+?):\s*(\d+)D\s+total\s*\|?\s*$", line)
            if match2:
                feat_name = match2.group(1).strip()
                total = int(match2.group(2))
                bid_data[feat_name] = {"total": total, "roles": {}}
    return bid_data


def _clean(val):
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def read_design_sheet(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Sorted by Epic"]
    rows = []
    current_epic = ""
    current_feat = ""
    current_story = ""

    for r in range(3, ws.max_row + 1):
        epic = _clean(ws.cell(r, 1).value)
        feat = _clean(ws.cell(r, 2).value)
        story = _clean(ws.cell(r, 3).value)
        task = _clean(ws.cell(r, 4).value)
        team = _clean(ws.cell(r, 5).value)
        desc = _clean(ws.cell(r, 6).value)
        dos = _clean(ws.cell(r, 7).value)

        if epic:
            current_epic = EPIC_NORMALISE.get(epic, epic)
            current_story = ""  # reset story on epic change
        if feat:
            current_feat = feat
            current_story = ""  # reset story on feature change
        if story:
            current_story = story

        if not feat and not story and not task:
            continue

        # 9 estimators x 2 (min/max) + Final = 19 values
        estimates = []
        for col in range(8, 26, 2):
            estimates.append(ws.cell(r, col).value)
            estimates.append(ws.cell(r, col + 1).value)
        estimates.append(ws.cell(r, 26).value)

        # For task rows with no story, inherit the current story
        effective_story = story if story else (current_story if task else None)

        rows.append({
            "epic": current_epic,
            "feature": current_feat,
            "feature_is_header": bool(feat and not story and not task),
            "story": effective_story,
            "task": task,
            "team": team,
            "department": TEAM_TO_DEPT.get(team, "") if team else "",
            "description": desc,
            "dos": dos,
            "source": "Design",
            "design_estimates": estimates,
            "eng_estimates": [None] * 21,
        })

    wb.close()
    return rows


def read_engineering_sheet(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Sheet1"]
    rows = []
    current_epic = ""
    current_feat = ""
    current_story = ""

    for r in range(3, ws.max_row + 1):
        epic = _clean(ws.cell(r, 1).value)
        feat = _clean(ws.cell(r, 2).value)
        story = _clean(ws.cell(r, 3).value)
        task = _clean(ws.cell(r, 4).value)
        team = _clean(ws.cell(r, 5).value)
        desc = _clean(ws.cell(r, 6).value)
        dos = _clean(ws.cell(r, 7).value)

        if epic:
            current_epic = EPIC_NORMALISE.get(epic, epic)
            current_story = ""
        if feat:
            current_feat = feat
            current_story = ""
        if story:
            current_story = story

        if not feat and not story and not task:
            continue

        # 9 estimators x 2 + Mustafa adj x 2 + Notes = 21 values
        estimates = []
        for col in range(8, 26, 2):
            estimates.append(ws.cell(r, col).value)
            estimates.append(ws.cell(r, col + 1).value)
        estimates.append(ws.cell(r, 26).value)  # Mustafa min
        estimates.append(ws.cell(r, 27).value)  # Mustafa max
        notes_val = ws.cell(r, 28).value
        estimates.append(str(notes_val).strip() if notes_val else None)

        effective_story = story if story else (current_story if task else None)

        rows.append({
            "epic": current_epic,
            "feature": current_feat,
            "feature_is_header": bool(feat and not story and not task),
            "story": effective_story,
            "task": task,
            "team": team,
            "department": TEAM_TO_DEPT.get(team, "") if team else "",
            "description": desc,
            "dos": dos,
            "source": "Engineering",
            "design_estimates": [None] * 19,
            "eng_estimates": estimates,
        })

    wb.close()
    return rows


# ============================================================
# MERGE & SORT
# ============================================================

def merge_rows(design_rows, eng_rows):
    eng_lookup = {}
    for row in eng_rows:
        key = (row["epic"], row["feature"], row["story"] or "", row["task"] or "", row["team"] or "")
        eng_lookup[key] = row

    merged = []
    merged_keys = set()

    for row in design_rows:
        key = (row["epic"], row["feature"], row["story"] or "", row["task"] or "", row["team"] or "")
        if key in eng_lookup:
            eng_row = eng_lookup[key]
            row["eng_estimates"] = eng_row["eng_estimates"]
            row["source"] = "Both"
            if not row["description"] and eng_row["description"]:
                row["description"] = eng_row["description"]
            if not row["dos"] and eng_row["dos"]:
                row["dos"] = eng_row["dos"]
            merged_keys.add(key)
        merged.append(row)

    for row in eng_rows:
        key = (row["epic"], row["feature"], row["story"] or "", row["task"] or "", row["team"] or "")
        if key not in merged_keys:
            merged.append(row)

    return merged


def sort_rows(rows):
    epic_order = [
        "Player Build", "World Systems", "Combat", "User Space",
        "Items & Inventory", "Player Economy", "Quest System",
        "Social & Multiplayer", "Game Bibles", "Platform",
        "Live Game", "Partner Build", "Product Publishing",
    ]
    epic_rank = {e: i for i, e in enumerate(epic_order)}

    def sort_key(row):
        return (
            epic_rank.get(row["epic"], 99),
            row["feature"] or "",
            0 if row["feature_is_header"] else 1,
            row["story"] or "",
            row["task"] or "",
        )
    return sorted(rows, key=sort_key)


def check_not_for_vs(feature_name):
    if not feature_name:
        return False
    for nfv in NOT_FOR_VS:
        if nfv.lower() in feature_name.lower() or feature_name.lower() in nfv.lower():
            return True
    return False


# Manual mapping: Excel feature name -> Miro hours_mapping feature name
EXCEL_TO_MIRO = {
    "In-game Tutorial": "FTUE (tutorial cave segment)",
    "Melee Combat Framework": "Combat framework",
    "Combat abilities": "Combat Abilities",
    "Basic Magic system": "Magic",
    "Ranged (Corruption Gun)": "Ranged combat",
    "PVP Combat": "PvP combat",
    "Combat AI behavior": "Combat AI behavior",
    "Guilds (S)": "Guilds System",
    "Co-op - Party": "Co-op/Party",
    "Loot Distribution": "Loot Distribution",
    "In-game Mail": "In-game mail",
    "Garment System": "Garment System",
    "Skill System (1 skill tree active)": "Skill System",
    "Player Progression": "Player Progression",
    "Anywhere door placement": "Anywhere Door",
    "User Space Decoration": "User space Decoration",
    "User Space Server & Instance": "User Space Server & Instance",
    "Pets (mechanical parrot reuse)": "Mounts/ Pets",
    "Auction House economy": "Auction House",
    "In-game economy": "In game Economy",
    "Trading System": "Trading System",
    "Various Quest Types Support": "Quests : Courier, Escort, Kill, Collect, research, time obj, Partner Quest, corruption",
    "Quest Integration": "Quest Narrative Content",
    "Quest Tools": "Quests Backend/Tooling",
    "Quest Rewards System": "Quest reward system",
    "Bugged Dungeon Integration": "The Dungeon",
    "Partner Loop": "Partner Portal",
    "In game RMT Store": "In-game Store",
    "Instancing": "Instancing",
    "Dungeons mechanics": "Dungeon Mechanics",
    "Corruption": "Enemies Art",
    "Downtime Final set dressing (L)": "User space Decoration",
    "Speedup Archway": "Partner Portal",
    "Full Motion Video (FMV)": "FTUE (tutorial cave segment)",
    "Static pictures intro + voiceover": "FTUE (tutorial cave segment)",
    "In-game cutscenes": "FTUE (tutorial cave segment)",
    "World Resources": "Equipment System",
    "Navigation": "Instancing",
    "Camera behavior": "Dynamic weather",
    "World/buiding scale design": "Dynamic weather",
    "Wolves": "Enemies Art",
    "Goblins": "Enemies Art",
    "Skeletons": "Enemies Art",
    "Slimes": "Enemies Art",
    "Corrupted Guardian": "Enemies Art",
    "Carapax (Hero, but one type)": "Enemies Art",
    "Fishing": "User space Decoration",
    "User Space Island (S)": "Building Houses",
    "Consumables": "Consumables",
    "Portal Peak (Medium)": "Day/Night cycle",
    "POI: The Two Hills (S)": "Day/Night cycle",
    "POI: Portal Dimension (XS)": "Day/Night cycle",
    "POI: Final Tower": "Day/Night cycle",
    "POI: Waterfall Beauty Spot (S)": "Day/Night cycle",
    "Puzzles": "Dungeon Mechanics",
    "Advanced Traversal": "NPCs",
    "Basic movement system": "NPCs",
    "Forge": "Equipment System",
    "Main Questline: Critical Path": "Quest Narrative Content",
    "Subscription page on the website": "RMT game Economy",
    "Sync hotbar to db": "Item Core System",
    "Chat system": "Game Chat System",
    "Dungeon finder": "LFG / Pick up groups",
    "Party system": "Co-op/Party",
    "Leaderboard system": "CombatBalance & Telemetry",
    "Combat component": "Combat framework",
    "Combat director": "Combat framework",
    "Session": "Instancing",
    "Sync last known player position": "Player Progression",
    "Corrupted enemies": "Enemies Art",
    "Portal Dimension (XS)": "Day/Night cycle",
}


def match_miro_feature(feat_name, miro_bids):
    """Try to match a feature name to a Miro bid entry."""
    if not feat_name:
        return None
    # Manual mapping first
    if feat_name in EXCEL_TO_MIRO:
        mapped = EXCEL_TO_MIRO[feat_name]
        if mapped in miro_bids:
            return miro_bids[mapped]
    # Exact match
    if feat_name in miro_bids:
        return miro_bids[feat_name]
    # Case-insensitive
    lower_lookup = {k.lower(): v for k, v in miro_bids.items()}
    if feat_name.lower() in lower_lookup:
        return lower_lookup[feat_name.lower()]
    # Partial match
    for mk, mv in miro_bids.items():
        if feat_name.lower() in mk.lower() or mk.lower() in feat_name.lower():
            return mv
    return None


# ============================================================
# WRITE OUTPUT
# ============================================================

def write_output(rows, template_path, output_path, template_metadata, miro_bids):
    wb = openpyxl.load_workbook(template_path)

    if "Plan" in wb.sheetnames:
        del wb["Plan"]
    ws = wb.create_sheet("Plan", 0)

    # --- Column layout ---
    META_HEADERS = [
        "Epic", "Feature", "Story", "Task",
        "Department", "Team", "Description", "Definition of Success",
        "Size", "Current Tier", "Notes",
    ]
    meta_count = len(META_HEADERS)                             # 11
    design_start = meta_count + 1                               # col 12
    design_count = len(DESIGN_ESTIMATORS) * 2 + 1              # 19
    eng_start = design_start + design_count                     # col 31
    eng_count = len(ENG_ESTIMATORS) * 2 + 2 + 1               # 21
    miro_start = eng_start + eng_count                          # col 52
    miro_count = len(MIRO_BID_ROLES) + 1                       # 17
    total_cols = miro_start + miro_count - 1

    # ---- ROW 1: Super-headers ----
    c = ws.cell(1, 1, "METADATA")
    c.font = HEADER_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=meta_count)

    c = ws.cell(1, design_start, "DESIGN ESTIMATES (Robin Jubber — Game Director)")
    c.font = HEADER_FONT_WHITE; c.fill = DESIGN_SUPER_FILL
    ws.merge_cells(start_row=1, start_column=design_start, end_row=1, end_column=design_start + design_count - 1)

    c = ws.cell(1, eng_start, "ENGINEERING ESTIMATES (Mustafa Sibai — Head of Engineering)")
    c.font = HEADER_FONT_WHITE; c.fill = ENG_SUPER_FILL
    ws.merge_cells(start_row=1, start_column=eng_start, end_row=1, end_column=eng_start + eng_count - 1)

    c = ws.cell(1, miro_start, "MIRO BID TOTALS — Agreed Days per Role (feature level)")
    c.font = HEADER_FONT_WHITE; c.fill = MIRO_SUPER_FILL
    ws.merge_cells(start_row=1, start_column=miro_start, end_row=1, end_column=miro_start + miro_count - 1)

    # ---- ROW 2: Person/role sub-headers ----
    col = design_start
    for name, role in DESIGN_ESTIMATORS:
        c = ws.cell(2, col, f"{name}\n({role})")
        c.font = Font(bold=True, size=9); c.fill = DESIGN_SUB_FILL
        c.alignment = Alignment(wrap_text=True, horizontal="center")
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 1)
        col += 2
    c = ws.cell(2, col, "Design\nFinal")
    c.font = Font(bold=True, size=9); c.fill = DESIGN_SUB_FILL
    c.alignment = Alignment(wrap_text=True, horizontal="center")

    col = eng_start
    for name, role in ENG_ESTIMATORS:
        c = ws.cell(2, col, f"{name}\n({role})")
        c.font = Font(bold=True, size=9); c.fill = ENG_SUB_FILL
        c.alignment = Alignment(wrap_text=True, horizontal="center")
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 1)
        col += 2
    c = ws.cell(2, col, "Mustafa Sibai\n(Head of Eng — Adjusted)")
    c.font = Font(bold=True, size=9); c.fill = ENG_SUB_FILL
    c.alignment = Alignment(wrap_text=True, horizontal="center")
    ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 1)
    col += 2
    c = ws.cell(2, col, "Eng\nNotes")
    c.font = Font(bold=True, size=9); c.fill = ENG_SUB_FILL
    c.alignment = Alignment(wrap_text=True, horizontal="center")

    # Miro row 2: role names (rotated)
    for i, role_name in enumerate(MIRO_BID_ROLES):
        c = ws.cell(2, miro_start + i, role_name)
        c.font = Font(bold=True, size=9); c.fill = MIRO_SUB_FILL
        c.alignment = Alignment(wrap_text=True, horizontal="center", text_rotation=90)
    c = ws.cell(2, miro_start + len(MIRO_BID_ROLES), "Bid\nTotal")
    c.font = Font(bold=True, size=9); c.fill = MIRO_SUB_FILL
    c.alignment = Alignment(wrap_text=True, horizontal="center")

    # ---- ROW 3: Column labels (Min/Max/Days/etc.) ----
    for i, h in enumerate(META_HEADERS, 1):
        c = ws.cell(3, i, h)
        c.font = HEADER_FONT_WHITE; c.fill = HEADER_FILL_BLUE; c.border = THIN_BORDER; c.alignment = WRAP

    col = design_start
    for _ in DESIGN_ESTIMATORS:
        for label in ("Min", "Max"):
            c = ws.cell(3, col, label)
            c.font = HEADER_FONT_WHITE; c.fill = HEADER_FILL_BLUE; c.border = THIN_BORDER
            col += 1
    c = ws.cell(3, col, "Final")
    c.font = HEADER_FONT_WHITE; c.fill = HEADER_FILL_BLUE; c.border = THIN_BORDER

    col = eng_start
    for _ in ENG_ESTIMATORS:
        for label in ("Min", "Max"):
            c = ws.cell(3, col, label)
            c.font = HEADER_FONT_WHITE; c.fill = ENG_SUPER_FILL; c.border = THIN_BORDER
            col += 1
    for label in ("Adj Min", "Adj Max"):
        c = ws.cell(3, col, label)
        c.font = HEADER_FONT_WHITE; c.fill = ENG_SUPER_FILL; c.border = THIN_BORDER
        col += 1
    c = ws.cell(3, col, "Notes")
    c.font = HEADER_FONT_WHITE; c.fill = ENG_SUPER_FILL; c.border = THIN_BORDER

    for i, role_name in enumerate(MIRO_BID_ROLES):
        c = ws.cell(3, miro_start + i, "Days")
        c.font = HEADER_FONT_WHITE; c.fill = MIRO_SUPER_FILL; c.border = THIN_BORDER
    c = ws.cell(3, miro_start + len(MIRO_BID_ROLES), "Days")
    c.font = HEADER_FONT_WHITE; c.fill = MIRO_SUPER_FILL; c.border = THIN_BORDER

    # ---- DATA ROWS (starting row 4) ----
    data_row = 4
    prev_epic = ""
    prev_feat = ""
    miro_features_matched = set()

    stats = {
        "total_rows": 0, "design_only": 0, "eng_only": 0, "both": 0,
        "with_estimates": 0, "without_estimates": 0, "not_for_vs": 0,
        "miro_matched": 0, "miro_unmatched": 0,
        "epics": set(), "features": set(),
    }

    for row in rows:
        epic = row["epic"]
        feat = row["feature"]
        story = row["story"]
        task = row["task"]

        is_epic_change = (epic != prev_epic)
        is_feat_change = (feat != prev_feat)

        # Epic name on first occurrence
        if is_epic_change:
            c = ws.cell(data_row, 1, epic)
            c.font = EPIC_FONT; c.fill = EPIC_FILL
            stats["epics"].add(epic)
            prev_epic = epic

        # Feature name on first occurrence
        if is_feat_change:
            c = ws.cell(data_row, 2, feat)
            c.font = FEAT_FONT; c.fill = FEAT_FILL

            meta_key = f"{epic}::{feat}"
            if meta_key in template_metadata:
                ws.cell(data_row, 9, template_metadata[meta_key]["size"])
                ws.cell(data_row, 10, template_metadata[meta_key]["tier"])

            # Write Miro bid totals at feature level
            miro_match = match_miro_feature(feat, miro_bids)
            if miro_match:
                # Track the MIRO feature name that was consumed
                if feat in EXCEL_TO_MIRO:
                    miro_features_matched.add(EXCEL_TO_MIRO[feat])
                miro_features_matched.add(feat)  # also track exact matches
                # Check all miro keys for partial matches too
                for mk in miro_bids:
                    if feat.lower() in mk.lower() or mk.lower() in feat.lower():
                        miro_features_matched.add(mk)
                stats["miro_matched"] += 1
                for i, role_name in enumerate(MIRO_BID_ROLES):
                    days = miro_match["roles"].get(role_name)
                    if days:
                        ws.cell(data_row, miro_start + i, days)
                ws.cell(data_row, miro_start + len(MIRO_BID_ROLES), miro_match["total"])

            stats["features"].add(f"{epic}::{feat}")
            prev_feat = feat

        # NOT FOR VS
        is_nfv = check_not_for_vs(feat)
        if is_nfv:
            stats["not_for_vs"] += 1
            ws.cell(data_row, 11, "NOT FOR VS")

        # Story / Task
        if story:
            c = ws.cell(data_row, 3, story)
            c.fill = STORY_FILL
        if task:
            ws.cell(data_row, 4, task)

        # Department, Team, Description, DoS
        if row["department"]:
            ws.cell(data_row, 5, row["department"])
        if row["team"]:
            ws.cell(data_row, 6, row["team"])
        if row["description"]:
            ws.cell(data_row, 7, row["description"]).alignment = WRAP
        if row["dos"]:
            ws.cell(data_row, 8, row["dos"]).alignment = WRAP

        # Design estimates
        col = design_start
        has_design = False
        for val in row["design_estimates"]:
            if val is not None:
                ws.cell(data_row, col, val)
                has_design = True
            col += 1

        # Engineering estimates
        col = eng_start
        has_eng = False
        for i, val in enumerate(row["eng_estimates"]):
            if val is not None:
                cell = ws.cell(data_row, col, val)
                if i == len(row["eng_estimates"]) - 1 and isinstance(val, str):
                    cell.alignment = WRAP
                has_eng = True
            col += 1

        # Stats
        stats["total_rows"] += 1
        if row["source"] == "Both":
            stats["both"] += 1
        elif row["source"] == "Design":
            stats["design_only"] += 1
        else:
            stats["eng_only"] += 1
        if has_design or has_eng:
            stats["with_estimates"] += 1
        else:
            stats["without_estimates"] += 1

        # Grey out NOT FOR VS
        if is_nfv:
            for c_idx in range(1, total_cols + 1):
                ws.cell(data_row, c_idx).fill = NFV_FILL

        # Borders
        for c_idx in range(1, total_cols + 1):
            ws.cell(data_row, c_idx).border = THIN_BORDER

        data_row += 1

    # --- Add Miro-only features (not in Excel estimation sheets) ---
    MIRO_ONLY_EPIC_MAP = {
        "Achievements & Trophies": "Player Build",
        "Banks": "Items & Inventory",
        "Character Creation & Customization": "Player Build",
        "Co-op/Raids": "Social & Multiplayer",
        "Combat AI behavior": "Combat",
        "Friends": "Social & Multiplayer",
        "Garment Shop": "Player Economy",
        "Guild House": "Social & Multiplayer",
        "Inventory System": "Items & Inventory",
        "Monetization Bible": "Game Bibles",
        "Notifications": "Social & Multiplayer",
        "P2P Item Transfer": "Items & Inventory",
        "P2P/P2G User Space Connection": "User Space",
        "Partner Shop": "Player Economy",
        "Player Account Authentification": "Platform",
        "Player Account Creation/Deletion": "Platform",
        "Player Account Privacy Settings & Legal": "Platform",
        "Professions": "Player Build",
        "Quest Item": "Quest System",
        "Quest Multiplayer": "Quest System",
        "Quest Tracking": "Platform",
        "Quest dialog GUI": "Quest System",
        "Quest log": "Quest System",
        "Quests as Events": "Quest System",
        "Red Books (emergency guidelines for Live Service)": "Game Bibles",
        "Saving House Layouts": "User Space",
        "Social Media Support": "Platform",
        "Targeted VO": "Quest System",
        "Virtual Currency (VC) System": "Player Economy",
        "[Developer Dashboard] Developer Account": "Platform",
        "[Developer Dashboard] Game Dashboard": "Platform",
        "[Player Dashboard] Friends": "Platform",
        "[Player Dashboard] Game Dashboard": "Platform",
        "[Player Dashboard] Marketplace": "Platform",
        "[Player Dashboard] Player Account Features": "Platform",
        "[Player Dashboard] Search & Notifications": "Platform",
    }

    # Find unmatched Miro features
    unmatched_miro = []
    for miro_feat in miro_bids:
        if miro_feat not in miro_features_matched:
            matched = False
            for mf in miro_features_matched:
                if miro_feat.lower() in mf.lower() or mf.lower() in miro_feat.lower():
                    matched = True
                    break
            if not matched:
                unmatched_miro.append(miro_feat)

    # Write Miro-only feature rows
    miro_only_fill = PatternFill(start_color="E8DAEF", end_color="E8DAEF", fill_type="solid")
    miro_added = 0
    for miro_feat in sorted(unmatched_miro):
        bid = miro_bids[miro_feat]
        if bid["total"] == 0:
            continue  # Skip 0D features
        epic = MIRO_ONLY_EPIC_MAP.get(miro_feat, "")
        if not epic:
            continue

        # Write epic if changed
        if epic != prev_epic:
            ws.cell(data_row, 1, epic).font = EPIC_FONT
            ws.cell(data_row, 1).fill = EPIC_FILL
            prev_epic = epic

        ws.cell(data_row, 2, miro_feat).font = FEAT_FONT
        ws.cell(data_row, 2).fill = miro_only_fill
        ws.cell(data_row, 11, "MIRO BID ONLY (not in VS estimation sheets)")

        # Write Miro bid columns
        for i, role_name in enumerate(MIRO_BID_ROLES):
            days = bid["roles"].get(role_name)
            if days:
                ws.cell(data_row, miro_start + i, days)
        ws.cell(data_row, miro_start + len(MIRO_BID_ROLES), bid["total"])

        # Check NOT FOR VS
        if check_not_for_vs(miro_feat):
            ws.cell(data_row, 11, "NOT FOR VS (MIRO BID ONLY)")
            for c_idx in range(1, total_cols + 1):
                ws.cell(data_row, c_idx).fill = NFV_FILL

        # Borders
        for c_idx in range(1, total_cols + 1):
            ws.cell(data_row, c_idx).border = THIN_BORDER

        # Template metadata
        for meta_key_try in [f"{epic}::{miro_feat}"]:
            if meta_key_try in template_metadata:
                ws.cell(data_row, 9, template_metadata[meta_key_try]["size"])
                ws.cell(data_row, 10, template_metadata[meta_key_try]["tier"])
                break

        data_row += 1
        miro_added += 1
        stats["total_rows"] += 1

    stats["miro_only_added"] = miro_added
    stats["miro_unmatched"] = len(unmatched_miro)
    stats["miro_unmatched_list"] = unmatched_miro

    # --- Column widths ---
    widths = {1: 20, 2: 35, 3: 40, 4: 45, 5: 14, 6: 22, 7: 50, 8: 50, 9: 8, 10: 18, 11: 15}
    for col_num, w in widths.items():
        ws.column_dimensions[get_column_letter(col_num)].width = w
    for c_idx in range(design_start, design_start + design_count):
        ws.column_dimensions[get_column_letter(c_idx)].width = 7
    ws.column_dimensions[get_column_letter(design_start + design_count - 1)].width = 8
    for c_idx in range(eng_start, eng_start + eng_count - 1):
        ws.column_dimensions[get_column_letter(c_idx)].width = 7
    ws.column_dimensions[get_column_letter(eng_start + eng_count - 1)].width = 40
    for c_idx in range(miro_start, miro_start + miro_count):
        ws.column_dimensions[get_column_letter(c_idx)].width = 7
    ws.column_dimensions[get_column_letter(miro_start + miro_count - 1)].width = 9

    # Row 2 height for rotated headers
    ws.row_dimensions[2].height = 120

    # Freeze first 3 header rows + first 4 metadata columns
    ws.freeze_panes = "E4"

    wb.save(output_path)
    wb.close()
    return stats


# ============================================================
# MAIN
# ============================================================

def main():
    print("=" * 60)
    print("COUCH HEROES PRODUCTION PLAN CONSOLIDATION")
    print("=" * 60)

    print("\n1. Loading template metadata...")
    template_meta = load_template_metadata(TEMPLATE_FILE)
    print(f"   {len(template_meta)} features with Size/Tier")

    print("\n2. Loading Miro bid totals...")
    miro_bids = load_miro_bid_totals(HOURS_MAPPING_FILE)
    print(f"   {len(miro_bids)} features with role-level day estimates")

    print("\n3. Reading Design estimation sheet...")
    design_rows = read_design_sheet(DESIGN_FILE)
    print(f"   {len(design_rows)} rows")

    print("\n4. Reading Engineering estimation sheet...")
    eng_rows = read_engineering_sheet(ENGINEERING_FILE)
    print(f"   {len(eng_rows)} rows")

    print("\n5. Merging...")
    merged = merge_rows(design_rows, eng_rows)
    dupes = len(design_rows) + len(eng_rows) - len(merged)
    print(f"   {len(merged)} rows ({dupes} duplicates merged)")

    print("\n6. Sorting...")
    sorted_rows = sort_rows(merged)

    print(f"\n7. Writing {os.path.basename(OUTPUT_FILE)}...")
    stats = write_output(sorted_rows, TEMPLATE_FILE, OUTPUT_FILE, template_meta, miro_bids)

    print("\n" + "=" * 60)
    print("CONSOLIDATION COMPLETE")
    print("=" * 60)
    print(f"Output:              {OUTPUT_FILE}")
    print(f"Total data rows:     {stats['total_rows']}")
    print(f"Unique epics:        {len(stats['epics'])}")
    print(f"Unique features:     {len(stats['features'])}")
    print(f"Design-only rows:    {stats['design_only']}")
    print(f"Engineering-only:    {stats['eng_only']}")
    print(f"Both sources:        {stats['both']}")
    print(f"With estimates:      {stats['with_estimates']}")
    print(f"Without estimates:   {stats['without_estimates']}")
    print(f"NOT FOR VS items:    {stats['not_for_vs']}")
    print(f"Miro bids matched:   {stats['miro_matched']}")
    print(f"Miro-only rows added:{stats.get('miro_only_added', 0)}")
    print(f"Miro bids unmatched: {stats['miro_unmatched']}")

    if stats.get("miro_unmatched_list"):
        print("\nUnmatched Miro features (no Excel equivalent):")
        for feat in stats["miro_unmatched_list"]:
            bid = miro_bids[feat]
            print(f"   {feat}: {bid['total']}D")

    # Save stats
    stats_out = {k: (sorted(list(v)) if isinstance(v, set) else v) for k, v in stats.items()}
    stats_path = os.path.join(BASE, "consolidation_stats.json")
    with open(stats_path, "w") as f:
        json.dump(stats_out, f, indent=2)
    print(f"\nStats: {stats_path}")


if __name__ == "__main__":
    main()
