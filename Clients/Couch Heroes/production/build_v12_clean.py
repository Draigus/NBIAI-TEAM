#!/usr/bin/env python3
"""
Build CouchHeroes_Man_Day_Work_Plan_v12.xlsx from:
  - v11 template (feature order, formatting reference)
  - VS Estimation Design sheet (design estimator min/max + Final)
  - VS Estimation Engineering sheet (engineering estimator min/max + Mustafa Adjusted)
  - hours_mapping.txt (Miro board feature-level role totals)

Output structure (matching Glen's screenshot):
  Cols A-G: Epic | Feature | Story | Team | Size | Current Tier | Notes
  Then phase columns with Early Prod having Min/Max per role.
"""

import json
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy

# ============================================================
# CONSTANTS
# ============================================================

ROLES = [
    'CEO', 'Creative Director', 'Head of Legal', 'COO', 'HR Manager',
    'Executive Producer', 'Producer', 'Game Producer', 'Tech Producer', 'Art Producer',
    'Game Director', 'Game Design Lead', 'Level Design Lead', 'Game Designer',
    'Combat Designer', 'Level Designer', 'Technical Designer', 'Writer',
    'Art Director', 'Concept Artist', 'UI/UX Designer', 'Graphic Designer',
    'Character Artist', 'Environment Artist', 'Prop Artist', 'Animator',
    'Character Rigger', 'VFX Artist', 'Technical Artist',
    'CTO', 'Tech Lead', 'Gameplay Developer', 'Full Stack Developer',
    'Backend Developer', 'Network Engineer', 'QA', 'Sound Designer', 'Platform Team'
]
NUM_ROLES = len(ROLES)  # 38

ROLE_INDEX = {name: i for i, name in enumerate(ROLES)}

# Discipline group fills for role headers (matching template exactly)
ROLE_FILLS = {}
_grey = PatternFill(start_color='FFD9D9D9', end_color='FFD9D9D9', fill_type='solid')
_yellow_prod = PatternFill(start_color='FFFFD966', end_color='FFFFD966', fill_type='solid')
_blue_design = PatternFill(start_color='FFBDD7EE', end_color='FFBDD7EE', fill_type='solid')
_peach_art = PatternFill(start_color='FFF8CBAD', end_color='FFF8CBAD', fill_type='solid')
_green_eng = PatternFill(start_color='FFC6E0B4', end_color='FFC6E0B4', fill_type='solid')
_purple_qa = PatternFill(start_color='FFE4DFEC', end_color='FFE4DFEC', fill_type='solid')
_yellow_sound = PatternFill(start_color='FFFFE699', end_color='FFFFE699', fill_type='solid')
_blue_platform = PatternFill(start_color='FFB4C7E7', end_color='FFB4C7E7', fill_type='solid')

for r in ['CEO', 'Creative Director', 'Head of Legal', 'COO', 'HR Manager']:
    ROLE_FILLS[r] = _grey
for r in ['Executive Producer', 'Producer', 'Game Producer', 'Tech Producer', 'Art Producer']:
    ROLE_FILLS[r] = _yellow_prod
for r in ['Game Director', 'Game Design Lead', 'Level Design Lead', 'Game Designer',
          'Combat Designer', 'Level Designer', 'Technical Designer', 'Writer']:
    ROLE_FILLS[r] = _blue_design
for r in ['Art Director', 'Concept Artist', 'UI/UX Designer', 'Graphic Designer',
          'Character Artist', 'Environment Artist', 'Prop Artist', 'Animator',
          'Character Rigger', 'VFX Artist', 'Technical Artist']:
    ROLE_FILLS[r] = _peach_art
for r in ['CTO', 'Tech Lead', 'Gameplay Developer', 'Full Stack Developer',
          'Backend Developer', 'Network Engineer']:
    ROLE_FILLS[r] = _green_eng
ROLE_FILLS['QA'] = _purple_qa
ROLE_FILLS['Sound Designer'] = _yellow_sound
ROLE_FILLS['Platform Team'] = _blue_platform

# Phase definitions
PHASES = [
    {'name': 'CONCEPT', 'gate_num': 0, 'sub_label': 'CONCEPT\n~2 months\nTarget: T0 Ideation',
     'super': 'CONCEPT', 'super_fill': 'FFBF8F00', 'min_max': False},
    {'name': 'PRE-PRODUCTION', 'gate_num': 1, 'sub_label': 'PRE-PRODUCTION\nTarget: T1-T2 R&D / GDD-TDD-Brief',
     'super': 'PRE-PRODUCTION', 'super_fill': 'FFC65911', 'min_max': False},
    {'name': 'EARLY PROD', 'gate_num': 2, 'sub_label': 'EARLY PROD\nTarget: T3 Prototype',
     'super': 'PRODUCTION', 'super_fill': 'FF2E75B6', 'min_max': True},
    {'name': 'MID PROD', 'gate_num': 3, 'sub_label': 'MID PRO\nTarget: T4 MVP',
     'super': 'PRODUCTION', 'super_fill': 'FF2E75B6', 'min_max': False},
    {'name': 'END PROD', 'gate_num': 4, 'sub_label': 'END PRO\nTarget: T5 Feature Complete',
     'super': 'PRODUCTION', 'super_fill': 'FF2E75B6', 'min_max': False},
    {'name': 'ALPHA', 'gate_num': 5, 'sub_label': 'ALPHA\nTarget: T6 Playtest',
     'super': 'RELEASE', 'super_fill': 'FF548235', 'min_max': False},
    {'name': 'BETA', 'gate_num': 6, 'sub_label': 'BETA\nTarget: T6 Playtest',
     'super': 'RELEASE', 'super_fill': 'FF548235', 'min_max': False},
    {'name': 'LAUNCH', 'gate_num': 7, 'sub_label': 'LAUNCH\nTarget: T6',
     'super': 'RELEASE', 'super_fill': 'FF548235', 'min_max': False},
    {'name': 'FIRST RELEASE', 'gate_num': 8, 'sub_label': 'FIRST RELEASE\nTarget: T6-T7',
     'super': 'RELEASE', 'super_fill': 'FF548235', 'min_max': False},
    {'name': 'LS 1', 'gate_num': 9, 'sub_label': 'LS 1\nLive\nT7',
     'super': 'LIVE SERVICE', 'super_fill': 'FF7030A0', 'min_max': False},
    {'name': 'LS 2', 'gate_num': 10, 'sub_label': 'LS 2\nLive\nT8',
     'super': 'LIVE SERVICE', 'super_fill': 'FF7030A0', 'min_max': False},
]

# Team -> Role mapping
TEAM_TO_ROLE = {
    'Game design': 'Game Designer',
    'Level design': 'Level Designer',
    'Gameplay Engineering': 'Gameplay Developer',
    'Backend Engineering': 'Backend Developer',
    'UX/UI': 'UI/UX Designer',
    'Concept Art': 'Concept Artist',
    'Environment Art': 'Environment Artist',
    'Character Art': 'Character Artist',
    'Animation & Rigging': 'Animator',
    'VFX': 'VFX Artist',
    'TechArt': 'Technical Artist',
    'Audio': 'Sound Designer',
    'QA': 'QA',
    'Narrative': 'Writer',
}

# Team -> Department mapping
TEAM_TO_DEPT = {
    'Game design': 'Design',
    'Level design': 'Design',
    'Narrative': 'Design',
    'UX/UI': 'Art',
    'Concept Art': 'Art',
    'Environment Art': 'Art',
    'Character Art': 'Art',
    'Animation & Rigging': 'Art',
    'VFX': 'Art',
    'TechArt': 'Art',
    'Gameplay Engineering': 'Engineering',
    'Backend Engineering': 'Engineering',
    'QA': 'QA',
    'Audio': 'Audio',
}

# Epic name normalisation (source -> template)
EPIC_NORM = {
    'Combat System': 'Combat',
    'Social': 'Social & Multiplayer',
    'Rename to in-game camera squancer during dialog': 'World Systems',
    'Phasing system': 'World Systems',
    'Items & Inventory': 'Items & Inventory',
    'Live Service Support': 'Live Game',
}

# Feature name normalisation (source -> template)
FEATURE_NORM = {
    # Player Build
    'In-game Tutorial': 'FTUE',
    'Skill System (1 skill tree active)': 'Skill System',
    # World Systems
    'Basic movement system': 'Basic Movement System',
    'In-game cutscenes': 'In Game Cinematic (IGC)',
    'Static pictures intro + voiceover': 'Full Motion Video (FMV)',
    'Full Motion Video (FMV)': 'Full Motion Video (FMV)',
    'Dungeons mechanics': 'Dungeon Mechanics',
    'Corruption': 'Foam Corruption',
    'Camera behavior': 'In Game Cinematic (IGC)',
    'Navigation': 'Navigation',
    'Instancing': 'Instancing',
    'Advanced Traversal': 'Advanced Traversal',
    'Portal Peak (Medium)': 'Biome - Portal Peak',
    'Downtime Final set dressing (L)': 'Biome - Downtime',
    'POI: Final Tower': 'Biome - Portal Peak',
    'POI: Portal Dimension (XS)': 'Biome - Portal Peak',
    'POI: The Two Hills (S)': 'Biome - Hidden Hills',
    'POI: Waterfall Beauty Spot (S)': 'Biome - Whispering Woods',
    'Portal Dimension (XS)': 'Biome - Portal Peak',
    'World/buiding scale design': 'Biome - Portal Peak',
    'Speedup Archway': 'Advanced Traversal',
    'Forge': 'Professions',
    'Puzzles': 'Dungeon Mechanics',
    'World Resources': 'Item Core System',
    'Session': 'Instancing',
    'Create a system to allow a player to interact with the world and have event only happen for that player in a multiplayer setting': 'Instancing',
    'Sync last known player position': 'Online Services & Backend',
    # Combat
    'Melee Combat Framework': 'Combat Framework',
    'Combat abilities': 'Combat Abilities',
    'Combat component': 'Combat Framework',
    'Combat director': 'Combat Framework',
    'Basic Magic system': 'Magic',
    'PVP Combat': 'PvP Combat',
    'Ranged (Corruption Gun)': 'Ranged Combat',
    'Carapax (Hero, but one type)': 'Enemies Art',
    'Corrupted Guardian': 'Enemies Art',
    'Corrupted enemies': 'Enemies Art',
    'Wolves': 'Enemies Art',
    'Goblins': 'Enemies Art',
    'Slimes': 'Enemies Art',
    'Skeletons': 'Enemies Art',
    # User Space
    'Anywhere door placement': 'Anywhere Door Placement',
    'User Space Island (S)': 'User Space Server & Instance',
    'Fishing': 'Professions',
    'Pets (mechanical parrot reuse)': 'Mounts / Pets',
    # Items & Inventory
    'Consumables': 'Consumables',
    'Sync hotbar to db': 'Equipment System',
    # Player Economy
    'Auction House economy': 'Auction House',
    'In-game economy': 'In-game Economy',
    'In game RMT Store': 'In-game Store',
    'Virtual currency balance and items': 'Currency System',
    # Quest System
    'Various Quest Types Support': 'Quest Types - Courier / Escort / Kill / Collect / Mystery / Time / Partner / Lockpicking',
    'Quest Integration': 'Quests Backend / Tooling',
    'Quest Rewards System': 'Quest Reward System',
    'Quest Tools': 'Quests Backend / Tooling',
    'Main Questline: Critical Path': 'Quest Narrative Content',
    # Social & Multiplayer
    'Co-op - Party': 'Co-op / Party',
    'Guilds (S)': 'Guilds',
    'Chat system': 'Game Chat System',
    'Party system': 'Co-op / Party',
    'Dungeon finder': 'LFG / Pick-up Groups',
    'Leaderboard system': 'Combat Balance & Telemetry',
    # Platform / Live
    'Bugged Dungeon Integration': 'Dungeon Mechanics',
    'Subscription page on the website': 'Website',
    'Partner Portal': 'Partner Portal',
    'Partner Loop': 'Partner Loop',
}

# Miro feature name mapping (hours_mapping.txt -> template)
MIRO_TO_TEMPLATE = {
    'FTUE (tutorial cave segment)': 'FTUE',
    'Combat framework': 'Combat Framework',
    'Combat AI behavior': 'Combat AI Behavior',
    'CombatBalance & Telemetry': 'Combat Balance & Telemetry',
    'Ranged combat': 'Ranged Combat',
    'PvP combat': 'PvP Combat',
    'Day/Night cycle': 'Day/Night Cycle',
    'Dynamic weather': 'Dynamic Weather',
    'Dungeon Mechanics': 'Dungeon Mechanics',
    'Enemies Art': 'Enemies Art',
    'Virtual Currency (VC) System': 'Currency System',
    'Guilds System': 'Guilds',
    'Co-op/Party': 'Co-op / Party',
    'Co-op/Raids': 'Co-op / Raids',
    'In-game mail': 'In-game Mail',
    'LFG / Pick up groups': 'LFG / Pick-up Groups',
    'Quests Backend/Tooling': 'Quests Backend / Tooling',
    'Quests as Events': 'Quests as Events',
    'Quest dialog GUI': 'Quest Dialog GUI',
    'Quest log': 'Quest Log',
    'Quest reward system': 'Quest Reward System',
    'Quest Multiplayer': 'Quest Multiplayer',
    'Quest Narrative Content': 'Quest Narrative Content',
    'Quests : Courier, Escort, Kill, Collect, research, time obj, Partner Quest, corruption':
        'Quest Types - Courier / Escort / Kill / Collect / Mystery / Time / Partner / Lockpicking',
    'Targeted VO': 'Targeted VO',
    'In game Economy': 'In-game Economy',
    'In-game Store': 'In-game Store',
    'RMT game Economy': 'RMT Game Economy',
    'Auction House': 'Auction House',
    'Trading System': 'Trading System',
    'Banks': 'Banks',
    'Consumables': 'Consumables',
    'Equipment System': 'Equipment System',
    'Inventory System': 'Inventory System',
    'Item Core System': 'Item Core System',
    'Loot Distribution': 'Loot Distribution',
    'P2P Item Transfer': 'P2P Item Transfer',
    'Mounts/ Pets': 'Mounts / Pets',
    'User Space Server & Instance': 'User Space Server & Instance',
    'User space Decoration': 'User Space Decoration',
    'P2P/P2G User Space Connection': 'P2P / P2G User Space Connection',
    'Building Houses': 'Building Houses',
    'Saving House Layouts': 'Saving House Layouts',
    'Anywhere Door': 'Anywhere Door Placement',
    'Partner Portal': 'Partner Portal',
    'Player Progression': 'Player Progression',
    'Skill System': 'Skill System',
    'Achievements & Trophies': 'Achievements & Trophies',
    'Character Creation & Customization': 'Character Creation & Customization',
    'Garment System': 'Garment System',
    'Friends': 'Friends',
    'Game Chat System': 'Game Chat System',
    'Notifications': 'Notifications',
    'Professions': 'Professions',
    'Monetization Bible': 'Monetization Bible',
    'Red Books (emergency guidelines for Live Service)': 'Red Books (Live Service emergency guidelines)',
    'The Dungeon': 'The Dungeon',
    'Instancing': 'Instancing',
    'Magic': 'Magic',
    'Combat Abilities': 'Combat Abilities',
    'Player Account Authentification': 'Player Account Authentication',
    'Player Account Creation/Deletion': 'Player Account Creation / Deletion',
    'Player Account Privacy Settings & Legal': 'Player Account Privacy Settings & Legal',
    'Social Media Support': 'Social Media Support',
    'Quest Tracking': 'Quest Tracking',
    '[Player Dashboard] Friends': 'Player Dashboard - Friends',
    '[Player Dashboard] Game Dashboard': 'Player Dashboard - Game Dashboard',
    '[Player Dashboard] Marketplace': 'Player Dashboard - Marketplace',
    '[Player Dashboard] Player Account Features': 'Player Dashboard - Account Features',
    '[Player Dashboard] Search & Notifications': 'Player Dashboard - Search & Notifications',
    '[Developer Dashboard] Developer Account': 'Developer Dashboard - Developer Account',
    '[Developer Dashboard] Game Dashboard': 'Developer Dashboard - Game Dashboard',
    'Garment Shop': 'In-game Store',
    'Partner Shop': 'In-game Store',
    'Guild House': 'Guilds',
    'NPCs': 'Enemies Art',
    'Quest Item': 'Item Core System',
    'For VS we need LG planned out and prototyped by the end of Early prod (tech, strategy etc)': None,
}

# Design estimator names (in sheet column order)
DESIGN_ESTIMATORS = [
    {'name': 'Yorgos Dritsas', 'short': 'Yorgos', 'role': 'Junior Game Designer', 'team': 'Game Design Team', 'dept': 'Design', 'min_col': 8, 'max_col': 9},
    {'name': 'Nadir Latif', 'short': 'Nadir', 'role': 'Combat Designer', 'team': 'Game Design Team', 'dept': 'Design', 'min_col': 10, 'max_col': 11},
    {'name': 'Kieron Naylor', 'short': 'Kieron', 'role': 'Technical Designer', 'team': 'Design', 'dept': 'Design', 'min_col': 12, 'max_col': 13},
    {'name': 'Panos Andriopoulos', 'short': 'Panos', 'role': 'Junior Level Designer', 'team': 'Level Design Team', 'dept': 'Design', 'min_col': 14, 'max_col': 15},
    {'name': 'Nikolas Tziotis', 'short': 'Nikolas', 'role': 'Junior Level Designer', 'team': 'Level Design Team', 'dept': 'Design', 'min_col': 16, 'max_col': 17},
    {'name': 'Seth Dahl', 'short': 'Seth', 'role': 'Junior Level Designer', 'team': 'Level Design Team', 'dept': 'Design', 'min_col': 18, 'max_col': 19},
    {'name': 'Maria Cibej', 'short': 'Maria', 'role': 'Junior Writer', 'team': 'Writing Team', 'dept': 'Design', 'min_col': 20, 'max_col': 21},
    {'name': 'Hannah Pickard', 'short': 'Hannah', 'role': 'QA', 'team': 'QA Team', 'dept': 'QA', 'min_col': 22, 'max_col': 23},
    {'name': 'Aggeliki Peroutsea', 'short': 'Aggeliki', 'role': 'Junior Sound Designer', 'team': 'Sound Design Team', 'dept': 'Audio', 'min_col': 24, 'max_col': 25},
]

# Engineering estimator names
ENG_ESTIMATORS = [
    {'name': 'Samer Balushi', 'short': 'Samer', 'role': 'Backend Developer', 'team': 'Backend Engineering Team', 'dept': 'Engineering', 'min_col': 8, 'max_col': 9},
    {'name': 'Nikos Gerontakis', 'short': 'Nikos', 'role': 'Senior Full Stack Developer', 'team': 'Backend Engineering Team', 'dept': 'Engineering', 'min_col': 10, 'max_col': 11},
    {'name': 'Daniel Negri', 'short': 'Daniel', 'role': 'Senior Full Stack Developer', 'team': 'Backend Engineering Team', 'dept': 'Engineering', 'min_col': 12, 'max_col': 13},
    {'name': 'Ilya Achmetov', 'short': 'Ilya', 'role': 'Gameplay Developer', 'team': 'Gameplay Team', 'dept': 'Engineering', 'min_col': 14, 'max_col': 15},
    {'name': 'Raynor D\'Souza', 'short': 'Raynor', 'role': 'Senior Gameplay Developer', 'team': 'Gameplay Team', 'dept': 'Engineering', 'min_col': 16, 'max_col': 17},
    {'name': 'Leon Huang', 'short': 'Leon', 'role': 'Gameplay Developer', 'team': 'Gameplay Team', 'dept': 'Engineering', 'min_col': 18, 'max_col': 19},
    {'name': 'Ignasi Ezpeleta', 'short': 'Ignasi', 'role': 'Junior Gameplay Developer', 'team': 'Gameplay Team', 'dept': 'Engineering', 'min_col': 20, 'max_col': 21},
    {'name': 'Roberto', 'short': 'Roberto', 'role': 'Gameplay Developer', 'team': 'Gameplay Team', 'dept': 'Engineering', 'min_col': 22, 'max_col': 23},
    {'name': 'Matt', 'short': 'Matt', 'role': 'Technical Animation Engineer', 'team': 'Gameplay Team', 'dept': 'Engineering', 'min_col': 24, 'max_col': 25},
]

# ============================================================
# COLUMN LAYOUT COMPUTATION
# ============================================================

METADATA_COLS = 7  # A-G

def compute_column_layout():
    """Compute start columns for each phase."""
    col = METADATA_COLS + 1  # Start after metadata (col 8)
    layout = {}
    for phase in PHASES:
        gate_col = col
        col += 1
        role_start = col
        if phase['min_max']:
            role_count = NUM_ROLES * 2
        else:
            role_count = NUM_ROLES
        role_end = col + role_count - 1
        col = role_end + 1
        layout[phase['name']] = {
            'gate_col': gate_col,
            'role_start': role_start,
            'role_end': role_end,
            'min_max': phase['min_max'],
        }
    layout['_total_cols'] = col - 1
    return layout

COL_LAYOUT = compute_column_layout()

def ep_role_col(role_name, is_min=True):
    """Get the column number for a role's Min or Max in Early Prod."""
    idx = ROLE_INDEX.get(role_name)
    if idx is None:
        return None
    ep = COL_LAYOUT['EARLY PROD']
    base = ep['role_start'] + idx * 2
    return base if is_min else base + 1


# ============================================================
# STYLING
# ============================================================

FILL_DARK_GREY = PatternFill(start_color='FF404040', end_color='FF404040', fill_type='solid')
FILL_MED_GREY = PatternFill(start_color='FF595959', end_color='FF595959', fill_type='solid')
FILL_GATE = PatternFill(start_color='FF1F4E78', end_color='FF1F4E78', fill_type='solid')
FILL_SUBPHASE = PatternFill(start_color='FF305496', end_color='FF305496', fill_type='solid')
FILL_EPIC = PatternFill(start_color='FFFFD966', end_color='FFFFD966', fill_type='solid')
FILL_MINMAX = PatternFill(start_color='FF305496', end_color='FF305496', fill_type='solid')

FONT_WHITE_BOLD_12 = Font(name='Calibri', size=12, bold=True, color='FFFFFFFF')
FONT_WHITE_BOLD_10 = Font(name='Calibri', size=10, bold=True, color='FFFFFFFF')
FONT_WHITE_BOLD_9 = Font(name='Calibri', size=9, bold=True, color='FFFFFFFF')
FONT_BLACK_BOLD_9 = Font(name='Calibri', size=9, bold=True, color='FF000000')
FONT_BLACK_BOLD_11 = Font(name='Calibri', size=11, bold=True, color='FF000000')
FONT_BLACK_10 = Font(name='Calibri', size=10, color='FF000000')
FONT_BLACK_BOLD_10 = Font(name='Calibri', size=10, bold=True, color='FF000000')
FONT_WHITE_9 = Font(name='Calibri', size=9, color='FFFFFFFF')

ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_ROTATE = Alignment(horizontal='center', vertical='bottom', text_rotation=90, wrap_text=False)
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

THIN_BORDER = Border(
    left=Side(style='thin', color='FFD9D9D9'),
    right=Side(style='thin', color='FFD9D9D9'),
    top=Side(style='thin', color='FFD9D9D9'),
    bottom=Side(style='thin', color='FFD9D9D9'),
)


# ============================================================
# DATA PARSING
# ============================================================

def parse_estimation_sheet(filepath, sheet_name, is_engineering=False):
    """Parse an estimation sheet into a hierarchical structure.
    Returns list of dicts: {epic, feature, story_or_task, team, estimates}
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[sheet_name]

    rows = []
    current_epic = None
    current_feature = None
    current_story = None

    for r in range(3, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value  # Epic
        b = ws.cell(row=r, column=2).value  # Feature
        c = ws.cell(row=r, column=3).value  # Story
        d = ws.cell(row=r, column=4).value  # Task
        e = ws.cell(row=r, column=5).value  # Team

        if a:
            current_epic = str(a).strip()
        if b:
            current_feature = str(b).strip()
            current_story = None
        if c:
            current_story = str(c).strip()

        # Skip rows with no useful data
        if not (c or d):
            continue
        if not e:
            continue

        team = str(e).strip()
        item_name = str(d).strip() if d else str(c).strip() if c else None
        if not item_name:
            continue

        # Get estimates
        estimates = {}
        if is_engineering:
            for est in ENG_ESTIMATORS:
                mn = ws.cell(row=r, column=est['min_col']).value
                mx = ws.cell(row=r, column=est['max_col']).value
                if mn is not None or mx is not None:
                    estimates[est['short']] = {'min': _to_float(mn), 'max': _to_float(mx)}
            # Mustafa adjusted
            mustafa_min = ws.cell(row=r, column=26).value
            mustafa_max = ws.cell(row=r, column=27).value
            estimates['_mustafa_min'] = _to_float(mustafa_min)
            estimates['_mustafa_max'] = _to_float(mustafa_max)
            notes = ws.cell(row=r, column=28).value
            estimates['_notes'] = str(notes).strip() if notes else ''
        else:
            for est in DESIGN_ESTIMATORS:
                mn = ws.cell(row=r, column=est['min_col']).value
                mx = ws.cell(row=r, column=est['max_col']).value
                if mn is not None or mx is not None:
                    estimates[est['short']] = {'min': _to_float(mn), 'max': _to_float(mx)}
            final = ws.cell(row=r, column=26).value
            estimates['_final'] = _to_float(final)

        # Normalize epic name
        epic_norm = EPIC_NORM.get(current_epic, current_epic) if current_epic else None

        rows.append({
            'epic': epic_norm,
            'feature': current_feature,
            'parent_story': current_story if d else None,
            'story': item_name,
            'team': team,
            'is_task': bool(d),
            'estimates': estimates,
            'source': 'engineering' if is_engineering else 'design',
        })

    wb.close()
    return rows


def _to_float(val):
    """Convert a value to float, handling strings like '~' and 'no idea'."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if s in ('~', 'no idea', '', '-', 'N/A', 'n/a', '?'):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def get_miro_art_estimates():
    """Miro board ENV art estimates (hours converted to 7h days).
    These are environment art production estimates from free text blocks on the Miro board.
    Returns {template_feature_name: {role: {'min': days, 'max': days}}}
    """
    H2D = 7.0
    raw = [
        # (miro_name, min_hours, max_hours, template_feature, role)
        ('Architecture Kit - Mythcore', 40, 80, 'Biome - Portal Peak', 'Environment Artist'),
        ('Architecture Kit - Mythcore Ruins', 40, 80, 'Biome - Portal Peak', 'Environment Artist'),
        ('Architecture Kit - Fantasy', 480, 800, 'Biome - Portal Peak', 'Environment Artist'),
        ('Architecture Kit - Guild Hall', 580, 945, 'Guilds', 'Environment Artist'),
        ('Architecture Kit - Player Housing', 460, 800, 'Building Houses', 'Environment Artist'),
        ('Architecture Kit - Arena', 240, 800, 'PvP Combat', 'Environment Artist'),
        ('Architecture Kit - Deck/Scaffolding', 80, 120, 'Biome - Portal Peak', 'Environment Artist'),
        ('Architecture Kit - Wooden Camp', 120, 160, 'Biome - Portal Peak', 'Environment Artist'),
        ('FTUE Cave', 224, 224, 'FTUE', 'Environment Artist'),
        ('Downtime Zone', 648, 800, 'Biome - Downtime', 'Environment Artist'),
        ('New Zone', 568, 648, 'Biome - Farmlands', 'Environment Artist'),
        ('Player Island 1', 488, 488, 'User Space Server & Instance', 'Environment Artist'),
        ('Player Island 2', 488, 488, 'User Space Server & Instance', 'Environment Artist'),
        ('Guild Island', 488, 488, 'Guilds', 'Environment Artist'),
        ('Arena', 400, 640, 'PvP Combat', 'Environment Artist'),
        ('Dungeon (1 medium)', 480, 480, 'The Dungeon', 'Environment Artist'),
        ('Portal Dimension', 200, 200, 'Biome - Portal Peak', 'Environment Artist'),
        ('Bug Dungeon', 80, 120, 'Dungeon Mechanics', 'Environment Artist'),
        ('Tower Interior', 320, 420, 'Biome - Portal Peak', 'Environment Artist'),
        ('Rock Kits', 192, 192, 'Biome - Portal Peak', 'Environment Artist'),
        ('Player Buildings', 480, 480, 'Building Houses', 'Environment Artist'),
        ('Guild Buildings', 240, 480, 'Guilds', 'Environment Artist'),
        ('Props - Gritcore', 200, 400, 'Biome - Portal Peak', 'Prop Artist'),
        ('Props - Fantasy Exterior', 200, 400, 'Biome - Portal Peak', 'Prop Artist'),
        ('Furniture - Fantasy', 40, 40, 'User Space Decoration', 'Prop Artist'),
        ('Furniture - Gritcore', 320, 320, 'User Space Decoration', 'Prop Artist'),
        ('Biomes - Cave', 350, 400, 'Biome - Portal Peak', 'Environment Artist'),
        ('Biomes - New Zone', 350, 400, 'Biome - Farmlands', 'Environment Artist'),
        ('Biomes - Downtime', 350, 400, 'Biome - Downtime', 'Environment Artist'),
        ('Tower Exterior (Epic POI)', 568, 640, 'Biome - Portal Peak', 'Environment Artist'),
        ('The Big Battle Hill (Epic POI)', 568, 640, 'Biome - Hidden Hills', 'Environment Artist'),
        ('Cave Interior (Med POI)', 224, 300, 'Biome - Portal Peak', 'Environment Artist'),
        ('Waterfall Beauty (Small POI)', 80, 120, 'Biome - Whispering Woods', 'Environment Artist'),
        ('Bridge (Small POI)', 80, 120, 'Biome - Portal Peak', 'Environment Artist'),
        ("Drifter's Cross (Small POI)", 320, 320, 'Biome - Farmlands', 'Environment Artist'),
    ]

    result = {}
    for name, min_h, max_h, feature, role in raw:
        min_d = round(min_h / H2D, 1)
        max_d = round(max_h / H2D, 1)
        if feature not in result:
            result[feature] = {}
        if role not in result[feature]:
            result[feature][role] = {'min': 0, 'max': 0}
        result[feature][role]['min'] += min_d
        result[feature][role]['max'] += max_d

    # Round accumulated totals
    for feat in result:
        for role in result[feat]:
            result[feat][role]['min'] = round(result[feat][role]['min'], 1)
            result[feat][role]['max'] = round(result[feat][role]['max'], 1)

    total_min = sum(v['min'] for feat in result.values() for v in feat.values())
    total_max = sum(v['max'] for feat in result.values() for v in feat.values())
    print(f"  Miro art: {len(raw)} items -> {len(result)} features, {total_min:.0f}-{total_max:.0f} total days")
    return result


def parse_hours_mapping(filepath):
    """Parse Miro hours_mapping.txt into {template_feature_name: {role: days}}."""
    mapping = {}
    with open(filepath, 'r') as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith('===') or line.startswith('Excluded') or line.startswith('Features'):
                continue
            if line.startswith('Unmatched'):
                break
            # Format: "Feature Name: NND total | Role1=NND, Role2=NND, ..."
            m = re.match(r'^(.+?):\s*(\d+)D total\s*\|\s*(.+)$', line)
            if not m:
                m2 = re.match(r'^(.+?):\s*(\d+)D total\s*\|\s*$', line)
                if m2:
                    continue
                continue
            feature_name = m.group(1).strip()
            roles_str = m.group(3).strip()

            # Map to template name
            template_name = MIRO_TO_TEMPLATE.get(feature_name, feature_name)
            if template_name is None:
                continue

            role_days = {}
            for pair in roles_str.split(','):
                pair = pair.strip()
                rm = re.match(r'^(.+?)=(\d+)D$', pair)
                if rm:
                    role_days[rm.group(1).strip()] = int(rm.group(2))

            if template_name in mapping:
                for role, days in role_days.items():
                    mapping[template_name][role] = mapping[template_name].get(role, 0) + days
            else:
                mapping[template_name] = role_days

    return mapping


def load_template_order():
    """Return the template feature order."""
    # Hardcoded from template analysis
    return [
        {'type': 'epic', 'name': 'Player Build'},
        {'type': 'feature', 'name': 'Player Progression', 'size': 'L', 'tier': 'T1: R&D'},
        {'type': 'feature', 'name': 'Skill System', 'size': 'L', 'tier': 'T1: R&D'},
        {'type': 'feature', 'name': 'Achievements & Trophies', 'size': 'M', 'tier': 'T1: R&D'},
        {'type': 'feature', 'name': 'Factions / Reputation', 'size': 'M', 'tier': 'T1: R&D'},
        {'type': 'feature', 'name': 'FTUE', 'size': 'M', 'tier': 'T2: GDD/TDD/Brief'},
        {'type': 'feature', 'name': 'Professions', 'size': 'L', 'tier': 'T1: R&D'},
        {'type': 'feature', 'name': 'Character Management', 'size': 'M', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Character Creation & Customization', 'size': 'L', 'tier': 'T1: R&D'},
        {'type': 'feature', 'name': 'Garment System', 'size': 'M', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Persona System', 'size': 'M', 'tier': 'T1: R&D'},
        {'type': 'feature', 'name': 'Emote System', 'size': 'S', 'tier': 'T0: Ideation'},
        {'type': 'epic', 'name': 'World Systems'},
        {'type': 'feature', 'name': 'Foam Corruption', 'size': 'M', 'tier': 'T2: GDD/TDD/Brief'},
        {'type': 'feature', 'name': 'Miasma Corruption', 'size': 'M', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Water Corruption', 'size': 'M', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Crystal Corruption', 'size': 'M', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Time Corruption', 'size': 'M', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Enemy Corruption', 'size': 'M', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Oil Corruption', 'size': 'M', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Texture Corruption', 'size': 'M', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Environmental Elemental Effects (combination states)', 'size': 'L', 'tier': 'T2: GDD/TDD/Brief'},
        {'type': 'feature', 'name': 'Basic Movement System', 'size': 'L', 'tier': 'T1: R&D'},
        {'type': 'feature', 'name': 'Advanced Traversal', 'size': 'L', 'tier': 'T1: R&D'},
        {'type': 'feature', 'name': 'Navigation', 'size': 'M', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Day/Night Cycle', 'size': 'M', 'tier': 'T1: R&D'},
        {'type': 'feature', 'name': 'Dynamic Weather', 'size': 'M', 'tier': 'T1: R&D'},
        {'type': 'feature', 'name': 'Instancing', 'size': 'L', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Dungeon Mechanics', 'size': 'L', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'The Dungeon', 'size': 'L', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Partner Portal', 'size': 'M', 'tier': 'T1: R&D'},
        {'type': 'feature', 'name': 'World Bosses', 'size': 'L', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Corruption Heatmap', 'size': 'M', 'tier': 'T1: R&D'},
        {'type': 'feature', 'name': 'Biome - Portal Peak', 'size': 'M', 'tier': 'T4: MVP'},
        {'type': 'feature', 'name': 'Biome - Hidden Hills', 'size': 'M', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Biome - Farmlands', 'size': 'M', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Biome - Mistrun Shore', 'size': 'M', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Biome - Whispering Woods', 'size': 'M', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Biome - Tranquil Sands', 'size': 'M', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Biome - Torrential Vale', 'size': 'M', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Biome - Clockwork', 'size': 'M', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Biome - Downtime', 'size': 'M', 'tier': 'T4: MVP'},
        {'type': 'epic', 'name': 'Combat'},
        {'type': 'feature', 'name': 'Combat Framework', 'size': 'L', 'tier': 'T2: GDD/TDD/Brief'},
        {'type': 'feature', 'name': 'Ranged Combat', 'size': 'M', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Combat AI Behavior', 'size': 'L', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Enemies Art', 'size': 'M', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'PvP Combat', 'size': 'L', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Combat Balance & Telemetry', 'size': 'M', 'tier': 'T2: GDD/TDD/Brief'},
        {'type': 'feature', 'name': 'Combat Abilities', 'size': 'L', 'tier': 'T1: R&D'},
        {'type': 'feature', 'name': 'Magic', 'size': 'L', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Full Motion Video (FMV)', 'size': 'M', 'tier': 'T1: R&D'},
        {'type': 'epic', 'name': 'User Space'},
        {'type': 'feature', 'name': 'User Space Server & Instance', 'size': 'L', 'tier': 'T2: GDD/TDD/Brief'},
        {'type': 'feature', 'name': 'User Space Decoration', 'size': 'M', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'P2P / P2G User Space Connection', 'size': 'M', 'tier': 'T2: GDD/TDD/Brief'},
        {'type': 'feature', 'name': 'Building Houses', 'size': 'L', 'tier': 'T1: R&D'},
        {'type': 'feature', 'name': 'Saving House Layouts', 'size': 'M', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Partner Portal System (player house)', 'size': 'M', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Arcade Cabinet', 'size': 'M', 'tier': 'T1: R&D'},
        {'type': 'feature', 'name': 'Anywhere Door Placement', 'size': 'S', 'tier': 'T0: Ideation'},
        {'type': 'epic', 'name': 'Items & Inventory'},
        {'type': 'feature', 'name': 'Item Core System', 'size': 'L', 'tier': 'T2: GDD/TDD/Brief'},
        {'type': 'feature', 'name': 'Inventory System', 'size': 'M', 'tier': 'T2: GDD/TDD/Brief'},
        {'type': 'feature', 'name': 'Equipment System', 'size': 'M', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Loot Distribution', 'size': 'M', 'tier': 'T2: GDD/TDD/Brief'},
        {'type': 'feature', 'name': 'Mounts / Pets', 'size': 'L', 'tier': 'T1: R&D'},
        {'type': 'feature', 'name': 'Banks', 'size': 'M', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'P2P Item Transfer', 'size': 'M', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Consumables', 'size': 'S', 'tier': 'T2: GDD/TDD/Brief'},
        {'type': 'feature', 'name': 'Durable Items', 'size': 'S', 'tier': 'T0: Ideation'},
        {'type': 'epic', 'name': 'Player Economy'},
        {'type': 'feature', 'name': 'Trading System', 'size': 'L', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Currency System', 'size': 'M', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Auction House', 'size': 'L', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'In-game Store', 'size': 'M', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'In-game Economy', 'size': 'L', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'RMT Game Economy', 'size': 'L', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'IG / RMT Monitoring System', 'size': 'M', 'tier': 'T0: Ideation'},
        {'type': 'epic', 'name': 'Quest System'},
        {'type': 'feature', 'name': 'Quests Backend / Tooling', 'size': 'L', 'tier': 'T2: GDD/TDD/Brief'},
        {'type': 'feature', 'name': 'Quests as Events', 'size': 'M', 'tier': 'T1: R&D'},
        {'type': 'feature', 'name': 'Quest Multiplayer', 'size': 'L', 'tier': 'T1: R&D'},
        {'type': 'feature', 'name': 'Quest Types - Courier / Escort / Kill / Collect / Mystery / Time / Partner / Lockpicking', 'size': 'L', 'tier': 'T2: GDD/TDD/Brief'},
        {'type': 'feature', 'name': 'Quest Reward System', 'size': 'M', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Quest Log', 'size': 'M', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Quest Dialog GUI', 'size': 'M', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Quest Narrative Content', 'size': 'L', 'tier': 'T2: GDD/TDD/Brief'},
        {'type': 'feature', 'name': 'In Game Cinematic (IGC)', 'size': 'L', 'tier': 'T2: GDD/TDD/Brief'},
        {'type': 'feature', 'name': 'Meta Quest System', 'size': 'M', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Quest Generator System', 'size': 'L', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Targeted VO', 'size': 'M', 'tier': 'T2: GDD/TDD/Brief'},
        {'type': 'epic', 'name': 'Social & Multiplayer'},
        {'type': 'feature', 'name': 'Game Chat System', 'size': 'M', 'tier': 'T1: R&D'},
        {'type': 'feature', 'name': 'Friends', 'size': 'M', 'tier': 'T1: R&D'},
        {'type': 'feature', 'name': 'Guilds', 'size': 'L', 'tier': 'T1: R&D'},
        {'type': 'feature', 'name': 'Co-op / Party', 'size': 'M', 'tier': 'T1: R&D'},
        {'type': 'feature', 'name': 'Co-op / Raids', 'size': 'L', 'tier': 'T1: R&D'},
        {'type': 'feature', 'name': 'Dueling', 'size': 'M', 'tier': 'T1: R&D'},
        {'type': 'feature', 'name': 'In-game Mail', 'size': 'S', 'tier': 'T2: GDD/TDD/Brief'},
        {'type': 'feature', 'name': 'Notifications', 'size': 'S', 'tier': 'T1: R&D'},
        {'type': 'feature', 'name': 'Player Config PvP Server', 'size': 'M', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'LFG / Pick-up Groups', 'size': 'M', 'tier': 'T0: Ideation'},
        {'type': 'epic', 'name': 'Game Bibles'},
        {'type': 'feature', 'name': 'UX/UI Bible', 'size': 'M', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Accessibility Guidelines', 'size': 'M', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Brand Bible', 'size': 'M', 'tier': 'T2: GDD/TDD/Brief'},
        {'type': 'feature', 'name': 'Art Bible', 'size': 'L', 'tier': 'T3: Prototype'},
        {'type': 'feature', 'name': 'Monetization Bible', 'size': 'M', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Red Books (Live Service emergency guidelines)', 'size': 'M', 'tier': 'T0: Ideation'},
        {'type': 'epic', 'name': 'Platform'},
        {'type': 'feature', 'name': 'Player Account Creation / Deletion', 'size': 'M', 'tier': 'T2: GDD/TDD/Brief'},
        {'type': 'feature', 'name': 'Player Account Authentication', 'size': 'M', 'tier': 'T2: GDD/TDD/Brief'},
        {'type': 'feature', 'name': 'Player Account Privacy Settings & Legal', 'size': 'M', 'tier': 'T1: R&D'},
        {'type': 'feature', 'name': 'Player Dashboard - Account Features', 'size': 'M', 'tier': 'T1: R&D'},
        {'type': 'feature', 'name': 'Player Dashboard - Game Dashboard', 'size': 'M', 'tier': 'T1: R&D'},
        {'type': 'feature', 'name': 'Player Dashboard - Friends', 'size': 'M', 'tier': 'T1: R&D'},
        {'type': 'feature', 'name': 'Player Dashboard - Search & Notifications', 'size': 'M', 'tier': 'T1: R&D'},
        {'type': 'feature', 'name': 'Player Dashboard - Marketplace', 'size': 'L', 'tier': 'T1: R&D'},
        {'type': 'feature', 'name': 'Developer Dashboard - Developer Account', 'size': 'M', 'tier': 'T1: R&D'},
        {'type': 'feature', 'name': 'Developer Dashboard - Game Dashboard', 'size': 'M', 'tier': 'T1: R&D'},
        {'type': 'feature', 'name': 'Developer Dashboard - Settings & Support', 'size': 'M', 'tier': 'T1: R&D'},
        {'type': 'feature', 'name': 'Quest Tracking', 'size': 'M', 'tier': 'T1: R&D'},
        {'type': 'feature', 'name': 'Payment Processing (incl. Refunds)', 'size': 'L', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Online Services & Backend', 'size': 'L', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Launcher & Distribution', 'size': 'L', 'tier': 'T2: GDD/TDD/Brief'},
        {'type': 'feature', 'name': 'Partner Loop', 'size': 'M', 'tier': 'T1: R&D'},
        {'type': 'feature', 'name': 'Game Tool Integration (incl. Game Key Generation)', 'size': 'M', 'tier': 'T1: R&D'},
        {'type': 'feature', 'name': 'Account Level Settings', 'size': 'S', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Social Media Support', 'size': 'M', 'tier': 'T0: Ideation'},
        {'type': 'epic', 'name': 'Live Game'},
        {'type': 'feature', 'name': 'Live GDD', 'size': 'M', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Content', 'size': 'L', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Forecast', 'size': 'M', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Analytics & BI', 'size': 'L', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Sales Dashboard', 'size': 'M', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Telemetry', 'size': 'L', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Live Service Event Plan', 'size': 'L', 'tier': 'T0: Ideation'},
        {'type': 'epic', 'name': 'Partner Build'},
        {'type': 'feature', 'name': 'Partner = Feature definition', 'size': 'S', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Sign Partner', 'size': 'S', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Partner Brief', 'size': 'M', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Partner Design Doc', 'size': 'M', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Partner TDD', 'size': 'M', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Partner Art Brief', 'size': 'M', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Partner Build', 'size': 'L', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Partner Integration', 'size': 'M', 'tier': 'T0: Ideation'},
        {'type': 'epic', 'name': 'Product Publishing'},
        {'type': 'feature', 'name': 'Web Presence', 'size': 'M', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Website', 'size': 'M', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Community Persona', 'size': 'M', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Title Identity', 'size': 'M', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Player Support', 'size': 'M', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': 'Go-To-Market Strategy & Implementation', 'size': 'L', 'tier': 'T0: Ideation'},
        {'type': 'feature', 'name': '3rd-party Store Publishing', 'size': 'M', 'tier': 'T0: Ideation'},
    ]


def match_feature(source_feature, source_epic, template_features):
    """Find the best matching template feature for a source feature."""
    if not source_feature:
        return None

    # Try direct normalisation first
    norm = FEATURE_NORM.get(source_feature, source_feature)

    # Try exact match (case-insensitive)
    for tf in template_features:
        if tf['name'].lower() == norm.lower():
            return tf['name']

    # Try substring match (longer substring first for better precision)
    norm_lower = norm.lower()
    best_match = None
    best_len = 0
    for tf in template_features:
        tf_lower = tf['name'].lower()
        if norm_lower in tf_lower or tf_lower in norm_lower:
            match_len = min(len(norm_lower), len(tf_lower))
            if match_len > best_len:
                best_len = match_len
                best_match = tf['name']
    if best_match:
        return best_match

    # Try word overlap (at least 2 significant words matching)
    stop_words = {'the', 'a', 'an', 'of', 'in', 'for', 'and', 'or', 'to', 'is', 'on', 'at', 'by', 's'}
    norm_words = set(w.lower() for w in re.split(r'[\s/\-\(\)]+', norm) if len(w) > 1 and w.lower() not in stop_words)
    for tf in template_features:
        tf_words = set(w.lower() for w in re.split(r'[\s/\-\(\)]+', tf['name']) if len(w) > 1 and w.lower() not in stop_words)
        overlap = norm_words & tf_words
        if len(overlap) >= 2:
            return tf['name']

    return None


# ============================================================
# WORKBOOK BUILDING (xlsxwriter — reliable merged cell fills)
# ============================================================

import xlsxwriter


def build_workbook(template_order, design_rows, eng_rows, miro_data, output_path):
    """Build the output workbook using xlsxwriter for correct merged cell formatting."""
    total_cols = COL_LAYOUT['_total_cols']
    wb = xlsxwriter.Workbook(output_path)
    ws = wb.add_worksheet('Plan')

    # --- Pre-create all format objects ---
    fmt_work_streams = wb.add_format({'bg_color': '#404040', 'font_color': 'white', 'bold': True, 'font_size': 12, 'align': 'center', 'valign': 'vcenter', 'text_wrap': True})
    fmt_meta_hdr = wb.add_format({'bg_color': '#595959', 'font_color': 'white', 'bold': True, 'font_size': 14, 'align': 'center', 'valign': 'vcenter'})
    fmt_gate = wb.add_format({'bg_color': '#1F4E78', 'font_color': 'white', 'bold': True, 'font_size': 10, 'align': 'center', 'valign': 'vcenter', 'text_wrap': True})
    fmt_subphase = wb.add_format({'bg_color': '#305496', 'font_color': 'white', 'bold': True, 'font_size': 10, 'align': 'center', 'valign': 'vcenter', 'text_wrap': True})
    fmt_epic = wb.add_format({'bg_color': '#FFD966', 'font_color': 'black', 'bold': True, 'font_size': 11, 'align': 'left', 'valign': 'vcenter'})
    fmt_minmax = wb.add_format({'bg_color': '#305496', 'font_color': 'white', 'font_size': 9, 'align': 'center', 'valign': 'vcenter'})
    # Off-white = #F2F2F2 (theme 0, tint -0.05 in Excel)
    OW = '#F2F2F2'
    THIN = 1  # xlsxwriter thin border

    # Feature row: metadata columns (off-white bg)
    fmt_feat_name = wb.add_format({'font_size': 10, 'bold': True, 'valign': 'vcenter', 'bg_color': OW})
    fmt_feat_data = wb.add_format({'font_size': 10, 'valign': 'vcenter', 'bg_color': OW})
    fmt_feat_blank = wb.add_format({'bg_color': OW})
    # Feature row: EP cells (12pt bold centered)
    fmt_feat_ep = wb.add_format({'font_size': 12, 'bold': True, 'align': 'center', 'num_format': '0.##'})
    fmt_feat_ep_stripe = wb.add_format({'font_size': 12, 'bold': True, 'align': 'center', 'num_format': '0.##',
                                         'bg_color': OW, 'border': THIN})
    fmt_feat_ep_blank = wb.add_format({'font_size': 12, 'bold': True, 'align': 'center'})
    fmt_feat_ep_blank_stripe = wb.add_format({'font_size': 12, 'bold': True, 'align': 'center',
                                               'bg_color': OW, 'border': THIN})

    # Story row: metadata columns (no bg)
    fmt_story_name = wb.add_format({'font_size': 10, 'valign': 'vcenter', 'text_wrap': True})
    fmt_story_team = wb.add_format({'font_size': 10, 'valign': 'vcenter'})
    # Story row: EP cells (12pt bold centered)
    fmt_story_ep = wb.add_format({'font_size': 12, 'bold': True, 'align': 'center', 'num_format': '0.##'})
    fmt_story_ep_stripe = wb.add_format({'font_size': 12, 'bold': True, 'align': 'center', 'num_format': '0.##',
                                          'bg_color': OW, 'border': THIN})
    fmt_story_ep_blank = wb.add_format({'font_size': 12, 'bold': True, 'align': 'center'})
    fmt_story_ep_blank_stripe = wb.add_format({'font_size': 12, 'bold': True, 'align': 'center',
                                                'bg_color': OW, 'border': THIN})

    # Gate column fill extending through data rows
    fmt_gate_data = wb.add_format({'bg_color': '#1F4E78'})

    # EP stripe columns: every other role pair starting from role 11 (odd indices >= 11)
    EP_STRIPE_ROLES = set()
    for ri in range(10, NUM_ROLES):
        if ri % 2 == 1:
            EP_STRIPE_ROLES.add(ri)

    # Super-header formats (one per phase group colour)
    super_fmts = {}
    for phase in PHASES:
        c = phase['super_fill'][2:]  # strip FF prefix
        if c not in super_fmts:
            super_fmts[c] = wb.add_format({'bg_color': f'#{c}', 'font_color': 'white', 'bold': True, 'font_size': 12, 'align': 'center', 'valign': 'vcenter'})

    # Role header formats (one per discipline colour)
    ROLE_COLOURS = {}
    for r in ['CEO', 'Creative Director', 'Head of Legal', 'COO', 'HR Manager']:
        ROLE_COLOURS[r] = '#D9D9D9'
    for r in ['Executive Producer', 'Producer', 'Game Producer', 'Tech Producer', 'Art Producer']:
        ROLE_COLOURS[r] = '#FFD966'
    for r in ['Game Director', 'Game Design Lead', 'Level Design Lead', 'Game Designer',
              'Combat Designer', 'Level Designer', 'Technical Designer', 'Writer']:
        ROLE_COLOURS[r] = '#BDD7EE'
    for r in ['Art Director', 'Concept Artist', 'UI/UX Designer', 'Graphic Designer',
              'Character Artist', 'Environment Artist', 'Prop Artist', 'Animator',
              'Character Rigger', 'VFX Artist', 'Technical Artist']:
        ROLE_COLOURS[r] = '#F8CBAD'
    for r in ['CTO', 'Tech Lead', 'Gameplay Developer', 'Full Stack Developer',
              'Backend Developer', 'Network Engineer']:
        ROLE_COLOURS[r] = '#C6E0B4'
    ROLE_COLOURS['QA'] = '#E4DFEC'
    ROLE_COLOURS['Sound Designer'] = '#FFE699'
    ROLE_COLOURS['Platform Team'] = '#B4C7E7'

    role_fmt_cache = {}
    def get_role_fmt(role_name):
        if role_name not in role_fmt_cache:
            bg = ROLE_COLOURS.get(role_name, '#D9D9D9')
            role_fmt_cache[role_name] = wb.add_format({
                'bg_color': bg, 'font_color': 'black', 'bold': True, 'font_size': 9,
                'rotation': 90, 'align': 'center', 'valign': 'bottom'
            })
        return role_fmt_cache[role_name]

    # xlsxwriter is 0-indexed; template uses 1-indexed. Convert: row-1, col-1

    # ---- ROW 1 (idx 0): Empty ----

    # ---- ROW 2 (idx 1): Super-headers ----
    # WORK STREAMS merged A2:G3 → (1,0) to (2, METADATA_COLS-1)
    ws.merge_range(1, 0, 2, METADATA_COLS - 1, 'WORK STREAMS', fmt_work_streams)

    # Phase super-header groups
    super_groups = {}
    for phase in PHASES:
        sname = phase['super']
        layout = COL_LAYOUT[phase['name']]
        sc = layout['gate_col'] - 1  # 0-indexed
        ec = layout['role_end'] - 1
        if sname not in super_groups:
            super_groups[sname] = {'start': sc, 'end': ec, 'fill': phase['super_fill'][2:]}
        else:
            super_groups[sname]['end'] = ec

    for sname, sg in super_groups.items():
        fmt = super_fmts[sg['fill']]
        ws.merge_range(1, sg['start'], 1, sg['end'], sname, fmt)

    # ---- ROW 3 (idx 2): Gate markers + sub-phase labels ----
    for phase in PHASES:
        layout = COL_LAYOUT[phase['name']]
        gc = layout['gate_col'] - 1
        ws.write(2, gc, f"GATE\n{phase['gate_num']}", fmt_gate)
        rs = layout['role_start'] - 1
        re_ = layout['role_end'] - 1
        if rs == re_:
            ws.write(2, rs, phase['sub_label'], fmt_subphase)
        else:
            ws.merge_range(2, rs, 2, re_, phase['sub_label'], fmt_subphase)

    # ---- ROW 4 (idx 3): Column headers ----
    meta_headers = ['Epic', 'Feature', 'Story', 'Team', 'Size', 'Current Tier', 'Notes']
    for i, h in enumerate(meta_headers):
        ws.write(3, i, h, fmt_meta_hdr)

    # Role headers per phase
    for phase in PHASES:
        layout = COL_LAYOUT[phase['name']]
        ws.write(3, layout['gate_col'] - 1, '', fmt_gate)

        if phase['min_max']:
            for ri, role_name in enumerate(ROLES):
                cm = layout['role_start'] - 1 + ri * 2
                cx = cm + 1
                fmt = get_role_fmt(role_name)
                ws.merge_range(3, cm, 3, cx, role_name, fmt)
        else:
            for ri, role_name in enumerate(ROLES):
                c = layout['role_start'] - 1 + ri
                ws.write(3, c, role_name, get_role_fmt(role_name))

    # ---- DATA ROWS ----
    feature_stories = build_feature_stories(design_rows, eng_rows, template_order)
    # Also stored externally for other sheets, but recomputed here for self-containment

    # Pre-compute EP stripe column set (0-indexed)
    ep_layout = COL_LAYOUT['EARLY PROD']
    ep_stripe_cols = set()
    for ri in EP_STRIPE_ROLES:
        c0 = ep_layout['role_start'] - 1 + ri * 2
        ep_stripe_cols.add(c0)
        ep_stripe_cols.add(c0 + 1)

    # Pre-compute gate column indices (0-indexed)
    gate_cols = set()
    for phase in PHASES:
        gate_cols.add(COL_LAYOUT[phase['name']]['gate_col'] - 1)

    def fill_row_backgrounds(ws, row, is_feature):
        """Write gate fills and EP cell formatting for a data row."""
        # Gate columns
        for gc in gate_cols:
            ws.write_blank(row, gc, None, fmt_gate_data)
        # EP columns: ALL get 12pt bold centered; stripe cols also get off-white + borders
        for ri in range(NUM_ROLES):
            c0 = ep_layout['role_start'] - 1 + ri * 2
            c1 = c0 + 1
            is_stripe = ri in EP_STRIPE_ROLES
            if is_feature:
                blank_fmt = fmt_feat_ep_blank_stripe if is_stripe else fmt_feat_ep_blank
            else:
                blank_fmt = fmt_story_ep_blank_stripe if is_stripe else fmt_story_ep_blank
            ws.write_blank(row, c0, None, blank_fmt)
            ws.write_blank(row, c1, None, blank_fmt)
        # Feature rows: off-white on metadata cols
        if is_feature:
            for c in range(METADATA_COLS):
                ws.write_blank(row, c, None, fmt_feat_blank)

    row = 4  # 0-indexed, starting at row 5 (idx 4)
    epic_row_ranges = {}  # {epic_name: (start_row_0idx, end_row_0idx)}
    current_epic = None
    current_epic_start = None

    for item in template_order:
        if item['type'] == 'epic':
            if current_epic:
                epic_row_ranges[current_epic] = (current_epic_start, row - 1)
            current_epic = item['name']
            ws.merge_range(row, 0, row, total_cols - 1, item['name'], fmt_epic)
            row += 1
            current_epic_start = row

            # Min/Max sub-header row
            for gc in gate_cols:
                ws.write_blank(row, gc, None, fmt_gate_data)
            for ri in range(NUM_ROLES):
                cm = ep_layout['role_start'] - 1 + ri * 2
                cx = cm + 1
                ws.write(row, cm, 'Min', fmt_minmax)
                ws.write(row, cx, 'Max', fmt_minmax)
            row += 1

        elif item['type'] == 'feature':
            fname = item['name']

            # Fill backgrounds first (gate cols, EP stripes, off-white metadata)
            fill_row_backgrounds(ws, row, is_feature=True)

            # Write feature data on top of backgrounds
            ws.write(row, 1, fname, fmt_feat_name)
            if item.get('size'):
                ws.write(row, 4, item['size'], fmt_feat_data)
            if item.get('tier'):
                ws.write(row, 5, item['tier'], fmt_feat_data)

            # Miro totals on feature row (in EP role Min/Max columns)
            miro_roles = miro_data.get(fname, {})
            for role_name, days in miro_roles.items():
                col_min = ep_role_col(role_name, is_min=True)
                col_max = ep_role_col(role_name, is_min=False)
                if isinstance(days, dict):
                    # Art data with min/max
                    if col_min and days.get('min'):
                        c0 = col_min - 1
                        ri = (c0 - (ep_layout['role_start'] - 1)) // 2
                        fmt = fmt_feat_ep_stripe if ri in EP_STRIPE_ROLES else fmt_feat_ep
                        ws.write(row, c0, days['min'], fmt)
                    if col_max and days.get('max'):
                        c0 = col_max - 1
                        ri = (c0 - (ep_layout['role_start'] - 1)) // 2
                        fmt = fmt_feat_ep_stripe if ri in EP_STRIPE_ROLES else fmt_feat_ep
                        ws.write(row, c0, days['max'], fmt)
                else:
                    # Game systems data (single value goes in Max)
                    if col_max:
                        c0 = col_max - 1
                        ri = (c0 - (ep_layout['role_start'] - 1)) // 2
                        fmt = fmt_feat_ep_stripe if ri in EP_STRIPE_ROLES else fmt_feat_ep
                        ws.write(row, c0, days, fmt)

            row += 1

            # Story rows
            stories = feature_stories.get(fname, [])
            for story in stories:
                team = story['team']
                dept = TEAM_TO_DEPT.get(team, team)
                team_label = f"{dept} | {team}"

                # Fill backgrounds first
                fill_row_backgrounds(ws, row, is_feature=False)

                ws.write(row, 2, story['name'], fmt_story_name)
                ws.write(row, 3, team_label, fmt_story_team)

                role_name = TEAM_TO_ROLE.get(team)
                if role_name:
                    col_min_ep = ep_role_col(role_name, is_min=True)
                    col_max_ep = ep_role_col(role_name, is_min=False)

                    def write_est(col_1idx, value):
                        c0 = col_1idx - 1
                        ri = (c0 - (ep_layout['role_start'] - 1)) // 2
                        fmt = fmt_story_ep_stripe if ri in EP_STRIPE_ROLES else fmt_story_ep
                        ws.write(row, c0, value, fmt)

                    # Art task shortcut (from Miro board)
                    if story.get('_art_min') is not None:
                        if col_min_ep:
                            write_est(col_min_ep, story['_art_min'])
                        if col_max_ep:
                            write_est(col_max_ep, story['_art_max'])
                    elif story.get('design_est'):
                        est = story['design_est']
                        final = est.get('_final')
                        ind_mins = [v['min'] for k, v in est.items()
                                   if not k.startswith('_') and isinstance(v, dict) and v.get('min') is not None]
                        if col_min_ep and ind_mins:
                            write_est(col_min_ep, min(ind_mins))
                        if col_max_ep and final is not None:
                            write_est(col_max_ep, final)
                        elif col_max_ep:
                            ind_maxes = [v['max'] for k, v in est.items()
                                        if not k.startswith('_') and isinstance(v, dict) and v.get('max') is not None]
                            if ind_maxes:
                                write_est(col_max_ep, max(ind_maxes))

                    if story.get('eng_est'):
                        est = story['eng_est']
                        m_min = est.get('_mustafa_min')
                        m_max = est.get('_mustafa_max')
                        if not m_min or not m_max:
                            ind_mins = [v['min'] for k, v in est.items()
                                       if not k.startswith('_') and isinstance(v, dict) and v.get('min') is not None]
                            ind_maxes = [v['max'] for k, v in est.items()
                                        if not k.startswith('_') and isinstance(v, dict) and v.get('max') is not None]
                            if not m_min and ind_mins:
                                m_min = min(ind_mins)
                            if not m_max and ind_maxes:
                                m_max = max(ind_maxes)
                        if col_min_ep and m_min is not None:
                            write_est(col_min_ep, m_min)
                        if col_max_ep and m_max is not None:
                            write_est(col_max_ep, m_max)

                row += 1

    # ---- Column widths + visibility in ONE pass (no conflicting set_column calls) ----
    # v14 DRAFT hidden columns (1-indexed): 8, 87-106, 164
    V14_HIDDEN = set([8] + list(range(87, 107)) + [164])

    ws.set_column(0, 0, 6.71)    # A: Epic
    ws.set_column(1, 1, 20.71)   # B: Feature
    ws.set_column(2, 2, 54.29)   # C: Story
    ws.set_column(3, 3, 30.0)    # D: Team
    ws.set_column(4, 4, 6.71)    # E: Size
    ws.set_column(5, 5, 16.71)   # F: Current Tier
    ws.set_column(6, 6, 15.71)   # G: Notes

    # Phase columns: each column gets ONE set_column call with both width and hidden
    for phase in PHASES:
        layout = COL_LAYOUT[phase['name']]
        gc = layout['gate_col']
        gc_hidden = gc in V14_HIDDEN
        ws.set_column(gc - 1, gc - 1, 5.71, None, {'hidden': gc_hidden} if gc_hidden else None)

        if phase['min_max']:
            for ri in range(NUM_ROLES):
                mc = layout['role_start'] + ri * 2
                xc = mc + 1
                mc_hidden = mc in V14_HIDDEN
                xc_hidden = xc in V14_HIDDEN
                ws.set_column(mc - 1, mc - 1, 4.0, None, {'hidden': True} if mc_hidden else None)
                ws.set_column(xc - 1, xc - 1, 4.14, None, {'hidden': True} if xc_hidden else None)
        else:
            # Non-EP phases: set each column individually to avoid range conflicts
            for ri in range(NUM_ROLES):
                c = layout['role_start'] + ri
                c_hidden = c in V14_HIDDEN
                ws.set_column(c - 1, c - 1, 13.0, None, {'hidden': True} if c_hidden else None)

    # Row heights (matching v14 DRAFT)
    ws.set_row(1, 15.75)
    ws.set_row(2, 25.5)
    ws.set_row(3, 84.0)
    ws.set_default_row(15.75)

    ws.freeze_panes(4, METADATA_COLS)
    ws.set_zoom(130)

    if current_epic:
        epic_row_ranges[current_epic] = (current_epic_start, row - 1)

    print(f"Built Plan sheet: {row} rows, {total_cols} columns")
    return wb, row, epic_row_ranges


def build_feature_stories(design_rows, eng_rows, template_order):
    """Build a dict of {template_feature_name: [story_dicts]}.
    No merging — every source row is a distinct work item, even if same name+team
    across design and engineering sheets. Different sheet = different task.
    Also includes Miro art tasks as individual story rows.
    """
    template_features = [t for t in template_order if t['type'] == 'feature']
    template_names = {t['name'] for t in template_features}

    feature_stories = {name: [] for name in template_names}

    def process_rows(rows, source_type):
        for r in rows:
            feat_name = match_feature(r['feature'], r['epic'], template_features)
            if feat_name is None:
                continue

            story = {
                'name': r['story'],
                'team': r['team'],
                'design_est': r['estimates'] if source_type == 'design' else None,
                'eng_est': r['estimates'] if source_type == 'engineering' else None,
            }
            feature_stories[feat_name].append(story)

    process_rows(design_rows, 'design')
    process_rows(eng_rows, 'engineering')

    # Add Miro art tasks as individual story rows
    H2D = 7.0
    ART_TASKS = [
        ('Architecture Kit - Mythcore', 40, 80, 'Biome - Portal Peak', 'Environment Art'),
        ('Architecture Kit - Mythcore Ruins', 40, 80, 'Biome - Portal Peak', 'Environment Art'),
        ('Architecture Kit - Fantasy', 480, 800, 'Biome - Portal Peak', 'Environment Art'),
        ('Architecture Kit - Guild Hall', 580, 945, 'Guilds', 'Environment Art'),
        ('Architecture Kit - Player Housing', 460, 800, 'Building Houses', 'Environment Art'),
        ('Architecture Kit - Arena', 240, 800, 'PvP Combat', 'Environment Art'),
        ('Architecture Kit - Deck/Scaffolding', 80, 120, 'Biome - Portal Peak', 'Environment Art'),
        ('Architecture Kit - Wooden Camp', 120, 160, 'Biome - Portal Peak', 'Environment Art'),
        ('FTUE Cave', 224, 224, 'FTUE', 'Environment Art'),
        ('Downtime Zone', 648, 800, 'Biome - Downtime', 'Environment Art'),
        ('New Zone', 568, 648, 'Biome - Farmlands', 'Environment Art'),
        ('Player Island 1', 488, 488, 'User Space Server & Instance', 'Environment Art'),
        ('Player Island 2', 488, 488, 'User Space Server & Instance', 'Environment Art'),
        ('Guild Island', 488, 488, 'Guilds', 'Environment Art'),
        ('Arena', 400, 640, 'PvP Combat', 'Environment Art'),
        ('Dungeon (1 medium)', 480, 480, 'The Dungeon', 'Environment Art'),
        ('Portal Dimension', 200, 200, 'Biome - Portal Peak', 'Environment Art'),
        ('Bug Dungeon', 80, 120, 'Dungeon Mechanics', 'Environment Art'),
        ('Tower Interior', 320, 420, 'Biome - Portal Peak', 'Environment Art'),
        ('Rock Kits', 192, 192, 'Biome - Portal Peak', 'Environment Art'),
        ('Player Buildings', 480, 480, 'Building Houses', 'Environment Art'),
        ('Guild Buildings', 240, 480, 'Guilds', 'Environment Art'),
        ('Props - Gritcore', 200, 400, 'Biome - Portal Peak', 'Environment Art'),
        ('Props - Fantasy Exterior', 200, 400, 'Biome - Portal Peak', 'Environment Art'),
        ('Furniture - Fantasy', 40, 40, 'User Space Decoration', 'Environment Art'),
        ('Furniture - Gritcore', 320, 320, 'User Space Decoration', 'Environment Art'),
        ('Biomes - Cave', 350, 400, 'Biome - Portal Peak', 'Environment Art'),
        ('Biomes - New Zone', 350, 400, 'Biome - Farmlands', 'Environment Art'),
        ('Biomes - Downtime', 350, 400, 'Biome - Downtime', 'Environment Art'),
        ('Tower Exterior (Epic POI)', 568, 640, 'Biome - Portal Peak', 'Environment Art'),
        ('The Big Battle Hill (Epic POI)', 568, 640, 'Biome - Hidden Hills', 'Environment Art'),
        ('Cave Interior (Med POI)', 224, 300, 'Biome - Portal Peak', 'Environment Art'),
        ('Waterfall Beauty (Small POI)', 80, 120, 'Biome - Whispering Woods', 'Environment Art'),
        ('Bridge (Small POI)', 80, 120, 'Biome - Portal Peak', 'Environment Art'),
        ("Drifter's Cross (Small POI)", 320, 320, 'Biome - Farmlands', 'Environment Art'),
    ]

    art_count = 0
    for name, min_h, max_h, feature, team in ART_TASKS:
        if feature in feature_stories:
            min_d = round(min_h / H2D, 1)
            max_d = round(max_h / H2D, 1)
            story = {
                'name': name,
                'team': team,
                'design_est': {'_final': max_d, '_art_min': min_d},
                'eng_est': None,
                '_art_min': min_d,
                '_art_max': max_d,
            }
            feature_stories[feature].append(story)
            art_count += 1

    print(f"  (includes {art_count} Miro art tasks as story rows)")

    # Miro board stories (from org_chart summaries — not in estimation sheets)
    MIRO_STORIES = [
        # (name, template_feature, dept, team, min_days_7h, max_days_7h)
        # Player Build
        ('Faction: Clothing Variations', 'Character Creation & Customization', 'Art', 'Character Art', None, None),
        ('Faction: Banners', 'Character Creation & Customization', 'Art', 'Concept Art', None, None),
        ('Faction: Icons', 'Character Creation & Customization', 'Art', 'Concept Art', None, None),
        ('Faction: Insignias', 'Character Creation & Customization', 'Art', 'Concept Art', None, None),
        ('Garment System Options', 'Garment System', 'Art', 'Character Art', None, None),
        ('Garment Sets', 'Garment System', 'Art', 'Character Art', None, None),
        ('Weapon Sets', 'Garment System', 'Art', 'Character Art', None, None),
        ('Weapon Parts', 'Garment System', 'Art', 'Character Art', None, None),
        ('Armour Sets', 'Garment System', 'Art', 'Character Art', None, None),
        ('Partner Skins', 'Garment System', 'Art', 'Character Art', None, None),
        ('Skins', 'Garment System', 'Art', 'Character Art', None, None),
        ('Base Body (MVP DONE)', 'Character Creation & Customization', 'Art', 'Character Art', None, None),
        ('Base Faces (MVP DONE)', 'Character Creation & Customization', 'Art', 'Character Art', None, None),
        ('Fantasy Races', 'Character Creation & Customization', 'Art', 'Character Art', None, None),
        ('Ethnicity Variations', 'Character Creation & Customization', 'Art', 'Character Art', None, None),
        ('Gathering', 'Professions', 'Design', 'Game design', None, None),
        ('Crafting', 'Professions', 'Design', 'Game design', None, None),
        ('Trophies', 'Achievements & Trophies', 'Art', 'Concept Art', None, None),
        ('Char Creation System', 'Character Creation & Customization', 'Art', 'Character Art', None, None),
        # World Systems ENV
        ('Mediterranean Biome', 'Biome - Portal Peak', 'Art', 'Environment Art', None, None),
        ('Lush Biome', 'Biome - Farmlands', 'Art', 'Environment Art', None, None),
        ('Lighting Tech', 'Navigation', 'Art', 'Environment Art', None, None),
        ('Atmospherics', 'Navigation', 'Art', 'Environment Art', None, None),
        ('Water System', 'Navigation', 'Art', 'Environment Art', None, None),
        ('VFX Tech', 'Foam Corruption', 'Art', 'VFX', None, None),
        ('Env Prop Kit - Tavern', 'Biome - Downtime', 'Art', 'Environment Art', None, None),
        ('Fluid Ninja Tech R&D', 'Foam Corruption', 'Art', 'TechArt', None, None),
        # World Systems REST - NPCs
        ('Green Faction NPCs', 'Enemies Art', 'Art', 'Character Art', None, None),
        ('Red Faction NPCs', 'Enemies Art', 'Art', 'Character Art', None, None),
        ('Yellow Faction NPCs', 'Enemies Art', 'Art', 'Character Art', None, None),
        ('Blue Faction NPCs', 'Enemies Art', 'Art', 'Character Art', None, None),
        ('Critter NPCs - Pigs (DONE)', 'Enemies Art', 'Art', 'Character Art', None, None),
        ('Critter NPCs - Hares (DONE)', 'Enemies Art', 'Art', 'Character Art', None, None),
        ('Critter NPCs - Sheep (DONE)', 'Enemies Art', 'Art', 'Character Art', None, None),
        ('Critter NPCs - Cows (DONE)', 'Enemies Art', 'Art', 'Character Art', None, None),
        ('Critter NPCs - Chickens (DONE)', 'Enemies Art', 'Art', 'Character Art', None, None),
        ('Critter NPCs - Horses (DONE)', 'Enemies Art', 'Art', 'Character Art', None, None),
        ('Hero NPC - Xyron (Upgrade)', 'Enemies Art', 'Art', 'Character Art', None, None),
        ('Hero NPC - Barbersmith (Upgrade)', 'Enemies Art', 'Art', 'Character Art', None, None),
        ('Hero NPC - Maeve/Mayor (Upgrade)', 'Enemies Art', 'Art', 'Character Art', None, None),
        ('Gliding System', 'Advanced Traversal', 'Art', 'Animation & Rigging', None, None),
        ('Grappling System', 'Advanced Traversal', 'Art', 'Animation & Rigging', None, None),
        # User Space
        ('Indoor Camera', 'User Space Decoration', 'Engineering', 'Gameplay Engineering', None, None),
        ('User Space Interior', 'User Space Decoration', 'Art', 'Environment Art', None, None),
        ('User Space Exterior', 'User Space Decoration', 'Art', 'Environment Art', None, None),
        ('Zone - User Space Island', 'User Space Server & Instance', 'Art', 'Environment Art', None, None),
        ('Modular Houses', 'Building Houses', 'Art', 'Environment Art', None, None),
        # Items & Inventory
        ('Potions', 'Consumables', 'Art', 'Concept Art', None, None),
        ('MECH PARROT', 'Mounts / Pets', 'Art', 'Character Art', None, None),
        # Quest System
        ('Lock Picking Quest', 'Quest Types - Courier / Escort / Kill / Collect / Mystery / Time / Partner / Lockpicking', 'Design', 'Game design', None, None),
        ('Hearthstone (Return Home Stone)', 'Quest Reward System', 'Art', 'Concept Art', None, None),
        ('Partner Crystal', 'Quest Reward System', 'Art', 'Concept Art', None, None),
        # Player Economy
        ('Currency System Art', 'Currency System', 'Art', 'Concept Art', None, None),
        ('Bugged Dungeon Shop', 'Dungeon Mechanics', 'Art', 'Environment Art', None, None),
        # Game Bibles (hours converted at 7h/day)
        ('Colour Palette', 'Brand Bible', 'Art', 'Concept Art', 5.7, 8.6),
        ('Typography', 'Brand Bible', 'Art', 'Concept Art', 5.7, 8.6),
        ('Wireframes', 'UX/UI Bible', 'Art', 'UX/UI', 5.7, 8.6),
        ('Game Logo', 'Brand Bible', 'Art', 'Concept Art', 11.4, 14.3),
        ('Environment Bible', 'Art Bible', 'Art', 'Environment Art', None, None),
        ('Animation Bible', 'Art Bible', 'Art', 'Animation & Rigging', None, None),
        ('Character Bible', 'Art Bible', 'Art', 'Character Art', None, None),
        ('VFX Bible', 'Art Bible', 'Art', 'VFX', None, None),
        ('Concept Bible', 'Art Bible', 'Art', 'Concept Art', None, None),
    ]

    miro_story_count = 0
    for name, feature, dept, team, min_d, max_d in MIRO_STORIES:
        if feature in feature_stories:
            story = {
                'name': name,
                'team': team,
                'design_est': None,
                'eng_est': None,
                '_art_min': min_d,
                '_art_max': max_d,
            }
            feature_stories[feature].append(story)
            miro_story_count += 1

    print(f"  (includes {miro_story_count} Miro board stories)")

    # Stats
    total_stories = sum(len(v) for v in feature_stories.values())
    features_with_stories = sum(1 for v in feature_stories.values() if v)
    print(f"Mapped {total_stories} stories across {features_with_stories} features")

    return feature_stories


# ============================================================
# ADDITIONAL SHEETS
# ============================================================

# All staff from both estimation sheets
ALL_STAFF = [
    # --- ENGINEERING DEPARTMENT ---
    {'name': 'Mustafa Sibai', 'role': 'CTO', 'team': 'Engineering', 'dept': 'Engineering'},
    # Web Development Team
    {'name': 'Daniel Negri', 'role': 'Full Stack Developer', 'team': 'Web Development', 'dept': 'Engineering'},
    {'name': 'Nikos Gerontakis', 'role': 'Full Stack Developer', 'team': 'Web Development', 'dept': 'Engineering'},
    {'name': 'Samer Balushi', 'role': 'Backend Developer', 'team': 'Web Development', 'dept': 'Engineering'},
    # Gameplay Team
    {'name': "Raynor D'Souza", 'role': 'Gameplay Developer', 'team': 'Gameplay', 'dept': 'Engineering'},
    {'name': 'Ilya Achmetov', 'role': 'Gameplay Developer', 'team': 'Gameplay', 'dept': 'Engineering'},
    {'name': 'Leon Huang', 'role': 'Gameplay Developer', 'team': 'Gameplay', 'dept': 'Engineering'},
    {'name': 'Ignasi Ezpeleta', 'role': 'Gameplay Developer', 'team': 'Gameplay', 'dept': 'Engineering'},
    # From estimation sheets, not on current org chart
    {'name': 'Roberto', 'role': 'Gameplay Developer', 'team': 'Gameplay', 'dept': 'Engineering'},
    {'name': 'Matt', 'role': 'Animator', 'team': 'Gameplay', 'dept': 'Engineering'},
    # --- ART DEPARTMENT ---
    {'name': 'David Luong', 'role': 'Art Director', 'team': 'Art Direction', 'dept': 'Art'},
    # VFX Team
    {'name': 'Wojciech Szon', 'role': 'VFX Artist', 'team': 'VFX', 'dept': 'Art'},
    {'name': 'Conor McLean', 'role': 'VFX Artist', 'team': 'VFX', 'dept': 'Art'},
    {'name': 'Michail Velissaris', 'role': 'VFX Artist', 'team': 'VFX', 'dept': 'Art'},
    # Technical Art Team
    {'name': 'Stefano Vietina', 'role': 'Technical Artist', 'team': 'Technical Art', 'dept': 'Art'},
    {'name': 'Samuele Piazzi', 'role': 'Technical Artist', 'team': 'Technical Art', 'dept': 'Art'},
    {'name': 'Andrea Rottini', 'role': 'Technical Artist', 'team': 'Technical Art', 'dept': 'Art'},
    # Concept Art
    {'name': 'Niko Gesell', 'role': 'Concept Artist', 'team': 'Concept Art', 'dept': 'Art'},
    {'name': 'Kerim Turay', 'role': 'Concept Artist', 'team': 'Concept Art', 'dept': 'Art'},
    {'name': 'Gerard Vicen', 'role': 'Concept Artist', 'team': 'Concept Art', 'dept': 'Art'},
    {'name': 'Julian Hartinger', 'role': 'Concept Artist', 'team': 'Concept Art', 'dept': 'Art'},
    {'name': 'Gianluca Perusi', 'role': 'UI/UX Designer', 'team': 'Concept Art', 'dept': 'Art'},
    {'name': 'Larisa Logofatu', 'role': 'Graphic Designer', 'team': 'Concept Art', 'dept': 'Art'},
    # 3D Animation Team
    {'name': 'Alon', 'role': 'Animator', 'team': '3D Animation', 'dept': 'Art'},
    {'name': 'Kunjal Asher', 'role': 'Animator', 'team': '3D Animation', 'dept': 'Art'},
    {'name': 'Maddalena Morello', 'role': 'Character Rigger', 'team': '3D Animation', 'dept': 'Art'},
    # 3D Characters Team
    {'name': 'Sasha Krieger', 'role': 'Character Artist', 'team': '3D Characters', 'dept': 'Art'},
    {'name': 'Dimitris Chrysanthakopoulos', 'role': 'Character Artist', 'team': '3D Characters', 'dept': 'Art'},
    {'name': 'Rebecca Tomasoni', 'role': 'Character Artist', 'team': '3D Characters', 'dept': 'Art'},
    # 3D Environment Team
    {'name': 'Michael Dunnam', 'role': 'Environment Artist', 'team': '3D Environment', 'dept': 'Art'},
    {'name': 'Ella Amatooni', 'role': 'Environment Artist', 'team': '3D Environment', 'dept': 'Art'},
    {'name': 'Marc Vives', 'role': 'Environment Artist', 'team': '3D Environment', 'dept': 'Art'},
    {'name': 'Iosif Alexiou', 'role': 'Environment Artist', 'team': '3D Environment', 'dept': 'Art'},
    {'name': 'Nick Chrysanthakopoulos', 'role': 'Prop Artist', 'team': '3D Environment', 'dept': 'Art'},
    {'name': 'Chris Brough', 'role': 'Environment Artist', 'team': '3D Environment', 'dept': 'Art'},
    # --- PRODUCT DEPARTMENT (Design) ---
    {'name': 'Robin Jubber', 'role': 'Game Director', 'team': 'Product', 'dept': 'Design'},
    # Game Design Team
    {'name': 'Yorgos Dritsas', 'role': 'Game Designer', 'team': 'Game Design', 'dept': 'Design'},
    {'name': 'Nadir Latif', 'role': 'Combat Designer', 'team': 'Game Design', 'dept': 'Design'},
    # Technical Designer (direct report)
    {'name': 'Kieron Naylor', 'role': 'Technical Designer', 'team': 'Design', 'dept': 'Design'},
    # Level Design Team
    {'name': 'Nikolas Tziotis', 'role': 'Level Designer', 'team': 'Level Design', 'dept': 'Design'},
    {'name': 'Seth Dahl', 'role': 'Level Designer', 'team': 'Level Design', 'dept': 'Design'},
    {'name': 'Panos Andriopoulos', 'role': 'Level Designer', 'team': 'Level Design', 'dept': 'Design'},
    # Writing Team
    {'name': 'Maria Cibej', 'role': 'Writer', 'team': 'Writing', 'dept': 'Design'},
    # Sound Design Team
    {'name': 'Dragan Vuckovic', 'role': 'Sound Designer', 'team': 'Sound Design', 'dept': 'Audio'},
    {'name': 'Aggeliki Peroutsea', 'role': 'Sound Designer', 'team': 'Sound Design', 'dept': 'Audio'},
    # QA Team
    {'name': 'Hannah Pickard', 'role': 'QA', 'team': 'QA', 'dept': 'QA'},
]


def build_staff_sheet(wb):
    """Staff reference sheet + named ranges for dropdowns.
    Dropdowns are DEPARTMENT-level: all designers together, all engineers together, etc.
    So a Game Designer story shows all Design dept staff, not just Game Designers.
    """
    ws = wb.add_worksheet('Staff')
    hdr = wb.add_format({'bg_color': '#404040', 'font_color': 'white', 'bold': True, 'font_size': 11})
    cell = wb.add_format({'font_size': 10, 'valign': 'vcenter'})

    ws.write(0, 0, 'Name', hdr)
    ws.write(0, 1, 'Role', hdr)
    ws.write(0, 2, 'Team', hdr)
    ws.write(0, 3, 'Department', hdr)

    for i, s in enumerate(ALL_STAFF, 1):
        ws.write(i, 0, s['name'], cell)
        ws.write(i, 1, s['role'], cell)
        ws.write(i, 2, s['team'], cell)
        ws.write(i, 3, s['dept'], cell)

    ws.set_column(0, 0, 22)
    ws.set_column(1, 1, 22)
    ws.set_column(2, 2, 22)
    ws.set_column(3, 3, 14)

    # Named ranges per DEPARTMENT (wider dropdown lists)
    from collections import defaultdict
    dept_staff = defaultdict(list)
    for i, s in enumerate(ALL_STAFF, 1):
        dept_staff[s['dept']].append(i)

    for dept, row_indices in dept_staff.items():
        range_name = re.sub(r'[^A-Za-z0-9_]', '_', dept) + '_Staff'
        first = min(row_indices)
        last = max(row_indices)
        wb.define_name(range_name, f"=Staff!$A${first+1}:$A${last+1}")

    # Also create an All_Staff range for roles that don't map to a department
    wb.define_name('All_Staff', f"=Staff!$A$2:$A${len(ALL_STAFF)+1}")

    # Map team -> department for dropdown lookups
    team_to_dept_dropdown = {}
    for s in ALL_STAFF:
        team_to_dept_dropdown[s['role']] = re.sub(r'[^A-Za-z0-9_]', '_', s['dept']) + '_Staff'

    print(f"  Staff sheet: {len(ALL_STAFF)} people, {len(dept_staff)} department ranges")
    return dept_staff, team_to_dept_dropdown


def build_reference_sheets(wb):
    """Restore Roles, Gates, Tiers, Legend from v11 data."""
    import json
    with open('_v11_sheets.json', 'r') as f:
        v11 = json.load(f)

    hdr = wb.add_format({'bg_color': '#404040', 'font_color': 'white', 'bold': True, 'font_size': 11})
    cell = wb.add_format({'font_size': 10, 'valign': 'vcenter'})

    for sheet_name in ['Roles', 'Gates', 'Tiers', 'Legend']:
        ws = wb.add_worksheet(sheet_name)
        rows = v11[sheet_name]
        for r, row_data in enumerate(rows):
            for c, val in enumerate(row_data):
                if val is not None:
                    fmt = hdr if r == 0 else cell
                    ws.write(r, c, val, fmt)
        ws.set_column(0, len(rows[0]) - 1 if rows else 0, 20)
        print(f"  {sheet_name}: {len(rows)} rows")


def build_summary_sheet(wb, epic_row_ranges, template_order):
    """Formula-driven summary — EP totals per epic per role."""
    ws = wb.add_worksheet('Summary')
    hdr = wb.add_format({'bg_color': '#404040', 'font_color': 'white', 'bold': True, 'font_size': 10})
    epic_fmt = wb.add_format({'bg_color': '#FFD966', 'bold': True, 'font_size': 10})
    num_fmt = wb.add_format({'font_size': 10, 'num_format': '0.##'})
    total_fmt = wb.add_format({'bg_color': '#A9D08E', 'bold': True, 'font_size': 10, 'num_format': '0'})

    # Visible EP roles (indices 10-37)
    visible_roles = ROLES[10:]

    ws.write(0, 0, 'Epic', hdr)
    ws.write(0, 1, 'Stories', hdr)
    for i, role in enumerate(visible_roles):
        ws.write(0, 2 + i * 2, f'{role} Min', hdr)
        ws.write(0, 3 + i * 2, f'{role} Max', hdr)
    total_col = 2 + len(visible_roles) * 2
    ws.write(0, total_col, 'Total Max', hdr)

    ep_layout = COL_LAYOUT['EARLY PROD']

    epics = [t['name'] for t in template_order if t['type'] == 'epic']
    for r, epic in enumerate(epics, 1):
        ws.write(r, 0, epic, epic_fmt)

        if epic in epic_row_ranges:
            sr, er = epic_row_ranges[epic]
            sr1 = sr + 1  # convert to 1-indexed for Excel
            er1 = er + 1
            ws.write_formula(r, 1, f'=COUNTA(Plan!C{sr1}:C{er1})', num_fmt)

            for i, role in enumerate(visible_roles):
                ri = 10 + i
                min_col_letter = get_column_letter(ep_layout['role_start'] + ri * 2)
                max_col_letter = get_column_letter(ep_layout['role_start'] + ri * 2 + 1)
                ws.write_formula(r, 2 + i * 2,
                    f'=SUM(Plan!{min_col_letter}{sr1}:{min_col_letter}{er1})', num_fmt)
                ws.write_formula(r, 3 + i * 2,
                    f'=SUM(Plan!{max_col_letter}{sr1}:{max_col_letter}{er1})', num_fmt)

            max_cols = [get_column_letter(ep_layout['role_start'] + (10 + i) * 2 + 1) for i in range(len(visible_roles))]
            sum_parts = '+'.join(f'SUM(Plan!{mc}{sr1}:{mc}{er1})' for mc in max_cols)
            ws.write_formula(r, total_col, f'={sum_parts}', total_fmt)
        else:
            ws.write(r, 1, 0, num_fmt)

    # Totals row
    tr = len(epics) + 1
    ws.write(tr, 0, 'TOTAL', hdr)
    for c in range(1, total_col + 1):
        col_letter = get_column_letter(c + 1)
        ws.write_formula(tr, c, f'=SUM({col_letter}2:{col_letter}{tr})', total_fmt)

    ws.set_column(0, 0, 25)
    ws.set_column(1, 1, 8)
    ws.set_column(2, total_col, 10)
    ws.freeze_panes(1, 1)
    print(f"  Summary: {len(epics)} epics x {len(visible_roles)} roles")


def build_epics_features_sheet(wb, template_order, epic_row_ranges):
    """Epics & Features with formula-driven story counts."""
    ws = wb.add_worksheet('Epics & Features')
    hdr = wb.add_format({'bg_color': '#404040', 'font_color': 'white', 'bold': True, 'font_size': 11})
    epic_fmt = wb.add_format({'bg_color': '#FFD966', 'bold': True, 'font_size': 10})
    cell = wb.add_format({'font_size': 10})
    num_fmt = wb.add_format({'font_size': 10, 'num_format': '0'})

    ws.write(0, 0, 'Epic', hdr)
    ws.write(0, 1, 'Feature', hdr)
    ws.write(0, 2, 'Sizing', hdr)
    ws.write(0, 3, 'Current Tier', hdr)
    ws.write(0, 4, 'Stories', hdr)

    r = 1
    for item in template_order:
        if item['type'] == 'epic':
            ws.write(r, 0, item['name'], epic_fmt)
        elif item['type'] == 'feature':
            ws.write(r, 1, item['name'], cell)
            if item.get('size'):
                ws.write(r, 2, item['size'], cell)
            if item.get('tier'):
                ws.write(r, 3, item['tier'], cell)
            ws.write_formula(r, 4,
                f'=COUNTIF(Plan!B:B,B{r+1})', num_fmt)
        r += 1

    ws.set_column(0, 0, 20)
    ws.set_column(1, 1, 55)
    ws.set_column(2, 2, 8)
    ws.set_column(3, 3, 18)
    ws.set_column(4, 4, 8)
    print(f"  Epics & Features: {r} rows")


def build_resource_planner(wb, template_order, feature_stories, team_to_dept_dropdown):
    """Pre-populated resource planner with role-filtered dropdowns."""
    ws = wb.add_worksheet('Resource Planner')

    hdr = wb.add_format({'bg_color': '#305496', 'font_color': 'white', 'bold': True,
                          'font_size': 11, 'align': 'center', 'valign': 'vcenter', 'text_wrap': True})
    # Zebra stripe: alternating off-white and white rows
    OW = '#F2F2F2'
    lock_w = wb.add_format({'font_size': 10, 'valign': 'vcenter'})
    lock_ow = wb.add_format({'font_size': 10, 'valign': 'vcenter', 'bg_color': OW})
    lock_bold_w = wb.add_format({'font_size': 10, 'valign': 'vcenter', 'bold': True})
    lock_bold_ow = wb.add_format({'font_size': 10, 'valign': 'vcenter', 'bold': True, 'bg_color': OW})
    epic_fmt = wb.add_format({'bg_color': '#FFD966', 'bold': True, 'font_size': 11})
    num_w = wb.add_format({'font_size': 10, 'valign': 'vcenter', 'num_format': '0.##'})
    num_ow = wb.add_format({'font_size': 10, 'valign': 'vcenter', 'num_format': '0.##', 'bg_color': OW})
    input_w = wb.add_format({'font_size': 10, 'valign': 'vcenter', 'bg_color': '#E2EFDA', 'num_format': '0.##'})
    input_ow = wb.add_format({'font_size': 10, 'valign': 'vcenter', 'bg_color': '#D5E8C7', 'num_format': '0.##'})
    formula_w = wb.add_format({'font_size': 10, 'valign': 'vcenter', 'num_format': '0.##', 'font_color': '#808080'})
    formula_ow = wb.add_format({'font_size': 10, 'valign': 'vcenter', 'num_format': '0.##', 'font_color': '#808080', 'bg_color': OW})
    alert_fmt = wb.add_format({'font_size': 10, 'valign': 'vcenter', 'num_format': '0.##', 'font_color': 'red', 'bold': True})

    headers = ['Epic', 'Feature', 'Story', 'Role', 'Team', 'Min Days', 'Max Days',
               'Primary Assignee', 'Primary Days', 'Secondary Assignee', 'Secondary Days', 'Unallocated']
    for i, h in enumerate(headers):
        ws.write(0, i, h, hdr)

    row = 1
    for item in template_order:
        if item['type'] == 'epic':
            ws.merge_range(row, 0, row, len(headers) - 1, item['name'], epic_fmt)
            row += 1
        elif item['type'] == 'feature':
            fname = item['name']
            stories = feature_stories.get(fname, [])

            story_idx = 0
            if stories:
                for story in stories:
                    team = story['team']
                    dept = TEAM_TO_DEPT.get(team, team)
                    role = TEAM_TO_ROLE.get(team, team)
                    is_odd = story_idx % 2 == 1  # zebra stripe

                    # Pick formats based on zebra
                    f_lock = lock_ow if is_odd else lock_w
                    f_bold = lock_bold_ow if is_odd else lock_bold_w
                    f_num = num_ow if is_odd else num_w
                    f_input = input_ow if is_odd else input_w
                    f_formula = formula_ow if is_odd else formula_w

                    est_min = est_max = None
                    # Art task shortcut (from Miro board)
                    if story.get('_art_min') is not None:
                        est_min = story['_art_min']
                        est_max = story['_art_max']
                    elif story.get('design_est'):
                        est = story['design_est']
                        ind_mins = [v['min'] for k, v in est.items()
                                   if not k.startswith('_') and isinstance(v, dict) and v.get('min') is not None]
                        if ind_mins:
                            est_min = min(ind_mins)
                        est_max = est.get('_final')
                        if est_max is None:
                            ind_maxes = [v['max'] for k, v in est.items()
                                        if not k.startswith('_') and isinstance(v, dict) and v.get('max') is not None]
                            if ind_maxes:
                                est_max = max(ind_maxes)
                    if story.get('eng_est'):
                        est = story['eng_est']
                        m_min = est.get('_mustafa_min')
                        m_max = est.get('_mustafa_max')
                        if not m_min or not m_max:
                            ind_mins = [v['min'] for k, v in est.items()
                                       if not k.startswith('_') and isinstance(v, dict) and v.get('min') is not None]
                            ind_maxes = [v['max'] for k, v in est.items()
                                        if not k.startswith('_') and isinstance(v, dict) and v.get('max') is not None]
                            if not m_min and ind_mins:
                                m_min = min(ind_mins)
                            if not m_max and ind_maxes:
                                m_max = max(ind_maxes)
                        if m_min is not None:
                            est_min = m_min
                        if m_max is not None:
                            est_max = m_max

                    ws.write(row, 0, item['name'], f_lock)
                    ws.write(row, 1, fname, f_bold)
                    ws.write(row, 2, story['name'], f_lock)
                    ws.write(row, 3, role, f_lock)
                    ws.write(row, 4, f"{dept} | {team}", f_lock)
                    if est_min is not None:
                        ws.write(row, 5, est_min, f_num)
                    else:
                        ws.write_blank(row, 5, None, f_lock)
                    if est_max is not None:
                        ws.write(row, 6, est_max, f_num)
                    else:
                        ws.write_blank(row, 6, None, f_lock)

                    # Department-level dropdown for assignee
                    dd_range = team_to_dept_dropdown.get(role, 'All_Staff')
                    ws.data_validation(row, 7, row, 7, {
                        'validate': 'list',
                        'source': f'={dd_range}',
                        'error_message': f'Select staff from {dept} department'
                    })
                    ws.write_blank(row, 7, None, f_input)

                    r1 = row + 1
                    ws.write_formula(row, 8, f'=IF(H{r1}="",0,G{r1})', f_input)

                    # Secondary assignee (same department dropdown)
                    ws.data_validation(row, 9, row, 9, {
                        'validate': 'list',
                        'source': f'={dd_range}',
                        'error_message': f'Select staff from {dept} department'
                    })
                    ws.write_blank(row, 9, None, f_input)
                    ws.write(row, 10, 0, f_input)

                    ws.write_formula(row, 11, f'=G{r1}-I{r1}-K{r1}', f_formula)

                    row += 1
                    story_idx += 1

    ws.set_column(0, 0, 18)
    ws.set_column(1, 1, 30)
    ws.set_column(2, 2, 45)
    ws.set_column(3, 3, 20)
    ws.set_column(4, 4, 28)
    ws.set_column(5, 5, 10)
    ws.set_column(6, 6, 10)
    ws.set_column(7, 7, 22)
    ws.set_column(8, 8, 12)
    ws.set_column(9, 9, 22)
    ws.set_column(10, 10, 12)
    ws.set_column(11, 11, 12)
    ws.freeze_panes(1, 3)
    ws.set_zoom(110)
    ws.autofilter(0, 0, row - 1, len(headers) - 1)

    # Conditional format: red if unallocated > 0 and someone is assigned
    ws.conditional_format(1, 11, row - 1, 11, {
        'type': 'formula',
        'criteria': '=AND(L2>0,H2<>"")',
        'format': alert_fmt
    })

    print(f"  Resource Planner: {row - 1} rows")
    return row


def build_person_dashboard(wb):
    """Per-person workload dashboard — formula-driven from Resource Planner."""
    ws = wb.add_worksheet('Person Dashboard')
    hdr = wb.add_format({'bg_color': '#305496', 'font_color': 'white', 'bold': True,
                          'font_size': 11, 'align': 'center', 'valign': 'vcenter'})
    cell = wb.add_format({'font_size': 10, 'valign': 'vcenter'})
    num = wb.add_format({'font_size': 10, 'valign': 'vcenter', 'num_format': '0'})
    pct = wb.add_format({'font_size': 10, 'valign': 'vcenter', 'num_format': '0%'})
    input_fmt = wb.add_format({'font_size': 10, 'valign': 'vcenter', 'bg_color': '#E2EFDA', 'num_format': '0'})
    alert = wb.add_format({'font_size': 10, 'bold': True, 'font_color': 'red'})

    headers = ['Person', 'Role', 'Department', 'Primary Stories', 'Primary Days',
               'Secondary Stories', 'Secondary Days', 'Total Days', 'Available Days', 'Utilisation', 'Overloaded?']
    for i, h in enumerate(headers):
        ws.write(0, i, h, hdr)

    for r, s in enumerate(ALL_STAFF, 1):
        ws.write(r, 0, s['name'], cell)
        ws.write(r, 1, s['role'], cell)
        ws.write(r, 2, s['dept'], cell)
        n = s['name']
        rp = "'Resource Planner'"
        ws.write_formula(r, 3, f'=COUNTIF({rp}!H:H,"{n}")', num)
        ws.write_formula(r, 4, f'=SUMPRODUCT(({rp}!H$2:H$9999="{n}")*{rp}!I$2:I$9999)', num)
        ws.write_formula(r, 5, f'=COUNTIF({rp}!J:J,"{n}")', num)
        ws.write_formula(r, 6, f'=SUMPRODUCT(({rp}!J$2:J$9999="{n}")*{rp}!K$2:K$9999)', num)
        ws.write_formula(r, 7, f'=E{r+1}+G{r+1}', num)
        ws.write(r, 8, 0, input_fmt)  # User inputs available days
        ws.write_formula(r, 9, f'=IF(I{r+1}=0,0,H{r+1}/I{r+1})', pct)
        ws.write_formula(r, 10, f'=IF(J{r+1}>1,"YES","")', alert)

    ws.set_column(0, 0, 22)
    ws.set_column(1, 1, 22)
    ws.set_column(2, 2, 14)
    ws.set_column(3, 6, 14)
    ws.set_column(7, 7, 12)
    ws.set_column(8, 8, 14)
    ws.set_column(9, 9, 12)
    ws.set_column(10, 10, 12)
    ws.freeze_panes(1, 1)
    print(f"  Person Dashboard: {len(ALL_STAFF)} staff")


def build_feature_timeline(wb, template_order, feature_stories):
    """Per-feature completion timeline — formula-driven from Resource Planner."""
    ws = wb.add_worksheet('Feature Timeline')
    hdr = wb.add_format({'bg_color': '#305496', 'font_color': 'white', 'bold': True,
                          'font_size': 11, 'align': 'center', 'valign': 'vcenter'})
    epic_fmt = wb.add_format({'bg_color': '#FFD966', 'bold': True, 'font_size': 10})
    cell = wb.add_format({'font_size': 10, 'valign': 'vcenter'})
    num = wb.add_format({'font_size': 10, 'valign': 'vcenter', 'num_format': '0'})

    headers = ['Epic', 'Feature', 'Total Stories', 'Total Max Days', 'Roles Involved']
    for i, h in enumerate(headers):
        ws.write(0, i, h, hdr)

    r = 1
    current_epic = None
    for item in template_order:
        if item['type'] == 'epic':
            current_epic = item['name']
            ws.write(r, 0, current_epic, epic_fmt)
            r += 1
        elif item['type'] == 'feature':
            fname = item['name']
            stories = feature_stories.get(fname, [])
            if not stories:
                continue

            teams = set()
            total_max = 0
            for s in stories:
                teams.add(TEAM_TO_ROLE.get(s['team'], s['team']))
                if s.get('design_est'):
                    v = s['design_est'].get('_final')
                    if v:
                        total_max += v
                if s.get('eng_est'):
                    v = s['eng_est'].get('_mustafa_max')
                    if v:
                        total_max += v

            ws.write(r, 0, current_epic, cell)
            ws.write(r, 1, fname, cell)
            ws.write(r, 2, len(stories), num)
            ws.write(r, 3, total_max, num)
            ws.write(r, 4, ', '.join(sorted(teams)), cell)
            r += 1

    ws.set_column(0, 0, 20)
    ws.set_column(1, 1, 55)
    ws.set_column(2, 2, 12)
    ws.set_column(3, 3, 14)
    ws.set_column(4, 4, 50)
    ws.freeze_panes(1, 2)
    print(f"  Feature Timeline: {r - 1} rows")


# ============================================================
# MAIN
# ============================================================

def main():
    import os
    base = os.path.dirname(os.path.abspath(__file__))

    print("=== Parsing Design estimation sheet ===")
    design_rows = parse_estimation_sheet(
        os.path.join(base, 'VS Estimation Design - General Sheet.xlsx'),
        'Sorted by Epic', is_engineering=False
    )
    print(f"  {len(design_rows)} story/task rows")

    print("=== Parsing Engineering estimation sheet ===")
    eng_rows = parse_estimation_sheet(
        os.path.join(base, 'VS Estimation Engineering - General Sheet (1).xlsx'),
        'Sheet1', is_engineering=True
    )
    print(f"  {len(eng_rows)} story/task rows")

    print("=== Parsing Miro hours mapping ===")
    miro_data = parse_hours_mapping(os.path.join(base, 'hours_mapping.txt'))
    print(f"  {len(miro_data)} features with role totals (game systems)")

    print("=== Parsing Miro art estimates ===")
    miro_art = get_miro_art_estimates()
    # Merge art into miro_data: art has min/max per role, miro_data has single values
    # For miro_data, the existing values become the Max; art adds Min and Max
    for feat, roles in miro_art.items():
        if feat not in miro_data:
            miro_data[feat] = {}
        for role, vals in roles.items():
            if role in miro_data[feat]:
                # Accumulate art days onto existing Miro totals
                existing = miro_data[feat][role]
                if isinstance(existing, dict):
                    existing['min'] = existing.get('min', 0) + vals['min']
                    existing['max'] = existing.get('max', 0) + vals['max']
                else:
                    miro_data[feat][role] = {'min': vals['min'], 'max': existing + vals['max']}
            else:
                miro_data[feat][role] = vals

    print("=== Loading template order ===")
    template_order = load_template_order()
    epics = sum(1 for t in template_order if t['type'] == 'epic')
    features = sum(1 for t in template_order if t['type'] == 'feature')
    print(f"  {epics} epics, {features} features")

    print("=== Building workbook ===")
    output_path = os.path.join(base, 'CouchHeroes_Man_Day_Work_Plan_v15.xlsx')

    # Build feature->stories lookup (used by multiple sheets)
    feature_stories = build_feature_stories(design_rows, eng_rows, template_order)

    # Plan sheet (main data)
    wb, plan_rows, epic_row_ranges = build_workbook(template_order, design_rows, eng_rows, miro_data, output_path)

    # Staff reference sheet + named ranges
    print("  Building additional sheets...")
    dept_staff, team_to_dept_dropdown = build_staff_sheet(wb)

    # Restored v11 reference sheets
    build_reference_sheets(wb)

    # Formula-driven Summary
    build_summary_sheet(wb, epic_row_ranges, template_order)

    # Formula-driven Epics & Features
    build_epics_features_sheet(wb, template_order, epic_row_ranges)

    # Resource Planner (pre-populated with dropdowns)
    build_resource_planner(wb, template_order, feature_stories, team_to_dept_dropdown)

    # Person Dashboard (formula-driven)
    build_person_dashboard(wb)

    # Feature Timeline (pre-computed)
    build_feature_timeline(wb, template_order, feature_stories)

    wb.close()
    print(f"\n=== SAVED: {output_path} ===")


if __name__ == '__main__':
    main()
