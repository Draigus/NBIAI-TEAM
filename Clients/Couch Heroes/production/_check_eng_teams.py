import openpyxl
wb = openpyxl.load_workbook('VS Estimation Engineering - General Sheet (1).xlsx', data_only=True)
ws = wb['Sheet1']
teams = {}
for r in range(3, ws.max_row + 1):
    e = ws.cell(row=r, column=5).value
    c = ws.cell(row=r, column=3).value
    d = ws.cell(row=r, column=4).value
    b = ws.cell(row=r, column=2).value
    if not e or not (b or c or d):
        continue
    team = str(e).strip()
    has_est = any(
        isinstance(ws.cell(row=r, column=ci).value, (int, float))
        and ws.cell(row=r, column=ci).value > 0
        for ci in range(6, 28)
    )
    if team not in teams:
        teams[team] = {'total': 0, 'with_est': 0}
    teams[team]['total'] += 1
    if has_est:
        teams[team]['with_est'] += 1

print("TEAMS IN ENGINEERING SHEET")
print("=" * 55)
for t, info in sorted(teams.items(), key=lambda x: -x[1]['total']):
    print(f"  {t:25s}: {info['total']:3d} rows, {info['with_est']:3d} with estimates")

eng_teams = {'Gameplay Engineering', 'Backend Engineering'}
non_eng = {t: info for t, info in teams.items() if t not in eng_teams}
total_non = sum(info['total'] for info in non_eng.values())
total_non_est = sum(info['with_est'] for info in non_eng.values())
print(f"\nNon-engineering rows: {total_non} ({total_non_est} with estimates)")
wb.close()
