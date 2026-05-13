"""
Goals Studio Pricing Engagement — Excel Project Tracker Generator
Output: goals_pricing_engagement_plan_v3.xlsx
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ── Colour palette ──────────────────────────────────────────────────────────
HEADER_BG   = "1F3864"   # dark navy
HEADER_FG   = "FFFFFF"   # white
TITLE_BG    = "2E74B5"   # mid-blue for section titles
TITLE_FG    = "FFFFFF"
ALT_ROW     = "EBF3FB"   # very light blue
WHITE       = "FFFFFF"
ACCENT1     = "D6E4F0"   # light accent
RED_BG      = "FCE4EC"
AMBER_BG    = "FFF8E1"
GREEN_BG    = "E8F5E9"
CRITICAL_BG = "FFCDD2"
HIGH_BG     = "FFE0B2"
MEDIUM_BG   = "FFF9C4"
LOW_BG      = "F1F8E9"

# Feature colours (subtle fills for Task Tracker)
FEAT_COLS = {
    "F1": "D6EAF8",
    "F2": "D5F5E3",
    "F3": "FAD7A0",
    "F4": "E8DAEF",
    "F5": "FDEBD0",
}

def header_style(ws, cell, text, size=11, bg=HEADER_BG, fg=HEADER_FG, bold=True, wrap=False):
    c = ws[cell]
    c.value = text
    c.font = Font(name="Calibri", size=size, bold=bold, color=fg)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    return c

def label_cell(ws, cell, text, bold=False, size=11, color="000000", bg=None, align="left", wrap=False):
    c = ws[cell]
    c.value = text
    c.font = Font(name="Calibri", size=size, bold=bold, color=color)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    return c

def thin_border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def set_col_widths(ws, widths):
    """widths: dict {col_letter: width}"""
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def apply_row_fill(ws, row, col_start, col_end, fill_color):
    fill = PatternFill("solid", fgColor=fill_color)
    for col in range(col_start, col_end + 1):
        ws.cell(row=row, column=col).fill = fill

def apply_all_borders(ws, min_row, max_row, min_col, max_col):
    bd = thin_border()
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(row=r, column=c).border = bd

def write_data_cell(ws, row, col, value, wrap=False, align="left", bold=False, bg=None, color="000000"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Calibri", size=11, bold=bold, color=color)
    c.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
    c.border = thin_border()
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    return c


# ════════════════════════════════════════════════════════════════════════════
# DATA
# ════════════════════════════════════════════════════════════════════════════

FEATURE_HOURS = [
    ("F1", "Competitive HC Pricing Benchmark", 38, 33),
    ("F2", "Regional Pricing Strategy",         34, 29),
    ("F3", "Risk Assessment, Scenarios & Recommendations", 16, 14),
    ("F4", "Value-Add (Framework + World Cup)",  8,  7),
    ("F5", "Deliverable Assembly & Client Review", 18, 16),
]

TASKS = [
    # Task ID, Feature, Story, Title, Owner, Hours, Priority, Dependencies, Notes
    ("T1.0.1", "F1", "1.0 Foundation — Map the HC Economy",
     "Map HC → Value Chain",
     "Glen", 3, "Critical",
     "None",
     "Document complete path from real money to in-game value. Map every HC store item to its coin cost. Identify 'minimum meaningful purchase'. Flag gaps for Julius (Mech/Pulse kit HC pricing unconfirmed)."),

    ("T1.0.2", "F1", "1.0 Foundation — Map the HC Economy",
     "Establish Payer Persona Benchmarks",
     "Glen", 5, "Critical",
     "T1.0.1",
     "Benchmark Goals' 3-persona model (10.4% overall payer rate, ARPPU $32.48) against published data from eFootball, EA FC, Sensor Tower. Assess realism for month-1 conversion."),

    ("T1.1.1", "F1", "1.1 Competitive Pricing Position",
     "Build Normalised Price Position Map",
     "Glen / AI", 8, "High",
     "T1.0.1",
     "Create percentile position map for each HC tier vs competitive range. One-line answer: 'Goals is at the Xth percentile — below/at/above median.' Spot-check 2-3 competitor prices for freshness."),

    ("T1.1.2", "F1", "1.1 Competitive Pricing Position",
     "Discount Curve Comparison",
     "Glen", 4, "High",
     "T1.1.1",
     "Compare Goals' 18% max discount against EA FC (39%), Fortnite (24%), Apex (13%). Assess whether discount curve incentivises whale spending. BLOCKER: which discount curve is final — confirm with Julius before starting."),

    ("T1.1.3", "F1", "1.1 Competitive Pricing Position",
     "SKU Structure and Entry Point Audit",
     "Glen", 5, "High",
     "T1.0.1, T1.1.1",
     "Assess structural health of 6-tier HC ladder ($5.99–$99.99). Entry point friction analysis: what does $5.99 (380 HC) actually buy? Whale ceiling analysis. Identify structural gaps."),

    ("T1.1.4", "F1", "1.1 Competitive Pricing Position",
     "Platform Fee Impact and Net Revenue Table",
     "Glen / AI", 2, "Medium",
     "T1.1.1",
     "Build net revenue table per SKU per platform (Steam 30%, Epic 12%, PS/Xbox 30%). Quantify Epic revenue advantage (~26% more per transaction at $5.99 tier)."),

    ("T1.2.1", "F1", "1.2 Conversion Rate and Elasticity Context",
     "Price Elasticity Scenario Modelling",
     "Glen", 8, "High",
     "T1.0.2, T1.1.1",
     "Model 4 scenarios: A) All prices +10%, B) All prices +20%, C) Raise entry tier only ($5.99→$6.99), D) Steepen discount curve. Present revenue impact ranges. Core principle: launch high, patch down."),

    ("T1.3.1", "F1", "1.3 Promotional and First-Purchase Pricing",
     "First-Purchase and Starter Pack Recommendations",
     "Glen", 5, "Medium",
     "T1.0.1, T1.1.3",
     "Research first-purchase incentives from EA FC, Fortnite, Apex, eFootball. Recommend: first-purchase bonus, limited starter pack (~$3.99 with 500 HC + kit), World Cup bundle. Frame as complementary to core pricing, not scope creep."),

    ("T2.1.1", "F2", "2.1 Regional Deviation and Risk Identification",
     "Categorise All Countries by Risk Level",
     "Devin / AI", 8, "Critical",
     "None",
     "Using Goals' existing deviation data, categorise every country: Green (<15% deviation, YES), Amber (15-25% or marked !/? by Goals), Red (>25% or Goals recommends change). Red candidates: Poland (+25%), Switzerland (+26%), Ukraine (-26%), Turkey, Brazil."),

    ("T2.1.2", "F2", "2.1 Regional Deviation and Risk Identification",
     "Competitor Regional Pricing for Red/Amber Markets",
     "Devin", 15, "Critical",
     "T2.1.1",
     "Research EA FC and Fortnite pricing in all red/amber countries. Sources: SteamDB (Steam), PS Store regional web, Epic Store. Priority markets: Brazil, Turkey, Poland, Russia/CIS. Build comparison table: Country | Goals proposed | EA FC actual | Fortnite actual | Goals position."),

    ("T2.2.1", "F2", "2.2 Platform Compliance and Submission Guidance",
     "Verify Platform Tier Compliance",
     "Devin", 3, "Critical",
     "T2.1.1",
     "Cross-reference Goals' PS/Xbox pricing against valid platform tier selections. Flag any non-standard prices (Sony/Xbox will reject). Document submission checklist and timeline (9-day review period from Julius's call). BLOCKER: confirm submission deadline with Julius."),

    ("T2.3.1", "F2", "2.3 Country-Specific Recommendations",
     "Write Country Recommendation Cards",
     "Glen / AI", 8, "High",
     "T2.1.2, T2.2.1",
     "Write 8-12 recommendation cards for all red/amber countries. Each card: current price, issue, recommended price, rationale, revenue impact, risk if unchanged. Group by action type: reduce, increase, or platform tier adjustment."),

    ("T3.1.1", "F3", "3.1 'Are We Too Low?' — Central Answer",
     "Build the Definitive Position Statement",
     "Glen", 4, "High",
     "T1.1.1, T1.0.2, T1.2.1",
     "One paragraph answering Jonas's core question with specific evidence and a specific recommendation. No hedging. State percentile position, conversion realism, and exact recommendation (e.g., 'maintain pricing but steepen top-tier discount')."),

    ("T3.2.1", "F3", "3.2 Risk Register",
     "Compile and Rank Risks",
     "Glen", 5, "High",
     "T1.0.1, T1.1.2, T2.1.1, T2.2.1",
     "Full 10-risk register covering: P2W perception, price increase impossibility, over-pricing in sensitive markets, platform rejection, shallow discount curve, entry-tier friction, Originals pricing gap, Netherlands gambling regulation, South Korea drop-rate disclosure, non-seasonal demand sustainability."),

    ("T3.3.1", "F3", "3.3 Top 5 Recommendations",
     "Write Top 5 Recommendations",
     "Glen", 7, "High",
     "T3.1.1, T3.2.1, T2.3.1",
     "Five specific, numbered, actionable recommendations. Each with: exact specification, evidence, revenue impact range, implementation steps (which cells in their matrix), risk if skipped. Must pass 'would a studio economist act on this immediately?' test."),

    ("T4.1.1", "F4", "4.1 Reusable Regional Pricing Methodology",
     "Write the 'How to Price a New SKU' Guide",
     "Glen / AI", 3, "Low",
     "T2.3.1",
     "One-page methodology for internal Goals use. Six steps from USD anchor to platform tier. Worked example: pricing a Celebration at $7.00 for Brazil. PPP factor lookup table. Position as 'manual method; Phase 2 automates this.' CONDITIONAL: only if F1-3 complete cleanly."),

    ("T4.2.1", "F4", "4.2 World Cup Pricing Opportunity Notes",
     "Write World Cup Quick-Hit Observations",
     "Glen / AI", 2, "Low",
     "T1.3.1",
     "3-5 bullet points: national bundle group-stage premium vs knockout discount, World Cup Coin Pack (limited 10% bonus), ARPPU spike research from EA FC/eFootball during WC 2022. Half a page. Teaser for Phase 2 live service engagement. CONDITIONAL."),

    ("T4.3.1", "F4", "4.3 Phase 2 Positioning",
     "Write 'What's Next' Section",
     "Tom", 3, "Low",
     "T3.3.1",
     "Half-page natural next steps: post-launch telemetry review, A/B test design, store optimisation, World Cup full plan, ongoing live service consulting. Close with check-in call recommendation. Helpful, not salesy. No internal NBI pricing included. CONDITIONAL."),

    ("T5.1.1", "F5", "5.1 Written Summary Document",
     "Assemble the Written Summary",
     "Glen", 10, "Critical",
     "T3.3.1, T2.3.1, T4.1.1, T4.2.1, T4.3.1",
     "Final client-facing document. 12-18 pages. Sections: Executive Summary, Competitive Position, Regional Assessment, Pricing Scenarios, Risk Register, Recommendations, Pricing Framework, World Cup Opportunities, Next Steps, Appendix. British English, no em dashes, numbers-backed."),

    ("T5.2.1", "F5", "5.2 Internal Review and Polish",
     "Quality Gate — Internal Review",
     "Glen + Tom", 4, "Critical",
     "T5.1.1",
     "Glen reviews: domain accuracy, tone, completeness vs Jonas's questions. Tom reviews: commercial positioning, NBI brand consistency, contractual/legal issues. Address all feedback, produce final version. Send to client 24h before review call."),

    ("T5.3.1", "F5", "5.3 Remote Review Session",
     "Conduct Client Review Call",
     "Glen", 4, "Critical",
     "T5.2.1",
     "45-min call with Jonas and Julius. Agenda: 5m scope recap, 10m top-5 recs, 10m regional highlights, 5m pricing framework, 5m next steps, 10m Q&A. Follow-up within 24h. Document Phase 2 interest level."),
]

TIMELINE = [
    # Day, Glen tasks, Devin tasks, Tom tasks, AI tasks
    ("Day 1",
     "T1.0.1 — Map HC value chain (3h)\nT1.0.2 — Payer persona benchmarks (5h)\nTotal: 8h",
     "Verify and QA AI research outputs\nCoordinate with Julius on data gaps",
     "—",
     "T1.1.1 — Data gathering for price position map\nT2.1.1 — Country categorisation from Goals matrix"),

    ("Day 2",
     "T1.1.2 — Discount curve comparison (4h)\nT1.1.3 — SKU structure audit (5h)\nT1.2.1 start — Scenario modelling begins (3h)\nTotal: 12h",
     "T2.1.2 — Store verification research (8h)\nPriority: Brazil, Turkey, Poland",
     "—",
     "T1.1.4 — Net revenue table\nContinue T2.1.2 scraping (SteamDB, PS Store)"),

    ("Day 3",
     "T1.2.1 finish — Scenario modelling (5h)\nT1.3.1 — First-purchase recommendations (5h)\nTotal: 10h",
     "T2.1.2 finish — Remaining markets (5h)\nT2.2.1 — Platform tier compliance (3h)\nTotal: 10h (with AI support)",
     "—",
     "T2.3.1 — Draft country recommendation cards for Glen review"),

    ("Day 4",
     "T3.1.1 — Definitive position statement (4h)\nT3.2.1 — Risk register (5h)\nT3.3.1 — Top 5 recommendations (7h)\nTotal: 16h",
     "QA + data verification across all research\nSupport Glen on Risk Register data",
     "T4.3.1 — 'What's Next' Phase 2 section (3h)",
     "T4.1.1 — 'How to Price a New SKU' guide draft\nT4.2.1 — World Cup observations draft"),

    ("Day 5",
     "T5.1.1 — Assemble written summary (10h)\nT5.2.1 — Glen quality review (2h)\nTotal: 14h",
     "Editorial support and data cross-checks",
     "T5.2.1 — Tom commercial review (2h)",
     "—"),

    ("Day 6",
     "T5.3.1 — Client review call prep + call + follow-up (4h)",
     "—",
     "—",
     "—"),
]

RISKS = [
    # ID, Description, Severity, Likelihood, Markets, Mitigation
    ("R01", "P2W Perception at High Price Points — Community data shows 'pay to win' is the #1 cited quit reason. If premium packs feel competitive rather than cosmetic, pricing becomes irrelevant because players leave first.",
     "Critical", "Medium", "All markets",
     "Frame all HC purchases as time-saving or cosmetic. Ensure store messaging emphasises 'customise' not 'dominate.' Monitor community sentiment in first 2 weeks post-launch."),

    ("R02", "Post-Launch Price Increase Impossibility — If current prices prove too low, raising them on console requires re-certification, triggers player backlash, and generates press coverage. This is a platform constraint, not just strategy.",
     "High", "Medium", "PS, Xbox",
     "Launch at the upper end of the acceptable competitive range. Explicitly label current pricing as 'Launch Pricing' so any future adjustments are framed as expected rather than a u-turn."),

    ("R03", "Over-Pricing in Price-Sensitive Markets — Large player bases with low purchasing power. Over-pricing reduces conversion and misses peak World Cup revenue. Poland is already +25% above FX.",
     "High", "High", "Poland, Turkey, Brazil, Ukraine",
     "Implement country-specific price adjustments per T2.3.1 recommendation cards before platform submission. Prioritise by player base size × deviation severity."),

    ("R04", "Platform Pricing Rejection — Sony or Xbox rejects submitted pricing due to non-standard tier selection, causing delay past the May 14 launch window.",
     "Critical", "Low (if T2.2.1 verified)", "PS, Xbox",
     "T2.2.1 verification is mandatory before submission. Submit minimum 2 weeks before deadline. Confirm Julius has a test submission result if available."),

    ("R05", "Discount Curve Too Shallow for Launch — 18% max discount may not drive enough whale commitment in the critical early-revenue window. New games need stronger incentive to break initial spending habit.",
     "Medium", "Medium", "Global — top spenders",
     "Consider steepening top-tier discount from 18% to 22-25% for the first 90 days post-launch, then normalise. Can be framed as a launch incentive that expires."),

    ("R06", "Missing Entry-Tier Conversion — $5.99 minimum may be too high a barrier for first-time buyers who need a low-risk trial purchase before committing to the economy.",
     "Medium", "Medium", "Global — new payers",
     "First-purchase bonus (T1.3.1) addresses this without adding a new micro-tier. A one-time $3.99 starter pack breaks the barrier while preserving the core ladder structure."),

    ("R07", "Originals/Celebrity Pack Pricing Gap — Premium HC-only items have no pricing framework. When these launch post-Day 1, Goals will need a method to price them without ad hoc guesswork.",
     "Low (not launch-critical)", "N/A (certain post-launch)", "Global",
     "Feature 4 pricing framework covers the methodology. Flag this explicitly in the deliverable as a Phase 2 priority. Do not leave Goals without a framework when the first Celebrity pack drops."),

    ("R08", "Netherlands Gambling Regulation (KSA) — Player packs with attribute variance may be classified as gambling-adjacent under Dutch Kansspelautoriteit rulings. EA FC has been fined in NL for similar mechanics.",
     "High", "Medium", "Netherlands (EUR zone)",
     "Legal review of pack mechanic before NL launch. Consider cosmetic-only packs in NL, or ensure all pack contents are predetermined and visible before purchase. Verify current KSA enforcement stance on F2P sports titles."),

    ("R09", "South Korea Drop Rate Disclosure — Korean law mandates exact probability disclosure for all randomised item purchases. Non-compliance is a legal blocker, not a risk — it will prevent the Korean launch.",
     "High", "Certain (if launching in KR)", "South Korea",
     "Confirm drop rate disclosure is built into the pack UI before KR submission. Goals' player tier distribution data provides the rates. Ensure these are displayed in-game in Korean."),

    ("R10", "Non-Seasonal Demand Sustainability — Without seasonal resets or a battle pass, there is no guaranteed per-player spend floor per quarter. Revenue will be flat and gradual rather than pulsing, which makes growth harder to model.",
     "Medium", "High", "Global — long-term",
     "Plan regular cosmetic refresh cadence (monthly kit and celebration drops). Leverage real-world football events as demand drivers. Use mode variety to drive squad depth demand. This is Phase 2 strategy but must be acknowledged in the deliverable as the fundamental tradeoff of the anti-P2W positioning."),
]

DELIVERABLE_SECTIONS = [
    ("Exec Summary", "Executive Summary",
     "1 page",
     "Headline finding, top 5 recommendations, and the single most urgent action required before platform submission. Written for Jonas (reads in 3 minutes on his phone)."),

    ("Section 1", "Your Competitive Position",
     "2-3 pages",
     "Where Goals sits vs the market at each HC tier. Percentile position map. Assessment of payer conversion and ARPPU targets against published comparable title data. Clear answer to 'are we too low?' Source: Feature 1 findings (T1.0.1, T1.1.1, T1.0.2)."),

    ("Section 2", "Regional Pricing Assessment",
     "4-5 pages",
     "Country risk map (green/amber/red). 8-12 recommendation cards for red/amber markets. Each card: issue, recommended price, rationale, revenue impact, risk if unchanged. Prioritised by player base × deviation severity. Source: Feature 2 findings (T2.1.1, T2.1.2, T2.3.1)."),

    ("Section 3", "Pricing Scenarios",
     "2 pages",
     "What-if modelling showing revenue impact of four pricing adjustments: +10% all prices, +20% all prices, entry tier increase only, steepened discount curve. Presented as conservative-to-optimistic ranges. Source: T1.2.1."),

    ("Section 4", "Risk Register",
     "2 pages",
     "All 10 identified pricing risks ranked by severity x likelihood. Top 3 with specific, step-by-step mitigation. Legal risks (NL, KR) called out explicitly as blockers not merely risks. Source: T3.2.1."),

    ("Section 5", "Recommendations",
     "2 pages",
     "Five numbered, specific, actionable recommendations. Format: what to change, why, revenue impact range, implementation steps, risk if skipped. Each must be executable by Julius without reading the full report. Source: T3.3.1."),

    ("Section 6", "Pricing Framework",
     "1 page (conditional)",
     "Reusable one-page methodology for pricing any future SKU. Six-step process from USD anchor to platform submission. Worked example for Brazil. PPP lookup table. Positioned as the manual method; Phase 2 automates it. Source: T4.1.1. CONDITIONAL: only included if F1-3 completed within budget."),

    ("Section 7", "World Cup Opportunities",
     "Half page (conditional)",
     "3-5 tactical pricing observations for the June World Cup window. National bundle group-stage premium vs knockout discount. World Cup Coin Pack structure. ARPPU spike context from comparable sporting event data. Teaser for Phase 2 live service engagement. Source: T4.2.1. CONDITIONAL."),

    ("Section 8", "Next Steps",
     "Half page",
     "Natural next steps framed as recommendations, not a sales pitch. Post-launch telemetry review (2-3 weeks after May 14), A/B test design, store optimisation, full World Cup plan, ongoing live service consulting. Closes with check-in call recommendation. Source: T4.3.1."),

    ("Appendix", "Supporting Data",
     "As required",
     "Full competitive comparison tables, country pricing matrix with red/amber/green categorisation, net revenue per SKU per platform, payer persona benchmark sources, competitor regional pricing data tables. Julius's working reference."),
]


# ════════════════════════════════════════════════════════════════════════════
# BUILD WORKBOOK
# ════════════════════════════════════════════════════════════════════════════

wb = Workbook()

# ── Remove default sheet ────────────────────────────────────────────────────
wb.remove(wb.active)


# ════════════════════════════════════════════════════════════════════════════
# SHEET 1: OVERVIEW
# ════════════════════════════════════════════════════════════════════════════

ws1 = wb.create_sheet("Overview")
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = None

set_col_widths(ws1, {
    "A": 28, "B": 45, "C": 18, "D": 18, "E": 5,
})

# Row 1: Main title
c = ws1["A1"]
c.value = "Goals Studio Pricing Engagement"
c.font = Font(name="Calibri", size=18, bold=True, color=HEADER_FG)
c.fill = PatternFill("solid", fgColor=HEADER_BG)
c.alignment = Alignment(horizontal="center", vertical="center")
ws1.merge_cells("A1:D1")
ws1.row_dimensions[1].height = 36

# Row 2: subtitle
c = ws1["A2"]
c.value = "Consulting Engagement Project Overview"
c.font = Font(name="Calibri", size=12, color=HEADER_FG, italic=True)
c.fill = PatternFill("solid", fgColor=TITLE_BG)
c.alignment = Alignment(horizontal="center", vertical="center")
ws1.merge_cells("A2:D2")
ws1.row_dimensions[2].height = 22

ws1.row_dimensions[3].height = 10  # spacer

# Project Details section title
r = 4
header_style(ws1, f"A{r}", "PROJECT DETAILS", size=11, bg=HEADER_BG)
ws1.merge_cells(f"A{r}:D{r}")
ws1.row_dimensions[r].height = 22

def detail_row(ws, row, label, value, label_bg=ACCENT1):
    ws.row_dimensions[row].height = 18
    lc = ws[f"A{row}"]
    lc.value = label
    lc.font = Font(name="Calibri", size=11, bold=True)
    lc.fill = PatternFill("solid", fgColor=label_bg)
    lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    lc.border = thin_border()

    vc = ws[f"B{row}"]
    vc.value = value
    vc.font = Font(name="Calibri", size=11)
    vc.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    vc.border = thin_border()
    ws.merge_cells(f"B{row}:D{row}")

details = [
    ("Project Title",    "HC Pricing Benchmark, Regional Strategy & Risk Assessment"),
    ("Client",           "Goals Studio"),
    ("Client Contacts",  "Jonas Rundberg — CEO  |  Julius — Live Ops Lead"),
    ("NBI Lead",         "Glen Pryer — Managing Director, NBI Gaming"),
    ("Budget",           "100,000 SEK (~$10,000 USD)"),
    ("Total Hours",      "~116 billable hours"),
    ("Platform Deadline","27 April 2026 (platform submission)"),
    ("Launch Date",      "14 May 2026"),
    ("Review Call",      "Day 6 — after document delivery (exact date TBC)"),
    ("Status",           "Active — Option A: AI-Assisted (5 Calendar Days)"),
]

for i, (lbl, val) in enumerate(details):
    bg = WHITE if i % 2 == 0 else ALT_ROW
    detail_row(ws1, 5 + i, lbl, val, label_bg=ACCENT1)
    ws1[f"B{5+i}"].fill = PatternFill("solid", fgColor=bg)

spacer_r = 5 + len(details)
ws1.row_dimensions[spacer_r].height = 10

# Team Allocation section
r = spacer_r + 1
header_style(ws1, f"A{r}", "TEAM ALLOCATION", size=11, bg=HEADER_BG)
ws1.merge_cells(f"A{r}:D{r}")
ws1.row_dimensions[r].height = 22

team_headers_r = r + 1
for col, txt in enumerate(["Team Member", "Role", "Hours", "Focus Area"], start=1):
    c = ws1.cell(row=team_headers_r, column=col, value=txt)
    c.font = Font(name="Calibri", size=11, bold=True, color=HEADER_FG)
    c.fill = PatternFill("solid", fgColor=TITLE_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin_border()
ws1.row_dimensions[team_headers_r].height = 18

team_data = [
    ("Glen Pryer", "Managing Director / Lead Consultant", "~45h",
     "Strategic synthesis, scenario modelling, deliverable authoring, quality review, client call"),
    ("Devin", "Senior Research Analyst", "~50h",
     "Competitive store research, country categorisation, regional pricing data gathering, data verification"),
    ("Tom", "Commercial Director", "~20h",
     "Commercial positioning, internal review, Phase 2 framing, contract and legal check"),
]

for i, (name, role, hrs, focus) in enumerate(team_data):
    row = team_headers_r + 1 + i
    bg = WHITE if i % 2 == 0 else ALT_ROW
    ws1.row_dimensions[row].height = 36
    for col, val in enumerate([name, role, hrs, focus], start=1):
        c = ws1.cell(row=row, column=col, value=val)
        c.font = Font(name="Calibri", size=11)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        c.border = thin_border()
    ws1.cell(row=row, column=1).font = Font(name="Calibri", size=11, bold=True)
    ws1.cell(row=row, column=3).alignment = Alignment(horizontal="center", vertical="center")

spacer_r2 = team_headers_r + len(team_data) + 1
ws1.row_dimensions[spacer_r2].height = 10

# Feature Hour Summary
r = spacer_r2 + 1
header_style(ws1, f"A{r}", "FEATURE HOUR SUMMARY", size=11, bg=HEADER_BG)
ws1.merge_cells(f"A{r}:D{r}")
ws1.row_dimensions[r].height = 22

fh_r = r + 1
for col, txt in enumerate(["Feature ID", "Feature Name", "Hours", "% of Budget"], start=1):
    c = ws1.cell(row=fh_r, column=col, value=txt)
    c.font = Font(name="Calibri", size=11, bold=True, color=HEADER_FG)
    c.fill = PatternFill("solid", fgColor=TITLE_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin_border()
ws1.row_dimensions[fh_r].height = 18

for i, (fid, fname, fhrs, fpct) in enumerate(FEATURE_HOURS):
    row = fh_r + 1 + i
    bg = FEAT_COLS.get(fid, WHITE)
    ws1.row_dimensions[row].height = 20
    for col, val in enumerate([fid, fname, f"{fhrs}h", f"{fpct}%"], start=1):
        c = ws1.cell(row=row, column=col, value=val)
        c.font = Font(name="Calibri", size=11)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="left" if col == 2 else "center", vertical="center", indent=1)
        c.border = thin_border()
    ws1.cell(row=row, column=1).font = Font(name="Calibri", size=11, bold=True)

# Total row
total_r = fh_r + len(FEATURE_HOURS) + 1
ws1.row_dimensions[total_r].height = 20
for col, val in enumerate(["", "TOTAL", "~116h", "100%"], start=1):
    c = ws1.cell(row=total_r, column=col, value=val)
    c.font = Font(name="Calibri", size=11, bold=True, color=HEADER_FG)
    c.fill = PatternFill("solid", fgColor=HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin_border()

spacer_r3 = total_r + 1
ws1.row_dimensions[spacer_r3].height = 10

# Key constraints
r = spacer_r3 + 1
header_style(ws1, f"A{r}", "KEY CONSTRAINTS AND BLOCKERS", size=11, bg=HEADER_BG)
ws1.merge_cells(f"A{r}:D{r}")
ws1.row_dimensions[r].height = 22

constraints = [
    ("BLOCKER", "Which discount curve is final — cannot proceed with T1.1.2 until Julius confirms."),
    ("BLOCKER", "Platform submission deadline — cannot finalise timeline until confirmed."),
    ("BLOCKER", "Day 1 store contents — required for T1.0.1 HC value chain mapping."),
    ("DECISION", "Option A (AI-Assisted, 5 days) vs Option B (Full Human, 8-10 days) — must decide before Day 1 begins."),
    ("NOTE", "116h exceeds the original 50h estimate by 2.3x. At 100,000 SEK contracted, the effective hourly rate drops materially. AI research support (Option A) is the mechanism to make this viable within 5 calendar days."),
]

for i, (tag, note) in enumerate(constraints):
    row = r + 1 + i
    bg = CRITICAL_BG if tag == "BLOCKER" else (HIGH_BG if tag == "DECISION" else ALT_ROW)
    ws1.row_dimensions[row].height = 28
    for col, val in enumerate([tag, note], start=1):
        c = ws1.cell(row=row, column=col, value=val)
        c.font = Font(name="Calibri", size=11, bold=(col == 1))
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        c.border = thin_border()
    ws1.merge_cells(f"B{row}:D{row}")


# ════════════════════════════════════════════════════════════════════════════
# SHEET 2: TASK TRACKER
# ════════════════════════════════════════════════════════════════════════════

ws2 = wb.create_sheet("Task Tracker")
ws2.sheet_view.showGridLines = False

# Column headers
TASK_COLS = [
    ("A", "Task ID",         10),
    ("B", "Feature",          8),
    ("C", "Story",           26),
    ("D", "Task Title",      28),
    ("E", "Owner",           14),
    ("F", "Hours Est.",       9),
    ("G", "Status",          14),
    ("H", "Priority",        10),
    ("I", "Dependencies",    16),
    ("J", "Notes",           55),
]

for col, (letter, header, width) in enumerate(TASK_COLS, start=1):
    ws2.column_dimensions[letter].width = width

# Title row
c = ws2["A1"]
c.value = "Goals Studio Pricing Engagement — Task Tracker"
c.font = Font(name="Calibri", size=14, bold=True, color=HEADER_FG)
c.fill = PatternFill("solid", fgColor=HEADER_BG)
c.alignment = Alignment(horizontal="center", vertical="center")
ws2.merge_cells("A1:J1")
ws2.row_dimensions[1].height = 30

# Subtitle / legend
c = ws2["A2"]
c.value = ("Status: Not Started | In Progress | Complete | Blocked          "
           "Priority: Critical | High | Medium | Low")
c.font = Font(name="Calibri", size=10, italic=True, color="444444")
c.fill = PatternFill("solid", fgColor=ALT_ROW)
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws2.merge_cells("A2:J2")
ws2.row_dimensions[2].height = 18

# Header row
for col, (letter, header, width) in enumerate(TASK_COLS, start=1):
    c = ws2.cell(row=3, column=col, value=header)
    c.font = Font(name="Calibri", size=11, bold=True, color=HEADER_FG)
    c.fill = PatternFill("solid", fgColor=HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin_border()
ws2.row_dimensions[3].height = 20

ws2.freeze_panes = "A4"

# Priority bg colours
PRIO_BG = {
    "Critical": CRITICAL_BG,
    "High":     HIGH_BG,
    "Medium":   MEDIUM_BG,
    "Low":      LOW_BG,
}

current_feature = None
row = 4

for task in TASKS:
    tid, feat, story, title, owner, hours, priority, deps, notes = task

    # Feature divider row when feature changes
    if feat != current_feature:
        current_feature = feat
        fname = next(f[1] for f in FEATURE_HOURS if f[0] == feat)
        fhrs  = next(f[2] for f in FEATURE_HOURS if f[0] == feat)
        c = ws2.cell(row=row, column=1,
                     value=f"  {feat}: {fname}  —  {fhrs}h")
        c.font = Font(name="Calibri", size=11, bold=True, color=HEADER_FG)
        c.fill = PatternFill("solid", fgColor=TITLE_BG)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        for col in range(1, 11):
            cell = ws2.cell(row=row, column=col)
            cell.fill = PatternFill("solid", fgColor=TITLE_BG)
            cell.border = thin_border()
        ws2.merge_cells(f"A{row}:J{row}")
        ws2.row_dimensions[row].height = 20
        row += 1

    # Data row
    feat_fill = FEAT_COLS.get(feat, WHITE)
    ws2.row_dimensions[row].height = 60

    values = [tid, feat, story, title, owner, f"{hours}h",
              "Not Started", priority, deps, notes]
    wraps  = [False, False, True, True, False, False, False, False, True, True]
    aligns = ["center", "center", "left", "left", "center", "center",
              "center", "center", "left", "left"]

    for col, (val, wrap, align) in enumerate(zip(values, wraps, aligns), start=1):
        c = ws2.cell(row=row, column=col, value=val)
        c.font = Font(name="Calibri", size=11)
        c.alignment = Alignment(horizontal=align, vertical="top",
                                wrap_text=wrap, indent=(1 if align == "left" else 0))
        c.border = thin_border()

        # Colouring logic
        if col == 8:  # Priority column
            c.fill = PatternFill("solid", fgColor=PRIO_BG.get(priority, WHITE))
            c.font = Font(name="Calibri", size=11, bold=True)
        elif col in (1, 2):  # Task ID, Feature
            c.fill = PatternFill("solid", fgColor=feat_fill)
            c.font = Font(name="Calibri", size=11, bold=True)
        else:
            c.fill = PatternFill("solid", fgColor=feat_fill)

    row += 1

# Total hours row
ws2.row_dimensions[row].height = 20
total_vals = ["", "", "", "TOTAL HOURS", "", "~116h", "", "", "", ""]
for col, val in enumerate(total_vals, start=1):
    c = ws2.cell(row=row, column=col, value=val)
    c.font = Font(name="Calibri", size=11, bold=True, color=HEADER_FG)
    c.fill = PatternFill("solid", fgColor=HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin_border()


# ════════════════════════════════════════════════════════════════════════════
# SHEET 3: TIMELINE
# ════════════════════════════════════════════════════════════════════════════

ws3 = wb.create_sheet("Timeline")
ws3.sheet_view.showGridLines = False

set_col_widths(ws3, {"A": 10, "B": 40, "C": 38, "D": 28, "E": 38})

# Title
c = ws3["A1"]
c.value = "Goals Studio Pricing Engagement — Execution Timeline (Option A: AI-Assisted, 5 Calendar Days)"
c.font = Font(name="Calibri", size=14, bold=True, color=HEADER_FG)
c.fill = PatternFill("solid", fgColor=HEADER_BG)
c.alignment = Alignment(horizontal="center", vertical="center")
ws3.merge_cells("A1:E1")
ws3.row_dimensions[1].height = 30

c = ws3["A2"]
c.value = ("Option A assumes AI handles the bulk of Devin's research workload "
           "(competitive store scraping, country categorisation, data gathering). "
           "Human team focuses on synthesis, judgement, and deliverable writing. "
           "Decision required before Day 1.")
c.font = Font(name="Calibri", size=10, italic=True)
c.fill = PatternFill("solid", fgColor=ALT_ROW)
c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
ws3.merge_cells("A2:E2")
ws3.row_dimensions[2].height = 28

# Headers
for col, txt in enumerate(["Day", "Glen", "Devin", "Tom", "AI Support"], start=1):
    c = ws3.cell(row=3, column=col, value=txt)
    c.font = Font(name="Calibri", size=11, bold=True, color=HEADER_FG)
    c.fill = PatternFill("solid", fgColor=HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin_border()
ws3.row_dimensions[3].height = 20
ws3.freeze_panes = "A4"

# Owner column colours
OWNER_COLS = {
    1: "D6EAF8",  # Glen - blue
    2: "D5F5E3",  # Devin - green
    3: "FAD7A0",  # Tom - amber
    4: "E8DAEF",  # AI - purple
}

for i, (day, glen, devin, tom, ai) in enumerate(TIMELINE):
    row = 4 + i
    bg = WHITE if i % 2 == 0 else ALT_ROW
    ws3.row_dimensions[row].height = 90

    for col, val in enumerate([day, glen, devin, tom, ai], start=1):
        c = ws3.cell(row=row, column=col, value=val)
        c.font = Font(name="Calibri", size=11, bold=(col == 1))
        c.fill = PatternFill("solid", fgColor=OWNER_COLS.get(col - 1, bg) if col > 1 else ALT_ROW)
        c.alignment = Alignment(horizontal="left" if col > 1 else "center",
                                vertical="top", wrap_text=True, indent=1)
        c.border = thin_border()

# Quality gates footer
gate_r = 4 + len(TIMELINE) + 1
header_style(ws3, f"A{gate_r}", "QUALITY GATES", size=11, bg=HEADER_BG)
ws3.merge_cells(f"A{gate_r}:E{gate_r}")
ws3.row_dimensions[gate_r].height = 22

gates = [
    ("Before Day 1", "Julius confirms: discount curve choice, submission deadline, Day 1 store contents. Without these, T1.0.1 and T1.1.2 cannot start properly."),
    ("End of Day 1", "Can we answer 'what does each HC tier buy?' If not, escalate to Julius immediately."),
    ("End of Day 2", "Competitive position map draft complete. Do we know where Goals sits vs market?"),
    ("End of Day 3 (CRITICAL)", "Glen reviews draft top 5 recommendations. Go/no-go on direction before synthesis begins."),
    ("End of Day 4", "Full deliverable content complete (pending assembly). Feature 4 status: complete or cut."),
    ("Day 5", "Final document to Glen and Tom for review. Send to client by end of Day 5 or first thing Day 6."),
]

for i, (milestone, gate) in enumerate(gates):
    row = gate_r + 1 + i
    bg = CRITICAL_BG if "CRITICAL" in milestone else (ALT_ROW if i % 2 == 0 else WHITE)
    ws3.row_dimensions[row].height = 30
    for col, val in enumerate([milestone, gate], start=1):
        c = ws3.cell(row=row, column=col, value=val)
        c.font = Font(name="Calibri", size=11, bold=(col == 1))
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        c.border = thin_border()
    ws3.merge_cells(f"B{row}:E{row}")


# ════════════════════════════════════════════════════════════════════════════
# SHEET 4: RISK REGISTER
# ════════════════════════════════════════════════════════════════════════════

ws4 = wb.create_sheet("Risk Register")
ws4.sheet_view.showGridLines = False

set_col_widths(ws4, {
    "A": 7, "B": 44, "C": 12, "D": 12, "E": 24, "F": 50,
})

# Title
c = ws4["A1"]
c.value = "Goals Studio Pricing Engagement — Risk Register"
c.font = Font(name="Calibri", size=14, bold=True, color=HEADER_FG)
c.fill = PatternFill("solid", fgColor=HEADER_BG)
c.alignment = Alignment(horizontal="center", vertical="center")
ws4.merge_cells("A1:F1")
ws4.row_dimensions[1].height = 30

c = ws4["A2"]
c.value = ("Ranked by Severity × Likelihood. Critical severity = launch-blocking. "
           "Top 3 risks have specific step-by-step mitigation. "
           "Legal risks (R08 Netherlands, R09 South Korea) are hard blockers, not merely risks.")
c.font = Font(name="Calibri", size=10, italic=True)
c.fill = PatternFill("solid", fgColor=ALT_ROW)
c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
ws4.merge_cells("A2:F2")
ws4.row_dimensions[2].height = 28

# Headers
for col, txt in enumerate(["Risk ID", "Description", "Severity", "Likelihood",
                            "Markets Affected", "Mitigation"], start=1):
    c = ws4.cell(row=3, column=col, value=txt)
    c.font = Font(name="Calibri", size=11, bold=True, color=HEADER_FG)
    c.fill = PatternFill("solid", fgColor=HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin_border()
ws4.row_dimensions[3].height = 20
ws4.freeze_panes = "A4"

SEV_BG = {
    "Critical": CRITICAL_BG,
    "High": HIGH_BG,
    "Medium": MEDIUM_BG,
    "Low": LOW_BG,
}

for i, (rid, desc, sev, lik, markets, mitigation) in enumerate(RISKS):
    row = 4 + i
    bg = SEV_BG.get(sev.split()[0], WHITE)
    ws4.row_dimensions[row].height = 80

    values = [rid, desc, sev, lik, markets, mitigation]
    for col, val in enumerate(values, start=1):
        c = ws4.cell(row=row, column=col, value=val)
        c.font = Font(name="Calibri", size=11, bold=(col in (1, 3)))
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center" if col in (1, 3, 4) else "left",
                                vertical="top", wrap_text=True, indent=(1 if col > 1 else 0))
        c.border = thin_border()

# Summary table
sum_r = 4 + len(RISKS) + 1
header_style(ws4, f"A{sum_r}", "RISK SUMMARY", size=11, bg=HEADER_BG)
ws4.merge_cells(f"A{sum_r}:F{sum_r}")
ws4.row_dimensions[sum_r].height = 22

for col, txt in enumerate(["Severity", "Count", "Risk IDs", "", "", ""], start=1):
    c = ws4.cell(row=sum_r + 1, column=col, value=txt)
    c.font = Font(name="Calibri", size=11, bold=True, color=HEADER_FG)
    c.fill = PatternFill("solid", fgColor=TITLE_BG)
    c.border = thin_border()
    c.alignment = Alignment(horizontal="center", vertical="center")

severity_groups = [
    ("Critical", "2", "R01, R04"),
    ("High",     "4", "R02, R03, R08, R09"),
    ("Medium",   "3", "R05, R06, R10"),
    ("Low",      "1", "R07"),
]

for i, (sev, count, ids) in enumerate(severity_groups):
    row = sum_r + 2 + i
    ws4.row_dimensions[row].height = 18
    for col, val in enumerate([sev, count, ids, "", "", ""], start=1):
        c = ws4.cell(row=row, column=col, value=val)
        c.font = Font(name="Calibri", size=11, bold=(col == 1))
        c.fill = PatternFill("solid", fgColor=SEV_BG.get(sev, WHITE))
        c.border = thin_border()
        c.alignment = Alignment(horizontal="center", vertical="center")


# ════════════════════════════════════════════════════════════════════════════
# SHEET 5: DELIVERABLE STRUCTURE
# ════════════════════════════════════════════════════════════════════════════

ws5 = wb.create_sheet("Deliverable Structure")
ws5.sheet_view.showGridLines = False

set_col_widths(ws5, {"A": 14, "B": 30, "C": 14, "D": 65})

# Title
c = ws5["A1"]
c.value = "Goals Studio Pricing Engagement — Final Deliverable Structure"
c.font = Font(name="Calibri", size=14, bold=True, color=HEADER_FG)
c.fill = PatternFill("solid", fgColor=HEADER_BG)
c.alignment = Alignment(horizontal="center", vertical="center")
ws5.merge_cells("A1:D1")
ws5.row_dimensions[1].height = 30

c = ws5["A2"]
c.value = ("Written summary: 12-18 pages. British English, no em dashes. "
           "Written simultaneously for Jonas (strategic reader) and Julius (practitioner). "
           "Assembled in T5.1.1. Internal review by Glen + Tom in T5.2.1. Sent 24h before review call.")
c.font = Font(name="Calibri", size=10, italic=True)
c.fill = PatternFill("solid", fgColor=ALT_ROW)
c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
ws5.merge_cells("A2:D2")
ws5.row_dimensions[2].height = 36

# Headers
for col, txt in enumerate(["Reference", "Section Title", "Est. Length", "Content Description"], start=1):
    c = ws5.cell(row=3, column=col, value=txt)
    c.font = Font(name="Calibri", size=11, bold=True, color=HEADER_FG)
    c.fill = PatternFill("solid", fgColor=HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin_border()
ws5.row_dimensions[3].height = 20
ws5.freeze_panes = "A4"

for i, (ref, title, length, desc) in enumerate(DELIVERABLE_SECTIONS):
    row = 4 + i
    bg = WHITE if i % 2 == 0 else ALT_ROW
    is_conditional = "CONDITIONAL" in desc or "conditional" in desc.lower()
    if is_conditional:
        bg = AMBER_BG

    ws5.row_dimensions[row].height = 70

    for col, val in enumerate([ref, title, length, desc], start=1):
        c = ws5.cell(row=row, column=col, value=val)
        c.font = Font(name="Calibri", size=11, bold=(col in (1, 2)))
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(
            horizontal="center" if col in (1, 3) else "left",
            vertical="top", wrap_text=True, indent=(1 if col > 1 else 0)
        )
        c.border = thin_border()

# Conditional note footer
note_r = 4 + len(DELIVERABLE_SECTIONS) + 1
ws5.row_dimensions[note_r].height = 10
note_r2 = note_r + 1
c = ws5.cell(row=note_r2, column=1,
             value="NOTE: Amber-highlighted sections (6 and 7) are conditional — only included if F1-3 complete cleanly within budget. Feature 4 scope is reduced first if any earlier feature runs over.")
c.font = Font(name="Calibri", size=10, italic=True, color="555555")
c.fill = PatternFill("solid", fgColor=AMBER_BG)
c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
c.border = thin_border()
ws5.merge_cells(f"A{note_r2}:D{note_r2}")
ws5.row_dimensions[note_r2].height = 36


# ════════════════════════════════════════════════════════════════════════════
# SAVE
# ════════════════════════════════════════════════════════════════════════════

output_path = r"d:\OneDrive\Claude_code\NBIAI_TEAM\Clients\Goals\goals_pricing_engagement_plan_v3.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
print(f"File size: {os.path.getsize(output_path):,} bytes")
print("Sheets:", [ws.title for ws in wb.worksheets])
print(f"Total tasks in tracker: {len(TASKS)}")
print(f"Total risks in register: {len(RISKS)}")
print(f"Deliverable sections: {len(DELIVERABLE_SECTIONS)}")
